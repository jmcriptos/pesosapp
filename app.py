import os
import re
import calendar
import secrets
import hmac
import base64
import json
from dotenv import load_dotenv
load_dotenv()
from flask import Flask, render_template, request, redirect, send_file, jsonify, session, url_for, flash, abort
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, and_, or_, cast, String
from sqlalchemy.exc import OperationalError, IntegrityError
from sqlalchemy.orm import joinedload, selectinload, load_only
import io
from flask import make_response
import csv
import tempfile
import logging
import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta, date, timezone
from dateutil.relativedelta import relativedelta
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from flask_migrate import Migrate
import xlsxwriter
import socket
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import locale
import traceback
from decimal import Decimal, InvalidOperation
from time import perf_counter
# from models.extensions import db  # Comentado para evitar conflictos
import requests
import threading
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
from functools import wraps
from flask_login import LoginManager, login_user, logout_user, login_required, current_user, UserMixin
from datetime import datetime, date, timedelta, timezone
from urllib.parse import urlparse, urljoin
from markupsafe import Markup
from zoneinfo import ZoneInfo
from utils.label_utils import (
    draw_order_label, draw_expiration_label, get_logo_path,
    create_single_label_pdf, create_letter_page_pdf, get_centered_x,
    LABEL_WIDTH, LABEL_HEIGHT,
    create_a4_page_pdf, get_a4_label_positions, draw_order_label_a4
)
# Flask-WTF es obligatorio: la protección CSRF debe fallar de forma cerrada.
# Si la dependencia falta, la app no debe arrancar sin CSRF.
from flask_wtf import CSRFProtect
try:
    from flask_talisman import Talisman
except ImportError:
    Talisman = None

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)


def _env_flag(name, default=False):
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def _excel_safe(value):
    """Neutraliza inyección de fórmulas (CSV/Excel injection): prefija con apóstrofe
    los valores de texto que empiezan con =, +, -, @, tab o CR para que la hoja
    los trate como texto y no como fórmula viva."""
    if value is None:
        return value
    if isinstance(value, str) and value[:1] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + value
    return value


def _webhook_headers():
    """Header de autenticación para webhooks salientes a N8N. Si N8N_OUTBOUND_SECRET
    está configurado, N8N puede validar X-Webhook-Token; si no, no rompe el flujo."""
    secret = os.environ.get('N8N_OUTBOUND_SECRET', '').strip()
    return {'X-Webhook-Token': secret} if secret else {}


def _firma_png_valida(firma):
    """Valida que la firma sea un data URL PNG base64 razonable (no SVG ni payloads)."""
    prefijo = 'data:image/png;base64,'
    if not firma.startswith(prefijo) or len(firma) > 600000:
        return False
    try:
        base64.b64decode(firma[len(prefijo):], validate=True)
        return True
    except Exception:
        return False


IS_HEROKU = bool(os.environ.get("DYNO"))
SECURE_COOKIES = _env_flag("SESSION_COOKIE_SECURE", default=IS_HEROKU)
FORCE_HTTPS = _env_flag("FORCE_HTTPS", default=IS_HEROKU)
DASHBOARD_PERF_LOG = _env_flag("DASHBOARD_PERF_LOG", default=True)

# --- SECRET KEY (con fallback seguro) ---
_secret_key = os.environ.get("SECRET_KEY")
if not _secret_key:
    import warnings
    warnings.warn(
        "SECRET_KEY not set! Using random key. Sessions will not persist across restarts. "
        "Set SECRET_KEY with: heroku config:set SECRET_KEY=$(python -c \"import secrets; print(secrets.token_hex(32))\")",
        RuntimeWarning
    )
    _secret_key = secrets.token_hex(32)
app.config["SECRET_KEY"] = _secret_key

basedir = os.path.abspath(os.path.dirname(__file__))

uri = os.environ.get("DATABASE_URL", "sqlite:///local.db")
if uri.startswith("postgres://"):
    uri = uri.replace("postgres://", "postgresql://", 1)
app.config['SQLALCHEMY_DATABASE_URI'] = uri
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Inicializar SQLAlchemy directamente (sin models.extensions)
db = SQLAlchemy(app)

# Cookies / sesión
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
# Usa cookies Secure en despliegues reales detrás de HTTPS.
app.config['SESSION_COOKIE_SECURE'] = SECURE_COOKIES
app.config['REMEMBER_COOKIE_SECURE'] = SECURE_COOKIES
app.config['REMEMBER_COOKIE_HTTPONLY'] = True
app.config['REMEMBER_COOKIE_SAMESITE'] = 'Lax'
# Duración de sesión y de la cookie "recordarme" (en vez del default de 365 días)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
app.config['REMEMBER_COOKIE_DURATION'] = timedelta(days=7)
# Límite global de tamaño de request (anti-DoS de memoria); aplica antes de procesar el body
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB

try:
    DASHBOARD_TIMEZONE = ZoneInfo(os.environ.get('BUSINESS_TIMEZONE', 'America/Curacao'))
except Exception:
    DASHBOARD_TIMEZONE = timezone.utc

# Fuente de ventas para dashboard:
# - local: calcula ventas desde la base de datos de PesosApp
# - quickbooks: usa endpoint N8N/QuickBooks para ventas
# - auto: usa QuickBooks solo si hay endpoint configurado, si no local
QB_SALES_SOURCE = os.environ.get('QB_SALES_SOURCE', 'auto').strip().lower()
N8N_QB_SALES_WEBHOOK_URL = os.environ.get('N8N_QB_SALES_WEBHOOK_URL', '').strip()
try:
    N8N_QB_SALES_TIMEOUT = int(os.environ.get('N8N_QB_SALES_TIMEOUT', 20))
except (TypeError, ValueError):
    N8N_QB_SALES_TIMEOUT = 20
try:
    N8N_QB_BLOCKING_TIMEOUT_MS = int(os.environ.get('N8N_QB_BLOCKING_TIMEOUT_MS', 2000))
except (TypeError, ValueError):
    N8N_QB_BLOCKING_TIMEOUT_MS = 2000
try:
    N8N_QB_STALE_CACHE_TTL = int(os.environ.get('N8N_QB_STALE_CACHE_TTL', 900))
except (TypeError, ValueError):
    N8N_QB_STALE_CACHE_TTL = 900
try:
    N8N_QB_REFRESH_THROTTLE_SEC = int(os.environ.get('N8N_QB_REFRESH_THROTTLE_SEC', 30))
except (TypeError, ValueError):
    N8N_QB_REFRESH_THROTTLE_SEC = 30
try:
    N8N_QB_CACHE_TTL = int(os.environ.get('N8N_QB_CACHE_TTL', 60))
except (TypeError, ValueError):
    N8N_QB_CACHE_TTL = 60
try:
    N8N_QB_FAILURE_CACHE_TTL = int(os.environ.get('N8N_QB_FAILURE_CACHE_TTL', 30))
except (TypeError, ValueError):
    N8N_QB_FAILURE_CACHE_TTL = 30
try:
    DASHBOARD_USD_TO_XCG_FALLBACK_RATE = float(
        os.environ.get('DASHBOARD_USD_TO_XCG_FALLBACK_RATE', 1.78)
    )
except (TypeError, ValueError):
    DASHBOARD_USD_TO_XCG_FALLBACK_RATE = 1.78
if DASHBOARD_USD_TO_XCG_FALLBACK_RATE <= 0:
    DASHBOARD_USD_TO_XCG_FALLBACK_RATE = 1.78
_qb_sales_cache = {
    'key': None,
    'value': None,
    'expires_at': 0.0,
    'stale_expires_at': 0.0,
    'failure_expires_at': 0.0,
    'last_refresh_attempt': 0.0,
}

# SQLAlchemy ya inicializado arriba

migrate = Migrate(app, db)


def _ensure_haccp_columns():
    """Crea tablas/columnas nuevas de HACCP de forma idempotente, en SQLite
    (local) y Postgres (Heroku), sin depender de `flask db upgrade`.
    Seguro de ejecutar en cada arranque: solo añade lo que falta."""
    try:
        from sqlalchemy import inspect as _sa_inspect, text as _sa_text
        insp = _sa_inspect(db.engine)
        existing_tables = set(insp.get_table_names())

        # Tabla nueva: evento_auditoria
        if 'evento_auditoria' not in existing_tables:
            EventoAuditoria.__table__.create(bind=db.engine, checkfirst=True)

        # Columnas nuevas por tabla -> (nombre, DDL del tipo)
        wanted = {
            'camara': [('responsable_id', 'INTEGER'), ('ronda_am', 'VARCHAR(5)'), ('ronda_pm', 'VARCHAR(5)')],
            'area_limpieza': [('responsable_id', 'INTEGER'), ('sanitizante_id', 'INTEGER')],
            'registro_limpieza': [('firma_png', 'TEXT'), ('concentracion_ppm', 'INTEGER'),
                                  ('verificado_por', 'INTEGER'), ('metodo_verificacion', 'VARCHAR(20)')],
        }
        for tabla, cols in wanted.items():
            if tabla not in existing_tables:
                continue
            have = {c['name'] for c in insp.get_columns(tabla)}
            for col, ddl in cols:
                if col not in have:
                    with db.engine.begin() as conn:
                        conn.execute(_sa_text(f'ALTER TABLE {tabla} ADD COLUMN {col} {ddl}'))
                    app.logger.info(f'[HACCP] columna añadida: {tabla}.{col}')
    except Exception as e:
        app.logger.warning(f'[HACCP] no se pudieron asegurar columnas: {e}')


# Catálogo oficial del programa de limpieza (PG-HACCP-LIMP-01).
_CAT_LIMP_DETERGENTES = ['Big Punch', 'POOFF']
_CAT_LIMP_SANITIZANTE = 'Sani-T-10 Plus'
_CAT_LIMP_EQUIPOS = ['Tanque de salmueras', 'Inyectadora Inject Star', 'Embutidora Vemag',
                     'Molino Torrey', 'Rebanadora Icone 700', 'Mezclador MPR 400',
                     'Horno Ahumador', 'Carros para horno']
_CAT_LIMP_ESPACIOS = ['Sala de Producción', 'Sala de Mezclado', 'Sala de Cocción y Ahumado',
                      'Almacenes', 'Pisos y drenajes', 'Camión de reparto']


def _seed_catalogo_limpieza():
    """Crea (idempotente) productos y áreas oficiales del programa de limpieza.
    No borra ni desactiva nada. A los equipos creados aquí les asigna Sani-T-10 Plus
    como sanitizante (activa el gate de ppm). Seguro de correr en cada arranque."""
    try:
        from sqlalchemy import inspect as _sa_inspect
        insp = _sa_inspect(db.engine)
        tables = set(insp.get_table_names())
        if 'producto_limpieza' not in tables or 'area_limpieza' not in tables:
            return

        def _get_or_create_producto(nombre):
            p = (ProductoLimpieza.query
                 .filter(func.lower(ProductoLimpieza.nombre) == nombre.lower()).first())
            if p is None:
                p = ProductoLimpieza(nombre=nombre, dilucion='Según ficha técnica', activo=True)
                db.session.add(p)
                db.session.flush()
            return p

        for nombre in _CAT_LIMP_DETERGENTES:
            _get_or_create_producto(nombre)
        sani = _get_or_create_producto(_CAT_LIMP_SANITIZANTE)

        def _ensure_area(nombre, tipo, sanitizante_id=None):
            existe = (AreaLimpieza.query
                      .filter(func.lower(AreaLimpieza.nombre) == nombre.lower()).first())
            if existe is None:
                db.session.add(AreaLimpieza(nombre=nombre, tipo=tipo,
                                            sanitizante_id=sanitizante_id, activa=True))

        for nombre in _CAT_LIMP_EQUIPOS:
            _ensure_area(nombre, 'equipo', sanitizante_id=sani.id)
        for nombre in _CAT_LIMP_ESPACIOS:
            _ensure_area(nombre, 'espacio')
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.warning(f'[HACCP] no se pudo sembrar catálogo de limpieza: {e}')


N8N_HACCP_ALERT_WEBHOOK_URL = os.environ.get('N8N_HACCP_ALERT_WEBHOOK_URL', '').strip()


def _haccp_alerta(tipo, titulo, detalle, accion=None):
    """Notifica una incidencia HACCP (fuera de rango / no conforme).
    Registra auditoría siempre y, si hay webhook N8N configurado, lo dispara
    en background sin bloquear la respuesta."""
    _audit(tipo, f'ALERTA: {titulo}', detalle)
    if not N8N_HACCP_ALERT_WEBHOOK_URL:
        return
    payload = {
        'tipo': tipo, 'titulo': titulo, 'detalle': detalle, 'accion': accion,
        'usuario': (current_user.nombre_completo or current_user.username) if isinstance(current_user, Vendedor) else None,
        'timestamp': datetime.utcnow().isoformat() + 'Z',
    }

    def _post():
        try:
            requests.post(N8N_HACCP_ALERT_WEBHOOK_URL, json=payload, timeout=8, headers=_webhook_headers())
        except Exception as e:
            app.logger.warning(f'[haccp-alert] webhook falló: {e}')
    try:
        threading.Thread(target=_post, daemon=True).start()
    except Exception as e:
        app.logger.warning(f'[haccp-alert] no se pudo lanzar hilo: {e}')


def _audit(tipo, accion, detalle=None):
    """Registra un evento de auditoría. No interrumpe el flujo si falla."""
    try:
        actor = None
        vid = None
        if isinstance(current_user, Vendedor):
            actor = current_user.nombre_completo or current_user.username
            vid = current_user.id
        db.session.add(EventoAuditoria(vendedor_id=vid, actor=actor, tipo=tipo,
                                       accion=accion, detalle=(detalle or None)))
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.warning(f'[audit] no se pudo registrar evento: {e}')


# CSRF (Flask-WTF) - obligatorio
csrf = CSRFProtect(app)

# Configuración de seguridad con Talisman (HSTS, CSP, etc.)
# Solo activa Talisman en producción (cuando uses HTTPS real)
if Talisman and os.environ.get("FLASK_ENV") == "production":
    # script-src usa nonces (sin 'unsafe-inline'): cada <script> inline lleva
    # nonce="{{ csp_nonce() }}". style-src mantiene 'unsafe-inline' (bajo riesgo).
    # OJO: requiere que Cloudflare Rocket Loader esté DESACTIVADO; si se reactiva,
    # reescribe/re-inyecta los scripts sin el nonce y la app deja de ejecutar JS en
    # web (volver entonces a 'unsafe-inline' + nonce_in=[]).
    talisman_policy = {
        'default-src': ["'self'"],
        'script-src': [
            "'self'",
            'https://cdn.jsdelivr.net',
            'https://code.jquery.com',
            'https://cdnjs.cloudflare.com'
        ],
        'style-src': [
            "'self'",
            "'unsafe-inline'",  # TODO: Migrar a nonces para mayor seguridad
            'https://cdn.jsdelivr.net',
            'https://cdnjs.cloudflare.com'
        ],
        'img-src': ["'self'", 'data:', 'blob:'],
        'font-src': ["'self'", 'data:', 'https://cdnjs.cloudflare.com'],
        'connect-src': [
            "'self'",
            'https://cdn.jsdelivr.net',
            'https://cdnjs.cloudflare.com',
            'https://code.jquery.com'
        ],
        'style-src-attr': ["'unsafe-inline'"],  # Necesario para estilos inline en atributos
        'frame-ancestors': ["'none'"],  # Prevenir clickjacking
        'base-uri': ["'self'"],  # Prevenir inyección de base URI
        'form-action': ["'self'"],  # Restringir destinos de formularios
        'object-src': ["'none'"]  # Bloquear plugins (Flash, Java, etc.)
    }
    Talisman(
        app,
        content_security_policy=talisman_policy,
        content_security_policy_nonce_in=['script-src'],
        # Configuración HSTS - forzar HTTPS por 1 año
        strict_transport_security=True,
        strict_transport_security_max_age=31536000,  # 1 año
        strict_transport_security_include_subdomains=True,
        strict_transport_security_preload=True,
        # Otras protecciones
        force_https=True,
        session_cookie_secure=True,
        session_cookie_http_only=True,
        # X-Frame-Options (adicional a CSP frame-ancestors)
        frame_options='DENY'
    )
else:
    # Sin Talisman (dev/testing): los templates usan {{ csp_nonce() }}; proveer un
    # fallback no-op para que rendericen. setdefault evita pisar el de Talisman.
    app.jinja_env.globals.setdefault('csp_nonce', lambda: '')


login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.session_protection = 'strong'


def _client_ip():
    """IP real del cliente detrás de Cloudflare/Heroku para rate limiting."""
    cf = request.headers.get('CF-Connecting-IP')
    if cf:
        return cf.strip()
    xff = request.headers.get('X-Forwarded-For')
    if xff:
        return xff.split(',')[0].strip()
    return request.remote_addr or '127.0.0.1'


# Rate limiting (defensa contra fuerza bruta). Degrada con aviso si la librería falta.
# Se desactiva en testing para no interferir con los logins repetidos de la suite.
app.config['RATELIMIT_ENABLED'] = (os.environ.get('FLASK_ENV') != 'testing')
try:
    from flask_limiter import Limiter
    limiter = Limiter(
        key_func=_client_ip,
        app=app,
        storage_uri=os.environ.get('RATELIMIT_STORAGE_URI', 'memory://'),
        default_limits=[],
        enabled=app.config['RATELIMIT_ENABLED'],
    )
except ImportError:
    limiter = None
    app.logger.warning("flask_limiter no instalado: /login sin rate limiting")


def _login_rate_limit(view):
    """Aplica el límite de intentos a /login solo si limiter está disponible."""
    if limiter is None:
        return view
    return limiter.limit('10 per minute; 60 per hour', methods=['POST'])(view)


@app.errorhandler(429)
def ratelimit_handler(e):
    if request.endpoint == 'login' or (request.path or '').startswith('/login'):
        flash('Demasiados intentos. Espera un momento antes de volver a intentar.', 'danger')
        return render_template('login.html'), 429
    return jsonify({'error': 'Demasiadas solicitudes. Intenta de nuevo en un momento.'}), 429


@app.before_request
def redirect_insecure_requests():
    if not FORCE_HTTPS or app.debug or app.testing or request.is_secure:
        return

    # Evita perder la cookie de sesión/CSRF cuando el navegador entra por HTTP.
    secure_url = request.url.replace("http://", "https://", 1)
    return redirect(secure_url, code=308)


@app.after_request
def disable_cache_for_auth_and_session_responses(response):
    should_disable_cache = (
        request.endpoint in {"login", "logout"} or
        "Set-Cookie" in response.headers
    )
    if should_disable_cache:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, private, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

from utils.filters import kpi_tag
app.jinja_env.filters['kpi_tag'] = kpi_tag

_PERMISOS_DEFAULT = {
    'super_admin': {
        'productos': ['leer', 'crear', 'editar', 'eliminar'],
        'clientes': ['leer', 'crear', 'editar', 'eliminar'],
        'pedidos': ['leer', 'crear', 'editar', 'eliminar'],
        'vendedores': ['leer', 'crear', 'editar', 'eliminar'],
        'precios': ['leer', 'crear', 'editar', 'eliminar'],
        'reportes': ['leer', 'crear', 'editar', 'eliminar'],
        'importaciones': ['leer', 'crear', 'editar', 'eliminar'],
        'facturacion': ['leer', 'crear', 'editar', 'eliminar'],
        'registros': ['leer', 'crear', 'editar', 'eliminar'],
    },
    'supervisor': {
        'productos': ['leer'],
        'clientes': ['leer', 'editar'],
        'pedidos': ['leer', 'crear', 'editar'],
        'vendedores': ['leer'],
        'precios': ['leer'],
        'reportes': ['leer'],
        'importaciones': [],
        'facturacion': ['leer'],
        'registros': ['leer', 'crear', 'editar'],
    },
    'vendedor': {
        'productos': ['leer'],
        'clientes': ['leer', 'editar'],
        'pedidos': ['leer', 'crear', 'editar'],
        'vendedores': [],
        'precios': ['leer'],
        'reportes': [],
        'importaciones': [],
        'facturacion': [],
        'registros': ['leer', 'crear'],
    },
}


def _permiso_default(rol_nombre, recurso, accion):
    """Defaults de permisos (fallback cuando no hay filas en RolPermiso)."""
    return accion in _PERMISOS_DEFAULT.get(rol_nombre, {}).get(recurso, [])


PERMISOS_RECURSOS = ['productos', 'clientes', 'pedidos', 'precios', 'registros']

PERMISOS_DEFAULTS = {
    'vendedor':    {'productos': ['leer'], 'clientes': ['leer', 'editar'],
                    'pedidos': ['leer', 'crear', 'editar'], 'precios': ['leer'],
                    'registros': ['leer', 'crear']},
    'supervisor':  {'productos': ['leer'], 'clientes': ['leer', 'editar'],
                    'pedidos': ['leer', 'crear', 'editar'], 'precios': ['leer'],
                    'registros': ['leer', 'crear', 'editar']},
    'super_admin': {r: ['leer', 'crear', 'editar', 'eliminar'] for r in PERMISOS_RECURSOS},
}


def _sembrar_permisos():
    """Crea (idempotente, no destructivo) las filas Permiso por recurso y las
    filas RolPermiso por rol con los defaults. No sobreescribe filas existentes."""
    for rec in PERMISOS_RECURSOS:
        if not Permiso.query.filter_by(recurso=rec).first():
            db.session.add(Permiso(nombre=rec, recurso=rec, categoria='recurso',
                                   descripcion=f'Recurso {rec}'))
    db.session.flush()
    permisos = {p.recurso: p for p in Permiso.query.all()}
    for rol_nombre, recursos in PERMISOS_DEFAULTS.items():
        rol = Rol.query.filter_by(nombre=rol_nombre).first()
        if rol is None:
            continue
        for rec, acciones in recursos.items():
            p = permisos.get(rec)
            if p is None:
                continue
            existe = RolPermiso.query.filter_by(rol_id=rol.id, permiso_id=p.id).first()
            if existe is None:
                db.session.add(RolPermiso(
                    rol_id=rol.id, permiso_id=p.id,
                    puede_leer='leer' in acciones, puede_crear='crear' in acciones,
                    puede_editar='editar' in acciones, puede_eliminar='eliminar' in acciones))
    db.session.commit()


class Vendedor(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    
    # Información personal
    nombre_completo = db.Column(db.String(200), nullable=False)
    telefono = db.Column(db.String(20))
    fecha_ingreso = db.Column(db.Date, default=datetime.utcnow)
    
    # Relaciones organizacionales
    rol_id = db.Column(db.Integer, db.ForeignKey('rol.id'), nullable=False)
    territorio_id = db.Column(db.Integer, db.ForeignKey('territorio.id'))
    supervisor_id = db.Column(db.Integer, db.ForeignKey('vendedor.id'))
    
    # Estado y actividad
    activo = db.Column(db.Boolean, default=True)
    ultimo_login = db.Column(db.DateTime)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    debe_cambiar_password = db.Column(db.Boolean, nullable=False, default=False,
                                      server_default=db.false())

    # Relaciones
    supervisor = db.relationship('Vendedor', remote_side=[id], backref='subordinados')
    clientes_asignados = db.relationship('ClienteVendedor', back_populates='vendedor', 
                                       cascade="all, delete-orphan")
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    def tiene_permiso(self, permiso_nombre, tipo_acceso='leer'):
        """Verifica un permiso leyendo de RolPermiso; super_admin siempre pasa;
        si no hay filas sembradas, cae a los defaults (_permiso_default)."""
        if not self.activo:
            return False
        if self.rol and self.rol.nombre == 'super_admin':
            return True
        rp = (RolPermiso.query.join(Permiso)
              .filter(RolPermiso.rol_id == self.rol_id,
                      Permiso.recurso == permiso_nombre).first())
        if rp is None:
            return _permiso_default(self.rol.nombre if self.rol else '', permiso_nombre, tipo_acceso)
        return bool({'leer': rp.puede_leer, 'crear': rp.puede_crear,
                     'editar': rp.puede_editar, 'eliminar': rp.puede_eliminar}.get(tipo_acceso, False))
    
    def puede_editar_pedido(self, pedido):
        """Verifica si el vendedor puede editar un pedido específico"""
        if self.rol.nombre == 'super_admin':
            return True
        
        # Verificar si el pedido es de un cliente asignado al vendedor
        return self.puede_ver_cliente(pedido.cliente_id)
    
    def puede_crear_pedido_para_cliente(self, cliente_id):
        """Verifica si el vendedor puede crear pedidos para un cliente específico"""
        if self.rol.nombre == 'super_admin':
            return True
        
        return self.puede_ver_cliente(cliente_id)

    
    def puede_ver_cliente(self, cliente_id):
        """Verifica si el vendedor puede ver un cliente específico"""
        if self.rol.nombre == 'super_admin':
            return True
            
        asignacion = ClienteVendedor.query.filter_by(
            cliente_id=cliente_id, 
            vendedor_id=self.id, 
            activo=True
        ).first()
        
        return asignacion is not None
    
    def obtener_clientes_visibles(self):
        """Obtiene todos los clientes que el vendedor puede ver"""
        if self.rol.nombre == 'super_admin':
            return Cliente.query.all()
        
        # Clientes asignados directamente
        clientes_directos = db.session.query(Cliente).join(ClienteVendedor).filter(
            ClienteVendedor.vendedor_id == self.id,
            ClienteVendedor.activo == True
        ).all()
        
        return clientes_directos
    
    def to_dict(self):
        return {
            'id': self.id,
            'username': self.username,
            'email': self.email,
            'nombre_completo': self.nombre_completo,
            'telefono': self.telefono,
            'rol': self.rol.nombre if self.rol else None,
            'territorio': self.territorio.nombre if self.territorio else None,
            'activo': self.activo,
            'ultimo_login': self.ultimo_login.isoformat() if self.ultimo_login else None
        }
    
    def __repr__(self):
        return f'<Vendedor {self.username}>'
    
@login_manager.user_loader
def load_user(user_id: str):
    """Reconstruye el usuario desde la cookie de sesión."""
    if not user_id:
        app.logger.warning("[login] user_loader: user_id vacío o None")
        return None

    # Solo se admiten usuarios Vendedor (id numérico). El usuario legacy fue eliminado.
    try:
        vid = int(user_id)
    except (TypeError, ValueError):
        app.logger.warning(f"[login] user_loader: user_id no numérico: {user_id!r}")
        return None

    vendedor = db.session.get(Vendedor, vid)
    if vendedor and vendedor.activo:
        app.logger.debug(f"[login] user_loader: Vendedor id={vid} OK")
        return vendedor

    app.logger.info(f"[login] user_loader: Vendedor id={vid} no encontrado o inactivo")
    return None


@app.route('/_csrf_ping', methods=['POST'])
def csrf_ping():
    # Endpoint de verificación usado por tests/UI para validar ciclo CSRF.
    # Se valida explícitamente para soportar entornos donde WTF_CSRF_ENABLED
    # puede haberse desactivado por configuración de pruebas.
    token = (
        request.headers.get('X-CSRFToken')
        or request.headers.get('X-CSRF-Token')
        or request.form.get('csrf_token')
    )
    if not token:
        return jsonify({'error': 'The CSRF token is missing.'}), 400

    try:
        from flask_wtf.csrf import validate_csrf
        validate_csrf(token)
    except Exception:
        return jsonify({'error': 'The CSRF token is invalid.'}), 400

    return jsonify({'ok': True}), 200


def _is_safe_next(target: str) -> bool:
    """Evita open redirects; acepta solo URLs del mismo host."""
    if not target:
        return False
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, target))
    return (test_url.scheme in ("http", "https")) and (ref_url.netloc == test_url.netloc)

@app.route('/login', methods=['GET', 'POST'])
@_login_rate_limit
def login():
    # Si ya está autenticado, respeta 'next' y si no, a la página de inicio (pedidos)
    if current_user.is_authenticated:
        next_url = request.args.get('next')
        if not _is_safe_next(next_url):
            next_url = url_for('index')
        return redirect(next_url)

    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        password = request.form.get('password') or ''
        remember_me = bool(request.form.get('remember_me'))  # checkbox → bool
        # 'next' puede venir por query o por form (hidden)
        next_url = request.form.get('next') or request.args.get('next')

        # 1) Intentar login como Vendedor
        vendedor = None
        try:
            vendedor = Vendedor.query.filter_by(username=username, activo=True).first()
        except OperationalError as e:
            # Entornos sin migración de multivendor aún aplicada.
            db.session.rollback()
            app.logger.warning(f"[login] No se pudo consultar tabla vendedor: {e}")
        if vendedor and vendedor.check_password(password):
            vendedor.ultimo_login = datetime.utcnow()
            db.session.commit()
            login_user(vendedor, remember=remember_me)
            try:
                _audit('auth', 'Inició sesión')
            except Exception:
                pass
            flash(f"¡Bienvenido, {vendedor.nombre_completo}!", "success")
            if not _is_safe_next(next_url):
                next_url = url_for('index')
            return redirect(next_url)

        # Credenciales inválidas
        flash("Credenciales inválidas", "danger")

    # GET o POST fallido → mostrar login
    # Mantén 'next' en la query para que el form lo preserve
    return render_template('login.html')


@app.route('/logout', methods=['POST'])
@login_required
def logout():
    logout_user()
    flash("Sesión cerrada", "success")
    return redirect(url_for('login'))

@app.route('/mi-cuenta/cambiar-contrasena', methods=['GET', 'POST'])
@login_required
def cambiar_password():
    # Solo para usuarios Vendedor; el usuario legacy usa variable de entorno.
    if not isinstance(current_user, Vendedor):
        flash('Esta función no está disponible para el usuario del sistema.', 'warning')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        actual    = request.form.get('actual') or ''
        nueva     = request.form.get('nueva') or ''
        confirmar = request.form.get('confirmar') or ''

        if not current_user.check_password(actual):
            flash('La contraseña actual es incorrecta.', 'danger')
            return render_template('cambiar_password.html')
        if nueva != confirmar:
            flash('La nueva contraseña y su confirmación no coinciden.', 'danger')
            return render_template('cambiar_password.html')
        if len(nueva) < 8:
            flash('La nueva contraseña debe tener al menos 8 caracteres.', 'danger')
            return render_template('cambiar_password.html')
        if nueva == actual:
            flash('La nueva contraseña debe ser distinta de la actual.', 'danger')
            return render_template('cambiar_password.html')

        current_user.set_password(nueva)
        current_user.debe_cambiar_password = False
        db.session.commit()
        app.logger.info(
            f'Contraseña cambiada para usuario {current_user.username} (id={current_user.id})'
        )
        logout_user()
        flash('Contraseña actualizada. Inicia sesión nuevamente.', 'success')
        return redirect(url_for('login'))

    return render_template('cambiar_password.html')

@app.before_request
def require_login():
    allowed_endpoints = ['login', 'logout', 'static']
    # Endpoints que se autentican por su cuenta (CSRF ping, webhook con token propio)
    if request.endpoint in ('csrf_ping', 'webhook_actualizacion_precios'):
        return

    if request.endpoint and not any(request.endpoint.startswith(ep) for ep in allowed_endpoints):
        if not current_user.is_authenticated:
            return redirect(url_for('login', next=request.url))

@app.before_request
def forzar_cambio_password():
    """Si el usuario tiene una contraseña temporal pendiente, lo obliga a cambiarla."""
    if not current_user.is_authenticated or not isinstance(current_user, Vendedor):
        return
    if not getattr(current_user, 'debe_cambiar_password', False):
        return
    ep = request.endpoint or ''
    if ep in ('cambiar_password', 'logout', 'login', 'csrf_ping') or ep.startswith('static'):
        return
    flash('Debes establecer una nueva contraseña para continuar.', 'warning')
    return redirect(url_for('cambiar_password'))

def log_vendedor_action():
    """Registra las acciones de los vendedores para auditoría"""
    if request.method in ['POST', 'PUT', 'DELETE'] and current_user.is_authenticated:
        if isinstance(current_user, Vendedor):
            app.logger.info(f"[AUDIT] Vendedor: {current_user.username} - "
                           f"Acción: {request.method} - URL: {request.url}")


# En lugar de @app.route('/dashboard')
# Reemplazar la función dashboard_vendedor en app.py con esta versión optimizada
@app.context_processor
def inject_permissions():
    """Inyecta funciones de verificación de permisos en los templates"""
    def puede_crear(recurso):
        if not current_user.is_authenticated:
            return False
        if not isinstance(current_user, Vendedor):
            return True  # Usuario legacy tiene todos los permisos
        return current_user.tiene_permiso(recurso, 'crear')
    
    def puede_editar(recurso):
        if not current_user.is_authenticated:
            return False
        if not isinstance(current_user, Vendedor):
            return True
        return current_user.tiene_permiso(recurso, 'editar')
    
    def puede_eliminar(recurso):
        if not current_user.is_authenticated:
            return False
        if not isinstance(current_user, Vendedor):
            return True
        return current_user.tiene_permiso(recurso, 'eliminar')
    
    return dict(
        puede_crear=puede_crear,
        puede_editar=puede_editar,
        puede_eliminar=puede_eliminar
    )


def _to_dashboard_date(dt):
    """Convierte datetime UTC/naive a fecha local del dashboard."""
    if not dt:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(DASHBOARD_TIMEZONE).date()


def _fmt_local(dt, fmt='%Y-%m-%d %H:%M'):
    """Formatea un datetime UTC/naive en la hora LOCAL de negocio.
    Los registros se guardan en UTC; este helper evita mostrarlos en UTC."""
    if not dt:
        return ''
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(DASHBOARD_TIMEZONE).strftime(fmt)


@app.template_filter('hora_local')
def _jinja_hora_local(dt, fmt='%Y-%m-%d %H:%M'):
    """Filtro Jinja: {{ x.registrado_en | hora_local('%d/%m · %H:%M') }}."""
    return _fmt_local(dt, fmt)


def _calcular_venta_pedido(pedido):
    """
    Venta facturada del pedido:
    - Usa líneas de preparación cuando existen
    - Para productos se_pesa, usa cajas pesadas si existen
    - Usa línea original solo si no hay preparación para ese producto
    - No aplica tipo_cambio: subtotal ya refleja el monto facturado de la línea
    """
    prep_products = set()
    prep_total = 0.0
    orig_total = 0.0
    productos_con_cajas = set()

    for d in pedido.detalles:
        if not d.es_linea_pedido or not getattr(d, 'producto', None) or not d.producto.se_pesa:
            continue
        if not d.cajas_pesadas_count:
            continue

        qty = float(d.peso_real)
        if qty <= 0:
            continue

        productos_con_cajas.add(d.producto_id)
        prep_products.add(d.producto_id)
        prep_total += float(d.precio_unitario or 0) * qty

    for d in pedido.detalles:
        if not d.es_linea_pedido and d.subtotal:
            if d.producto and d.producto.se_pesa and d.producto_id in productos_con_cajas:
                continue
            prep_products.add(d.producto_id)
            prep_total += float(d.subtotal)

    for d in pedido.detalles:
        if d.es_linea_pedido and d.subtotal and d.producto_id not in prep_products:
            orig_total += float(d.subtotal)

    return prep_total + orig_total


def _pedido_facturado_en_periodo_local(pedido, fecha_inicio, fecha_fin=None):
    """True si el pedido está facturado y su fecha_facturacion local cae en el rango."""
    if pedido.estado != 'facturado' or not pedido.fecha_facturacion:
        return False

    fecha_local = _to_dashboard_date(pedido.fecha_facturacion)
    if not fecha_local:
        return False

    if fecha_fin is None:
        return fecha_local >= fecha_inicio
    return fecha_inicio <= fecha_local <= fecha_fin


def _coerce_float(value, default=0.0):
    if value in (None, ''):
        return default
    try:
        if isinstance(value, str):
            value = value.replace(',', '').strip()
            if value == '':
                return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _coerce_int(value, default=0):
    if value in (None, ''):
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _pick_row_value(row, keys):
    for key in keys:
        if key in row and row[key] not in (None, ''):
            return row[key]
    return None


def _normalizar_codigo_moneda(value):
    if value in (None, ''):
        return ''

    if isinstance(value, dict):
        value = value.get('value') or value.get('code') or value.get('name')
        if value in (None, ''):
            return ''

    code = str(value).strip().upper()
    if not code:
        return ''

    # Ejemplos esperados: "USD", "usd", "USD - US Dollar", "ANG", "XCG"
    code = code.replace('-', ' ').split()[0]
    aliases = {
        'ANG': 'XCG',
        'NLG': 'XCG',
    }
    return aliases.get(code, code)


def _monto_qb_a_xcg(row):
    """
    Normaliza montos de QuickBooks a XCG:
    1) Si el payload ya trae monto en moneda base, usarlo.
    2) Si viene en USD, convertir por exchange_rate.
    3) Si USD trae exchange_rate inválido (1/vacío), usar fallback fijo.
    """
    monto_base = _coerce_float(_pick_row_value(
        row,
        [
            'home_amount',
            'amount_home',
            'home_total',
            'home_total_amt',
            'home_total_amount',
            'amount_xcg',
            'amount_ang',
            'xcg_amount',
            'ang_amount',
            'total_xcg',
            'total_ang',
        ]
    ), None)
    if monto_base is not None:
        return monto_base

    monto = _coerce_float(_pick_row_value(row, ['amount', 'sales', 'total', 'venta', 'subtotal']), None)
    if monto is None:
        return None

    moneda = _normalizar_codigo_moneda(_pick_row_value(
        row,
        ['currency', 'currency_code', 'currency_ref', 'currencyref', 'moneda']
    ))
    if moneda == 'USD':
        tasa_original = _coerce_float(_pick_row_value(
            row,
            ['exchange_rate', 'exchangeRate', 'fx_rate', 'rate', 'tipo_cambio']
        ), None)
        tasa = tasa_original
        if tasa is None or tasa <= 1:
            tasa = DASHBOARD_USD_TO_XCG_FALLBACK_RATE
            app.logger.debug(
                f'[QBO→XCG] USD monto={monto:.2f} tasa_original={tasa_original} '
                f'→ usando fallback {DASHBOARD_USD_TO_XCG_FALLBACK_RATE}'
            )
        return monto * tasa

    if not moneda:
        app.logger.debug(
            f'[QBO→XCG] monto={monto:.2f} sin campo moneda — asumido XCG'
        )

    return monto


def _monto_qb_summary_a_xcg(summary, keys):
    """Normaliza montos de summary QBO a XCG usando metadata de moneda/tasa cuando exista."""
    monto = _coerce_float(_pick_row_value(summary, keys), None)
    if monto is None:
        return None

    moneda = _normalizar_codigo_moneda(_pick_row_value(
        summary,
        ['currency', 'currency_code', 'currency_ref', 'currencyref', 'moneda']
    ))
    if moneda == 'USD':
        tasa_original = _coerce_float(_pick_row_value(
            summary,
            ['exchange_rate', 'exchangeRate', 'fx_rate', 'rate', 'tipo_cambio']
        ), None)
        tasa = tasa_original
        if tasa is None or tasa <= 1:
            tasa = DASHBOARD_USD_TO_XCG_FALLBACK_RATE
            app.logger.debug(
                f'[QBO→XCG] summary USD monto={monto:.2f} tasa_original={tasa_original} '
                f'→ usando fallback {DASHBOARD_USD_TO_XCG_FALLBACK_RATE}'
            )
        return monto * tasa

    if not moneda:
        app.logger.debug(
            f'[QBO→XCG] summary monto={monto:.2f} sin campo moneda — asumido XCG'
        )

    return monto


def _parse_dashboard_date_value(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return _to_dashboard_date(value)
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    # Strings de solo-fecha (sin componente de hora): interpretar como fecha local
    # del dashboard. Evita que `datetime.fromisoformat("2026-04-01")` se trate como
    # UTC y luego se convierta a America/Curacao (UTC-4), corriendo la fecha al día
    # anterior y excluyendo las ventas del día 1 del mes.
    if 'T' not in text and ':' not in text:
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%Y/%m/%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(text[:10], fmt).date()
            except ValueError:
                continue

    # ISO datetime/date (con o sin zona horaria)
    try:
        dt = datetime.fromisoformat(text.replace('Z', '+00:00'))
        return _to_dashboard_date(dt)
    except ValueError:
        pass

    # Fechas comunes exportadas por QuickBooks
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%Y/%m/%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue

    return None


def _inicio_semana_local(fecha):
    return fecha - timedelta(days=fecha.weekday())


def _quickbooks_sales_enabled():
    if QB_SALES_SOURCE == 'local':
        return False
    if QB_SALES_SOURCE == 'quickbooks':
        return bool(N8N_QB_SALES_WEBHOOK_URL)
    return bool(N8N_QB_SALES_WEBHOOK_URL)


def _build_rankings_periodos_from_rows(client_rows, product_rows, hoy, period_starts):
    """Construye rankings de clientes/productos por periodo desde filas normalizadas."""
    rankings = {}

    for period_key, start_date in period_starts.items():
        clientes_ventas = {}
        productos_ventas = {}

        for row in client_rows or []:
            if not isinstance(row, dict):
                continue
            fecha = row.get('date')
            if not isinstance(fecha, date):
                continue
            if fecha < start_date or fecha > hoy:
                continue

            monto = _coerce_float(row.get('amount'), 0.0)
            if monto <= 0:
                continue
            invoice_key = str(row.get('invoice_key') or f'{fecha.isoformat()}-{monto}')
            customer_name = str(row.get('customer') or 'Sin cliente')

            cli = clientes_ventas.setdefault(
                customer_name,
                {'total': 0.0, 'ultimo_pedido': None, '_invoices': set()}
            )
            cli['total'] += monto
            cli['_invoices'].add(invoice_key)
            if not cli['ultimo_pedido'] or fecha > cli['ultimo_pedido']:
                cli['ultimo_pedido'] = fecha

        for row in product_rows or []:
            if not isinstance(row, dict):
                continue
            fecha = row.get('date')
            if not isinstance(fecha, date):
                continue
            if fecha < start_date or fecha > hoy:
                continue

            monto = _coerce_float(row.get('amount'), 0.0)
            if monto <= 0:
                continue
            nombre = str(row.get('product') or '').strip()
            if not nombre:
                continue

            invoice_key = str(row.get('invoice_key') or f'{fecha.isoformat()}-{monto}')
            qty = _coerce_float(row.get('quantity'), 0.0)
            peso = _coerce_float(row.get('weight'), 0.0)

            prod = productos_ventas.setdefault(
                nombre,
                {'ingresos': 0.0, 'cajas': 0.0, 'peso': 0.0, '_invoices': set()}
            )
            prod['ingresos'] += monto
            if qty > 0:
                prod['cajas'] += qty
            if peso > 0:
                prod['peso'] += peso
            prod['_invoices'].add(invoice_key)

        top_clientes = []
        for nombre, datos in sorted(
            clientes_ventas.items(),
            key=lambda x: x[1].get('total', 0),
            reverse=True
        )[:5]:
            top_clientes.append((
                nombre,
                {
                    'pedidos': len(datos.get('_invoices', set())),
                    'total': round(_coerce_float(datos.get('total'), 0.0), 2),
                    'ultimo_pedido': datos.get('ultimo_pedido')
                }
            ))

        top_productos = []
        for nombre, datos in sorted(
            productos_ventas.items(),
            key=lambda x: x[1].get('ingresos', 0),
            reverse=True
        )[:5]:
            ingresos = round(_coerce_float(datos.get('ingresos'), 0.0), 2)
            cajas = round(_coerce_float(datos.get('cajas'), 0.0), 2)
            peso = round(_coerce_float(datos.get('peso'), 0.0), 2)
            top_productos.append({
                'nombre': nombre,
                'total_vendido': ingresos,
                'cajas': cajas,
                'peso': peso,
                'ingresos': ingresos,
                'pedidos': len(datos.get('_invoices', set()))
            })

        rankings[period_key] = {
            'top_clientes': top_clientes,
            'top_productos': top_productos,
            'max_ventas': max((p['ingresos'] for p in top_productos), default=0) or 1,
            'max_total_clientes': max(
                (c[1].get('total', 0) for c in top_clientes),
                default=0
            ) or 1,
        }

    return rankings


def _serialize_rankings_periodos(rankings_periodos):
    """Convierte rankings por periodo a un formato JSON-safe para el template."""
    serializado = {}
    for period_key, payload in (rankings_periodos or {}).items():
        if not isinstance(payload, dict):
            continue

        productos = []
        for p in payload.get('top_productos', []) or []:
            if not isinstance(p, dict):
                continue
            productos.append({
                'nombre': str(p.get('nombre') or ''),
                'ingresos': round(_coerce_float(p.get('ingresos'), 0.0), 2),
                'cajas': round(_coerce_float(p.get('cajas'), 0.0), 2),
                'peso': round(_coerce_float(p.get('peso'), 0.0), 2),
                'pedidos': _coerce_int(p.get('pedidos'), 0),
            })

        clientes = []
        for item in payload.get('top_clientes', []) or []:
            if isinstance(item, (list, tuple)) and len(item) == 2:
                nombre, datos = item
            elif isinstance(item, dict):
                nombre = item.get('nombre') or 'Sin cliente'
                datos = item
            else:
                continue

            datos = datos or {}
            ultimo = _parse_dashboard_date_value(datos.get('ultimo_pedido'))
            clientes.append({
                'nombre': str(nombre or 'Sin cliente'),
                'total': round(_coerce_float(datos.get('total'), 0.0), 2),
                'pedidos': _coerce_int(datos.get('pedidos'), 0),
                'ultimo_pedido': ultimo.strftime('%d/%m') if ultimo else 'N/A',
            })

        serializado[period_key] = {
            'top_productos': productos,
            'top_clientes': clientes,
            'max_ventas': round(_coerce_float(payload.get('max_ventas'), 1.0), 2) or 1,
            'max_total_clientes': round(_coerce_float(payload.get('max_total_clientes'), 1.0), 2) or 1,
        }

    return serializado


def _normalizar_metricas_ventas_quickbooks(
    raw_data,
    hoy,
    inicio_mes,
    inicio_semana,
    inicio_mes_anterior,
    fin_mes_anterior,
    inicio_tendencia,
    inicio_ultimos_7_dias,
):
    ventas_mes = 0.0
    ventas_semana = 0.0
    ventas_mes_anterior = 0.0

    ventas_diarias_idx = {}
    ventas_semanales_idx = {}
    clientes_ventas = {}
    productos_ventas = {}
    ranking_client_rows = []
    ranking_product_rows = []

    filas = raw_data.get('transactions') or raw_data.get('invoices') or raw_data.get('rows') or []
    if isinstance(filas, dict):
        filas = filas.get('items', [])
    if not isinstance(filas, list):
        filas = []

    filas_procesadas = 0
    for row in filas:
        if not isinstance(row, dict):
            continue

        fecha = _parse_dashboard_date_value(_pick_row_value(
            row,
            ['date', 'invoice_date', 'transaction_date', 'fecha', 'fecha_facturacion']
        ))
        if not fecha:
            continue
        if fecha < inicio_tendencia or fecha > hoy:
            continue

        monto = _monto_qb_a_xcg(row)
        if monto is None:
            continue
        filas_procesadas += 1

        invoice_key = str(_pick_row_value(
            row,
            ['invoice_id', 'invoice_number', 'invoice_no', 'num', 'numero_factura', 'id']
        ) or f'{fecha.isoformat()}-{monto}')

        cliente_nombre = str(_pick_row_value(
            row,
            ['customer', 'customer_name', 'cliente', 'cliente_nombre']
        ) or 'Sin cliente')

        producto_nombre = _pick_row_value(
            row,
            ['product', 'product_name', 'producto', 'producto_nombre', 'item_name']
        )
        qty = _coerce_float(_pick_row_value(row, ['quantity', 'qty', 'cajas', 'cantidad']), 0.0)
        peso = _coerce_float(_pick_row_value(row, ['weight', 'peso']), 0.0)

        if fecha >= inicio_mes:
            ventas_mes += monto
        if fecha >= inicio_semana:
            ventas_semana += monto
        if inicio_mes_anterior <= fecha <= fin_mes_anterior:
            ventas_mes_anterior += monto

        if fecha >= inicio_ultimos_7_dias:
            bucket_d = ventas_diarias_idx.setdefault(fecha, {'ventas': 0.0, '_invoices': set()})
            bucket_d['ventas'] += monto
            bucket_d['_invoices'].add(invoice_key)

        semana_inicio = _inicio_semana_local(fecha)
        if inicio_tendencia <= semana_inicio <= inicio_semana:
            bucket_w = ventas_semanales_idx.setdefault(semana_inicio, {'ventas': 0.0, '_invoices': set()})
            bucket_w['ventas'] += monto
            bucket_w['_invoices'].add(invoice_key)

        cli = clientes_ventas.setdefault(
            cliente_nombre,
            {'total': 0.0, 'ultimo_pedido': None, '_invoices': set()}
        )
        cli['total'] += monto
        cli['_invoices'].add(invoice_key)
        if not cli['ultimo_pedido'] or fecha > cli['ultimo_pedido']:
            cli['ultimo_pedido'] = fecha

        if producto_nombre:
            prod = productos_ventas.setdefault(
                str(producto_nombre),
                {'ingresos': 0.0, 'cajas': 0.0, 'peso': 0.0, '_invoices': set()}
            )
            prod['ingresos'] += monto
            if qty > 0:
                prod['cajas'] += qty
            if peso > 0:
                prod['peso'] += peso
            prod['_invoices'].add(invoice_key)

        ranking_client_rows.append({
            'date': fecha,
            'invoice_key': invoice_key,
            'customer': cliente_nombre,
            'amount': monto,
        })
        if producto_nombre:
            ranking_product_rows.append({
                'date': fecha,
                'invoice_key': invoice_key,
                'product': str(producto_nombre),
                'amount': monto,
                'quantity': qty,
                'weight': peso,
            })

    # Fallback cuando N8N devuelve agregados ya resumidos
    if filas_procesadas == 0:
        for row in (raw_data.get('daily') or raw_data.get('sales_by_day') or []):
            if not isinstance(row, dict):
                continue
            fecha = _parse_dashboard_date_value(_pick_row_value(row, ['date', 'day', 'fecha']))
            if not fecha:
                continue
            ventas = _monto_qb_a_xcg(row)
            if ventas is None:
                ventas = 0.0
            pedidos = _coerce_int(_pick_row_value(row, ['invoices', 'orders', 'pedidos']), 0)
            ventas_diarias_idx[fecha] = {'ventas': ventas, '_invoices': set(range(pedidos))}
            if fecha >= inicio_mes:
                ventas_mes += ventas
            if fecha >= inicio_semana:
                ventas_semana += ventas
            if inicio_mes_anterior <= fecha <= fin_mes_anterior:
                ventas_mes_anterior += ventas

        for row in (raw_data.get('weekly') or raw_data.get('sales_by_week') or []):
            if not isinstance(row, dict):
                continue
            semana_inicio = _parse_dashboard_date_value(_pick_row_value(
                row,
                ['week_start', 'start_date', 'week', 'fecha_inicio']
            ))
            if not semana_inicio:
                continue
            ventas = _monto_qb_a_xcg(row)
            if ventas is None:
                ventas = 0.0
            pedidos = _coerce_int(_pick_row_value(row, ['invoices', 'orders', 'pedidos']), 0)
            ventas_semanales_idx[semana_inicio] = {'ventas': ventas, '_invoices': set(range(pedidos))}

        for row in (raw_data.get('customers') or raw_data.get('sales_by_customer') or []):
            if not isinstance(row, dict):
                continue
            nombre = str(_pick_row_value(row, ['name', 'customer', 'cliente']) or 'Sin cliente')
            clientes_ventas[nombre] = {
                'total': _monto_qb_a_xcg(row) or 0.0,
                'ultimo_pedido': _parse_dashboard_date_value(_pick_row_value(
                    row,
                    ['last_date', 'last_invoice_date', 'ultimo_pedido']
                )),
                '_invoices': set(range(_coerce_int(_pick_row_value(row, ['invoices', 'orders', 'pedidos']), 0)))
            }

        for row in (raw_data.get('products') or raw_data.get('sales_by_product') or []):
            if not isinstance(row, dict):
                continue
            nombre = str(_pick_row_value(row, ['name', 'product', 'producto']) or '').strip()
            if not nombre:
                continue
            productos_ventas[nombre] = {
                'ingresos': _monto_qb_a_xcg(row) or 0.0,
                'cajas': _coerce_float(_pick_row_value(row, ['quantity', 'qty', 'cajas', 'cantidad']), 0.0),
                'peso': _coerce_float(_pick_row_value(row, ['weight', 'peso']), 0.0),
                '_invoices': set(range(_coerce_int(_pick_row_value(row, ['invoices', 'orders', 'pedidos']), 0)))
            }

    summary = raw_data.get('summary') if isinstance(raw_data.get('summary'), dict) else {}
    # Solo aplicar summary cuando no hay datos procesables en filas/agregados.
    # Evita sobreescribir montos ya normalizados (ej. filas USD convertidas a XCG).
    if (
        filas_procesadas == 0
        and not ventas_diarias_idx
        and not ventas_semanales_idx
        and not clientes_ventas
        and not productos_ventas
    ):
        resumen_mes = _monto_qb_summary_a_xcg(summary, ['ventas_mes', 'sales_month', 'month_sales'])
        if resumen_mes is not None:
            ventas_mes = resumen_mes

        resumen_semana = _monto_qb_summary_a_xcg(summary, ['ventas_semana', 'sales_week', 'week_sales'])
        if resumen_semana is not None:
            ventas_semana = resumen_semana

        resumen_mes_anterior = _monto_qb_summary_a_xcg(
            summary,
            ['ventas_mes_anterior', 'sales_previous_month', 'previous_month_sales']
        )
        if resumen_mes_anterior is not None:
            ventas_mes_anterior = resumen_mes_anterior

    # Normalizar buckets para el dashboard (pedidos = facturas únicas)
    for bucket in ventas_diarias_idx.values():
        bucket['pedidos'] = len(bucket.pop('_invoices', set()))

    for bucket in ventas_semanales_idx.values():
        bucket['pedidos'] = len(bucket.pop('_invoices', set()))

    # Fallback de ranking general usando agregados globales
    top_clientes_fallback = []
    for nombre, datos in sorted(
        clientes_ventas.items(),
        key=lambda x: x[1].get('total', 0),
        reverse=True
    )[:5]:
        top_clientes_fallback.append((
            nombre,
            {
                'pedidos': len(datos.get('_invoices', set())),
                'total': round(_coerce_float(datos.get('total'), 0.0), 2),
                'ultimo_pedido': datos.get('ultimo_pedido')
            }
        ))

    top_productos_fallback = []
    for nombre, datos in sorted(
        productos_ventas.items(),
        key=lambda x: x[1].get('ingresos', 0),
        reverse=True
    )[:5]:
        ingresos = round(_coerce_float(datos.get('ingresos'), 0.0), 2)
        cajas = round(_coerce_float(datos.get('cajas'), 0.0), 2)
        peso = round(_coerce_float(datos.get('peso'), 0.0), 2)
        top_productos_fallback.append({
            'nombre': nombre,
            'total_vendido': ingresos,
            'cajas': cajas,
            'peso': peso,
            'ingresos': ingresos,
            'pedidos': len(datos.get('_invoices', set()))
        })

    period_starts = {
        'month': inicio_mes,
        '6m': inicio_tendencia,
        '3m': inicio_semana - timedelta(weeks=12),
        '4w': inicio_semana - timedelta(weeks=3),
    }

    if ranking_client_rows or ranking_product_rows:
        rankings_periodos = _build_rankings_periodos_from_rows(
            ranking_client_rows,
            ranking_product_rows,
            hoy=hoy,
            period_starts=period_starts,
        )
    else:
        max_ventas_fallback = max((p['ingresos'] for p in top_productos_fallback), default=0) or 1
        max_clientes_fallback = max(
            (c[1].get('total', 0) for c in top_clientes_fallback),
            default=0
        ) or 1
        rankings_periodos = {
            key: {
                'top_clientes': top_clientes_fallback,
                'top_productos': top_productos_fallback,
                'max_ventas': max_ventas_fallback,
                'max_total_clientes': max_clientes_fallback,
            }
            for key in period_starts.keys()
        }

    month_rankings = rankings_periodos.get('month', {})
    top_clientes = month_rankings.get('top_clientes', top_clientes_fallback)
    top_productos = month_rankings.get('top_productos', top_productos_fallback)
    max_ventas = month_rankings.get('max_ventas', 1)

    fuente = (
        'transacciones' if filas_procesadas > 0
        else 'agregados' if ventas_diarias_idx or ventas_semanales_idx
        else 'summary' if ventas_mes > 0
        else 'vacío'
    )
    app.logger.info(
        f'[QBO normalizado] fuente={fuente} filas={filas_procesadas} '
        f'ventas_mes={ventas_mes:.2f} ventas_semana={ventas_semana:.2f} '
        f'ventas_mes_anterior={ventas_mes_anterior:.2f}'
    )

    return {
        'ventas_mes': round(ventas_mes, 2),
        'ventas_semana': round(ventas_semana, 2),
        'ventas_mes_anterior': round(ventas_mes_anterior, 2),
        'ventas_diarias_idx': ventas_diarias_idx,
        'ventas_semanales_idx': ventas_semanales_idx,
        'top_clientes': top_clientes,
        'top_productos': top_productos,
        'max_ventas': max_ventas if max_ventas > 0 else 1,
        'rankings_periodos': rankings_periodos,
    }


def _obtener_metricas_ventas_quickbooks(
    hoy,
    inicio_mes,
    inicio_semana,
    inicio_mes_anterior,
    fin_mes_anterior,
    inicio_tendencia,
    inicio_ultimos_7_dias,
):
    global _qb_sales_cache

    if not _quickbooks_sales_enabled():
        return None

    if not N8N_QB_SALES_WEBHOOK_URL:
        if QB_SALES_SOURCE == 'quickbooks':
            app.logger.warning('QB_SALES_SOURCE=quickbooks pero N8N_QB_SALES_WEBHOOK_URL no está configurada')
        return None

    payload = {
        'from_date': inicio_tendencia.isoformat(),
        'to_date': hoy.isoformat(),
        'timezone': str(DASHBOARD_TIMEZONE),
        'group_by': ['day', 'week', 'customer', 'product'],
        'include_summary': True
    }
    cache_key = (
        N8N_QB_SALES_WEBHOOK_URL,
        payload['from_date'],
        payload['to_date'],
        payload['timezone'],
    )
    now_ts = datetime.now(timezone.utc).timestamp()
    cache_hit_for_key = _qb_sales_cache.get('key') == cache_key
    cached_value = _qb_sales_cache.get('value')
    has_stale_value = (
        cache_hit_for_key
        and cached_value is not None
        and _qb_sales_cache.get('stale_expires_at', 0.0) > now_ts
    )

    if (
        N8N_QB_CACHE_TTL > 0
        and cache_hit_for_key
        and _qb_sales_cache.get('expires_at', 0.0) > now_ts
    ):
        return _qb_sales_cache.get('value')
    if (
        N8N_QB_FAILURE_CACHE_TTL > 0
        and cache_hit_for_key
        and _qb_sales_cache.get('failure_expires_at', 0.0) > now_ts
    ):
        return cached_value if has_stale_value else None
    if (
        has_stale_value
        and N8N_QB_REFRESH_THROTTLE_SEC > 0
        and (_qb_sales_cache.get('last_refresh_attempt', 0.0) + N8N_QB_REFRESH_THROTTLE_SEC) > now_ts
    ):
        return cached_value

    blocking_timeout = N8N_QB_SALES_TIMEOUT
    if N8N_QB_BLOCKING_TIMEOUT_MS > 0:
        blocking_timeout = min(
            float(N8N_QB_SALES_TIMEOUT),
            float(N8N_QB_BLOCKING_TIMEOUT_MS) / 1000.0
        )
    if blocking_timeout <= 0:
        blocking_timeout = float(N8N_QB_SALES_TIMEOUT)

    _qb_sales_cache['last_refresh_attempt'] = now_ts

    try:
        qb_fetch_start = perf_counter()
        resp = requests.post(
            N8N_QB_SALES_WEBHOOK_URL,
            json=payload,
            timeout=blocking_timeout,
            headers=_webhook_headers()
        )
        resp.raise_for_status()
        data = resp.json()
        if not isinstance(data, dict):
            app.logger.warning('Respuesta QuickBooks inválida: se esperaba JSON objeto')
            return cached_value if has_stale_value else None
        normalized = _normalizar_metricas_ventas_quickbooks(
            data,
            hoy=hoy,
            inicio_mes=inicio_mes,
            inicio_semana=inicio_semana,
            inicio_mes_anterior=inicio_mes_anterior,
            fin_mes_anterior=fin_mes_anterior,
            inicio_tendencia=inicio_tendencia,
            inicio_ultimos_7_dias=inicio_ultimos_7_dias,
        )
        cache_now = datetime.now(timezone.utc).timestamp()
        _qb_sales_cache = {
            'key': cache_key,
            'value': normalized,
            'expires_at': cache_now + N8N_QB_CACHE_TTL if N8N_QB_CACHE_TTL > 0 else 0.0,
            'stale_expires_at': cache_now + N8N_QB_STALE_CACHE_TTL if N8N_QB_STALE_CACHE_TTL > 0 else 0.0,
            'failure_expires_at': 0.0,
            'last_refresh_attempt': cache_now,
        }
        qb_elapsed_ms = (perf_counter() - qb_fetch_start) * 1000
        if qb_elapsed_ms >= 1000:
            app.logger.info(
                f'QuickBooks ventas respondió en {qb_elapsed_ms:.0f}ms '
                f'(timeout efectivo {blocking_timeout:.2f}s)'
            )
        return normalized
    except requests.Timeout:
        app.logger.warning(
            f'Timeout obteniendo ventas de QuickBooks ({blocking_timeout:.2f}s). '
            f'Fallback a {"cache stale" if has_stale_value else "datos locales"}.'
        )
    except requests.RequestException as e:
        app.logger.warning(
            f'Error consultando ventas QuickBooks: {e}. '
            f'Fallback a {"cache stale" if has_stale_value else "datos locales"}.'
        )
    except ValueError:
        app.logger.warning(
            f'Error parseando JSON de ventas QuickBooks. '
            f'Fallback a {"cache stale" if has_stale_value else "datos locales"}.'
        )
    except Exception as e:
        app.logger.warning(
            f'Error inesperado ventas QuickBooks: {e}. '
            f'Fallback a {"cache stale" if has_stale_value else "datos locales"}.'
        )
    if N8N_QB_FAILURE_CACHE_TTL > 0:
        _qb_sales_cache = {
            'key': cache_key,
            'value': cached_value if has_stale_value else None,
            'expires_at': 0.0,
            'stale_expires_at': _qb_sales_cache.get('stale_expires_at', 0.0) if has_stale_value else 0.0,
            'failure_expires_at': datetime.now(timezone.utc).timestamp() + N8N_QB_FAILURE_CACHE_TTL,
            'last_refresh_attempt': now_ts,
        }
    return cached_value if has_stale_value else None


@app.route('/dashboard_vendedor')
@login_required
def dashboard_vendedor():
    """Dashboard optimizado con funcionalidades específicas para administradores y vendedores"""
    
    # Verificar si es vendedor del nuevo sistema
    if not isinstance(current_user, Vendedor):
        # Usuario del sistema anterior - redirigir a index
        return redirect(url_for('index'))
    
    try:
        # Fechas para análisis
        hoy = datetime.now(DASHBOARD_TIMEZONE).date()
        inicio_mes = hoy.replace(day=1)
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        hace_30_dias = hoy - timedelta(days=30)
        
        # Variables base
        context = {
            'vendedor': current_user,
            'fecha_actual': hoy,
            'fecha_sistema': 'Enero 2025',
            # Valores por defecto para evitar Undefined en templates
            'tendencia_semanal': [],
            'estados_pedidos': [],
            'kpi_weekly': []
        }
        
        # ===== MÉTRICAS ESPECÍFICAS PARA ADMINISTRADOR =====
        if current_user.rol.nombre == 'super_admin':
            
            # 1. MÉTRICAS GENERALES DEL SISTEMA
            total_vendedores = Vendedor.query.filter_by(activo=True).count()
            total_clientes = Cliente.query.count()
            total_productos = Producto.query.count()

            # 2. MÉTRICAS DE VENTAS GLOBALES (QuickBooks como fuente de verdad cuando está disponible)
            inicio_mes_anterior = (inicio_mes - timedelta(days=1)).replace(day=1)
            fin_mes_anterior = inicio_mes - timedelta(days=1)
            metricas_ventas_qb_admin = _obtener_metricas_ventas_quickbooks(
                hoy=hoy,
                inicio_mes=inicio_mes,
                inicio_semana=inicio_semana,
                inicio_mes_anterior=inicio_mes_anterior,
                fin_mes_anterior=fin_mes_anterior,
                inicio_tendencia=inicio_semana - timedelta(weeks=8),
                inicio_ultimos_7_dias=hoy - timedelta(days=6),
            )

            if metricas_ventas_qb_admin:
                ventas_totales = metricas_ventas_qb_admin['ventas_mes']
                ventas_hoy = metricas_ventas_qb_admin['ventas_diarias_idx'].get(
                    hoy,
                    {'ventas': 0.0, 'pedidos': 0}
                )['ventas']
                ventas_mes_anterior = metricas_ventas_qb_admin['ventas_mes_anterior']
            else:
                pedidos_facturados_data = Pedido.query.filter(
                    Pedido.estado == 'facturado',
                    Pedido.fecha_facturacion.isnot(None)
                ).all()

                ventas_totales = sum(
                    _calcular_venta_pedido(p)
                    for p in pedidos_facturados_data
                    if _pedido_facturado_en_periodo_local(p, inicio_mes)
                )

                ventas_hoy = sum(
                    _calcular_venta_pedido(p)
                    for p in pedidos_facturados_data
                    if _pedido_facturado_en_periodo_local(p, hoy, hoy)
                )

                ventas_mes_anterior = sum(
                    _calcular_venta_pedido(p)
                    for p in pedidos_facturados_data
                    if _pedido_facturado_en_periodo_local(p, inicio_mes_anterior, fin_mes_anterior)
                )
            
            # Conteo de pedidos optimizado
            pedidos_mes_count = Pedido.query.filter(Pedido.fecha_pedido >= inicio_mes).count()
            pedidos_hoy_count = Pedido.query.filter(
                Pedido.fecha_pedido >= hoy,
                Pedido.fecha_pedido < hoy + timedelta(days=1)
            ).count()
            
            # 3. MÉTRICAS DE EFICIENCIA
            pedidos_pendientes = Pedido.query.filter_by(estado='pendiente').count()
            pedidos_facturados = Pedido.query.filter_by(estado='facturado').count()
            pedidos_totales = Pedido.query.count()
            
            # Calcular eficiencia del sistema (% de pedidos completados)
            if pedidos_totales > 0:
                eficiencia_sistema = (pedidos_facturados / pedidos_totales) * 100
            else:
                eficiencia_sistema = 95.8  # Valor por defecto
            
            # 4. ANÁLISIS DE TENDENCIAS
            # Pedidos por vendedor (top 5)
            vendedores_performance = db.session.query(
                Vendedor.nombre_completo,
                func.count(Pedido.id).label('total_pedidos'),
                func.coalesce(func.sum(DetallePedido.subtotal), 0).label('ventas_total')
            ).outerjoin(
                ClienteVendedor, Vendedor.id == ClienteVendedor.vendedor_id
            ).outerjoin(
                Pedido, ClienteVendedor.cliente_id == Pedido.cliente_id
            ).outerjoin(
                DetallePedido, Pedido.id == DetallePedido.pedido_id
            ).filter(
                Pedido.fecha_pedido >= hace_30_dias,
                db.or_(DetallePedido.es_linea_pedido == True, DetallePedido.id.is_(None))
            ).group_by(
                Vendedor.id, Vendedor.nombre_completo
            ).order_by(
                func.coalesce(func.sum(DetallePedido.subtotal), 0).desc()
            ).limit(5).all()
            
            # 5. ALERTAS DEL SISTEMA
            alertas_sistema = []
            
            # Alerta por pedidos pendientes
            if pedidos_pendientes > 10:
                alertas_sistema.append({
                    'tipo': 'warning',
                    'titulo': 'Alto volumen de pedidos pendientes',
                    'mensaje': f'{pedidos_pendientes} pedidos requieren atención',
                    'accion': '/pedidos?estado=pendiente'
                })
            
            # Alerta por productos sin precio
            productos_sin_precio = Producto.query.outerjoin(PrecioProducto).filter(
                PrecioProducto.id.is_(None)
            ).count()
            
            if productos_sin_precio > 0:
                alertas_sistema.append({
                    'tipo': 'info',
                    'titulo': 'Productos sin precios configurados',
                    'mensaje': f'{productos_sin_precio} productos requieren configuración de precios',
                    'accion': '/precios'
                })
            
            # 6. MÉTRICAS FINANCIERAS ADICIONALES
            # Calcular crecimiento
            if ventas_mes_anterior > 0:
                crecimiento_ventas = ((ventas_totales - ventas_mes_anterior) / ventas_mes_anterior) * 100
            else:
                crecimiento_ventas = 0
            
            # Actualizar contexto para administrador
            context.update({
                'total_vendedores': total_vendedores,
                'total_clientes': total_clientes,
                'total_productos': total_productos,
                'ventas_totales': ventas_totales,
                'ventas_hoy': ventas_hoy,
                'ventas_mes': ventas_totales,
                'pedidos_mes': pedidos_mes_count,
                'total_pedidos': pedidos_pendientes,
                'pedidos_hoy': pedidos_hoy_count,
                'eficiencia_sistema': round(eficiencia_sistema, 1),
                'crecimiento_ventas': round(crecimiento_ventas, 1),
                'vendedores_performance': vendedores_performance,
                'alertas_sistema': alertas_sistema,
                'pedidos_facturados': pedidos_facturados,
                'pedidos_totales': pedidos_totales
            })
            
        else:
            # ===== MÉTRICAS ESPECÍFICAS PARA VENDEDOR REGULAR =====
            
            # 1. Obtener clientes asignados al vendedor
            clientes_vendedor = current_user.obtener_clientes_visibles()
            clientes_ids = [c.id for c in clientes_vendedor]
            
            # 2. Métricas del vendedor (OPTIMIZADO)
            if clientes_ids:
                # Ventas facturadas del vendedor (fecha de facturación local)
                pedidos_vend_facturados = Pedido.query.filter(
                    Pedido.cliente_id.in_(clientes_ids),
                    Pedido.estado == 'facturado',
                    Pedido.fecha_facturacion.isnot(None)
                ).all()

                ventas_vendedor_hoy = sum(
                    _calcular_venta_pedido(p)
                    for p in pedidos_vend_facturados
                    if _pedido_facturado_en_periodo_local(p, hoy, hoy)
                )

                pedidos_vend_mes = Pedido.query.filter(
                    Pedido.cliente_id.in_(clientes_ids),
                    Pedido.fecha_pedido >= inicio_mes
                ).all()
                ventas_vendedor_mes = sum(
                    _calcular_venta_pedido(p)
                    for p in pedidos_vend_facturados
                    if _pedido_facturado_en_periodo_local(p, inicio_mes)
                )
                
                # Conteo de pedidos del día
                pedidos_hoy_count = Pedido.query.filter(
                    Pedido.cliente_id.in_(clientes_ids),
                    Pedido.fecha_pedido >= hoy,
                    Pedido.fecha_pedido < hoy + timedelta(days=1)
                ).count()
            else:
                ventas_vendedor_hoy = 0
                ventas_vendedor_mes = 0
                pedidos_hoy_count = 0
            
            # 4. Estadísticas del vendedor
            pedidos_pendientes_vendedor = 0
            if clientes_ids:
                pedidos_pendientes_vendedor = Pedido.query.filter(
                    Pedido.cliente_id.in_(clientes_ids),
                    Pedido.estado == 'pendiente'
                ).count()
            
            # Actualizar contexto para vendedor
            context.update({
                'clientes_asignados': len(clientes_vendedor),
                'pedidos_hoy': pedidos_hoy_count,
                'pedidos_mes': len(pedidos_vend_mes) if clientes_ids else 0,
                'ventas_hoy': ventas_vendedor_hoy,
                'ventas_mes': ventas_vendedor_mes,
                'pedidos_pendientes': pedidos_pendientes_vendedor,
                'total_pedidos': pedidos_pendientes_vendedor
            })
        
        # Renderizar template apropiado
        return render_template('dashboard_vendedor.html', **context)
        
    except Exception as e:
        app.logger.error(f"Error en dashboard_vendedor: {e}", exc_info=True)
        
        flash('Error al cargar el dashboard. Contacte al administrador.', 'error')
        
        # Contexto mínimo en caso de error
        context = {
            'vendedor': current_user,
            'fecha_actual': datetime.now(DASHBOARD_TIMEZONE).date(),
            'total_vendedores': 0,
            'ventas_totales': 0,
            'total_pedidos': 0,
            'eficiencia_sistema': 0,
            'pedidos_hoy': 0,
            'ventas_hoy': 0
        }
        
        return render_template('dashboard_vendedor.html', **context)

# Función auxiliar para obtener métricas del sistema (opcional)
def obtener_metricas_sistema():
    """Función auxiliar para obtener métricas del sistema de forma centralizada"""
    try:
        hoy = datetime.now(DASHBOARD_TIMEZONE).date()
        inicio_mes = hoy.replace(day=1)
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        inicio_mes_anterior = (inicio_mes - timedelta(days=1)).replace(day=1)
        fin_mes_anterior = inicio_mes - timedelta(days=1)
        
        metricas = {
            'total_vendedores': Vendedor.query.filter_by(activo=True).count(),
            'total_clientes': Cliente.query.count(),
            'total_productos': Producto.query.count(),
            'pedidos_pendientes': Pedido.query.filter_by(estado='pendiente').count(),
            'pedidos_facturados': Pedido.query.filter_by(estado='facturado').count(),
        }

        metricas_qb = _obtener_metricas_ventas_quickbooks(
            hoy=hoy,
            inicio_mes=inicio_mes,
            inicio_semana=inicio_semana,
            inicio_mes_anterior=inicio_mes_anterior,
            fin_mes_anterior=fin_mes_anterior,
            inicio_tendencia=inicio_semana - timedelta(weeks=8),
            inicio_ultimos_7_dias=hoy - timedelta(days=6),
        )
        if metricas_qb:
            metricas['ventas_mes'] = metricas_qb['ventas_mes']
        else:
            # Fallback local cuando QuickBooks no está disponible
            pedidos_facturados_data = Pedido.query.filter(
                Pedido.estado == 'facturado',
                Pedido.fecha_facturacion.isnot(None)
            ).all()
            metricas['ventas_mes'] = sum(
                _calcular_venta_pedido(p)
                for p in pedidos_facturados_data
                if _pedido_facturado_en_periodo_local(p, inicio_mes)
            )

        metricas['total_pedidos'] = Pedido.query.count()
        
        if metricas['total_pedidos'] > 0:
            metricas['eficiencia'] = (metricas['pedidos_facturados'] / metricas['total_pedidos']) * 100
        else:
            metricas['eficiencia'] = 95.8
        
        return metricas
        
    except Exception as e:
        app.logger.error(f"Error obteniendo métricas del sistema: {e}")
        return {
            'total_vendedores': 0,
            'total_clientes': 0,
            'total_productos': 0,
            'pedidos_pendientes': 0,
            'pedidos_facturados': 0,
            'ventas_mes': 0,
            'total_pedidos': 0,
            'eficiencia': 0
        }

# Ruta adicional para API de métricas (útil para actualizaciones en tiempo real)
@app.route('/api/dashboard/metricas')
@login_required
def api_dashboard_metricas():
    """API para obtener métricas del dashboard en tiempo real"""
    
    if not isinstance(current_user, Vendedor):
        return jsonify({'error': 'No autorizado'}), 403
    
    try:
        if current_user.rol.nombre == 'super_admin':
            metricas = obtener_metricas_sistema()
        else:
            # Métricas específicas del vendedor
            clientes_vendedor = current_user.obtener_clientes_visibles()
            clientes_ids = [c.id for c in clientes_vendedor]
            
            hoy = datetime.now(DASHBOARD_TIMEZONE).date()
            pedidos_hoy = 0
            ventas_hoy = 0
            
            if clientes_ids:
                pedidos_vendedor_hoy = Pedido.query.filter(
                    Pedido.cliente_id.in_(clientes_ids),
                    Pedido.fecha_pedido >= hoy
                ).all()
                
                pedidos_hoy = len(pedidos_vendedor_hoy)
                pedidos_facturados_hoy = Pedido.query.filter(
                    Pedido.cliente_id.in_(clientes_ids),
                    Pedido.estado == 'facturado',
                    Pedido.fecha_facturacion.isnot(None)
                ).all()
                ventas_hoy = sum(
                    _calcular_venta_pedido(p)
                    for p in pedidos_facturados_hoy
                    if _pedido_facturado_en_periodo_local(p, hoy, hoy)
                )
            
            metricas = {
                'clientes_asignados': len(clientes_vendedor),
                'pedidos_hoy': pedidos_hoy,
                'ventas_hoy': ventas_hoy,
                'pedidos_pendientes': Pedido.query.filter(
                    Pedido.cliente_id.in_(clientes_ids),
                    Pedido.estado == 'pendiente'
                ).count() if clientes_ids else 0
            }
        
        return jsonify(metricas)
        
    except Exception as e:
        app.logger.error(f"Error en API de métricas: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500

############################################
# MODELOS Y BASE DE DATOS
############################################

class Producto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    descripcion = db.Column(db.String(200))
    temperatura = db.Column(db.String(50))
    facturaciones = db.relationship('Facturacion', back_populates='producto', cascade="all, delete-orphan")
    recepciones = db.relationship('Recepcion', back_populates='producto', lazy=True, cascade="all, delete-orphan")
    importaciones = db.relationship('Importacion', back_populates='producto', lazy=True, cascade="all, delete-orphan")
    qbo_id = db.Column(db.String(20), unique=True)
    tax_rate = db.Column(db.Float, nullable=False, default=0.0)  # Nueva columna para tasa de impuesto
    se_pesa = db.Column(db.Boolean, default=False, nullable=False)
    proveedor = db.Column(db.String(100))

    def to_dict(self):
        return {
            'id': self.id,
            'nombre': self.nombre,
            'descripcion': self.descripcion,
            'temperatura': self.temperatura,
            'qbo_id': self.qbo_id,
            'tax_rate': self.tax_rate,
            'se_pesa': self.se_pesa,
            'proveedor': self.proveedor
        }

class Cliente(db.Model):
    __tablename__ = 'cliente'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    territorio_id = db.Column(db.Integer,
                              db.ForeignKey('territorio.id'))      
    territorio  = db.relationship('Territorio', backref='clientes') 
    facturaciones = db.relationship('Facturacion', back_populates='cliente', cascade="all, delete-orphan")
    pedidos = db.relationship('Pedido', back_populates='cliente', cascade="all, delete-orphan")
    qbo_id = db.Column(db.String(20), unique=True)
    moneda = db.Column(db.String(3), default='XCG', nullable=False)

    def to_dict(self):
        return {'id': self.id, 'nombre': self.nombre, 'qbo_id': self.qbo_id, 'moneda': self.moneda}

class Facturacion(db.Model):
    __tablename__ = 'facturacion'
    id = db.Column(db.Integer, primary_key=True)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    peso = db.Column(db.Float, nullable=False)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    lote = db.Column(db.String(50), nullable=False)
    fecha_fabricacion = db.Column(db.String(10), nullable=False)
    fecha_expiracion = db.Column(db.String(10), nullable=False)
    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)
    producto = db.relationship('Producto', back_populates='facturaciones')
    cliente = db.relationship('Cliente', back_populates='facturaciones')

    def to_dict(self):
        return {
            'id': self.id,
            'producto_id': self.producto_id,
            'producto': self.producto.nombre if self.producto else None,
            'peso': self.peso,
            'cliente_id': self.cliente_id,
            'cliente': self.cliente.nombre if self.cliente else None,
            'lote': self.lote,
            'fecha_fabricacion': self.fecha_fabricacion,
            'fecha_expiracion': self.fecha_expiracion,
            'fecha_registro': self.fecha_registro.strftime('%Y-%m-%d %H:%M:%S') if self.fecha_registro else None
        }

class Recepcion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    peso = db.Column(db.Float, nullable=False)
    proveedor = db.Column(db.String(100), nullable=False)
    numero_factura = db.Column(db.String(100), nullable=False)
    recibido_en = db.Column(db.Date, nullable=False)
    producto = db.relationship('Producto', back_populates='recepciones')
    
    def to_dict(self):
        return {
            'id': self.id,
            'producto_id': self.producto_id,
            'producto': self.producto.nombre if self.producto else 'undefined',
            'peso': self.peso,
            'recibido_en': self.recibido_en.strftime('%Y-%m-%d') if self.recibido_en else 'No disponible',
            'proveedor': self.proveedor,
            'numero_factura': self.numero_factura
        }

class Importacion(db.Model):
    __tablename__ = 'importacion'
    id = db.Column(db.Integer, primary_key=True)
    numero_factura = db.Column(db.String(100), nullable=False)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    cantidad_cajas = db.Column(db.Integer, nullable=False)
    cantidad_total = db.Column(db.Float, nullable=False)
    precio_fob_unidad = db.Column(db.Float, nullable=False)
    flete = db.Column(db.Float, nullable=False)
    arancel = db.Column(db.Float, nullable=False)
    costo_aduana = db.Column(db.Float, nullable=False)
    precio_jomar = db.Column(db.Float, nullable=False)
    precio_retail = db.Column(db.Float, nullable=False)
    cif_ang = db.Column(db.Float, nullable=True)
    ob_ang = db.Column(db.Float, nullable=True)
    fecha_importacion = db.Column(db.DateTime, default=datetime.utcnow)
    flete_local = db.Column(db.Float, nullable=True)
    costo_total_almacen = db.Column(db.Float, nullable=True)
    producto = db.relationship('Producto', back_populates='importaciones')

# MODELOS NUEVOS PARA PEDIDOS

class Pedido(db.Model):
    __tablename__ = 'pedido'
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    fecha_pedido = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    fecha_facturacion = db.Column(db.DateTime, nullable=True)
    estado = db.Column(db.String(30), default="pendiente", nullable=False)
    notas = db.Column(db.Text, nullable=True)
    invoice_id_qbo = db.Column(db.String(100), nullable=True)
    doc_number_qbo = db.Column(db.String(20), nullable=True)
    tipo_cambio = db.Column(db.Float, default=1.0, nullable=False)
    cliente = db.relationship('Cliente', back_populates='pedidos')
    detalles = db.relationship('DetallePedido', back_populates='pedido', cascade="all, delete-orphan")

    def __repr__(self):
        return f'<Pedido {self.id} - Cliente {self.cliente.nombre} - Estado {self.estado}>'

class DetallePedido(db.Model):
    __tablename__ = 'detalle_pedido'
    id = db.Column(db.Integer, primary_key=True)
    pedido_id = db.Column(db.Integer, db.ForeignKey('pedido.id'), nullable=False)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    cajas = db.Column(db.Integer, nullable=False, default=0)  # NUEVO
    peso = db.Column(db.Float, nullable=False, default=0)
    lote = db.Column(db.String(50), nullable=True)
    fecha_fabricacion = db.Column(db.String(10), nullable=True)
    fecha_expiracion = db.Column(db.String(10), nullable=True)
    pedido = db.relationship('Pedido', back_populates='detalles')
    producto = db.relationship('Producto')
    precio_unitario = db.Column(db.Numeric(10,2), nullable=False, default=0)
    subtotal        = db.Column(db.Numeric(10,2), nullable=False)
    es_linea_pedido = db.Column(db.Boolean, default=True, nullable=False)
    cajas_pedidas   = db.Column(db.Integer, nullable=False, default=0)

    @property
    def cajas_objetivo(self):
        return int(self.cajas_pedidas or self.cajas or 0)

    @property
    def cajas_pesadas_count(self):
        return len(getattr(self, 'cajas_pesadas', []) or [])

    @property
    def peso_real(self):
        if self.cajas_pesadas_count:
            total = Decimal('0')
            for caja in self.cajas_pesadas:
                total += Decimal(str(caja.peso or 0))
            return total
        return Decimal(str(self.peso or 0))

    @property
    def pesaje_completo(self):
        return self.cajas_pesadas_count >= self.cajas_objetivo

    @property
    def lote_principal(self):
        if self.cajas_pesadas_count:
            lotes = {caja.lote.strip() for caja in self.cajas_pesadas if caja.lote}
            return next(iter(lotes)) if len(lotes) == 1 else None
        return self.lote

    @property
    def fecha_elaboracion_principal(self):
        if self.cajas_pesadas_count:
            fechas = {caja.fecha_elaboracion for caja in self.cajas_pesadas if caja.fecha_elaboracion}
            return next(iter(fechas)) if len(fechas) == 1 else None
        return self.fecha_fabricacion

    @property
    def fecha_vencimiento_principal(self):
        if self.cajas_pesadas_count:
            fechas = {caja.fecha_vencimiento for caja in self.cajas_pesadas if caja.fecha_vencimiento}
            return next(iter(fechas)) if len(fechas) == 1 else None
        return self.fecha_expiracion

    @property
    def categoria_code(self):
        return _producto_categoria_code(self.producto) if self.producto else 'RES'

    def __repr__(self):
        return f'<DetallePedido {self.id} - Producto {self.producto.nombre} - Peso {self.peso}>'


class CajaPesada(db.Model):
    __tablename__ = 'caja_pesada'

    id = db.Column(db.Integer, primary_key=True)
    detalle_pedido_id = db.Column(
        db.Integer,
        db.ForeignKey('detalle_pedido.id', ondelete='CASCADE'),
        nullable=False,
        index=True,
    )
    numero = db.Column(db.Integer, nullable=False)
    peso = db.Column(db.Numeric(8, 3), nullable=False)
    lote = db.Column(db.String(50), nullable=False, index=True)
    fecha_elaboracion = db.Column(db.Date, nullable=False)
    fecha_vencimiento = db.Column(db.Date, nullable=False)
    pesado_por = db.Column(db.Integer, db.ForeignKey('vendedor.id'), nullable=True)
    pesado_en = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    detalle_pedido = db.relationship(
        'DetallePedido',
        backref=db.backref(
            'cajas_pesadas',
            cascade='all, delete-orphan',
            order_by='CajaPesada.numero',
        ),
    )
    pesado_por_vendedor = db.relationship('Vendedor')

    __table_args__ = (
        db.UniqueConstraint('detalle_pedido_id', 'numero', name='uq_caja_detalle_numero'),
    )

    def __repr__(self):
        return f'<CajaPesada {self.id} detalle={self.detalle_pedido_id} numero={self.numero}>'


class Camara(db.Model):
    __tablename__ = 'camara'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(120), nullable=False)
    tipo = db.Column(db.String(20), nullable=False, default='refrigeracion')  # refrigeracion|congelacion
    temp_min = db.Column(db.Numeric(5, 2), nullable=False)
    temp_max = db.Column(db.Numeric(5, 2), nullable=False)
    activa = db.Column(db.Boolean, nullable=False, default=True)
    responsable_id = db.Column(db.Integer, db.ForeignKey('vendedor.id'), nullable=True)
    ronda_am = db.Column(db.String(5), nullable=True)   # 'HH:MM'
    ronda_pm = db.Column(db.String(5), nullable=True)   # 'HH:MM'
    creado_en = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    responsable = db.relationship('Vendedor')

    def fuera_de_rango(self, temperatura):
        """True si la temperatura está fuera del rango aceptable [min, max]."""
        t = Decimal(str(temperatura))
        return t < self.temp_min or t > self.temp_max

    def __repr__(self):
        return f'<Camara {self.id} {self.nombre}>'


class LecturaTemperatura(db.Model):
    __tablename__ = 'lectura_temperatura'
    id = db.Column(db.Integer, primary_key=True)
    camara_id = db.Column(db.Integer, db.ForeignKey('camara.id'), nullable=False, index=True)
    temperatura = db.Column(db.Numeric(5, 2), nullable=False)
    registrado_por = db.Column(db.Integer, db.ForeignKey('vendedor.id'), nullable=True)
    registrado_en = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    fuera_de_rango = db.Column(db.Boolean, nullable=False, default=False)
    accion_correctiva = db.Column(db.Text, nullable=True)
    accion_causa = db.Column(db.Text, nullable=True)
    accion_tomada = db.Column(db.Text, nullable=True)
    accion_responsable = db.Column(db.String(120), nullable=True)
    accion_disposicion = db.Column(db.Text, nullable=True)

    camara = db.relationship('Camara')
    registrado_por_vendedor = db.relationship('Vendedor')

    def __repr__(self):
        return f'<LecturaTemperatura {self.id} camara={self.camara_id} {self.temperatura}>'


class RegistroConfig(db.Model):
    """Configuración (fila única) del registro de temperaturas para el PDF."""
    __tablename__ = 'registro_config'
    id = db.Column(db.Integer, primary_key=True)
    codigo_documento = db.Column(db.String(60), nullable=False, default='FR-HACCP-TEMP-01')
    version = db.Column(db.String(20), nullable=False, default='1')
    frecuencia_texto = db.Column(db.String(120), nullable=False, default='2 veces al día')
    termometro = db.Column(db.String(120), nullable=True)
    termometro_calibrado_en = db.Column(db.Date, nullable=True)
    actualizado_en = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    def __repr__(self):
        return f'<RegistroConfig {self.codigo_documento} v{self.version}>'


class RevisionRegistro(db.Model):
    """Verificación HACCP: un responsable revisa los registros de un período."""
    __tablename__ = 'revision_registro'
    id = db.Column(db.Integer, primary_key=True)
    revisado_por = db.Column(db.Integer, db.ForeignKey('vendedor.id'), nullable=True)
    revisado_en = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    periodo_desde = db.Column(db.Date, nullable=True)
    periodo_hasta = db.Column(db.Date, nullable=True)
    nota = db.Column(db.Text, nullable=True)

    revisado_por_vendedor = db.relationship('Vendedor')

    def __repr__(self):
        return f'<RevisionRegistro {self.id} por={self.revisado_por}>'


class ProductoLimpieza(db.Model):
    """Catálogo consultable de productos de limpieza: dilución y procedimiento (SSOP)."""
    __tablename__ = 'producto_limpieza'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(120), nullable=False)
    dilucion = db.Column(db.Text, nullable=False)
    procedimiento = db.Column(db.Text, nullable=True)
    notas_seguridad = db.Column(db.Text, nullable=True)
    activo = db.Column(db.Boolean, nullable=False, default=True)
    creado_en = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    def __repr__(self):
        return f'<ProductoLimpieza {self.id} {self.nombre}>'


class AreaLimpieza(db.Model):
    """Equipo o espacio a limpiar, con su producto/método/frecuencia (ficha fija)."""
    __tablename__ = 'area_limpieza'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(120), nullable=False)
    tipo = db.Column(db.String(20), nullable=False, default='equipo')  # equipo|espacio
    producto_id = db.Column(db.Integer, db.ForeignKey('producto_limpieza.id'), nullable=True)
    metodo = db.Column(db.Text, nullable=True)
    frecuencia_texto = db.Column(db.String(120), nullable=True)
    responsable_id = db.Column(db.Integer, db.ForeignKey('vendedor.id'), nullable=True)
    # Proceso de inocuidad en 2 pasos: 1) limpiar con detergente (producto),
    # 2) sanitizar (sanitizante). Ambos opcionales.
    sanitizante_id = db.Column(db.Integer, db.ForeignKey('producto_limpieza.id'), nullable=True)
    activa = db.Column(db.Boolean, nullable=False, default=True)
    creado_en = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    producto = db.relationship('ProductoLimpieza', foreign_keys=[producto_id])
    sanitizante = db.relationship('ProductoLimpieza', foreign_keys=[sanitizante_id])
    responsable = db.relationship('Vendedor')

    def __repr__(self):
        return f'<AreaLimpieza {self.id} {self.nombre}>'


class RegistroLimpieza(db.Model):
    """Registro de una limpieza ejecutada (HACCP/SSOP)."""
    __tablename__ = 'registro_limpieza'
    id = db.Column(db.Integer, primary_key=True)
    area_id = db.Column(db.Integer, db.ForeignKey('area_limpieza.id'), nullable=False, index=True)
    registrado_por = db.Column(db.Integer, db.ForeignKey('vendedor.id'), nullable=True)
    registrado_en = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)
    conforme = db.Column(db.Boolean, nullable=False, default=True)
    observacion = db.Column(db.Text, nullable=True)
    # Acción correctiva estructurada (cuando conforme=False). A diferencia de
    # LecturaTemperatura, no hay columna legacy `accion_correctiva`: esta tabla
    # nace nueva, sin registros antiguos de texto libre que mantener.
    accion_causa = db.Column(db.Text, nullable=True)
    accion_tomada = db.Column(db.Text, nullable=True)
    accion_responsable = db.Column(db.String(120), nullable=True)
    accion_disposicion = db.Column(db.Text, nullable=True)
    firma_png = db.Column(db.Text, nullable=True)  # data URL PNG de la firma del responsable
    # Ajustes auditoría de inocuidad (FR-HACCP-LIMP-01):
    concentracion_ppm = db.Column(db.Integer, nullable=True)   # ppm de Sani-T-10 Plus
    verificado_por = db.Column(db.Integer, db.ForeignKey('vendedor.id'), nullable=True)
    metodo_verificacion = db.Column(db.String(20), nullable=True)  # visual|atp|hisopado

    area = db.relationship('AreaLimpieza')
    # Dos FKs a vendedor -> foreign_keys explícito en ambas relaciones.
    registrado_por_vendedor = db.relationship('Vendedor', foreign_keys=[registrado_por])
    verificado_por_vendedor = db.relationship('Vendedor', foreign_keys=[verificado_por])

    def __repr__(self):
        return f'<RegistroLimpieza {self.id} area={self.area_id} conforme={self.conforme}>'


class EventoAuditoria(db.Model):
    """Log de auditoría real: quién hizo qué y cuándo (HACCP + seguridad)."""
    __tablename__ = 'evento_auditoria'
    id = db.Column(db.Integer, primary_key=True)
    vendedor_id = db.Column(db.Integer, db.ForeignKey('vendedor.id'), nullable=True)
    actor = db.Column(db.String(120), nullable=True)   # nombre cacheado
    tipo = db.Column(db.String(20), nullable=False)     # temp|clean|user|auth|config
    accion = db.Column(db.String(160), nullable=False)
    detalle = db.Column(db.String(255), nullable=True)
    creado_en = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)

    vendedor = db.relationship('Vendedor')

    def __repr__(self):
        return f'<EventoAuditoria {self.id} {self.tipo}:{self.accion}>'


class LimpiezaConfig(db.Model):
    """Configuración (fila única) del registro de limpieza para el PDF."""
    __tablename__ = 'limpieza_config'
    id = db.Column(db.Integer, primary_key=True)
    codigo_documento = db.Column(db.String(60), nullable=False, default='FR-HACCP-LIMP-01')
    version = db.Column(db.String(20), nullable=False, default='1')
    frecuencia_texto = db.Column(db.String(120), nullable=False, default='Según programa de limpieza')
    responsable_verificacion = db.Column(db.String(120), nullable=True)
    actualizado_en = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    def __repr__(self):
        return f'<LimpiezaConfig {self.codigo_documento} v{self.version}>'


class RevisionLimpieza(db.Model):
    """Verificación HACCP: un responsable revisa los registros de limpieza de un período."""
    __tablename__ = 'revision_limpieza'
    id = db.Column(db.Integer, primary_key=True)
    revisado_por = db.Column(db.Integer, db.ForeignKey('vendedor.id'), nullable=True)
    revisado_en = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    periodo_desde = db.Column(db.Date, nullable=True)
    periodo_hasta = db.Column(db.Date, nullable=True)
    nota = db.Column(db.Text, nullable=True)

    revisado_por_vendedor = db.relationship('Vendedor')

    def __repr__(self):
        return f'<RevisionLimpieza {self.id} por={self.revisado_por}>'


class PedidoEvento(db.Model):
    __tablename__ = 'pedido_evento'

    id = db.Column(db.Integer, primary_key=True)
    pedido_id = db.Column(
        db.Integer,
        db.ForeignKey('pedido.id', ondelete='CASCADE'),
        nullable=False,
        index=True,
    )
    tipo = db.Column(db.String(50), nullable=False)
    descripcion = db.Column(db.String(255), nullable=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('vendedor.id'), nullable=True)
    meta = db.Column('metadata_json', db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False, index=True)

    pedido = db.relationship(
        'Pedido',
        backref=db.backref(
            'eventos',
            cascade='all, delete-orphan',
            order_by='PedidoEvento.created_at.desc()',
        ),
    )
    usuario = db.relationship('Vendedor')

    def __repr__(self):
        return f'<PedidoEvento {self.id} pedido={self.pedido_id} tipo={self.tipo}>'

############################################
# MODELOS MULTI-VENDEDOR (AGREGAR AL FINAL DE APP.PY)
############################################

class Rol(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(50), unique=True, nullable=False)
    descripcion = db.Column(db.Text)
    nivel_jerarquia = db.Column(db.Integer, default=0)
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relaciones
    vendedores = db.relationship('Vendedor', backref='rol', lazy=True)
    permisos = db.relationship('RolPermiso', back_populates='rol', cascade="all, delete-orphan")
    
    def __repr__(self):
        return f'<Rol {self.nombre}>'

class Permiso(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), unique=True, nullable=False)
    descripcion = db.Column(db.Text)
    categoria = db.Column(db.String(50))
    recurso = db.Column(db.String(100))
    
    def __repr__(self):
        return f'<Permiso {self.nombre}>'

class RolPermiso(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    rol_id = db.Column(db.Integer, db.ForeignKey('rol.id'), nullable=False)
    permiso_id = db.Column(db.Integer, db.ForeignKey('permiso.id'), nullable=False)
    
    # Tipos de acceso CRUD
    puede_leer = db.Column(db.Boolean, default=True)
    puede_crear = db.Column(db.Boolean, default=False)
    puede_editar = db.Column(db.Boolean, default=False)
    puede_eliminar = db.Column(db.Boolean, default=False)
    
    # Relaciones
    rol = db.relationship('Rol', back_populates='permisos')
    permiso = db.relationship('Permiso')
    
    __table_args__ = (db.UniqueConstraint('rol_id', 'permiso_id', name='unique_rol_permiso'),)

class Territorio(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    descripcion = db.Column(db.Text)
    tipo = db.Column(db.String(50))
    coordenadas = db.Column(db.JSON)
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relaciones
    vendedores = db.relationship('Vendedor', backref='territorio', lazy=True)
    
    def __repr__(self):
        return f'<Territorio {self.nombre}>'



class ClienteVendedor(db.Model):
    __tablename__ = 'cliente_vendedor'        #  ← bueno especificarlo

    id          = db.Column(db.Integer, primary_key=True)
    cliente_id  = db.Column(db.Integer, db.ForeignKey('cliente.id'),  nullable=False)
    vendedor_id = db.Column(db.Integer, db.ForeignKey('vendedor.id'), nullable=False)

    # Config asignación
    tipo_asignacion     = db.Column(db.String(50), default='principal')
    fecha_inicio        = db.Column(db.Date, default=datetime.utcnow)
    fecha_fin           = db.Column(db.Date)
    activo              = db.Column(db.Boolean, default=True)
    porcentaje_comision = db.Column(db.Float, default=100.0)

    # Relaciones
    cliente  = db.relationship('Cliente',
                               backref='vendedores_asignados')
    vendedor = db.relationship('Vendedor',
                               back_populates='clientes_asignados')

    # ───────── helpers ─────────
    @classmethod
    def asignar(cls, cliente_id: int, vendedor_id: int):
        """Asigna (o reactiva) un cliente a un vendedor."""
        asign = cls.query.filter_by(cliente_id=cliente_id,
                                    vendedor_id=vendedor_id).first()
        if asign:
            asign.activo       = True
            asign.fecha_inicio = date.today()
        else:
            asign = cls(cliente_id=cliente_id,
                        vendedor_id=vendedor_id,
                        activo=True,
                        fecha_inicio=date.today())
            db.session.add(asign)
        db.session.commit()
        return asign

    @classmethod
    def desasignar(cls, asign_id: int):
        asign = cls.query.get_or_404(asign_id)
        asign.activo    = False
        asign.fecha_fin = date.today()
        db.session.commit()
        return asign

    def __repr__(self):
        return f'<ClienteVendedor {self.cliente_id}-{self.vendedor_id}>'


############################################
# DECORADORES Y FUNCIONES DE SEGURIDAD
############################################

# 2. DECORADOR MEJORADO PARA VERIFICAR PERMISOS
def requiere_permiso_recurso(recurso, tipo_acceso='leer'):
    """Decorador mejorado para verificar permisos sobre recursos específicos"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                flash('Debes iniciar sesión para acceder a esta página', 'warning')
                return redirect(url_for('login'))

            # Fail-closed: solo usuarios Vendedor tienen permisos definidos
            if not isinstance(current_user, Vendedor):
                abort(403)

            # Verificar permiso sobre el recurso
            if not current_user.tiene_permiso(recurso, tipo_acceso):
                flash(f'No tienes permisos para {tipo_acceso} {recurso}', 'error')
                return redirect(url_for('index'))
                
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def requiere_rol(roles_permitidos):
    """Decorador para verificar roles específicos"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))

            # Fail-closed: solo usuarios Vendedor tienen roles definidos
            if not isinstance(current_user, Vendedor):
                abort(403)

            if current_user.rol.nombre not in roles_permitidos:
                flash("No tienes autorización para acceder a esta función", "error")
                return redirect(url_for('index'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def obtener_vendedor_actual():
    """Obtiene el vendedor actual o None si es usuario legacy"""
    if isinstance(current_user, Vendedor):
        return current_user
    return None
# Agregar estos modelos al archivo app.py, después de los modelos existentes

class ListaPrecio(db.Model):
    __tablename__ = 'lista_precio'
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    descripcion = db.Column(db.String(200))
    es_default = db.Column(db.Boolean, default=False, nullable=False)
    activa = db.Column(db.Boolean, default=True, nullable=False)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relaciones
    precios_productos = db.relationship('PrecioProducto', back_populates='lista_precio', cascade="all, delete-orphan")
    clientes = db.relationship('ClienteListaPrecio', back_populates='lista_precio', cascade="all, delete-orphan")
    
    def to_dict(self):
        return {
            'id': self.id,
            'nombre': self.nombre,
            'descripcion': self.descripcion,
            'es_default': self.es_default,
            'activa': self.activa,
            'fecha_creacion': self.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if self.fecha_creacion else None
        }


class PrecioProducto(db.Model):
    __tablename__ = 'precio_producto'
    id = db.Column(db.Integer, primary_key=True)
    lista_precio_id = db.Column(db.Integer, db.ForeignKey('lista_precio.id'), nullable=False)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    precio_base = db.Column(db.Float, nullable=False)
    precio_jomar = db.Column(db.Float)
    precio_retail = db.Column(db.Float)
    margen_jomar = db.Column(db.Float, default=1.0)
    margen_retail = db.Column(db.Float, default=1.2)
    activo = db.Column(db.Boolean, default=True, nullable=False)
    fecha_actualizacion = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relaciones
    lista_precio = db.relationship('ListaPrecio', back_populates='precios_productos')
    producto = db.relationship('Producto')
    
    __table_args__ = (db.UniqueConstraint('lista_precio_id', 'producto_id', name='uk_lista_producto'),)
    
    def calcular_precios(self):
        """Calcula precios Jomar y Retail basados en precio base y márgenes"""
        self.precio_jomar = self.precio_base * (self.margen_jomar or 1.0)
        self.precio_retail = self.precio_base * (self.margen_retail or 1.2)
    
    def to_dict(self):
        return {
            'id': self.id,
            'lista_precio_id': self.lista_precio_id,
            'producto_id': self.producto_id,
            'producto_nombre': self.producto.nombre if self.producto else None,
            'precio_base': self.precio_base,
            'precio_jomar': self.precio_jomar,
            'precio_retail': self.precio_retail,
            'margen_jomar': self.margen_jomar,
            'margen_retail': self.margen_retail,
            'activo': self.activo,
            'fecha_actualizacion': self.fecha_actualizacion.strftime('%Y-%m-%d %H:%M:%S') if self.fecha_actualizacion else None
        }


class ClienteListaPrecio(db.Model):
    __tablename__ = 'cliente_lista_precio'
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    lista_precio_id = db.Column(db.Integer, db.ForeignKey('lista_precio.id'), nullable=False)
    fecha_asignacion = db.Column(db.DateTime, default=datetime.utcnow)
    activa = db.Column(db.Boolean, default=True, nullable=False)
    
    # Relaciones
    cliente = db.relationship('Cliente')
    lista_precio = db.relationship('ListaPrecio', back_populates='clientes')
    
    __table_args__ = (db.UniqueConstraint('cliente_id', 'lista_precio_id', name='uk_cliente_lista'),)


class PrecioClienteProducto(db.Model):
    __tablename__ = 'precio_cliente_producto'
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    precio_base = db.Column(db.Float, nullable=False)
    precio_jomar = db.Column(db.Float)
    precio_retail = db.Column(db.Float)
    margen_jomar = db.Column(db.Float, default=1.0)
    margen_retail = db.Column(db.Float, default=1.2)
    activo = db.Column(db.Boolean, default=True, nullable=False)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    fecha_actualizacion = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relaciones
    cliente = db.relationship('Cliente')
    producto = db.relationship('Producto')
    
    __table_args__ = (db.UniqueConstraint('cliente_id', 'producto_id', name='uk_cliente_producto_precio'),)
    
    def calcular_precios(self):
        """Calcula precios Jomar y Retail basados en precio base y márgenes"""
        self.precio_jomar = self.precio_base * (self.margen_jomar or 1.0)
        self.precio_retail = self.precio_base * (self.margen_retail or 1.2)
    
    def to_dict(self):
        return {
            'id': self.id,
            'cliente_id': self.cliente_id,
            'cliente_nombre': self.cliente.nombre if self.cliente else None,
            'producto_id': self.producto_id,
            'producto_nombre': self.producto.nombre if self.producto else None,
            'precio_base': self.precio_base,
            'precio_jomar': self.precio_jomar,
            'precio_retail': self.precio_retail,
            'margen_jomar': self.margen_jomar,
            'margen_retail': self.margen_retail,
            'activo': self.activo,
            'fecha_creacion': self.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if self.fecha_creacion else None,
            'fecha_actualizacion': self.fecha_actualizacion.strftime('%Y-%m-%d %H:%M:%S') if self.fecha_actualizacion else None
        }


# Función auxiliar para obtener el precio de un producto para un cliente específico
def obtener_precio_producto_cliente(cliente_id, producto_id, tipo_precio='jomar'):
    """
    Obtiene el precio de un producto para un cliente específico.
    Prioridad: Precio específico cliente-producto > Lista de precios del cliente > Lista default
    
    Args:
        cliente_id: ID del cliente
        producto_id: ID del producto
        tipo_precio: 'base', 'jomar' o 'retail'
    
    Returns:
        Float: Precio encontrado o None si no existe
    """
    # 1. Verificar precio específico cliente-producto
    precio_especifico = PrecioClienteProducto.query.filter_by(
        cliente_id=cliente_id,
        producto_id=producto_id,
        activo=True
    ).first()

    if precio_especifico:
        if tipo_precio == 'base':
            precio = precio_especifico.precio_base
        elif tipo_precio == 'jomar':
            precio = precio_especifico.precio_jomar
        elif tipo_precio == 'retail':
            precio = precio_especifico.precio_retail
        return precio

    # 2. Verificar lista de precios del cliente
    cliente_lista = ClienteListaPrecio.query.filter_by(
        cliente_id=cliente_id,
        activa=True
    ).first()

    if cliente_lista:
        precio_lista = PrecioProducto.query.filter_by(
            lista_precio_id=cliente_lista.lista_precio_id,
            producto_id=producto_id,
            activo=True
        ).first()

        if precio_lista:
            if tipo_precio == 'base':
                precio = precio_lista.precio_base
            elif tipo_precio == 'jomar':
                precio = precio_lista.precio_jomar
            elif tipo_precio == 'retail':
                precio = precio_lista.precio_retail
            return precio

    # 3. Usar lista de precios por defecto
    lista_default = ListaPrecio.query.filter_by(es_default=True, activa=True).first()
    if lista_default:
        precio_default = PrecioProducto.query.filter_by(
            lista_precio_id=lista_default.id,
            producto_id=producto_id,
            activo=True
        ).first()

        if precio_default:
            if tipo_precio == 'base':
                precio = precio_default.precio_base
            elif tipo_precio == 'jomar':
                precio = precio_default.precio_jomar
            elif tipo_precio == 'retail':
                precio = precio_default.precio_retail
            return precio

    return None

def _fila_precio_vigente(cliente_id, producto_id):
    """Fila de precio que rige para (cliente, producto), y de dónde salió.

    Misma precedencia que `obtener_precio_producto_cliente`; existe solo para
    exponer los márgenes junto al precio. El precio siempre se toma del
    resolutor, no de acá: si alguna vez se separan, lo cachea
    `test_api_precios_coincide_con_el_resolutor`.
    """
    especifico = PrecioClienteProducto.query.filter_by(
        cliente_id=cliente_id, producto_id=producto_id, activo=True).first()
    if especifico:
        return especifico, 'específico'

    cliente_lista = ClienteListaPrecio.query.filter_by(
        cliente_id=cliente_id, activa=True).first()
    if cliente_lista:
        de_lista = PrecioProducto.query.filter_by(
            lista_precio_id=cliente_lista.lista_precio_id,
            producto_id=producto_id, activo=True).first()
        if de_lista:
            return de_lista, 'lista_asignada'

    lista_default = ListaPrecio.query.filter_by(es_default=True, activa=True).first()
    if lista_default:
        de_default = PrecioProducto.query.filter_by(
            lista_precio_id=lista_default.id,
            producto_id=producto_id, activo=True).first()
        if de_default:
            return de_default, 'lista_default'

    return None, 'sin_precio'


def obtener_precio_default_producto(producto_id, tipo_precio='base'):
    """
    Devuelve el precio de un producto tomado de la lista marcada como es_default.
    Si no existe, devuelve None.
    """
    lista_default = ListaPrecio.query.filter_by(es_default=True, activa=True).first()
    if not lista_default:
        return None

    precio = PrecioProducto.query.filter_by(
        lista_precio_id=lista_default.id,
        producto_id=producto_id,
        activo=True
    ).first()

    if not precio:
        return None

    if   tipo_precio == 'base':
        return precio.precio_base
    elif tipo_precio == 'jomar':
        return precio.precio_jomar
    elif tipo_precio == 'retail':
        return precio.precio_retail
    return None


def _parse_peso_caja(value):
    raw = str(value or '').strip().replace(',', '.')
    if not raw:
        return None, 'El peso es obligatorio'
    try:
        peso = Decimal(raw)
    except (InvalidOperation, ValueError):
        return None, 'Peso inválido'

    if peso <= 0:
        return None, 'El peso debe ser mayor que cero'

    if abs(peso.as_tuple().exponent) > 3:
        return None, 'El peso admite máximo 3 decimales'

    return peso.quantize(Decimal('0.001')), None


def _parse_iso_date_field(value, label):
    raw = str(value or '').strip()
    if not raw:
        return None, f'{label} es obligatoria'
    try:
        return datetime.strptime(raw, '%Y-%m-%d').date(), None
    except ValueError:
        return None, f'{label} inválida'


def _date_to_iso(value):
    if not value:
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    return str(value)


def _date_like_to_date(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.strptime(str(value), '%Y-%m-%d').date()
    except ValueError:
        return None


def _producto_categoria_code(producto):
    nombre = f'{getattr(producto, "nombre", "")} {getattr(producto, "descripcion", "")}'.lower()
    if 'lomo' in nombre:
        return 'LOM'
    if 'falda' in nombre:
        return 'FAL'
    if 'pollo' in nombre:
        return 'POL'
    if 'molida' in nombre:
        return 'MOL'
    if 'cerdo' in nombre or 'chuleta' in nombre or 'pork' in nombre:
        return 'CER'
    return 'RES'


def _detalle_cajas_por_lote(detalle):
    grupos = []
    lookup = {}
    for caja in detalle.cajas_pesadas:
        key = (caja.lote, caja.fecha_elaboracion, caja.fecha_vencimiento)
        grupo = lookup.get(key)
        if grupo is None:
            grupo = {
                'lote': caja.lote,
                'fecha_elaboracion': caja.fecha_elaboracion,
                'fecha_vencimiento': caja.fecha_vencimiento,
                'cajas': [],
            }
            lookup[key] = grupo
            grupos.append(grupo)
        grupo['cajas'].append(caja)
    return grupos


def _pedido_detalles_pesables(pedido):
    detalles = [
        detalle for detalle in pedido.detalles
        if detalle.es_linea_pedido and detalle.producto and detalle.producto.se_pesa
    ]
    return sorted(detalles, key=lambda detalle: (detalle.producto.nombre.lower(), detalle.id))


def _legacy_prep_lines_for_producto(pedido, producto_id):
    return [
        detalle for detalle in pedido.detalles
        if not detalle.es_linea_pedido and detalle.producto_id == producto_id
    ]


def _pedido_tiene_productos_pesables(pedido):
    return any(detalle.producto and detalle.producto.se_pesa for detalle in pedido.detalles if detalle.es_linea_pedido)


def _pedido_peso_total(pedido):
    total = Decimal('0')
    for detalle in _pedido_detalles_pesables(pedido):
        total += detalle.peso_real
    return total


def _pedido_cajas_pesadas_total(pedido):
    return sum(detalle.cajas_pesadas_count for detalle in _pedido_detalles_pesables(pedido))


def _pedido_cajas_objetivo_total(pedido):
    return sum(detalle.cajas_objetivo for detalle in _pedido_detalles_pesables(pedido))


def _pedido_can_finalize_pesar(pedido):
    return len(_validar_preparacion_pedido(pedido)) == 0


def _log_pedido_evento(pedido, tipo, descripcion=None, meta=None, commit=False):
    """Append an audit event to pedido_evento. Does NOT commit by default —
    callers should commit alongside their own writes."""
    evento = PedidoEvento(
        pedido_id=pedido.id,
        tipo=tipo,
        descripcion=descripcion,
        usuario_id=current_user.id if isinstance(current_user, Vendedor) else None,
        meta=json.dumps(meta, default=str) if meta else None,
    )
    db.session.add(evento)
    if commit:
        db.session.commit()
    return evento


def _renumerar_cajas_pesadas(detalle):
    for idx, caja in enumerate(sorted(detalle.cajas_pesadas, key=lambda item: (item.numero, item.id)), start=1):
        caja.numero = idx


def _user_can_manage_pedido(pedido):
    if not isinstance(current_user, Vendedor):
        return True
    if current_user.rol.nombre == 'super_admin':
        return True
    return current_user.puede_editar_pedido(pedido)


def _user_can_view_cliente(cliente_id):
    """True si el usuario actual puede ver este cliente (super_admin o usuario
    legacy ven todo; un Vendedor solo sus clientes asignados)."""
    if not isinstance(current_user, Vendedor):
        return True
    if current_user.rol.nombre == 'super_admin':
        return True
    return current_user.puede_ver_cliente(cliente_id)


def _is_ios_request():
    """True si la petición viene de un dispositivo iOS (iPhone/iPad/iPod).
    Cubre TODOS los navegadores de iOS (Safari, Chrome/CriOS, Firefox/FxiOS)
    porque Apple obliga a usar WebKit. iOS no descarga PDFs con
    Content-Disposition: attachment desde un visor sin controles, así que los
    PDFs de etiquetas se sirven 'inline' para que aparezcan en el visor nativo
    (con opción de compartir/guardar). Android/escritorio reciben 'attachment'."""
    ua = request.headers.get('User-Agent', '')
    return ('iPhone' in ua) or ('iPad' in ua) or ('iPod' in ua)


def _caja_pesada_to_label_item(caja):
    detalle = caja.detalle_pedido
    producto = detalle.producto if detalle else None
    return {
        'producto_nombre': producto.nombre if producto else 'N/A',
        'temperatura': getattr(producto, 'temperatura', None) or 'N/A',
        'peso_label': f'{Decimal(str(caja.peso or 0)).quantize(Decimal("0.01")):.2f} kg',
        'lote': caja.lote or 'N/A',
        'fecha_fabricacion': _date_to_iso(caja.fecha_elaboracion) or 'N/A',
        'fecha_expiracion': _date_to_iso(caja.fecha_vencimiento) or 'N/A',
    }


def _detalle_legacy_to_label_item(detalle):
    peso_float = float(detalle.peso or 0)
    if peso_float > 0:
        peso_label = f'{peso_float:.2f} kg'
    else:
        peso_label = f'{int(detalle.cajas or 0)} uds'

    return {
        'producto_nombre': detalle.producto.nombre if getattr(detalle, 'producto', None) else 'N/A',
        'temperatura': getattr(detalle.producto, 'temperatura', None) or 'N/A',
        'peso_label': peso_label,
        'lote': detalle.lote or 'N/A',
        'fecha_fabricacion': _date_to_iso(detalle.fecha_fabricacion) or 'N/A',
        'fecha_expiracion': _date_to_iso(detalle.fecha_expiracion) or 'N/A',
    }


def _build_label_items_for_pedido(pedido, fecha_ini, fecha_fin):
    inicio = _date_like_to_date(fecha_ini)
    fin = _date_like_to_date(fecha_fin)
    if not inicio or not fin:
        return []

    items = []
    productos_con_cajas = set()

    for detalle in _pedido_detalles_pesables(pedido):
        if not detalle.cajas_pesadas_count:
            continue
        productos_con_cajas.add(detalle.producto_id)
        for caja in detalle.cajas_pesadas:
            if not (inicio <= caja.fecha_elaboracion <= fin):
                continue
            items.append(_caja_pesada_to_label_item(caja))

    legacy_detalles = (
        DetallePedido.query
        .filter_by(pedido_id=pedido.id)
        .filter(DetallePedido.es_linea_pedido == False)
        .order_by(DetallePedido.id.asc())
        .all()
    )
    for detalle in legacy_detalles:
        if detalle.producto and detalle.producto.se_pesa and detalle.producto_id in productos_con_cajas:
            continue
        fecha_fab = _date_like_to_date(detalle.fecha_fabricacion)
        if not fecha_fab or not (inicio <= fecha_fab <= fin):
            continue
        items.append(_detalle_legacy_to_label_item(detalle))

    return items


def _validar_preparacion_pedido(pedido):
    errores = []
    prep_lines_por_producto = {}

    for detalle in pedido.detalles:
        if detalle.es_linea_pedido:
            continue
        prep_lines_por_producto.setdefault(detalle.producto_id, []).append(detalle)

    for detalle in [item for item in pedido.detalles if item.es_linea_pedido]:
        producto = detalle.producto
        if not producto:
            continue

        if producto.se_pesa and detalle.cajas_pesadas_count:
            if detalle.cajas_pesadas_count < detalle.cajas_objetivo:
                faltan = detalle.cajas_objetivo - detalle.cajas_pesadas_count
                errores.append(f'{producto.nombre}: faltan {faltan} cajas por pesar')

            for caja in detalle.cajas_pesadas:
                campos_faltantes = []
                if not caja.lote:
                    campos_faltantes.append('lote')
                if not caja.fecha_elaboracion:
                    campos_faltantes.append('fecha elaboración')
                if not caja.fecha_vencimiento:
                    campos_faltantes.append('fecha vencimiento')
                elif caja.fecha_vencimiento < caja.fecha_elaboracion:
                    campos_faltantes.append('vencimiento anterior a elaboración')
                if campos_faltantes:
                    errores.append(
                        f'{producto.nombre} caja #{caja.numero}: falta {", ".join(campos_faltantes)}'
                    )
            continue

        prep_lines = prep_lines_por_producto.get(detalle.producto_id, [])
        if not prep_lines:
            tipo = 'peso' if producto.se_pesa else 'cajas'
            errores.append(f'{producto.nombre}: sin líneas de preparación ({tipo})')
            continue

        if producto.se_pesa:
            for prep in prep_lines:
                campos_faltantes = []
                if not prep.lote:
                    campos_faltantes.append('lote')
                if not prep.fecha_fabricacion:
                    campos_faltantes.append('fecha fabricación')
                if not prep.fecha_expiracion:
                    campos_faltantes.append('fecha expiración')
                if campos_faltantes:
                    errores.append(f'{producto.nombre}: falta {", ".join(campos_faltantes)}')

    return errores


def _load_pedido_for_pesar(pedido_id):
    return (
        Pedido.query.options(
            joinedload(Pedido.cliente),
            selectinload(Pedido.detalles).joinedload(DetallePedido.producto),
            selectinload(Pedido.detalles).selectinload(DetallePedido.cajas_pesadas),
        )
        .filter_by(id=pedido_id)
        .first_or_404()
    )


def _get_active_pesable_detail(pedido, active_detalle_id=None):
    detalles = _pedido_detalles_pesables(pedido)
    if not detalles:
        return None

    if active_detalle_id:
        for detalle in detalles:
            if detalle.id == active_detalle_id:
                return detalle

    for detalle in detalles:
        if not detalle.pesaje_completo:
            return detalle

    return detalles[0]


def _build_pesar_context(pedido, active_detalle_id=None):
    detalles = _pedido_detalles_pesables(pedido)
    active_detalle = _get_active_pesable_detail(pedido, active_detalle_id=active_detalle_id)
    return {
        'pedido': pedido,
        'detalles_pesables': detalles,
        'active_detalle': active_detalle,
        'cajas_por_detalle': {detalle.id: _detalle_cajas_por_lote(detalle) for detalle in detalles},
        'peso_total_pedido': _pedido_peso_total(pedido),
        'cajas_pesadas_total': _pedido_cajas_pesadas_total(pedido),
        'cajas_objetivo_total': _pedido_cajas_objetivo_total(pedido),
        'puede_finalizar_pesar': _pedido_can_finalize_pesar(pedido),
    }


def _render_pesar_cajas_partial(pedido, detalle):
    return render_template(
        'partials/pesar_cajas_lista.html',
        pedido=pedido,
        detalle=detalle,
        active_detalle_id=detalle.id,
        grupos=_detalle_cajas_por_lote(detalle),
        peso_total_pedido=_pedido_peso_total(pedido),
        cajas_pesadas_total=_pedido_cajas_pesadas_total(pedido),
        cajas_objetivo_total=_pedido_cajas_objetivo_total(pedido),
        puede_finalizar_pesar=_pedido_can_finalize_pesar(pedido),
        oob_echoes=True,
    )


def _htmx_error_response(message, status=422):
    response = make_response(message, status)
    response.headers['HX-Retarget'] = '#pesar-feedback'
    response.headers['HX-Reswap'] = 'innerHTML'
    return response

############################################
# FUNCIONES Y RUTAS DE PEDIDOS
############################################

# --------------------------------------------------
# Helper: convierte un pedido y sus detalles en JSON
# --------------------------------------------------
# Actualizar la función pedido_a_json en app.py

def pedido_a_json(pedido: Pedido) -> dict:
    lineas = []
    total  = 0
    productos_con_cajas = set()

    for detalle in _pedido_detalles_pesables(pedido):
        if not detalle.cajas_pesadas_count:
            continue

        productos_con_cajas.add(detalle.producto_id)

        # Una línea del payload por cada CajaPesada: N8N agrupa por
        # (product_qbo_id, unit_price) y acumula cada qty individual en
        # descriptions[] para escribirlo en Line.Description de QBO.
        cajas_ordenadas = sorted(
            detalle.cajas_pesadas, key=lambda c: (c.numero, c.id)
        )
        for caja in cajas_ordenadas:
            qty = float(caja.peso or 0)
            if qty == 0:
                continue
            subtotal = float(detalle.precio_unitario) * qty
            total += subtotal
            lineas.append({
                "product_qbo_id": detalle.producto.qbo_id,
                "descripcion": detalle.producto.nombre,
                "qty": qty,
                "unit_price": float(detalle.precio_unitario),
                "amount": round(subtotal, 2),
                "tax_rate": detalle.producto.tax_rate,
            })

    # Productos que aún tienen línea original (lo que pidió el cliente). Una
    # línea de preparación sin su línea original es huérfana (el producto fue
    # eliminado del pedido) y NO debe facturarse.
    productos_con_linea_original = {
        d.producto_id for d in pedido.detalles if d.es_linea_pedido
    }

    for d in pedido.detalles:
        # Solo usar líneas de preparación (tanto manufactura como importación)
        if d.es_linea_pedido:
            continue
        if d.producto_id not in productos_con_linea_original:
            continue
        if d.producto and d.producto.se_pesa and d.producto_id in productos_con_cajas:
            continue

        # Omitir líneas con cantidad cero (producto no disponible)
        qty = float(d.cajas or d.peso or 0)
        if qty == 0:
            continue

        descripcion = d.producto.nombre
        if d.lote:
            descripcion += f" (Lote {d.lote})"

        subtotal = float(d.precio_unitario) * qty
        total   += subtotal

        lineas.append({
            "product_qbo_id": d.producto.qbo_id,
            "descripcion"   : descripcion,
            "qty"           : qty,
            "unit_price"    : float(d.precio_unitario),
            "amount"        : round(subtotal, 2),
            "tax_rate"      : d.producto.tax_rate
        })

    return {
        "order_id"        : pedido.id,
        "order_date"      : pedido.fecha_pedido.isoformat(),
        "customer_qbo_id" : pedido.cliente.qbo_id,
        "currency"        : pedido.cliente.moneda or 'XCG',
        "notes"           : pedido.notas,
        "lines"           : lineas,
        "total"           : round(total, 2)
    }


def _validar_datos_facturacion(payload):
    """Valida el payload que se enviará a QBO: cada línea necesita item y precio.

    Una línea sin product_qbo_id o con unit_price 0 produce una factura
    incorrecta en QuickBooks que hay que corregir a mano. Pasó el 2026-08-14
    con un producto nuevo creado sin precio en ninguna lista y con un qbo_id
    que no correspondía al item real.
    """
    errores = []
    vistos = set()

    def _agregar(mensaje):
        if mensaje not in vistos:
            vistos.add(mensaje)
            errores.append(mensaje)

    if not payload.get('customer_qbo_id'):
        _agregar('El cliente no tiene QBO ID configurado.')

    lineas = payload.get('lines') or []
    if not lineas:
        _agregar('El pedido no tiene líneas para facturar.')

    for linea in lineas:
        nombre = linea.get('descripcion') or 'Producto sin nombre'
        if not linea.get('product_qbo_id'):
            _agregar(f'{nombre}: el producto no tiene QBO ID configurado.')
        try:
            precio = float(linea.get('unit_price') or 0)
        except (TypeError, ValueError):
            precio = 0.0
        if precio <= 0:
            _agregar(f'{nombre}: precio en 0 — falta cargarlo en la lista de precios.')

    return errores


def _extraer_invoice_id(resp_data):
    """Devuelve (invoice_id, doc_number) desde la respuesta de N8N.

    El webhook está en modo 'Last Node', así que N8N responde con el objeto
    crudo de QuickBooks: {"Invoice": {"Id": "47349", "DocNumber": "5816"}}.
    Se acepta además la forma plana {"invoice_id": ...} por si el workflow
    vuelve a usar un nodo 'Respond to Webhook' propio.
    """
    if isinstance(resp_data, list):
        resp_data = resp_data[0] if resp_data else None
    if not isinstance(resp_data, dict):
        return None, None

    invoice = resp_data.get('Invoice')
    if isinstance(invoice, dict) and invoice.get('Id'):
        doc = invoice.get('DocNumber')
        return str(invoice['Id']), (str(doc) if doc else None)

    plano = resp_data.get('invoice_id')
    if plano:
        doc = resp_data.get('doc_number')
        return str(plano), (str(doc) if doc else None)

    return None, None


def obtener_ip_servidor():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('8.8.8.8', 80))
        ip_servidor = s.getsockname()[0]
    except Exception:
        ip_servidor = '127.0.0.1'
    finally:
        s.close()
    return ip_servidor

@app.route('/')
@login_required
def index():
    # Página de inicio tras login: listado de pedidos.
    # Si el usuario no tiene permiso para ver pedidos, cae al dashboard
    # (evita un bucle de redirección con el decorador de lista_pedidos).
    if isinstance(current_user, Vendedor) and not current_user.tiene_permiso('pedidos', 'leer'):
        return redirect(url_for('dashboard'))
    return redirect(url_for('lista_pedidos'))


@app.route('/home')
@login_required
def home():
    """Ruta de inicio unificada tras login.
    Ahora redirige a '/'.
    """
    return redirect(url_for('index'))


def _es_ultimo_super_admin(vendedor):
    """True si 'vendedor' es super_admin activo y es el único super_admin activo."""
    if not vendedor.rol or vendedor.rol.nombre != 'super_admin' or not vendedor.activo:
        return False
    activos = (Vendedor.query.join(Rol)
               .filter(Rol.nombre == 'super_admin', Vendedor.activo.is_(True)).count())
    return activos <= 1


def _generar_password_temporal():
    """Genera una contraseña temporal legible y aleatoria."""
    return secrets.token_urlsafe(8)


@app.route('/admin/vendedores')
@login_required
@requiere_rol(['super_admin'])
def gestionar_vendedores():
    vendedores = Vendedor.query.all()
    roles = Rol.query.filter_by(activo=True).all()
    territorios = Territorio.query.filter_by(activo=True).all()

    # Matriz de permisos (mismos roles gestionables que gestionar_permisos).
    acciones = ['leer', 'crear', 'editar', 'eliminar']
    roles_matriz = (Rol.query.filter(Rol.nombre.in_(['supervisor', 'vendedor']))
                    .order_by(Rol.nombre).all())
    permisos = {p.recurso: p for p in Permiso.query.all()}
    matriz = {}
    for rol in roles_matriz:
        matriz[rol.id] = {}
        for rec in PERMISOS_RECURSOS:
            p = permisos.get(rec)
            rp = RolPermiso.query.filter_by(rol_id=rol.id, permiso_id=p.id).first() if p else None
            matriz[rol.id][rec] = {
                a: (getattr(rp, f'puede_{a}') if rp else _permiso_default(rol.nombre, rec, a))
                for a in acciones
            }

    # Tiempo relativo de último acceso + actividad reciente (datos reales).
    def _hace(dt):
        if not dt:
            return None
        seg = (datetime.utcnow() - dt).total_seconds()
        if seg < 60:
            return 'Hace instantes'
        if seg < 3600:
            return f'Hace {int(seg // 60)} min'
        if seg < 86400:
            return f'Hace {int(seg // 3600)} h'
        dias = int(seg // 86400)
        return 'Ayer' if dias == 1 else f'Hace {dias} días'

    ultimo_rel = {v.id: _hace(v.ultimo_login) for v in vendedores}

    # Auditoría real: últimos eventos registrados (con fallback a últimos accesos).
    eventos = EventoAuditoria.query.order_by(EventoAuditoria.creado_en.desc()).limit(15).all()
    actividad = [{
        'actor': e.actor or (e.vendedor.nombre_completo if e.vendedor else 'Sistema'),
        'accion': e.accion, 'detalle': e.detalle, 'tipo': e.tipo,
        'hora': _hace(e.creado_en),
    } for e in eventos]
    if not actividad:
        actividad = [{
            'actor': v.nombre_completo or v.username, 'accion': 'Último acceso',
            'detalle': (v.rol.nombre.replace('_', ' ').title() if v.rol else None),
            'tipo': 'auth', 'hora': _hace(v.ultimo_login),
        } for v in sorted([x for x in vendedores if x.ultimo_login],
                          key=lambda x: x.ultimo_login, reverse=True)[:8]]

    return render_template('admin/vendedores.html',
                           vendedores=vendedores,
                           roles=roles,
                           territorios=territorios,
                           roles_matriz=roles_matriz,
                           recursos=PERMISOS_RECURSOS,
                           acciones=acciones,
                           matriz=matriz,
                           ultimo_rel=ultimo_rel,
                           actividad=actividad)

@app.route('/admin/roles-permisos', methods=['GET', 'POST'])
@login_required
@requiere_rol(['super_admin'])
def gestionar_permisos():
    acciones = ['leer', 'crear', 'editar', 'eliminar']
    roles = (Rol.query.filter(Rol.nombre.in_(['supervisor', 'vendedor']))
             .order_by(Rol.nombre).all())
    if request.method == 'POST':
        _sembrar_permisos()  # garantiza filas Permiso
        permisos = {p.recurso: p for p in Permiso.query.all()}
        for rol in roles:
            for rec in PERMISOS_RECURSOS:
                p = permisos.get(rec)
                if p is None:
                    continue
                rp = RolPermiso.query.filter_by(rol_id=rol.id, permiso_id=p.id).first()
                if rp is None:
                    rp = RolPermiso(rol_id=rol.id, permiso_id=p.id)
                    db.session.add(rp)
                rp.puede_leer = bool(request.form.get(f'perm_{rol.id}_{rec}_leer'))
                rp.puede_crear = bool(request.form.get(f'perm_{rol.id}_{rec}_crear'))
                rp.puede_editar = bool(request.form.get(f'perm_{rol.id}_{rec}_editar'))
                rp.puede_eliminar = bool(request.form.get(f'perm_{rol.id}_{rec}_eliminar'))
        db.session.commit()
        flash('Permisos actualizados.', 'success')
        return redirect(url_for('gestionar_permisos'))

    permisos = {p.recurso: p for p in Permiso.query.all()}
    matriz = {}
    for rol in roles:
        matriz[rol.id] = {}
        for rec in PERMISOS_RECURSOS:
            p = permisos.get(rec)
            rp = RolPermiso.query.filter_by(rol_id=rol.id, permiso_id=p.id).first() if p else None
            matriz[rol.id][rec] = {
                a: (getattr(rp, f'puede_{a}') if rp else _permiso_default(rol.nombre, rec, a))
                for a in acciones
            }
    return render_template('admin/roles_permisos.html',
                           roles=roles, recursos=PERMISOS_RECURSOS,
                           acciones=acciones, matriz=matriz)

@app.route('/admin/vendedores/nuevo', methods=['GET', 'POST'])
@login_required
@requiere_rol(['super_admin'])
def crear_vendedor():
    """Crear nuevo vendedor - maneja GET (mostrar formulario) y POST (procesar datos)"""
    
    # ====== MANEJAR GET REQUEST (mostrar formulario) ======
    if request.method == 'GET':
        try:
            # Obtener datos necesarios para el formulario
            roles = Rol.query.filter_by(activo=True).all()
            territorios = Territorio.query.filter_by(activo=True).all()
            vendedores = Vendedor.query.filter_by(activo=True).all()  # Para supervisor dropdown
            
            # Si tienes un template específico, úsalo:
            # return render_template('admin/vendedor_form.html', 
            #                      roles=roles, 
            #                      territorios=territorios,
            #                      vendedores=vendedores)
            
            # OPCIÓN 1: Redirigir a gestión con mensaje informativo
            flash('Utilice el formulario en la página de gestión de vendedores para crear un nuevo vendedor', 'info')
            return redirect(url_for('gestionar_vendedores'))
            
            # OPCIÓN 2: Si quieres mostrar un formulario simple aquí
            # return f"""
            # <h1>Crear Nuevo Vendedor</h1>
            # <form method="POST">
            #     <p>Username: <input type="text" name="username" required></p>
            #     <p>Email: <input type="email" name="email" required></p>
            #     <p>Nombre: <input type="text" name="nombre_completo" required></p>
            #     <p>Teléfono: <input type="text" name="telefono"></p>
            #     <p>Password: <input type="password" name="password" required></p>
            #     <p>Rol: <select name="rol_id" required>
            #         {''.join([f'<option value="{r.id}">{r.nombre}</option>' for r in roles])}
            #     </select></p>
            #     <p>Territorio: <select name="territorio_id">
            #         <option value="">Sin territorio</option>
            #         {''.join([f'<option value="{t.id}">{t.nombre}</option>' for t in territorios])}
            #     </select></p>
            #     <p><button type="submit">Crear Vendedor</button></p>
            # </form>
            # <a href="{url_for('gestionar_vendedores')}">← Volver</a>
            # """
            
        except Exception as e:
            app.logger.error(f'Error al cargar el formulario: {e}')
            flash('Error al cargar el formulario. Intente de nuevo.', 'error')
            return redirect(url_for('gestionar_vendedores'))
    
    # ====== MANEJAR POST REQUEST (procesar formulario) ======
    elif request.method == 'POST':
        try:
            # Obtener datos del formulario
            username = request.form.get('username')
            email = request.form.get('email')
            nombre_completo = request.form.get('nombre_completo')
            telefono = request.form.get('telefono')
            password = request.form.get('password')
            rol_id = request.form.get('rol_id')
            territorio_id = request.form.get('territorio_id')
            
            # Validaciones básicas
            if not username or not email or not nombre_completo or not password or not rol_id:
                flash("Todos los campos obligatorios deben ser completados", "error")
                return redirect(url_for('crear_vendedor'))
            
            # Verificar que el username no exista
            vendedor_existente = Vendedor.query.filter_by(username=username).first()
            if vendedor_existente:
                flash("El nombre de usuario ya existe", "error")
                return redirect(url_for('crear_vendedor'))
            
            # Verificar que el email no exista
            email_existente = Vendedor.query.filter_by(email=email).first()
            if email_existente:
                flash("El email ya está registrado", "error")
                return redirect(url_for('crear_vendedor'))
            
            # Verificar que el rol existe
            rol_valido = db.session.get(Rol, rol_id)
            if not rol_valido:
                flash("El rol seleccionado no es válido", "error")
                return redirect(url_for('crear_vendedor'))
            
            # Verificar territorio si se proporcionó
            if territorio_id:
                territorio_valido = db.session.get(Territorio, territorio_id)
                if not territorio_valido:
                    flash("El territorio seleccionado no es válido", "error")
                    return redirect(url_for('crear_vendedor'))
            
            # Crear nuevo vendedor
            nuevo_vendedor = Vendedor(
                username=username,
                email=email,
                nombre_completo=nombre_completo,
                telefono=telefono,
                rol_id=int(rol_id),
                territorio_id=int(territorio_id) if territorio_id else None,
                fecha_ingreso=date.today(),
                activo=True,
                debe_cambiar_password=True
            )
            
            # Establecer password hasheado
            nuevo_vendedor.set_password(password)
            
            # Guardar en base de datos
            db.session.add(nuevo_vendedor)
            db.session.commit()
            
            flash(f'Vendedor {nombre_completo} creado exitosamente', 'success')
            return redirect(url_for('gestionar_vendedores'))
            
        except ValueError as ve:
            db.session.rollback()
            app.logger.warning(f'Error de validación al crear vendedor: {ve}')
            flash('Error en los datos proporcionados. Verifique los campos.', 'error')
            return redirect(url_for('crear_vendedor'))
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Error al crear vendedor: {e}')
            flash('Error al crear vendedor. Intente de nuevo.', 'error')
            return redirect(url_for('crear_vendedor'))
    
    # ====== FALLBACK (no debería llegar aquí) ======
    else:
        flash("Método no permitido", "error")
        return redirect(url_for('gestionar_vendedores'))

@app.route('/admin/asignar_cliente', methods=['GET', 'POST'])
@login_required
@requiere_rol(['super_admin'])
def asignar_cliente():
    
    # Obtener vendedores activos
    vendedores = (Vendedor.query
                  .filter_by(activo=True)
                  .order_by(Vendedor.nombre_completo)
                  .all())
    
    # Obtener clientes SIN ASIGNAR
    clientes_sin_asignar = db.session.query(Cliente).outerjoin(
        ClienteVendedor,
        db.and_(
            Cliente.id == ClienteVendedor.cliente_id,
            ClienteVendedor.activo == True
        )
    ).filter(
        ClienteVendedor.id.is_(None)
    ).order_by(Cliente.nombre).all()

    if request.method == 'POST':
        try:
            ClienteVendedor.asignar(
                cliente_id=int(request.form['cliente_id']),
                vendedor_id=int(request.form['vendedor_id'])
            )
            flash('Cliente asignado correctamente', 'success')
            return redirect(url_for('asignar_cliente'))
        except Exception as e:
            app.logger.error(f'Error al asignar cliente: {e}')
            flash('Error al asignar cliente. Intente de nuevo.', 'error')
    
    return render_template('admin/clientes_vendedores.html',
                           clientes_sin_asignar=clientes_sin_asignar,  # ← Variable correcta
                           vendedores=vendedores)



@app.route('/api/clientes/sin-asignar')
@login_required
@requiere_rol(['super_admin'])
def api_clientes_sin_asignar():
    """
    Devuelve los clientes que no tienen asignación activa
    """
    try:
        # Obtener todos los clientes
        todos_los_clientes = Cliente.query.all()

        # Obtener IDs de clientes que SÍ tienen asignación activa
        clientes_asignados_ids = db.session.query(ClienteVendedor.cliente_id).filter_by(activo=True).distinct().all()
        clientes_asignados_ids = [row[0] for row in clientes_asignados_ids]

        # Filtrar clientes que NO están en la lista de asignados
        clientes_sin_asignar = [c for c in todos_los_clientes if c.id not in clientes_asignados_ids]

        resultado = [{'id': c.id, 'nombre': c.nombre} for c in clientes_sin_asignar]
        return jsonify(resultado)

    except Exception as e:
        app.logger.error(f"Error en api_clientes_sin_asignar: {e}", exc_info=True)
        return jsonify([]), 200  



# ---------- Asignar ----------
@app.route('/api/asignaciones', methods=['POST'])
@login_required
@requiere_rol(['super_admin'])
def api_asignar_cliente():
    data = request.get_json(force=True)
    try:
        asign = ClienteVendedor.asignar(
            cliente_id  = int(data['cliente_id']),
            vendedor_id = int(data['vendedor_id'])
        )
        return jsonify({'success': True, 'asign_id': asign.id})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error en operación: {e}')
        return jsonify({'success': False, 'error': 'Error interno del servidor'}), 400


# ---------- Desasignar ----------
@app.route('/api/asignaciones/<int:asign_id>', methods=['DELETE'])
@login_required
@requiere_rol(['super_admin'])
def api_desasignar_cliente(asign_id):
    try:
        ClienteVendedor.desasignar(asign_id)
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error en operación: {e}')
        return jsonify({'success': False, 'error': 'Error interno del servidor'}), 400

@app.route('/api/vendedores/<int:v_id>/clientes')
@login_required
@requiere_rol(['super_admin'])
def api_clientes_del_vendedor(v_id):
    """
    Devuelve (JSON) los clientes activos asignados a un vendedor.
    """
    try:
        # Verificar que el vendedor existe
        vendedor = db.session.get(Vendedor, v_id)
        if not vendedor:
            return jsonify({'error': 'Vendedor no encontrado'}), 404

        # Obtener asignaciones activas del vendedor con información del cliente
        asignaciones = db.session.query(
            ClienteVendedor, Cliente
        ).join(
            Cliente, ClienteVendedor.cliente_id == Cliente.id
        ).filter(
            ClienteVendedor.vendedor_id == v_id,
            ClienteVendedor.activo == True
        ).order_by(Cliente.nombre).all()

        # Formatear respuesta - CORREGIDO: usar fecha_inicio en lugar de fecha_asignacion
        resultado = []
        for asignacion, cliente in asignaciones:
            resultado.append({
                'id': cliente.id,
                'nombre': cliente.nombre,
                'asign_id': asignacion.id,
                'fecha_asignacion': asignacion.fecha_inicio.strftime('%Y-%m-%d') if asignacion.fecha_inicio else None
            })

        return jsonify(resultado)

    except Exception as e:
        app.logger.error(f"Error en api_clientes_del_vendedor: {e}", exc_info=True)
        return jsonify({'error': 'Error interno del servidor'}), 500

        # ===== RUTAS ADMINISTRATIVAS ADICIONALES =====

@app.route('/admin/reportes')
@login_required
@requiere_rol(['super_admin'])
def admin_reportes():
    """Página de reportes avanzados para administradores"""
    try:
        # Obtener datos para reportes
        hoy = datetime.now().date()
        inicio_mes = hoy.replace(day=1)
        hace_30_dias = hoy - timedelta(days=30)
        
        # Reporte de ventas por vendedor
        ventas_por_vendedor = db.session.query(
            Vendedor.nombre_completo,
            Vendedor.username,
            func.count(Pedido.id).label('total_pedidos'),
            func.coalesce(func.sum(DetallePedido.subtotal), 0).label('ventas_total')
        ).outerjoin(
            ClienteVendedor, Vendedor.id == ClienteVendedor.vendedor_id
        ).outerjoin(
            Pedido, ClienteVendedor.cliente_id == Pedido.cliente_id
        ).outerjoin(
            DetallePedido, Pedido.id == DetallePedido.pedido_id
        ).filter(
            Vendedor.activo == True,
            Pedido.fecha_pedido >= hace_30_dias,
            db.or_(DetallePedido.es_linea_pedido == True, DetallePedido.id.is_(None))
        ).group_by(
            Vendedor.id, Vendedor.nombre_completo, Vendedor.username
        ).order_by(
            func.coalesce(func.sum(DetallePedido.subtotal), 0).desc()
        ).all()
        
        # Reporte de productos más vendidos
        productos_mas_vendidos = db.session.query(
            Producto.nombre,
            func.sum(DetallePedido.cajas).label('total_cajas'),
            func.coalesce(func.sum(DetallePedido.subtotal), 0).label('ingresos_total')
        ).join(
            DetallePedido, Producto.id == DetallePedido.producto_id
        ).join(
            Pedido, DetallePedido.pedido_id == Pedido.id
        ).filter(
            Pedido.fecha_pedido >= hace_30_dias,
            DetallePedido.es_linea_pedido == True
        ).group_by(
            Producto.id, Producto.nombre
        ).order_by(
            func.sum(DetallePedido.cajas).desc()
        ).limit(10).all()
        
        # Reporte de clientes más activos
        clientes_mas_activos = db.session.query(
            Cliente.nombre,
            func.count(Pedido.id).label('total_pedidos'),
            func.coalesce(func.sum(DetallePedido.subtotal), 0).label('compras_total')
        ).join(
            Pedido, Cliente.id == Pedido.cliente_id
        ).join(
            DetallePedido, Pedido.id == DetallePedido.pedido_id
        ).filter(
            Pedido.fecha_pedido >= hace_30_dias,
            DetallePedido.es_linea_pedido == True
        ).group_by(
            Cliente.id, Cliente.nombre
        ).order_by(
            func.coalesce(func.sum(DetallePedido.subtotal), 0).desc()
        ).limit(10).all()
        
        return render_template('admin/reportes.html',
                             ventas_por_vendedor=ventas_por_vendedor,
                             productos_mas_vendidos=productos_mas_vendidos,
                             clientes_mas_activos=clientes_mas_activos,
                             fecha_inicio=hace_30_dias,
                             fecha_fin=hoy)
        
    except Exception as e:
        app.logger.error(f"Error en admin_reportes: {e}")
        flash('Error al cargar los reportes', 'error')
        return redirect(url_for('index'))

@app.route('/admin/analytics')
@login_required
@requiere_rol(['super_admin'])
def admin_analytics():
    """Página de analytics avanzados"""
    try:
        # Datos para gráficos y análisis
        hoy = datetime.now().date()
        hace_90_dias = hoy - timedelta(days=90)
        
        # Tendencia de ventas últimos 90 días (por semanas)
        # Usa líneas facturadas reales del pedido (sin reconversión por tipo_cambio)
        def _venta_pedido_a(pedido):
            return _calcular_venta_pedido(pedido)

        tendencia_ventas = []
        for i in range(12):  # 12 semanas
            inicio_semana = hace_90_dias + timedelta(weeks=i)
            fin_semana = inicio_semana + timedelta(days=6)

            pedidos_sem = Pedido.query.filter(
                Pedido.fecha_pedido >= inicio_semana,
                Pedido.fecha_pedido <= fin_semana
            ).all()
            ventas_semana = sum(_venta_pedido_a(p) for p in pedidos_sem)

            tendencia_ventas.append({
                'semana': inicio_semana.strftime('%d/%m'),
                'ventas': float(ventas_semana or 0)
            })
        
        # Distribución de pedidos por estado
        estados_pedidos = db.session.query(
            Pedido.estado,
            func.count(Pedido.id).label('cantidad')
        ).group_by(Pedido.estado).all()
        
        # Eficiencia por vendedor
        eficiencia_vendedores = db.session.query(
            Vendedor.nombre_completo,
            func.count(Pedido.id).label('pedidos_totales'),
            func.sum(
                db.case([(Pedido.estado == 'facturado', 1)], else_=0)
            ).label('pedidos_facturados')
        ).outerjoin(
            ClienteVendedor, Vendedor.id == ClienteVendedor.vendedor_id
        ).outerjoin(
            Pedido, ClienteVendedor.cliente_id == Pedido.cliente_id
        ).filter(
            Vendedor.activo == True
        ).group_by(
            Vendedor.id, Vendedor.nombre_completo
        ).having(
            func.count(Pedido.id) > 0
        ).all()
        
        return render_template('admin/analytics.html',
                             tendencia_ventas=tendencia_ventas,
                             estados_pedidos=estados_pedidos,
                             eficiencia_vendedores=eficiencia_vendedores)
        
    except Exception as e:
        app.logger.error(f"Error en admin_analytics: {e}")
        flash('Error al cargar analytics', 'error')
        return redirect(url_for('index'))

@app.route('/admin/configuracion')
@login_required
@requiere_rol(['super_admin'])
def admin_configuracion():
    """Página de configuración del sistema"""
    return render_template('admin/configuracion.html')

@app.route('/admin/clientes-vendedores')
@login_required
@requiere_rol(['super_admin'])
def admin_clientes_vendedores():
    """Gestión de asignaciones cliente-vendedor"""
    try:
        # Obtener todas las asignaciones activas
        asignaciones = db.session.query(
            ClienteVendedor,
            Cliente.nombre.label('cliente_nombre'),
            Vendedor.nombre_completo.label('vendedor_nombre')
        ).join(
            Cliente, ClienteVendedor.cliente_id == Cliente.id
        ).join(
            Vendedor, ClienteVendedor.vendedor_id == Vendedor.id
        ).filter(
            ClienteVendedor.activo == True
        ).all()
        
        # Clientes sin asignar
        clientes_sin_asignar = db.session.query(Cliente).outerjoin(
            ClienteVendedor,
            db.and_(
                Cliente.id == ClienteVendedor.cliente_id,
                ClienteVendedor.activo == True
            )
        ).filter(
            ClienteVendedor.id.is_(None)
        ).all()
        
        # Vendedores activos
        vendedores = Vendedor.query.filter_by(activo=True).all()
        
        return render_template('admin/clientes_vendedores.html',
                             asignaciones=asignaciones,
                             clientes_sin_asignar=clientes_sin_asignar,
                             vendedores=vendedores)
        
    except Exception as e:
        app.logger.error(f"Error en admin_clientes_vendedores: {e}")
        flash('Error al cargar asignaciones', 'error')
        return redirect(url_for('index'))

@app.route('/admin/exportar')
@login_required
@requiere_rol(['super_admin'])
def admin_exportar():
    """Página de exportación de datos"""
    return render_template('admin/exportar.html')

@app.route('/admin/exportar/ventas')
@login_required
@requiere_rol(['super_admin'])
def exportar_ventas():
    """Exportar datos de ventas a Excel"""
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            flash('Debe especificar fechas de inicio y fin', 'error')
            return redirect(url_for('admin_exportar'))
        
        inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        
        # Consulta de ventas detalladas
        ventas_detalle = db.session.query(
            Pedido.id.label('pedido_id'),
            Pedido.fecha_pedido,
            Cliente.nombre.label('cliente'),
            Producto.nombre.label('producto'),
            DetallePedido.cajas,
            DetallePedido.precio_unitario,
            DetallePedido.subtotal,
            Vendedor.nombre_completo.label('vendedor')
        ).join(
            Cliente, Pedido.cliente_id == Cliente.id
        ).join(
            DetallePedido, Pedido.id == DetallePedido.pedido_id
        ).join(
            Producto, DetallePedido.producto_id == Producto.id
        ).outerjoin(
            ClienteVendedor, Cliente.id == ClienteVendedor.cliente_id
        ).outerjoin(
            Vendedor, ClienteVendedor.vendedor_id == Vendedor.id
        ).filter(
            Pedido.fecha_pedido >= inicio,
            Pedido.fecha_pedido <= fin
        ).order_by(
            Pedido.fecha_pedido.desc()
        ).all()
        
        # Crear Excel
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'strings_to_formulas': False, 'strings_to_urls': False})
        worksheet = workbook.add_worksheet('Reporte de Ventas')
        
        # Formatos
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#3498db',
            'color': 'white',
            'align': 'center'
        })
        
        date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
        money_format = workbook.add_format({'num_format': '$#,##0.00'})
        
        # Encabezados
        headers = ['Pedido ID', 'Fecha', 'Cliente', 'Producto', 'Cajas', 
                  'Precio Unitario', 'Subtotal', 'Vendedor']
        
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        # Datos
        total_ventas = 0
        for row, venta in enumerate(ventas_detalle, 1):
            worksheet.write(row, 0, venta.pedido_id)
            worksheet.write(row, 1, venta.fecha_pedido, date_format)
            worksheet.write(row, 2, venta.cliente)
            worksheet.write(row, 3, venta.producto)
            worksheet.write(row, 4, venta.cajas or 0)
            worksheet.write(row, 5, float(venta.precio_unitario or 0), money_format)
            worksheet.write(row, 6, float(venta.subtotal or 0), money_format)
            worksheet.write(row, 7, venta.vendedor or 'Sin asignar')
            
            total_ventas += float(venta.subtotal or 0)
        
        # Total
        last_row = len(ventas_detalle) + 1
        worksheet.write(last_row, 5, 'TOTAL:', header_format)
        worksheet.write(last_row, 6, total_ventas, money_format)
        
        workbook.close()
        output.seek(0)
        
        filename = f"ventas_{fecha_inicio}_a_{fecha_fin}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        app.logger.error(f"Error exportando ventas: {e}")
        flash('Error al exportar datos de ventas', 'error')
        return redirect(url_for('admin_exportar'))

@app.route('/admin/backup')
@login_required
@requiere_rol(['super_admin'])
def admin_backup():
    """Página de gestión de backups"""
    return render_template('admin/backup.html')

@app.route('/admin/logs')
@login_required
@requiere_rol(['super_admin'])
def admin_logs():
    """Página de visualización de logs del sistema"""
    try:
        # Logs simulados - en producción conectarías con archivos de log reales
        logs_ejemplo = [
            {
                'timestamp': datetime.now() - timedelta(minutes=5),
                'nivel': 'INFO',
                'mensaje': 'Usuario admin inició sesión',
                'usuario': current_user.username
            },
            {
                'timestamp': datetime.now() - timedelta(hours=1),
                'nivel': 'SUCCESS',
                'mensaje': 'Pedido #123 facturado correctamente',
                'usuario': 'sistema'
            },
            {
                'timestamp': datetime.now() - timedelta(hours=2),
                'nivel': 'WARNING',
                'mensaje': 'Intento de acceso con credenciales incorrectas',
                'usuario': 'desconocido'
            }
        ]
        
        return render_template('admin/logs.html', logs=logs_ejemplo)
        
    except Exception as e:
        app.logger.error(f"Error en admin_logs: {e}")
        flash('Error al cargar logs', 'error')
        return redirect(url_for('index'))

@app.route('/admin/roles')
@login_required
@requiere_rol(['super_admin'])
def admin_roles():
    """Gestión de roles y permisos"""
    try:
        roles = Rol.query.filter_by(activo=True).all()
        permisos = Permiso.query.all()
        
        return render_template('admin/roles.html', roles=roles, permisos=permisos)
        
    except Exception as e:
        app.logger.error(f"Error en admin_roles: {e}")
        flash('Error al cargar roles', 'error')
        return redirect(url_for('index'))

@app.route('/admin/roles/nuevo', methods=['GET', 'POST'])
@login_required
@requiere_rol(['super_admin'])
def crear_rol():
    """Crear nuevo rol"""
    if request.method == 'POST':
        try:
            nombre = request.form['nombre']
            descripcion = request.form.get('descripcion', '')
            nivel_jerarquia = int(request.form.get('nivel_jerarquia', 0))
            
            nuevo_rol = Rol(
                nombre=nombre,
                descripcion=descripcion,
                nivel_jerarquia=nivel_jerarquia
            )
            
            db.session.add(nuevo_rol)
            db.session.commit()
            
            flash(f'Rol {nombre} creado exitosamente', 'success')
            return redirect(url_for('admin_roles'))
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Error al crear rol: {e}')
            flash('Error al crear rol. Intente de nuevo.', 'error')

    return render_template('admin/rol_form.html')

# ===== API ENDPOINTS PARA DASHBOARD =====

@app.route('/api/admin/stats')
@login_required
@requiere_rol(['super_admin'])
def api_admin_stats():
    """API para obtener estadísticas en tiempo real para el dashboard admin"""
    try:
        hoy = datetime.now(DASHBOARD_TIMEZONE).date()
        inicio_mes = hoy.replace(day=1)
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        inicio_mes_anterior = (inicio_mes - timedelta(days=1)).replace(day=1)
        fin_mes_anterior = inicio_mes - timedelta(days=1)
        
        stats = {
            'vendedores_activos': Vendedor.query.filter_by(activo=True).count(),
            'clientes_totales': Cliente.query.count(),
            'pedidos_hoy': Pedido.query.filter(
                Pedido.fecha_pedido >= hoy
            ).count(),
            'pedidos_pendientes': Pedido.query.filter_by(estado='pendiente').count(),
            'ventas_mes': 0
        }

        metricas_qb = _obtener_metricas_ventas_quickbooks(
            hoy=hoy,
            inicio_mes=inicio_mes,
            inicio_semana=inicio_semana,
            inicio_mes_anterior=inicio_mes_anterior,
            fin_mes_anterior=fin_mes_anterior,
            inicio_tendencia=inicio_semana - timedelta(weeks=8),
            inicio_ultimos_7_dias=hoy - timedelta(days=6),
        )
        if metricas_qb:
            stats['ventas_mes'] = metricas_qb['ventas_mes']
        else:
            # Fallback local cuando QuickBooks no está disponible
            pedidos_facturados_data = Pedido.query.filter(
                Pedido.estado == 'facturado',
                Pedido.fecha_facturacion.isnot(None)
            ).all()
            stats['ventas_mes'] = sum(
                _calcular_venta_pedido(p)
                for p in pedidos_facturados_data
                if _pedido_facturado_en_periodo_local(p, inicio_mes)
            )
        
        return jsonify(stats)
        
    except Exception as e:
        app.logger.error(f"Error en API admin stats: {e}")
        return jsonify({'error': 'Error interno'}), 500

@app.route('/api/admin/performance')
@login_required
@requiere_rol(['super_admin'])
def api_admin_performance():
    """API para obtener datos de performance de vendedores"""
    try:
        hace_30_dias = datetime.now().date() - timedelta(days=30)
        
        performance = db.session.query(
            Vendedor.nombre_completo,
            Vendedor.username,
            func.count(Pedido.id).label('pedidos'),
            func.coalesce(func.sum(DetallePedido.subtotal), 0).label('ventas')
        ).outerjoin(
            ClienteVendedor, Vendedor.id == ClienteVendedor.vendedor_id
        ).outerjoin(
            Pedido, ClienteVendedor.cliente_id == Pedido.cliente_id
        ).outerjoin(
            DetallePedido, Pedido.id == DetallePedido.pedido_id
        ).filter(
            Vendedor.activo == True,
            Pedido.fecha_pedido >= hace_30_dias,
            db.or_(DetallePedido.es_linea_pedido == True, DetallePedido.id.is_(None))
        ).group_by(
            Vendedor.id, Vendedor.nombre_completo, Vendedor.username
        ).order_by(
            func.coalesce(func.sum(DetallePedido.subtotal), 0).desc()
        ).limit(5).all()
        
        resultado = []
        for p in performance:
            resultado.append({
                'vendedor': p.nombre_completo,
                'username': p.username,
                'pedidos': p.pedidos,
                'ventas': float(p.ventas or 0)
            })
        
        return jsonify(resultado)
        
    except Exception as e:
        app.logger.error(f"Error en API performance: {e}")
        return jsonify({'error': 'Error interno'}), 500

# ===== RUTAS DE GESTIÓN DE TERRITORIOS =====

@app.route('/admin/territorios')
@login_required
@requiere_rol(['super_admin'])
def admin_territorios():
    """Gestión de territorios"""
    territorios = Territorio.query.filter_by(activo=True).all()
    return render_template('admin/territorios.html', territorios=territorios)

@app.route('/admin/territorios/nuevo', methods=['GET', 'POST'])
@login_required
@requiere_rol(['super_admin'])
def crear_territorio():
    """Crear nuevo territorio"""
    if request.method == 'POST':
        try:
            nombre = request.form['nombre']
            descripcion = request.form.get('descripcion', '')
            tipo = request.form.get('tipo', 'geografico')
            
            nuevo_territorio = Territorio(
                nombre=nombre,
                descripcion=descripcion,
                tipo=tipo
            )
            
            db.session.add(nuevo_territorio)
            db.session.commit()
            
            flash(f'Territorio {nombre} creado exitosamente', 'success')
            return redirect(url_for('admin_territorios'))
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Error al crear territorio: {e}')
            flash('Error al crear territorio. Intente de nuevo.', 'error')

    return render_template('admin/territorio_form.html')

@app.route('/admin/territorios/asignar_cliente', methods=['POST'])
@login_required
@requiere_rol(['super_admin'])
def asignar_cliente_territorio():
    try:
        cliente_id    = int(request.form['cliente_id'])
        territorio_id = int(request.form['territorio_id'])

        cliente = Cliente.query.get_or_404(cliente_id)
        cliente.territorio_id = territorio_id
        db.session.commit()

        flash('Territorio asignado al cliente', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al asignar territorio al cliente: {e}')
        flash('Error al asignar territorio. Intente de nuevo.', 'danger')
    return redirect(url_for('admin_clientes_vendedores'))

@app.route('/admin/vendedores/<int:v_id>/territorio', methods=['POST'])
@login_required
@requiere_rol(['super_admin'])
def actualizar_territorio_vendedor(v_id):
    """
    Cambia el territorio de un vendedor.
    Se recibe territorio_id desde el formulario.
    """
    try:
        territorio_id = request.form.get('territorio_id') or None  # '' → None
        vendedor = Vendedor.query.get_or_404(v_id)

        # actualizar (permite dejarlo sin territorio)
        vendedor.territorio_id = int(territorio_id) if territorio_id else None
        db.session.commit()

        flash('Territorio del vendedor actualizado', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al actualizar territorio del vendedor: {e}')
        flash('Error al actualizar territorio. Intente de nuevo.', 'danger')

    return redirect(url_for('gestionar_vendedores'))


@app.route('/admin/vendedores/<int:v_id>/toggle', methods=['POST'])
@login_required
@requiere_rol(['super_admin'])
def toggle_vendedor(v_id):
    v = Vendedor.query.get_or_404(v_id)
    if v.id == current_user.id:
        flash('No podés desactivar tu propia cuenta.', 'danger')
        return redirect(url_for('gestionar_vendedores'))
    if v.activo and _es_ultimo_super_admin(v):
        flash('No se puede desactivar al único super_admin activo.', 'danger')
        return redirect(url_for('gestionar_vendedores'))
    try:
        v.activo = not v.activo
        db.session.commit()
        _audit('user', f"{'Activó' if v.activo else 'Desactivó'} usuario", v.nombre_completo or v.username)
        flash(f"Usuario {'activado' if v.activo else 'desactivado'}.", 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al activar/desactivar vendedor {v_id}: {e}')
        flash('Error al actualizar el usuario. Intente de nuevo.', 'danger')
    return redirect(url_for('gestionar_vendedores'))


@app.route('/admin/vendedores/<int:v_id>/editar', methods=['POST'])
@login_required
@requiere_rol(['super_admin'])
def editar_vendedor(v_id):
    v = Vendedor.query.get_or_404(v_id)
    nombre = (request.form.get('nombre_completo') or '').strip()
    email = (request.form.get('email') or '').strip()
    telefono = (request.form.get('telefono') or '').strip() or None
    rol_id = request.form.get('rol_id', type=int)
    territorio_id = request.form.get('territorio_id', type=int) or None

    if not nombre or not email:
        flash('Nombre y email son obligatorios.', 'danger')
        return redirect(url_for('gestionar_vendedores'))
    rol = db.session.get(Rol, rol_id) if rol_id else None
    if rol is None:
        flash('El rol seleccionado no es válido.', 'danger')
        return redirect(url_for('gestionar_vendedores'))
    if territorio_id and db.session.get(Territorio, territorio_id) is None:
        flash('El territorio seleccionado no es válido.', 'danger')
        return redirect(url_for('gestionar_vendedores'))
    otro = Vendedor.query.filter(Vendedor.email == email, Vendedor.id != v.id).first()
    if otro:
        flash('Ese email ya está en uso por otro usuario.', 'danger')
        return redirect(url_for('gestionar_vendedores'))
    if rol.nombre != 'super_admin' and _es_ultimo_super_admin(v):
        flash('No se puede quitar el rol super_admin al único administrador activo.', 'danger')
        return redirect(url_for('gestionar_vendedores'))

    try:
        v.nombre_completo = nombre
        v.email = email
        v.telefono = telefono
        v.rol_id = rol.id
        v.territorio_id = territorio_id
        db.session.commit()
        _audit('user', 'Editó usuario', f'{nombre} → {rol.nombre}')
        flash('Usuario actualizado.', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al editar vendedor {v_id}: {e}')
        flash('Error al actualizar el usuario. Intente de nuevo.', 'danger')
    return redirect(url_for('gestionar_vendedores'))


@app.route('/admin/vendedores/<int:v_id>/reset-password', methods=['POST'])
@login_required
@requiere_rol(['super_admin'])
def reset_password_vendedor(v_id):
    v = Vendedor.query.get_or_404(v_id)
    temp = (request.form.get('password_temporal') or '').strip()
    if temp and len(temp) < 8:
        flash('La contraseña temporal debe tener al menos 8 caracteres.', 'danger')
        return redirect(url_for('gestionar_vendedores'))
    if not temp:
        temp = _generar_password_temporal()
    try:
        v.set_password(temp)
        v.debe_cambiar_password = True
        db.session.commit()
        _audit('user', 'Reseteó contraseña', v.nombre_completo or v.username)
        flash(f'Contraseña temporal de {v.nombre_completo}: {temp} — comunicásela; '
              f'deberá cambiarla al ingresar.', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al resetear contraseña de vendedor {v_id}: {e}')
        flash('Error al restablecer la contraseña. Intente de nuevo.', 'danger')
    return redirect(url_for('gestionar_vendedores'))

# ===== WEBHOOKS Y INTEGRACIONES =====

@app.route('/webhook/actualizacion-precios', methods=['POST'])
@csrf.exempt
def webhook_actualizacion_precios():
    """Webhook para actualizaciones automáticas de precios desde sistemas externos.

    Autenticado con un secreto compartido (WEBHOOK_SECRET) enviado en el header
    X-Webhook-Token. Sin secreto configurado, el endpoint queda deshabilitado.
    """
    expected = os.environ.get('WEBHOOK_SECRET', '').strip()
    if not expected:
        app.logger.error("Webhook precios deshabilitado: WEBHOOK_SECRET no configurado")
        return jsonify({'error': 'No autorizado'}), 401

    provided = (request.headers.get('X-Webhook-Token') or '').strip()
    if not provided or not hmac.compare_digest(provided, expected):
        app.logger.warning(f"Webhook precios: token inválido desde IP {request.remote_addr}")
        return jsonify({'error': 'No autorizado'}), 401

    try:
        data = request.get_json()

        if not data or 'productos' not in data:
            return jsonify({'error': 'Datos inválidos'}), 400
        
        productos_actualizados = 0
        
        for item in data['productos']:
            producto_id = item.get('producto_id')
            nuevo_precio = item.get('precio_jomar')
            
            if producto_id and nuevo_precio:
                # Actualizar precio en lista por defecto
                lista_default = ListaPrecio.query.filter_by(es_default=True).first()
                if lista_default:
                    precio = PrecioProducto.query.filter_by(
                        lista_precio_id=lista_default.id,
                        producto_id=producto_id
                    ).first()
                    
                    if precio:
                        precio.precio_jomar = nuevo_precio
                        precio.fecha_actualizacion = datetime.utcnow()
                        productos_actualizados += 1
        
        if productos_actualizados > 0:
            db.session.commit()
        
        return jsonify({
            'success': True,
            'productos_actualizados': productos_actualizados
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error en webhook precios: {e}")
        return jsonify({'error': 'Error interno'}), 500


@app.route('/dashboard')
@login_required
def dashboard():
    app.logger.info("[/dashboard] entrando")
    """Dashboard optimizado con KPIs de ventas y nivel de servicio"""
    dashboard_perf_start = perf_counter()
    dashboard_perf_marks = {}
    dashboard_perf_last = dashboard_perf_start

    def mark_dashboard_perf(block_name):
        nonlocal dashboard_perf_last
        now = perf_counter()
        dashboard_perf_marks[block_name] = round((now - dashboard_perf_last) * 1000, 2)
        dashboard_perf_last = now

    def log_dashboard_perf(stage):
        if not DASHBOARD_PERF_LOG:
            return
        total_ms = (perf_counter() - dashboard_perf_start) * 1000
        breakdown = " | ".join(f"{k}={v:.2f}ms" for k, v in dashboard_perf_marks.items())
        if breakdown:
            app.logger.info(f"[/dashboard] perf stage={stage} total={total_ms:.2f}ms :: {breakdown}")
        else:
            app.logger.info(f"[/dashboard] perf stage={stage} total={total_ms:.2f}ms")

    try:
        # Verificación de dependencias críticas
        if not db or not Pedido:
            app.logger.error("Dependencias críticas no disponibles")
            raise Exception("Base de datos no inicializada")
            
        app.logger.info("Iniciando cálculo de dashboard...")
        # === FECHAS DE REFERENCIA ===
        hoy = datetime.now(DASHBOARD_TIMEZONE).date()
        inicio_mes = hoy.replace(day=1)
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        hace_30_dias = hoy - timedelta(days=30)
        # Mes anterior (mismo rango de días para comparación justa)
        fin_mes_anterior = inicio_mes - timedelta(days=1)
        inicio_mes_anterior = fin_mes_anterior.replace(day=1)
        inicio_tendencia = inicio_semana - timedelta(weeks=25)
        inicio_ultimos_7_dias = hoy - timedelta(days=6)
        pedidos_period_starts = {
            '6m': hoy - timedelta(days=181),
            '3m': hoy - timedelta(days=90),
            '4w': hoy - timedelta(days=27),
        }

        # === CLIENTE EXCLUIDO (ALMACÉN INTERNO) ===
        # AL01 es usado para registrar pesos e imprimir etiquetas de almacén, no son pedidos reales
        cliente_almacen = Cliente.query.filter_by(nombre='AL01').first()
        cliente_almacen_id = cliente_almacen.id if cliente_almacen else None

        # Helper para filtrar pedidos excluyendo AL01
        def filtro_sin_almacen():
            if cliente_almacen_id:
                return Pedido.cliente_id != cliente_almacen_id
            return True  # Si no existe AL01, no filtrar nada
        mark_dashboard_perf('setup')

        # === CONSULTAS ROBUSTAS PARA PRODUCCIÓN ===
        try:
            # Cargar una sola vez pedidos de los últimos ~6 meses con relaciones necesarias,
            # para evitar N+1 y múltiples consultas solapadas por período.
            inicio_historico_mensual = (inicio_mes - relativedelta(months=5)).replace(day=1)
            inicio_carga_pedidos = min(inicio_historico_mensual, pedidos_period_starts['6m'])
            pedido_load_options = (
                load_only(
                    Pedido.id,
                    Pedido.cliente_id,
                    Pedido.fecha_pedido,
                    Pedido.fecha_facturacion,
                    Pedido.estado,
                ),
                joinedload(Pedido.cliente).load_only(Cliente.id, Cliente.nombre),
                selectinload(Pedido.detalles).load_only(
                    DetallePedido.id,
                    DetallePedido.pedido_id,
                    DetallePedido.producto_id,
                    DetallePedido.es_linea_pedido,
                    DetallePedido.cajas,
                    DetallePedido.cajas_pedidas,
                    DetallePedido.peso,
                    DetallePedido.subtotal,
                ).selectinload(DetallePedido.producto).load_only(
                    Producto.id,
                    Producto.nombre,
                    Producto.se_pesa,
                ),
                selectinload(Pedido.detalles).selectinload(DetallePedido.cajas_pesadas).load_only(
                    CajaPesada.id,
                    CajaPesada.detalle_pedido_id,
                    CajaPesada.numero,
                    CajaPesada.peso,
                ),
            )

            # Una sola query: pedidos por fecha_pedido O facturados por fecha_facturacion
            fecha_corte_facturados = inicio_tendencia - timedelta(days=1)
            pedidos_base_list = (
                Pedido.query.options(*pedido_load_options)
                .filter(
                    filtro_sin_almacen(),
                    db.or_(
                        Pedido.fecha_pedido >= inicio_carga_pedidos,
                        db.and_(
                            Pedido.estado == 'facturado',
                            Pedido.fecha_facturacion.isnot(None),
                            Pedido.fecha_facturacion >= fecha_corte_facturados,
                        )
                    )
                )
                .all()
            )

            pedidos_base_with_dates = []
            for pedido in pedidos_base_list:
                fecha_local = _to_dashboard_date(pedido.fecha_pedido)
                if not fecha_local or fecha_local > hoy:
                    continue
                pedidos_base_with_dates.append((pedido, fecha_local))

            pedidos_mes_list = [p for p, fecha in pedidos_base_with_dates if fecha >= inicio_mes]
            pedidos_semana_list = [p for p, fecha in pedidos_base_with_dates if fecha >= inicio_semana]
            pedidos_30_dias = [p for p, fecha in pedidos_base_with_dates if fecha >= hace_30_dias]
            pedidos_mes_anterior_list = [
                p for p, fecha in pedidos_base_with_dates
                if inicio_mes_anterior <= fecha <= fin_mes_anterior
            ]
            pedidos_6m_list = [
                p for p, fecha in pedidos_base_with_dates
                if fecha >= pedidos_period_starts['6m']
            ]

            pedidos_facturados_list = [
                p for p in pedidos_base_list
                if (p.estado or '').strip().lower() == 'facturado'
                and p.fecha_facturacion is not None
                and _to_dashboard_date(p.fecha_facturacion) is not None
                and _to_dashboard_date(p.fecha_facturacion) >= fecha_corte_facturados
            ]

            app.logger.info(
                f"Datos cargados: {len(pedidos_mes_list)} pedidos mes, "
                f"{len(pedidos_semana_list)} semana, {len(pedidos_30_dias)} últimos 30 días, "
                f"{len(pedidos_facturados_list)} facturados"
            )
        except Exception as e:
            app.logger.error(f"Error en consultas dashboard: {e}")
            # Fallback con datos vacíos
            pedidos_base_with_dates = []
            pedidos_mes_list = []
            pedidos_semana_list = []
            pedidos_30_dias = []
            pedidos_facturados_list = []
            pedidos_mes_anterior_list = []
            pedidos_6m_list = []
        mark_dashboard_perf('carga_datos')

        pedido_metricas_cache = {}

        def obtener_metricas_pedido(pedido):
            cache_key = pedido.id if pedido.id is not None else id(pedido)
            metricas = pedido_metricas_cache.get(cache_key)
            if metricas is not None:
                return metricas

            fecha_pedido_local = _to_dashboard_date(pedido.fecha_pedido)
            fecha_fact_local = _to_dashboard_date(pedido.fecha_facturacion) if pedido.fecha_facturacion else None
            estado_normalizado = (pedido.estado or '').strip().lower()
            es_facturado = estado_normalizado == 'facturado' and bool(fecha_fact_local)

            venta = _coerce_float(_calcular_venta_pedido(pedido), 0.0)

            lead_time_days = None
            if fecha_fact_local and fecha_pedido_local:
                dias = (fecha_fact_local - fecha_pedido_local).days
                if dias >= 0:
                    lead_time_days = dias

            total_pedidas = 0
            total_entregadas = 0
            if es_facturado:
                prep_count_por_producto = {}
                prep_cajas_por_producto = {}
                lineas_pedido = []

                for det in pedido.detalles:
                    if det.es_linea_pedido:
                        lineas_pedido.append(det)
                    elif det.producto_id is not None:
                        prep_count_por_producto[det.producto_id] = prep_count_por_producto.get(det.producto_id, 0) + 1
                        prep_cajas_por_producto[det.producto_id] = prep_cajas_por_producto.get(det.producto_id, 0) + (det.cajas or 0)

                for det in lineas_pedido:
                    pedidas = det.cajas_pedidas or det.cajas or 0
                    if pedidas <= 0:
                        continue

                    total_pedidas += pedidas

                    if det.producto and det.producto.se_pesa:
                        entregadas = det.cajas_pesadas_count or prep_count_por_producto.get(det.producto_id, 0)
                    else:
                        prep_cajas = prep_cajas_por_producto.get(det.producto_id, 0)
                        entregadas = prep_cajas if prep_cajas > 0 else (det.cajas or 0)

                    total_entregadas += min(entregadas, pedidas)

            fecha_pedido_ref = pedido.fecha_pedido.date() if pedido.fecha_pedido else None
            is_pending_overdue = (
                estado_normalizado in ('pendiente', 'preparado')
                and bool(fecha_pedido_ref)
                and (hoy - fecha_pedido_ref).days > 2
            )

            metricas = {
                'fecha_pedido_local': fecha_pedido_local,
                'fecha_fact_local': fecha_fact_local,
                'estado': estado_normalizado,
                'es_facturado': es_facturado,
                'venta': venta,
                'lead_time_days': lead_time_days,
                'total_pedidas': total_pedidas,
                'total_entregadas': total_entregadas,
                'is_pending_overdue': is_pending_overdue,
            }
            pedido_metricas_cache[cache_key] = metricas
            return metricas

        try:
            # Ventas reconocidas por fecha de facturación local
            ventas_mes = 0
            ventas_semana = 0
            ventas_mes_anterior = 0
            ventas_semanales_idx = {}
            ventas_diarias_idx = {}

            for p in pedidos_facturados_list:
                try:
                    pedido_metricas = obtener_metricas_pedido(p)
                    fecha_fact_local = pedido_metricas['fecha_fact_local']
                    if not fecha_fact_local:
                        continue
                    venta = pedido_metricas['venta']
                except (AttributeError, ValueError, TypeError) as e:
                    app.logger.warning(f"Error en cálculo ventas pedido {p.id}: {e}")
                    continue

                if fecha_fact_local >= inicio_mes:
                    ventas_mes += venta

                if fecha_fact_local >= inicio_semana:
                    ventas_semana += venta

                if inicio_mes_anterior <= fecha_fact_local <= fin_mes_anterior:
                    ventas_mes_anterior += venta

                if fecha_fact_local >= inicio_ultimos_7_dias:
                    bucket = ventas_diarias_idx.setdefault(
                        fecha_fact_local,
                        {'ventas': 0.0, 'pedidos': 0}
                    )
                    bucket['ventas'] += venta
                    bucket['pedidos'] += 1

                semana_inicio = fecha_fact_local - timedelta(days=fecha_fact_local.weekday())
                if inicio_tendencia <= semana_inicio <= inicio_semana:
                    bucket = ventas_semanales_idx.setdefault(
                        semana_inicio,
                        {'ventas': 0.0, 'pedidos': 0}
                    )
                    bucket['ventas'] += venta
                    bucket['pedidos'] += 1

            pedidos_mes_anterior = len(pedidos_mes_anterior_list)

            # Derivar pendientes de pedidos ya cargados (evita query adicional)
            pedidos_pendientes = sum(
                1 for p in pedidos_base_list
                if (p.estado or '').strip().lower() == 'pendiente'
            )
            
        except Exception as e:
            app.logger.error(f"Error en cálculos de ventas: {e}")
            ventas_mes = 0
            ventas_semana = 0
            ventas_mes_anterior = 0
            pedidos_mes_anterior = 0
            pedidos_pendientes = 0
            ventas_semanales_idx = {}
            ventas_diarias_idx = {}
        mark_dashboard_perf('ventas_locales')

        # === OVERRIDE OPCIONAL: ventas desde QuickBooks (fuente de verdad) ===
        metricas_ventas_qb = _obtener_metricas_ventas_quickbooks(
            hoy=hoy,
            inicio_mes=inicio_mes,
            inicio_semana=inicio_semana,
            inicio_mes_anterior=inicio_mes_anterior,
            fin_mes_anterior=fin_mes_anterior,
            inicio_tendencia=inicio_tendencia,
            inicio_ultimos_7_dias=inicio_ultimos_7_dias,
        )
        if metricas_ventas_qb:
            ventas_mes = metricas_ventas_qb['ventas_mes']
            ventas_semana = metricas_ventas_qb['ventas_semana']
            ventas_mes_anterior = metricas_ventas_qb['ventas_mes_anterior']
            ventas_diarias_idx = metricas_ventas_qb['ventas_diarias_idx']
            ventas_semanales_idx = metricas_ventas_qb['ventas_semanales_idx']
        mark_dashboard_perf('fuente_ventas')

        # === KPIs OPTIMIZADOS DE NIVEL DE SERVICIO (MES EN CURSO) ===
        # NOTA: Todos los KPIs ahora se calculan sobre el mes en curso

        # Función helper para calcular KPIs de un período
        def calcular_kpis_periodo(pedidos_periodo):
            """Calcula todos los KPIs para un período dado"""
            if not pedidos_periodo:
                return {
                    'order_completion_rate': 0, 'otd_rate': 0,
                    'lead_time': 0, 'total_pedidos': 0, 'ventas': 0
                }

            # Pedidos facturados del período
            facturados_metricas = []
            pendientes_vencidos = 0
            for p in pedidos_periodo:
                pedido_metricas = obtener_metricas_pedido(p)
                if pedido_metricas['es_facturado']:
                    facturados_metricas.append(pedido_metricas)
                if pedido_metricas['is_pending_overdue']:
                    pendientes_vencidos += 1

            # Lead times
            lead_times_p = [
                m['lead_time_days']
                for m in facturados_metricas
                if m['lead_time_days'] is not None
            ]

            # OFR real — línea por línea: cajas entregadas / cajas pedidas
            total_pedidas = sum(m['total_pedidas'] for m in facturados_metricas)
            total_entregadas = sum(m['total_entregadas'] for m in facturados_metricas)
            order_completion_rate_p = (total_entregadas / total_pedidas * 100) if total_pedidas > 0 else 100

            # OTD Rate corregido — incluye pendientes vencidos como "fuera de tiempo"
            a_tiempo = sum(1 for lt in lead_times_p if lt <= 2)
            total_otd = len(lead_times_p) + pendientes_vencidos
            otd_p = (a_tiempo / total_otd * 100) if total_otd > 0 else 100

            # Ventas del período
            total_p = len(pedidos_periodo)
            ventas_p = sum(m['venta'] for m in facturados_metricas)

            return {
                'order_completion_rate': round(order_completion_rate_p, 1),
                'otd_rate': round(otd_p, 1),
                'lead_time': round(sum(lead_times_p) / len(lead_times_p), 1) if lead_times_p else 0,
                'total_pedidos': total_p,
                'ventas': ventas_p
            }

        # Calcular KPIs del mes en curso
        kpis_mes_actual = calcular_kpis_periodo(pedidos_mes_list)

        # Extraer valores para compatibilidad con template existente
        pedidos_facturados = []
        lead_times = []
        for p in pedidos_mes_list:
            pedido_metricas = obtener_metricas_pedido(p)
            if not pedido_metricas['es_facturado']:
                continue
            pedidos_facturados.append(p)
            if pedido_metricas['lead_time_days'] is not None:
                lead_times.append(pedido_metricas['lead_time_days'])
        lead_time_promedio = kpis_mes_actual['lead_time']
        order_completion_rate = kpis_mes_actual['order_completion_rate']
        otd_rate = kpis_mes_actual['otd_rate']

        # === HISTÓRICO DE KPIs (ÚLTIMOS 6 MESES) ===
        kpis_historicos = []
        for months_back in range(5, -1, -1):
            mes_inicio = (inicio_mes - relativedelta(months=months_back)).replace(day=1)
            if months_back == 0:
                mes_fin = hoy
            else:
                mes_fin = (mes_inicio + relativedelta(months=1)) - timedelta(days=1)

            pedidos_mes_hist = [
                p for p, fecha_local in pedidos_base_with_dates
                if mes_inicio <= fecha_local <= mes_fin
            ]

            kpis = calcular_kpis_periodo(pedidos_mes_hist)
            kpis['mes'] = mes_inicio.strftime('%b %Y')
            kpis['mes_num'] = mes_inicio.month
            kpis['año'] = mes_inicio.year
            kpis_historicos.append(kpis)

        # Perfect order rate del mes actual (simplificado: solo OTD)
        perfect_orders = sum(1 for lt in lead_times if lt <= 2)
        pendientes_vencidos_mes = sum(
            1 for p in pedidos_mes_list if obtener_metricas_pedido(p)['is_pending_overdue']
        )
        total_evaluado_por = len(pedidos_facturados) + pendientes_vencidos_mes
        perfect_order_rate = (
            (perfect_orders / total_evaluado_por * 100)
            if total_evaluado_por > 0 else 0
        )

        # 6. Customer engagement optimizado
        clientes_activos_ids = {p.cliente_id for p in pedidos_mes_list if p.cliente_id}
        clientes_activos_mes = len(clientes_activos_ids)
        
        # Cache de total de clientes para evitar query innecesaria (excluyendo AL01)
        clientes_query = Cliente.query
        if cliente_almacen_id:
            clientes_query = clientes_query.filter(Cliente.id != cliente_almacen_id)
        total_clientes = clientes_query.count()
        customer_engagement = (
            (clientes_activos_mes / total_clientes * 100) 
            if total_clientes > 0 else 0
        )
        mark_dashboard_perf('kpis_servicio')

        # === RANKINGS DE PRODUCTOS/CLIENTES (MES + 6M/3M/4S) ===
        period_starts_rankings = {
            'month': inicio_mes,
            '6m': inicio_tendencia,
            '3m': inicio_semana - timedelta(weeks=12),
            '4w': inicio_semana - timedelta(weeks=3),
        }

        ranking_client_rows_local = []
        ranking_product_rows_local = []

        for p in pedidos_facturados_list:
            pedido_metricas = obtener_metricas_pedido(p)
            fecha_fact_local = pedido_metricas['fecha_fact_local']
            if not fecha_fact_local or fecha_fact_local < inicio_tendencia or fecha_fact_local > hoy:
                continue

            invoice_key = str(p.id)
            if p.cliente and p.cliente.nombre:
                cliente_nombre = p.cliente.nombre
            else:
                app.logger.warning(f'Pedido sin cliente: {p.id}')
                cliente_nombre = 'Sin cliente'

            ranking_client_rows_local.append({
                'date': fecha_fact_local,
                'invoice_key': invoice_key,
                'customer': cliente_nombre,
                'amount': pedido_metricas['venta'],
            })

            productos_con_prep = set()
            lineas_pedido_ranking = []
            for d in p.detalles:
                if not d.producto or not d.producto.nombre:
                    continue
                if not d.es_linea_pedido:
                    productos_con_prep.add(d.producto_id)
                    ranking_product_rows_local.append({
                        'date': fecha_fact_local,
                        'invoice_key': invoice_key,
                        'product': d.producto.nombre,
                        'amount': float(d.subtotal or 0),
                        'quantity': d.cajas or 0,
                        'weight': d.peso or 0,
                    })
                else:
                    lineas_pedido_ranking.append(d)

            for d in lineas_pedido_ranking:
                if d.producto_id in productos_con_prep:
                    continue
                ranking_product_rows_local.append({
                    'date': fecha_fact_local,
                    'invoice_key': invoice_key,
                    'product': d.producto.nombre,
                    'amount': float(d.subtotal or 0),
                    'quantity': d.cajas or 0,
                    'weight': d.peso or 0,
                })

        rankings_periodos_local = _build_rankings_periodos_from_rows(
            ranking_client_rows_local,
            ranking_product_rows_local,
            hoy=hoy,
            period_starts=period_starts_rankings,
        )
        rankings_periodos = rankings_periodos_local

        # Si QuickBooks está disponible, priorizar su ranking por periodos
        if metricas_ventas_qb and metricas_ventas_qb.get('rankings_periodos'):
            rankings_periodos_qb = metricas_ventas_qb['rankings_periodos']
            rankings_periodos = {}
            for period_key in period_starts_rankings.keys():
                local_payload = rankings_periodos_local.get(period_key, {})
                qb_payload = rankings_periodos_qb.get(period_key, {})
                top_productos_qb = qb_payload.get('top_productos') or []
                top_clientes_qb = qb_payload.get('top_clientes') or []

                rankings_periodos[period_key] = {
                    'top_productos': top_productos_qb or local_payload.get('top_productos', []),
                    'top_clientes': top_clientes_qb or local_payload.get('top_clientes', []),
                    'max_ventas': qb_payload.get('max_ventas') or local_payload.get('max_ventas', 1),
                    'max_total_clientes': qb_payload.get('max_total_clientes') or local_payload.get('max_total_clientes', 1),
                }

        month_rankings = rankings_periodos.get('month', {})
        top_productos = month_rankings.get('top_productos', [])
        top_clientes = month_rankings.get('top_clientes', [])
        max_ventas = month_rankings.get('max_ventas', 1)
        rankings_periodos_json = _serialize_rankings_periodos(rankings_periodos)
        mark_dashboard_perf('rankings')

        # === TENDENCIA SEMANAL (excluyendo AL01) — 26 semanas (6 meses) ===
        tendencia_semanal = []
        for i in range(25, -1, -1):
            inicio_i = inicio_semana - timedelta(weeks=i)
            bucket = ventas_semanales_idx.get(inicio_i, {'ventas': 0.0, 'pedidos': 0})
            tendencia_semanal.append(
                {
                    'semana': inicio_i.strftime('%d/%m'),
                    'ventas': bucket['ventas'],
                    'pedidos': bucket['pedidos'],
                }
            )

        # === ESTADOS DE PEDIDOS ===
        estados_count = {}
        for p in pedidos_30_dias:
            estado = p.estado or 'sin_estado'  # Manejar estados nulos
            estados_count[estado] = estados_count.get(estado, 0) + 1

        # Asegurar que siempre tengamos datos básicos
        estados_pedidos = {
            'pendiente': estados_count.get('pendiente', 0),
            'preparado': estados_count.get('preparado', 0),
            'facturado': estados_count.get('facturado', 0),
            **{k: v for k, v in estados_count.items() if k not in ['pendiente', 'preparado', 'facturado']}
        }

        # === PEDIDOS VISUAL (DIARIO + ESTADOS 6M / 3M / 4S) ===
        pedidos_diarios_periodos = {k: [] for k in pedidos_period_starts.keys()}
        pedidos_resumen_periodos = {
            k: {'total': 0, 'facturados': 0, 'pendientes': 0, 'otros': 0}
            for k in pedidos_period_starts.keys()
        }

        try:
            pedidos_diarios_idx = {}
            for p in pedidos_6m_list:
                pedido_metricas = obtener_metricas_pedido(p)
                fecha_pedido_local = pedido_metricas['fecha_pedido_local']
                if (
                    not fecha_pedido_local
                    or fecha_pedido_local < pedidos_period_starts['6m']
                    or fecha_pedido_local > hoy
                ):
                    continue

                pedidos_diarios_idx[fecha_pedido_local] = pedidos_diarios_idx.get(fecha_pedido_local, 0) + 1
                estado = (p.estado or '').strip().lower()

                for period_key, start_date in pedidos_period_starts.items():
                    if fecha_pedido_local < start_date:
                        continue
                    resumen = pedidos_resumen_periodos[period_key]
                    resumen['total'] += 1
                    if estado == 'facturado':
                        resumen['facturados'] += 1
                    elif estado in ('pendiente', 'preparado'):
                        resumen['pendientes'] += 1
                    else:
                        resumen['otros'] += 1

            for period_key, start_date in pedidos_period_starts.items():
                cursor = start_date
                serie = []
                while cursor <= hoy:
                    serie.append({
                        'fecha': cursor.strftime('%d/%m'),
                        'pedidos': pedidos_diarios_idx.get(cursor, 0),
                    })
                    cursor += timedelta(days=1)
                pedidos_diarios_periodos[period_key] = serie

        except Exception as e:
            app.logger.error(f'Error calculando visual de pedidos: {e}')

        # === PEDIDOS RECIENTES (excluyendo AL01) ===
        recientes_query = Pedido.query.options(joinedload(Pedido.cliente))
        if cliente_almacen_id:
            recientes_query = recientes_query.filter(Pedido.cliente_id != cliente_almacen_id)
        pedidos_recientes_data = recientes_query.order_by(
            Pedido.fecha_pedido.desc()
        ).limit(10).all()

        # === OPERACIÓN DE PEDIDOS (TAB PEDIDOS) ===
        pedidos_facturados_hoy = sum(
            1 for p in pedidos_facturados_list
            if obtener_metricas_pedido(p)['fecha_fact_local'] == hoy
        )

        pedidos_operativos = []
        pedidos_vencidos = 0
        pedidos_preparados_activos = 0

        estado_priority = {
            'pendiente': 0,
            'preparado': 1,
            'facturado': 2,
        }

        for p in pedidos_30_dias:
            pedido_metricas = obtener_metricas_pedido(p)
            estado = pedido_metricas['estado'] or 'sin_estado'
            fecha_pedido_local = pedido_metricas['fecha_pedido_local']
            edad_dias = (hoy - fecha_pedido_local).days if fecha_pedido_local else 0
            total_xcg = round(pedido_metricas['venta'], 2)
            cliente_nombre = p.cliente.nombre if p.cliente and p.cliente.nombre else 'Sin cliente'

            es_urgente = estado in ('pendiente', 'preparado') and edad_dias > 2
            if es_urgente:
                pedidos_vencidos += 1
            if estado == 'preparado':
                pedidos_preparados_activos += 1

            if estado == 'facturado':
                sla_text = 'Facturado'
                sla_class = 'sla-ok'
            elif edad_dias <= 1:
                sla_text = 'En tiempo'
                sla_class = 'sla-ok'
            elif edad_dias == 2:
                sla_text = 'Límite hoy'
                sla_class = 'sla-warn'
            else:
                sla_text = f'Vencido {edad_dias - 2}d'
                sla_class = 'sla-danger'

            pedidos_operativos.append({
                'id': p.id,
                'cliente': cliente_nombre,
                'estado': estado,
                'fecha_pedido': fecha_pedido_local.strftime('%d/%m') if fecha_pedido_local else 'N/A',
                'fecha_pedido_full': p.fecha_pedido.strftime('%d/%m %H:%M') if p.fecha_pedido else '',
                'edad_dias': max(0, edad_dias),
                'total_xcg': total_xcg,
                'sla_text': sla_text,
                'sla_class': sla_class,
                'es_urgente': es_urgente,
                'puede_preparar': estado == 'pendiente',
                'puede_facturar': estado == 'preparado',
                'puede_editar': estado != 'facturado',
            })

        pedidos_operativos.sort(
            key=lambda x: (
                estado_priority.get(x['estado'], 3),
                -x['edad_dias'],
                -x['total_xcg'],
                -x['id']
            )
        )
        pedidos_operativos = pedidos_operativos[:20]

        # === VENTAS DIARIAS (excluyendo AL01) ===
        ventas_dias = []
        for i in range(6, -1, -1):
            dia = hoy - timedelta(days=i)
            bucket = ventas_diarias_idx.get(dia, {'ventas': 0.0, 'pedidos': 0})
            ventas_dias.append({
                'fecha': dia.strftime('%d/%m'),
                'ventas': bucket['ventas'],
                'pedidos': bucket['pedidos']
            })
        mark_dashboard_perf('operacion_pedidos')

        # === CALCULAR PORCENTAJE DE META ===
        try:
            meta_mensual = float(os.environ.get('MONTHLY_SALES_TARGET', '120000'))
        except (ValueError, TypeError):
            meta_mensual = 120000.0
        porcentaje_meta = (ventas_mes / meta_mensual * 100) if meta_mensual > 0 else 0

        # === PROYECCIÓN DE VENTAS A FIN DE MES ===
        dias_transcurridos = hoy.day
        dias_total_mes = calendar.monthrange(hoy.year, hoy.month)[1]
        if dias_transcurridos >= 2:
            proyeccion_ventas = ventas_mes / dias_transcurridos * dias_total_mes
        else:
            proyeccion_ventas = 0
        porcentaje_proyeccion = (proyeccion_ventas / meta_mensual * 100) if meta_mensual > 0 else 0

        # === TIEMPO DE RESPUESTA POR CLIENTE ===
        tiempos_respuesta_cliente = {}
        for p in pedidos_30_dias:
            if p.cliente and p.fecha_facturacion:
                nombre_cliente = p.cliente.nombre
                tiempo = (p.fecha_facturacion.date() - p.fecha_pedido.date()).days
                if nombre_cliente not in tiempos_respuesta_cliente:
                    tiempos_respuesta_cliente[nombre_cliente] = []
                tiempos_respuesta_cliente[nombre_cliente].append(tiempo)
        
        # Calcular promedios
        tiempo_respuesta_data = []
        for cliente, tiempos in tiempos_respuesta_cliente.items():
            tiempo_respuesta_data.append({
                'cliente': cliente,
                'promedio': round(sum(tiempos) / len(tiempos), 1),
                'pedidos': len(tiempos)
            })
        tiempo_respuesta_data = sorted(tiempo_respuesta_data, key=lambda x: x['promedio'])[:10]
        mark_dashboard_perf('metas_y_respuesta')

        # === RENDER ===
        dashboard_html = render_template(
            'dashboard.html',
            # Métricas principales
            ventas_mes=ventas_mes,
            pedidos_mes=len(pedidos_mes_list),
            ventas_semana=ventas_semana,
            pedidos_semana=len(pedidos_semana_list),
            ventas_mes_anterior=ventas_mes_anterior,
            pedidos_mes_anterior=pedidos_mes_anterior,
            pedidos_pendientes=pedidos_pendientes,
            meta_mensual=meta_mensual,
            porcentaje_meta=porcentaje_meta,
            proyeccion_ventas=round(proyeccion_ventas, 2),
            porcentaje_proyeccion=round(porcentaje_proyeccion, 1),
            dias_total_mes=dias_total_mes,
            
            # KPIs de servicio (actualizados)
            lead_time_promedio=round(lead_time_promedio, 1),
            order_completion_rate=round(order_completion_rate, 1),
            otd_rate=round(otd_rate, 1),
            perfect_order_rate=round(perfect_order_rate, 1),
            customer_engagement=round(customer_engagement, 1),
            
            # Datos para gráficos
            top_clientes=top_clientes,
            top_productos=top_productos,
            max_ventas=max_ventas,
            rankings_periodos_json=rankings_periodos_json,
            estados_pedidos=estados_pedidos,
            pedidos_diarios_periodos=pedidos_diarios_periodos,
            pedidos_resumen_periodos=pedidos_resumen_periodos,
            tendencia_semanal=tendencia_semanal,
            pedidos_recientes=pedidos_recientes_data,
            pedidos_operativos=pedidos_operativos,
            pedidos_vencidos=pedidos_vencidos,
            pedidos_preparados_activos=pedidos_preparados_activos,
            pedidos_facturados_hoy=pedidos_facturados_hoy,
            fecha_actual=hoy,
            ventas_dias=ventas_dias,
            tiempo_respuesta_data=tiempo_respuesta_data,

            # Histórico de KPIs (últimos 6 meses)
            kpis_historicos=kpis_historicos,

            # Configuración
            moneda='XCG'
        )
        mark_dashboard_perf('render_template')
        log_dashboard_perf('ok')
        return dashboard_html

    except Exception as e:
        app.logger.exception(f'Error crítico en /dashboard: {e}')
        log_dashboard_perf('error')
        
        # Datos de fallback para evitar error 500
        try:
            _fallback_meta = float(os.environ.get('MONTHLY_SALES_TARGET', '120000'))
        except (ValueError, TypeError):
            _fallback_meta = 120000.0
        fallback_data = {
            'ventas_mes': 0,
            'pedidos_mes': 0,
            'ventas_semana': 0,
            'pedidos_semana': 0,
            'pedidos_pendientes': 0,
            'meta_mensual': _fallback_meta,
            'porcentaje_meta': 0,
            'proyeccion_ventas': 0,
            'porcentaje_proyeccion': 0,
            'ventas_mes_anterior': 0,
            'pedidos_mes_anterior': 0,
            'dias_total_mes': calendar.monthrange(datetime.now().year, datetime.now().month)[1],
            'lead_time_promedio': 0,
            'order_completion_rate': 0,
            'otd_rate': 0,
            'perfect_order_rate': 0,
            'customer_engagement': 0,
            'top_clientes': [],
            'top_productos': [],
            'max_ventas': 1,
            'rankings_periodos_json': {},
            'estados_pedidos': {'pendiente': 0, 'preparado': 0, 'facturado': 0},
            'pedidos_diarios_periodos': {'6m': [], '3m': [], '4w': []},
            'pedidos_resumen_periodos': {
                '6m': {'total': 0, 'facturados': 0, 'pendientes': 0, 'otros': 0},
                '3m': {'total': 0, 'facturados': 0, 'pendientes': 0, 'otros': 0},
                '4w': {'total': 0, 'facturados': 0, 'pendientes': 0, 'otros': 0},
            },
            'tendencia_semanal': [],
            'pedidos_recientes': [],
            'pedidos_operativos': [],
            'pedidos_vencidos': 0,
            'pedidos_preparados_activos': 0,
            'pedidos_facturados_hoy': 0,
            'fecha_actual': datetime.now().date(),
            'ventas_dias': [],
            'tiempo_respuesta_data': [],
            'kpis_historicos': [],
            'moneda': 'XCG'
        }
        
        try:
            fallback_html = render_template('dashboard.html', **fallback_data)
            mark_dashboard_perf('render_fallback')
            log_dashboard_perf('fallback')
            return fallback_html
        except Exception as template_error:
            app.logger.error(f'Error incluso con datos de fallback: {template_error}')
            from flask import abort
            abort(500)


@app.route('/pedidos')
@login_required
@requiere_permiso_recurso('pedidos', 'leer')
def lista_pedidos():
    # Parámetros de paginación
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    per_page = min(per_page, 100)  # Máximo 100 por página
    estado = (request.args.get('estado', 'todos', type=str) or 'todos').strip().lower()
    estado = estado if estado in {'todos', 'pendiente', 'preparado', 'facturado'} else 'todos'
    q = (request.args.get('q', '', type=str) or '').strip()
    solo_notas = (request.args.get('solo_notas', '', type=str) or '').strip().lower() in {'1', 'true', 'on', 'yes'}

    filtros = {
        'estado': estado,
        'q': q,
        'solo_notas': solo_notas,
    }

    # Subquery: total de líneas de preparación (cantidades reales)
    prep_subq = db.session.query(
        DetallePedido.pedido_id,
        func.coalesce(func.sum(DetallePedido.subtotal), 0).label('total_prep')
    ).filter(
        DetallePedido.es_linea_pedido == False
    ).group_by(DetallePedido.pedido_id).subquery()

    # Subquery: total de líneas originales del pedido
    orig_subq = db.session.query(
        DetallePedido.pedido_id,
        func.coalesce(func.sum(DetallePedido.subtotal), 0).label('total_orig')
    ).filter(
        DetallePedido.es_linea_pedido == True
    ).group_by(DetallePedido.pedido_id).subquery()

    # Query base: usar preparación si existe, si no original
    base_query = db.session.query(
        Pedido,
        db.case(
            (prep_subq.c.total_prep.isnot(None), prep_subq.c.total_prep),
            else_=func.coalesce(orig_subq.c.total_orig, 0)
        ).label('total_calculado')
    ).outerjoin(
        prep_subq, Pedido.id == prep_subq.c.pedido_id
    ).outerjoin(
        orig_subq, Pedido.id == orig_subq.c.pedido_id
    ).options(
        joinedload(Pedido.cliente),
        # Eager-load detalles + cajas_pesadas + producto so the post-query
        # call to _calcular_venta_pedido(pedido) doesn't fire N+1 queries
        # for the listed page.
        selectinload(Pedido.detalles).selectinload(DetallePedido.cajas_pesadas),
        selectinload(Pedido.detalles).selectinload(DetallePedido.producto),
    ).filter(
        Pedido.estado != 'entregado'
    )

    # Orden: 1) Estado (pendiente → preparado → facturado), 2) ID desc
    orden_optimizado = [
        db.case(
            (Pedido.estado == 'pendiente', 0),
            (Pedido.estado == 'preparado', 1),
            (Pedido.estado == 'facturado', 2),
            else_=3
        ),
        Pedido.id.desc(),
    ]

    # Filtrar por permisos del usuario
    if isinstance(current_user, Vendedor) and current_user.rol.nombre != 'super_admin':
        clientes_ids = [c.id for c in current_user.obtener_clientes_visibles()]
        if not clientes_ids:
            # Sin clientes asignados - retornar vacío con paginación mock
            return render_template(
                'pedidos.html',
                pedidos=[],
                pagination=None,
                filtros=filtros,
                status_counts={
                    'total': 0,
                    'pendiente': 0,
                    'preparado': 0,
                    'facturado': 0,
                },
            )
        base_query = base_query.filter(Pedido.cliente_id.in_(clientes_ids))

    # Conteos globales de la bandeja visible antes de aplicar filtros
    raw_status_counts = base_query.with_entities(
        Pedido.estado,
        func.count(Pedido.id)
    ).group_by(Pedido.estado).all()

    status_counts = {
        'total': 0,
        'pendiente': 0,
        'preparado': 0,
        'facturado': 0,
    }
    for estado_nombre, cantidad in raw_status_counts:
        if estado_nombre in status_counts:
            status_counts[estado_nombre] = cantidad
            status_counts['total'] += cantidad

    # Filtros de bandeja operativa
    if estado != 'todos':
        base_query = base_query.filter(Pedido.estado == estado)

    if solo_notas:
        base_query = base_query.filter(Pedido.notas.isnot(None), Pedido.notas != '')

    if q:
        q_like = f'%{q}%'
        base_query = base_query.filter(or_(
            cast(Pedido.id, String).ilike(q_like),
            Pedido.notas.ilike(q_like),
            Pedido.cliente.has(Cliente.nombre.ilike(q_like)),
        ))

    # Aplicar ordenamiento
    base_query = base_query.order_by(*orden_optimizado)

    # Contar total para paginación
    total_count = base_query.count()
    total_pages = (total_count + per_page - 1) // per_page

    # Aplicar paginación manual (offset/limit)
    pedidos_query = base_query.offset((page - 1) * per_page).limit(per_page).all()

    # El SQL computa el total como subtotal (cajas pedidas × precio_unitario).
    # Para reflejar el avance real, sustituimos por _calcular_venta_pedido,
    # que para productos pesables usa peso_real × precio_unitario y cae a la
    # línea original solo cuando aún no hay cajas pesadas. La eager-load
    # de detalles + cajas_pesadas + producto evita el N+1.
    pedidos = []
    for pedido, total in pedidos_query:
        venta_real = _calcular_venta_pedido(pedido)
        if venta_real and venta_real > 0:
            pedido.total_calculado = float(venta_real)
        else:
            pedido.total_calculado = float(total)
        pedidos.append(pedido)

    # Info de paginación para el template
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total_count,
        'pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages,
        'prev_num': page - 1 if page > 1 else None,
        'next_num': page + 1 if page < total_pages else None,
    }

    return render_template(
        'pedidos.html',
        pedidos=pedidos,
        pagination=pagination,
        filtros=filtros,
        status_counts=status_counts,
    )



def _resolver_precio_unitario_pedido(cliente_id, producto_id, precio_form=None):
    """Precio unitario de una línea del pedido, SIEMPRE resuelto por jerarquía.

    El formulario manda `productos[i][precio]`, sembrado con el precio de la
    lista default y sobreescrito por JS recién cuando se elige el cliente. Si ese
    JS no alcanza a correr —producto agregado antes de elegir cliente, fetch que
    no vuelve, cliente cambiado al editar— lo que llega es el precio default y no
    el del cliente. Antes ese valor le ganaba a la jerarquía y el pedido salía mal
    cobrado en silencio: le pasó al pedido 1270 (cobrado a 14,00 con precio de
    cliente 13,00), y obligó a corregir la factura a mano en QuickBooks.

    Como nunca se carga un precio a mano, el valor del formulario dejó de ganarle
    a la jerarquía: ahora es el ÚLTIMO recurso, para cuando el producto no tiene
    precio en ninguna lista. Descartarlo del todo sería peor — dejaría la línea en
    0 y borraría el precio de un pedido que ya lo tenía.

    Cuando el formulario manda algo distinto de lo resuelto se registra en el log:
    es la señal de que el JS de precios por cliente no llegó a correr.
    """
    precio = obtener_precio_producto_cliente(cliente_id, producto_id, 'base')
    if precio is None:
        precio = obtener_precio_default_producto(producto_id, 'base')

    if precio is None:
        # Sin precio configurado en ningún lado: mejor lo que trajo el formulario
        # que un 0 que además bloquea la facturación.
        if precio_form not in (None, ''):
            try:
                return Decimal(str(precio_form))
            except (ArithmeticError, TypeError, ValueError):
                return Decimal('0')
        return Decimal('0')

    resuelto = Decimal(str(precio))

    if precio_form not in (None, ''):
        try:
            if Decimal(str(precio_form)) != resuelto:
                app.logger.warning(
                    '[precio-form] cliente=%s producto=%s form=%s resuelto=%s '
                    '(se usa el resuelto)',
                    cliente_id, producto_id, precio_form, resuelto,
                )
        except (ArithmeticError, TypeError, ValueError):
            pass

    return resuelto


class _PedidoFormError(ValueError):
    """Raised when the pedido form payload contains an invalid line."""
    pass


def _extraer_lineas_pedido_form(form_data, cliente_id):
    """Normaliza las líneas enviadas desde el formulario de pedido.

    Validates each line:
    - producto_id is a positive integer
    - cajas is an integer between 1 and 9999 (no negative, no zero, no
      runaway values that would warp subtotals or the QBO payload)

    De-duplicates lines that share a producto_id by summing their cajas
    so a JS bug or double-submit can't silently drop quantity.

    Raises _PedidoFormError on bad input — callers should catch and
    flash + redirect rather than 500.
    """
    lineas_por_producto = {}
    idx = 0

    while f'productos[{idx}][id]' in form_data:
        try:
            prod_id = int(form_data.get(f'productos[{idx}][id]'))
        except (TypeError, ValueError):
            raise _PedidoFormError(f'Producto inválido en línea {idx + 1}')

        try:
            cajas = int(form_data.get(f'productos[{idx}][cajas]', 0) or 0)
        except (TypeError, ValueError):
            raise _PedidoFormError(f'Cantidad de cajas inválida en línea {idx + 1}')

        if cajas <= 0:
            raise _PedidoFormError(f'La cantidad de cajas debe ser mayor que 0 (línea {idx + 1})')
        if cajas > 9999:
            raise _PedidoFormError(f'La cantidad de cajas no puede exceder 9999 (línea {idx + 1})')
        if prod_id <= 0:
            raise _PedidoFormError(f'Producto inválido en línea {idx + 1}')

        precio_unitario = _resolver_precio_unitario_pedido(
            cliente_id,
            prod_id,
            form_data.get(f'productos[{idx}][precio]'),
        )

        if prod_id in lineas_por_producto:
            lineas_por_producto[prod_id]['cajas'] += cajas
            lineas_por_producto[prod_id]['subtotal'] = (
                precio_unitario * lineas_por_producto[prod_id]['cajas']
            )
        else:
            lineas_por_producto[prod_id] = {
                'producto_id': prod_id,
                'cajas': cajas,
                'precio_unitario': precio_unitario,
                'subtotal': precio_unitario * cajas,
            }
        idx += 1

    return list(lineas_por_producto.values())


def _cantidad_detalle_facturable(detalle):
    """Cantidad real utilizada para subtotalizar líneas de preparación."""
    cantidad = detalle.peso if detalle.peso else detalle.cajas
    return Decimal(str(cantidad or 0))


@app.route('/pedidos/nuevo', methods=['GET', 'POST'])
@login_required
@requiere_permiso_recurso('pedidos', 'crear')
def nuevo_pedido():
    # Obtener clientes según el vendedor
    if not isinstance(current_user, Vendedor):
        # Usuario del sistema anterior - todos los clientes
        clientes = Cliente.query.all()
    else:
        if current_user.rol.nombre == 'super_admin':
            # Super admin ve todos los clientes
            clientes = Cliente.query.all()
        else:
            # Vendedor regular: solo sus clientes asignados
            clientes = current_user.obtener_clientes_visibles()

    productos = Producto.query.all()

    # Enviamos al front-end cada producto con su precio (lista default)
    productos_dicts = [{
        'id'    : p.id,
        'nombre': p.nombre,
        'precio': float(
            obtener_precio_default_producto(p.id, 'base') or 0
        )
    } for p in productos]

    if request.method == 'POST':
        # 1) Cabecera
        cliente_id = int(request.form['cliente_id'])
        
        # VERIFICAR QUE EL VENDEDOR PUEDE CREAR PEDIDOS PARA ESTE CLIENTE
        if isinstance(current_user, Vendedor) and current_user.rol.nombre != 'super_admin':
            clientes_permitidos_ids = [c.id for c in current_user.obtener_clientes_visibles()]
            if cliente_id not in clientes_permitidos_ids:
                flash('No tienes permisos para crear pedidos para este cliente', 'error')
                return redirect(url_for('nuevo_pedido'))
        
        notas = request.form.get('notas')
        tipo_cambio = float(request.form.get('tipo_cambio', 1.0) or 1.0)

        # Validate form lines BEFORE creating the pedido row so a bad
        # payload doesn't leave an orphan empty pedido in the DB.
        try:
            lineas_form = _extraer_lineas_pedido_form(request.form, cliente_id)
        except _PedidoFormError as e:
            flash(str(e), 'error')
            return redirect(url_for('nuevo_pedido'))
        if not lineas_form:
            flash('Agrega al menos un producto al pedido', 'error')
            return redirect(url_for('nuevo_pedido'))

        pedido = Pedido(cliente_id=cliente_id, notas=notas, tipo_cambio=tipo_cambio)
        db.session.add(pedido)
        db.session.commit()
        _log_pedido_evento(pedido, 'creado', 'Pedido creado', commit=True)

        # 2) Detalle (resto del código igual)
        for linea in lineas_form:
            detalle = DetallePedido(
                pedido_id=pedido.id,
                producto_id=linea['producto_id'],
                cajas=linea['cajas'],
                cajas_pedidas=linea['cajas'],
                precio_unitario=linea['precio_unitario'],
                subtotal=linea['subtotal']
            )
            db.session.add(detalle)

        db.session.commit()

        # 2b) Auto-generar líneas de preparación para productos de importación
        for det in DetallePedido.query.filter_by(pedido_id=pedido.id, es_linea_pedido=True).all():
            prod = db.session.get(Producto, det.producto_id)
            if prod and not prod.se_pesa:
                prep = DetallePedido(
                    pedido_id=pedido.id,
                    producto_id=det.producto_id,
                    cajas=det.cajas,
                    cajas_pedidas=0,
                    peso=0,
                    precio_unitario=det.precio_unitario,
                    subtotal=round(float(det.precio_unitario) * det.cajas, 2),
                    es_linea_pedido=False,
                )
                db.session.add(prep)
        db.session.commit()

        # 3) Total del pedido (solo líneas originales)
        total = db.session.query(
            func.coalesce(func.sum(DetallePedido.subtotal), 0)
        ).filter_by(pedido_id=pedido.id, es_linea_pedido=True).scalar()

        if hasattr(pedido, 'total'):
            pedido.total = total
            db.session.commit()

        flash('Pedido creado con precios registrados.', 'success')
        return redirect(url_for('lista_pedidos'))

    return render_template(
        'pedido_form.html',
        clientes=clientes,
        productos=productos_dicts,
        pedido=None,
        productos_pedido=[]
    )

@app.route('/pedidos/<int:pedido_id>/editar', methods=['GET', 'POST'])
@login_required
@requiere_permiso_recurso('pedidos', 'editar')
def editar_pedido(pedido_id):
    pedido    = Pedido.query.get_or_404(pedido_id)

    # ── Verificación de autorización IDOR (cubre GET y POST) ───
    if not _user_can_manage_pedido(pedido):
        flash('No tienes permisos para editar este pedido', 'error')
        return redirect(url_for('lista_pedidos'))

    clientes  = Cliente.query.all()
    productos = Producto.query.all()

    # Para editar, sí tenemos el cliente del pedido, así que podemos usar precios específicos
    productos_dicts = [{
        'id'    : p.id,
        'nombre': p.nombre,
        # precio mostrado = el que vería ESTE cliente específico
        'precio': float(
            obtener_precio_producto_cliente(
                pedido.cliente_id,  # Ahora sí existe pedido
                p.id, 'base'
            ) or obtener_precio_default_producto(p.id, 'base') or 0
        )
    } for p in productos]

    
    if request.method == 'POST':
        if isinstance(current_user, Vendedor) and not current_user.puede_editar_pedido(pedido):
            flash('No tienes permisos para editar este pedido', 'error')
            return redirect(url_for('lista_pedidos'))

        # Pedidos facturados son inmutables — un cambio aquí divergiría
        # la DB local del invoice ya enviado a QuickBooks.
        if pedido.estado == 'facturado':
            flash('No se puede editar un pedido facturado', 'error')
            return redirect(url_for('detalles_pedido', pedido_id=pedido.id))

        nuevo_cliente_id = int(request.form['cliente_id'])
        if isinstance(current_user, Vendedor) and current_user.rol.nombre != 'super_admin':
            if not current_user.puede_crear_pedido_para_cliente(nuevo_cliente_id):
                flash('No tienes permisos para reasignar este pedido a ese cliente', 'error')
                return redirect(url_for('editar_pedido', pedido_id=pedido.id))

        try:
            lineas_form = _extraer_lineas_pedido_form(request.form, nuevo_cliente_id)
        except _PedidoFormError as e:
            flash(str(e), 'error')
            return redirect(url_for('editar_pedido', pedido_id=pedido.id))
        if not lineas_form:
            flash('Agrega al menos un producto al pedido', 'error')
            return redirect(url_for('editar_pedido', pedido_id=pedido.id))

        lineas_por_producto = {
            linea['producto_id']: linea for linea in lineas_form
        }
        productos_en_form = set(lineas_por_producto.keys())

        # Actualizar cabecera
        pedido.cliente_id = nuevo_cliente_id
        pedido.notas = request.form.get('notas')

        # Preservar líneas de preparación ya capturadas y solo resincronizar precios.
        productos_con_prep = set()
        detalles_prep = DetallePedido.query.filter_by(
            pedido_id=pedido.id,
            es_linea_pedido=False,
        ).all()
        for detalle in detalles_prep:
            if detalle.producto_id not in productos_en_form:
                db.session.delete(detalle)
                continue

            linea = lineas_por_producto[detalle.producto_id]
            detalle.precio_unitario = linea['precio_unitario']
            detalle.subtotal = linea['precio_unitario'] * _cantidad_detalle_facturable(detalle)
            productos_con_prep.add(detalle.producto_id)

        # Upsert líneas originales — NUNCA delete masivo: la FK
        # CajaPesada.detalle_pedido_id es ondelete=CASCADE, así que borrar
        # la línea original elimina en cascada todas las cajas pesadas
        # asociadas. Mantener el id del detalle preserva los pesos.
        existing_lineas = {
            d.producto_id: d
            for d in DetallePedido.query.filter_by(
                pedido_id=pedido.id,
                es_linea_pedido=True,
            ).all()
        }

        # Borra solo las líneas cuyo producto ya no está en el form.
        for producto_id, detalle in list(existing_lineas.items()):
            if producto_id not in productos_en_form:
                db.session.delete(detalle)
                del existing_lineas[producto_id]

        # Actualiza las que siguen, crea las nuevas.
        for linea in lineas_form:
            existente = existing_lineas.get(linea['producto_id'])
            if existente is not None:
                existente.cajas = linea['cajas']
                existente.cajas_pedidas = linea['cajas']
                existente.precio_unitario = linea['precio_unitario']
                existente.subtotal = linea['subtotal']
            else:
                detalle = DetallePedido(
                    pedido_id=pedido.id,
                    producto_id=linea['producto_id'],
                    cajas=linea['cajas'],
                    cajas_pedidas=linea['cajas'],
                    precio_unitario=linea['precio_unitario'],
                    subtotal=linea['subtotal']
                )
                db.session.add(detalle)

        # Auto-generar líneas de preparación para productos de importación
        for linea in lineas_form:
            prod = db.session.get(Producto, linea['producto_id'])
            if prod and not prod.se_pesa and linea['producto_id'] not in productos_con_prep:
                prep = DetallePedido(
                    pedido_id=pedido.id,
                    producto_id=linea['producto_id'],
                    cajas=linea['cajas'],
                    cajas_pedidas=0,
                    peso=0,
                    precio_unitario=linea['precio_unitario'],
                    subtotal=linea['precio_unitario'] * linea['cajas'],
                    es_linea_pedido=False,
                )
                db.session.add(prep)
                productos_con_prep.add(linea['producto_id'])
        db.session.commit()

        # Actualizar total del pedido (solo líneas originales)
        total = db.session.query(
            func.coalesce(func.sum(DetallePedido.subtotal), 0)
        ).filter_by(pedido_id=pedido.id, es_linea_pedido=True).scalar()

        if hasattr(pedido, 'total'):
            pedido.total = total
            db.session.commit()

        flash('Pedido actualizado.', 'success')
        return redirect(url_for('lista_pedidos'))

    # ----------- pre-cargar detalles (solo líneas originales) -----------
    productos_pedido = [{
        'id'     : d.producto.id,
        'nombre' : d.producto.nombre,
        'cajas'  : d.cajas,
        'precio' : float(d.precio_unitario)
    } for d in pedido.detalles if d.es_linea_pedido]

    return render_template(
        'pedido_form.html',
        clientes        = clientes,
        productos       = productos_dicts,
        pedido          = pedido,
        productos_pedido= productos_pedido
    )




@app.route('/pedidos/<int:pedido_id>/eliminar', methods=['POST'])
@login_required
@requiere_permiso_recurso('pedidos', 'eliminar')
def eliminar_pedido(pedido_id):
    pedido = Pedido.query.get_or_404(pedido_id)

    # IDOR guard: a vendedor with eliminar permission can't wipe
    # another territory's pedidos. super_admin passes through.
    if not _user_can_manage_pedido(pedido):
        flash('No tienes permisos para eliminar este pedido', 'error')
        return redirect(url_for('lista_pedidos'))

    # Estado guard: facturado pedidos must not be deleted — the QBO
    # invoice on the other side would be orphaned and the audit
    # trail destroyed.
    if pedido.estado == 'facturado':
        flash('No se puede eliminar un pedido facturado', 'error')
        return redirect(url_for('lista_pedidos'))

    # Audit before the cascade kicks in so we can reconstruct what
    # was lost from the event log.
    cajas_count = sum(d.cajas_pesadas_count for d in pedido.detalles)
    _log_pedido_evento(
        pedido,
        'eliminado',
        f'Pedido eliminado (estado={pedido.estado}, líneas={len(pedido.detalles)}, cajas_pesadas={cajas_count})',
        commit=True,
    )

    db.session.delete(pedido)
    db.session.commit()
    flash('Pedido eliminado.', 'success')
    return redirect(url_for('lista_pedidos'))


# ------------------------------------------------------------
#  Detalles de un pedido (agregar / ver / eliminar líneas)
# ------------------------------------------------------------
@app.route('/pedidos/<int:pedido_id>/detalles', methods=['GET', 'POST'])
@login_required
def detalles_pedido(pedido_id):
    # ── 1) Traer cabecera y productos ─────────────────────────
    pedido    = Pedido.query.get_or_404(pedido_id)

    # ── Verificación de autorización IDOR ─────────────────────
    if isinstance(current_user, Vendedor) and current_user.rol.nombre != 'super_admin':
        if not current_user.puede_ver_cliente(pedido.cliente_id):
            flash('No tienes permisos para acceder a este pedido', 'error')
            return redirect(url_for('lista_pedidos'))

    productos = Producto.query.all()
    detalle_context_key = f'last_detalle:{pedido.id}'

    # ── 2) Alta de un nuevo detalle ───────────────────────────
    if request.method == 'POST':
        # Agregar líneas requiere permiso de edición de pedidos
        if isinstance(current_user, Vendedor) and not current_user.tiene_permiso('pedidos', 'editar'):
            flash('No tienes permisos para editar pedidos', 'error')
            return redirect(url_for('detalles_pedido', pedido_id=pedido.id))
        # ── Inmutabilidad post-facturación ──────────────────────
        if pedido.estado == 'facturado':
            flash('No se puede agregar detalles a un pedido facturado', 'error')
            return redirect(url_for('detalles_pedido', pedido_id=pedido.id))

        producto_id       = int(request.form['producto_id'])
        peso              = float(request.form.get('peso', 0)  or 0)
        cajas             = int  (request.form.get('cajas', 0) or 0)   # reservado
        lote              = request.form.get('lote', '').strip()
        fecha_fabricacion = request.form.get('fecha_fabricacion', '').strip()
        fecha_expiracion  = request.form.get('fecha_expiracion', '').strip()
        session[detalle_context_key] = {
            'producto_id': producto_id,
            'lote': lote,
            'fecha_fabricacion': fecha_fabricacion,
            'fecha_expiracion': fecha_expiracion,
        }
        session.modified = True

        # ── Validación de trazabilidad obligatoria ──────────────
        producto_obj = db.session.get(Producto, producto_id)
        if producto_obj and producto_obj.se_pesa:
            if not all([lote, fecha_fabricacion, fecha_expiracion]):
                flash('Lote, fecha fabricación y fecha expiración son obligatorios', 'error')
                return redirect(url_for('detalles_pedido', pedido_id=pedido.id))

        # Importación: el form envía cajas en el campo peso
        if producto_obj and not producto_obj.se_pesa:
            cajas = int(peso)
            peso = 0

        # -------- Obtener precio unitario según la jerarquía --------
        precio_unitario = obtener_precio_producto_cliente(
                              pedido.cliente_id,   # 1️⃣ precio específico / lista cliente
                              producto_id,
                              'base'
                          )
        if precio_unitario is None:
            # 2️⃣ (fallback) lista de precios por defecto
            precio_unitario = obtener_precio_default_producto(
                                  producto_id, 'base'
                              ) or 0
        
        # ------------------------------------------------------------

        cantidad = cajas if cajas else peso   # si se usan cajas en el futuro
        subtotal = round(precio_unitario * cantidad, 2)

        # -------- Crear y guardar el detalle --------
        detalle = DetallePedido(
            pedido_id        = pedido.id,
            producto_id      = producto_id,
            cajas            = cajas,
            peso             = peso,
            lote             = lote,
            fecha_fabricacion= fecha_fabricacion,
            fecha_expiracion = fecha_expiracion,
            precio_unitario  = precio_unitario,
            subtotal         = subtotal,
            es_linea_pedido  = False
        )
        db.session.add(detalle)
        producto_nombre = producto_obj.nombre if producto_obj else '—'
        descripcion = (
            f'Línea agregada: {producto_nombre} ({peso} kg, lote {lote})'
            if producto_obj and producto_obj.se_pesa
            else f'Línea agregada: {producto_nombre} ({cajas} cajas)'
        )
        _log_pedido_evento(
            pedido,
            'linea_agregada',
            descripcion,
            meta={'producto_id': producto_id, 'peso': peso, 'cajas': cajas, 'lote': lote},
        )
        db.session.commit()

        flash('Detalle agregado.', 'success')
        return redirect(url_for('detalles_pedido', pedido_id=pedido.id))

    # ── 3) GET: mostrar plantilla ─────────────────────────────
    # Ordenar detalles por nombre de producto y luego por id
    detalles_ordenados = sorted(pedido.detalles, key=lambda d: (d.producto.nombre, d.id))

    # Calcular conteo de registros por producto para mostrar "#/Total"
    # Excluir líneas originales del pedido (es_linea_pedido=True)
    conteo_por_producto = {}
    for detalle in detalles_ordenados:
        if detalle.es_linea_pedido:
            continue
        pid = detalle.producto_id
        conteo_por_producto[pid] = conteo_por_producto.get(pid, 0) + 1

    # Asignar índice a cada detalle para mostrar "1/3", "2/3", etc.
    indice_detalle = {}
    contador_temp = {}
    for detalle in detalles_ordenados:
        if detalle.es_linea_pedido:
            continue
        pid = detalle.producto_id
        contador_temp[pid] = contador_temp.get(pid, 0) + 1
        indice_detalle[detalle.id] = contador_temp[pid]

    saved_detalle_context = session.get(detalle_context_key, {})

    eventos = (
        PedidoEvento.query
        .filter_by(pedido_id=pedido.id)
        .order_by(PedidoEvento.created_at.desc())
        .all()
    )

    # Mapeo: producto_id → primera línea de preparación (para importación no-pesable)
    prep_by_producto = {}
    for detalle in detalles_ordenados:
        if not detalle.es_linea_pedido and detalle.producto_id not in prep_by_producto:
            prep_by_producto[detalle.producto_id] = detalle

    # Sólo líneas originales del pedido (lo que pidió el cliente)
    lineas_originales = [d for d in detalles_ordenados if d.es_linea_pedido]

    return render_template('detalles_pedido.html',
                           pedido   = pedido,
                           productos= productos,
                           saved_detalle_context=saved_detalle_context,
                           conteo_por_producto=conteo_por_producto,
                           indice_detalle=indice_detalle,
                           detalles_ordenados=detalles_ordenados,
                           lineas_originales=lineas_originales,
                           prep_by_producto=prep_by_producto,
                           eventos=eventos,
                           pedido_total_xcg=_calcular_venta_pedido(pedido),
                           tiene_productos_pesables=_pedido_tiene_productos_pesables(pedido))


@app.route('/pedidos/<int:pedido_id>/pesar', methods=['GET'])
@login_required
@requiere_permiso_recurso('pedidos', 'editar')
def pesar_pedido(pedido_id):
    pedido = _load_pedido_for_pesar(pedido_id)

    if not _user_can_manage_pedido(pedido):
        flash('No tienes permisos para pesar este pedido', 'error')
        return redirect(url_for('lista_pedidos'))

    if pedido.estado == 'facturado':
        flash('No se puede pesar un pedido facturado', 'error')
        return redirect(url_for('detalles_pedido', pedido_id=pedido.id))

    if not _pedido_tiene_productos_pesables(pedido):
        flash('Este pedido no tiene productos que se pesen', 'info')
        return redirect(url_for('detalles_pedido', pedido_id=pedido.id))

    active_detalle_id = request.args.get('detalle_id', type=int)
    return render_template('pesar.html', **_build_pesar_context(pedido, active_detalle_id=active_detalle_id))


@app.route('/pedidos/<int:pedido_id>/pesar/caja', methods=['POST'])
@login_required
@requiere_permiso_recurso('pedidos', 'editar')
def registrar_caja_pesada(pedido_id):
    pedido = _load_pedido_for_pesar(pedido_id)

    if not _user_can_manage_pedido(pedido):
        return _htmx_error_response('No tienes permisos para pesar este pedido', status=403)

    if pedido.estado == 'facturado':
        return _htmx_error_response('No se puede registrar cajas en un pedido facturado', status=409)

    detalle_id = request.form.get('detalle_pedido_id', type=int)
    detalle = next((
        item for item in pedido.detalles
        if item.id == detalle_id and item.es_linea_pedido and item.producto and item.producto.se_pesa
    ), None)
    if detalle is None:
        return _htmx_error_response('Detalle de pedido inválido', status=404)

    peso, peso_error = _parse_peso_caja(request.form.get('peso'))
    if peso_error:
        return _htmx_error_response(peso_error)

    lote = (request.form.get('lote') or '').strip()
    if not lote or len(lote) > 50:
        return _htmx_error_response('El lote es obligatorio y debe tener máximo 50 caracteres')

    fecha_elaboracion, fecha_error = _parse_iso_date_field(
        request.form.get('fecha_elaboracion'),
        'La fecha de elaboración',
    )
    if fecha_error:
        return _htmx_error_response(fecha_error)

    fecha_vencimiento, fecha_venc_error = _parse_iso_date_field(
        request.form.get('fecha_vencimiento'),
        'La fecha de vencimiento',
    )
    if fecha_venc_error:
        return _htmx_error_response(fecha_venc_error)

    if fecha_vencimiento < fecha_elaboracion:
        return _htmx_error_response('La fecha de vencimiento no puede ser anterior a la elaboración')

    siguiente_numero = max((caja.numero for caja in detalle.cajas_pesadas), default=0) + 1
    caja = CajaPesada(
        detalle_pedido_id=detalle.id,
        numero=siguiente_numero,
        peso=peso,
        lote=lote,
        fecha_elaboracion=fecha_elaboracion,
        fecha_vencimiento=fecha_vencimiento,
        pesado_por=current_user.id if isinstance(current_user, Vendedor) else None,
    )
    db.session.add(caja)
    _log_pedido_evento(
        pedido,
        'caja_pesada',
        f'Caja #{siguiente_numero:02d} de {detalle.producto.nombre}: {peso} kg (lote {lote})',
        meta={'detalle_id': detalle.id, 'numero': siguiente_numero, 'peso': float(peso), 'lote': lote},
    )
    db.session.commit()

    pedido = _load_pedido_for_pesar(pedido_id)
    detalle = next(item for item in pedido.detalles if item.id == detalle_id)
    return _render_pesar_cajas_partial(pedido, detalle)


@app.route('/cajas/<int:caja_id>/edit', methods=['GET'])
@login_required
@requiere_permiso_recurso('pedidos', 'editar')
def editar_caja_pesada_modal(caja_id):
    caja = (
        CajaPesada.query.options(
            joinedload(CajaPesada.detalle_pedido).joinedload(DetallePedido.pedido),
            joinedload(CajaPesada.detalle_pedido).joinedload(DetallePedido.producto),
        )
        .filter_by(id=caja_id)
        .first_or_404()
    )
    pedido = caja.detalle_pedido.pedido

    if not _user_can_manage_pedido(pedido):
        return _htmx_error_response('No tienes permisos para editar esta caja', status=403)

    if pedido.estado == 'facturado':
        return _htmx_error_response('No se puede editar una caja de un pedido facturado', status=409)

    return render_template(
        'partials/pesar_caja_edit_modal.html',
        caja=caja,
        detalle=caja.detalle_pedido,
        pedido=pedido,
    )


@app.route('/cajas/<int:caja_id>', methods=['PATCH'])
@login_required
@requiere_permiso_recurso('pedidos', 'editar')
def actualizar_caja_pesada(caja_id):
    caja = (
        CajaPesada.query.options(
            joinedload(CajaPesada.detalle_pedido).joinedload(DetallePedido.pedido),
            joinedload(CajaPesada.detalle_pedido).joinedload(DetallePedido.producto),
        )
        .filter_by(id=caja_id)
        .first_or_404()
    )
    pedido = caja.detalle_pedido.pedido

    if not _user_can_manage_pedido(pedido):
        return _htmx_error_response('No tienes permisos para editar esta caja', status=403)

    if pedido.estado == 'facturado':
        return _htmx_error_response('No se puede editar una caja de un pedido facturado', status=409)

    peso, peso_error = _parse_peso_caja(request.form.get('peso'))
    if peso_error:
        return _htmx_error_response(peso_error)

    lote = (request.form.get('lote') or '').strip()
    if not lote or len(lote) > 50:
        return _htmx_error_response('El lote es obligatorio y debe tener máximo 50 caracteres')

    fecha_elaboracion, fecha_error = _parse_iso_date_field(
        request.form.get('fecha_elaboracion'),
        'La fecha de elaboración',
    )
    if fecha_error:
        return _htmx_error_response(fecha_error)

    fecha_vencimiento, fecha_venc_error = _parse_iso_date_field(
        request.form.get('fecha_vencimiento'),
        'La fecha de vencimiento',
    )
    if fecha_venc_error:
        return _htmx_error_response(fecha_venc_error)

    if fecha_vencimiento < fecha_elaboracion:
        return _htmx_error_response('La fecha de vencimiento no puede ser anterior a la elaboración')

    caja.peso = peso
    caja.lote = lote
    caja.fecha_elaboracion = fecha_elaboracion
    caja.fecha_vencimiento = fecha_vencimiento
    _log_pedido_evento(
        pedido,
        'caja_editada',
        f'Caja #{caja.numero:02d} de {caja.detalle_pedido.producto.nombre} editada ({peso} kg, lote {lote})',
        meta={'caja_id': caja.id, 'numero': caja.numero, 'peso': float(peso), 'lote': lote},
    )
    db.session.commit()

    pedido = _load_pedido_for_pesar(pedido.id)
    detalle = next(item for item in pedido.detalles if item.id == caja.detalle_pedido_id)
    return _render_pesar_cajas_partial(pedido, detalle)


@app.route('/cajas/<int:caja_id>', methods=['DELETE'])
@login_required
@requiere_permiso_recurso('pedidos', 'editar')
def eliminar_caja_pesada(caja_id):
    caja = (
        CajaPesada.query.options(
            joinedload(CajaPesada.detalle_pedido).joinedload(DetallePedido.pedido),
            joinedload(CajaPesada.detalle_pedido).joinedload(DetallePedido.producto),
        )
        .filter_by(id=caja_id)
        .first_or_404()
    )
    detalle_id = caja.detalle_pedido_id
    pedido = caja.detalle_pedido.pedido

    if not _user_can_manage_pedido(pedido):
        return _htmx_error_response('No tienes permisos para eliminar esta caja', status=403)

    if pedido.estado == 'facturado':
        return _htmx_error_response('No se puede eliminar una caja de un pedido facturado', status=409)

    detalle = caja.detalle_pedido
    numero_eliminada = caja.numero
    producto_nombre = detalle.producto.nombre
    db.session.delete(caja)
    db.session.flush()
    _renumerar_cajas_pesadas(detalle)
    _log_pedido_evento(
        pedido,
        'caja_eliminada',
        f'Caja #{numero_eliminada:02d} de {producto_nombre} eliminada',
        meta={'numero': numero_eliminada, 'detalle_id': detalle.id},
    )
    db.session.commit()

    pedido = _load_pedido_for_pesar(pedido.id)
    detalle = next(item for item in pedido.detalles if item.id == detalle_id)
    return _render_pesar_cajas_partial(pedido, detalle)


@app.route('/pedidos/<int:pedido_id>/pesar/finalizar', methods=['POST'])
@login_required
@requiere_permiso_recurso('pedidos', 'editar')
def finalizar_pesaje_pedido(pedido_id):
    pedido = _load_pedido_for_pesar(pedido_id)

    if not _user_can_manage_pedido(pedido):
        flash('No tienes permisos para finalizar este pesaje', 'error')
        return redirect(url_for('lista_pedidos'))

    if pedido.estado == 'facturado':
        flash('No se puede finalizar el pesaje de un pedido facturado', 'error')
        return redirect(url_for('detalles_pedido', pedido_id=pedido.id))

    errores = _validar_preparacion_pedido(pedido)
    if errores:
        flash('No se puede finalizar el pesaje. Hay datos incompletos:', 'error')
        for error in errores:
            flash(error, 'error')
        return redirect(url_for('pesar_pedido', pedido_id=pedido.id))

    pedido.estado = 'preparado'
    _log_pedido_evento(
        pedido,
        'pesaje_finalizado',
        'Pesaje finalizado y pedido marcado como preparado',
        meta={
            'cajas_pesadas': _pedido_cajas_pesadas_total(pedido),
            'peso_total': float(_pedido_peso_total(pedido)),
        },
    )
    db.session.commit()
    flash('Pesaje finalizado y pedido marcado como preparado', 'success')
    return redirect(url_for('detalles_pedido', pedido_id=pedido.id))


@app.route('/detalles_pedido/<int:detalle_id>/eliminar', methods=['POST'])
@login_required
@requiere_permiso_recurso('pedidos', 'eliminar')
def eliminar_detalle_pedido(detalle_id):
    detalle = DetallePedido.query.get_or_404(detalle_id)
    pedido = Pedido.query.get_or_404(detalle.pedido_id)

    # ── Verificación de autorización IDOR ─────────────────────
    if isinstance(current_user, Vendedor) and current_user.rol.nombre != 'super_admin':
        if not current_user.puede_ver_cliente(pedido.cliente_id):
            flash('No tienes permisos para eliminar este detalle', 'error')
            return redirect(url_for('lista_pedidos'))

    # ── Inmutabilidad post-facturación ─────────────────────────
    if pedido.estado == 'facturado':
        flash('No se puede eliminar un detalle de un pedido facturado', 'error')
        return redirect(url_for('detalles_pedido', pedido_id=pedido.id))

    producto_nombre = detalle.producto.nombre if detalle.producto else '—'

    # Eliminar el producto del pedido = borrar TODAS sus filas: la línea
    # original (es_linea_pedido=True) y la(s) línea(s) de preparación
    # (es_linea_pedido=False). Si solo se borrara la línea original, la prep
    # quedaría huérfana y pedido_a_json la seguiría facturando.
    filas = DetallePedido.query.filter_by(
        pedido_id=pedido.id,
        producto_id=detalle.producto_id,
    ).all()
    ids_borrados = [d.id for d in filas]
    for fila in filas:
        db.session.delete(fila)
    _log_pedido_evento(
        pedido,
        'linea_eliminada',
        f'Línea eliminada: {producto_nombre}',
        meta={'detalle_ids': ids_borrados},
    )
    db.session.commit()
    flash('Detalle eliminado.', 'success')
    return redirect(url_for('detalles_pedido', pedido_id=pedido.id))


@app.route('/detalles_pedido/<int:detalle_id>/editar', methods=['POST'])
@login_required
@requiere_permiso_recurso('pedidos', 'editar')
def editar_detalle_pedido(detalle_id):
    """Edita un detalle de pedido existente."""
    detalle = DetallePedido.query.get_or_404(detalle_id)
    pedido = Pedido.query.get_or_404(detalle.pedido_id)

    # ── Verificación de autorización IDOR ─────────────────────
    if isinstance(current_user, Vendedor) and current_user.rol.nombre != 'super_admin':
        if not current_user.puede_ver_cliente(pedido.cliente_id):
            flash('No tienes permisos para editar este detalle', 'error')
            return redirect(url_for('lista_pedidos'))

    # ── Inmutabilidad post-facturación ─────────────────────────
    if pedido.estado == 'facturado':
        flash('No se puede editar un pedido facturado', 'error')
        return redirect(url_for('detalles_pedido', pedido_id=pedido.id))

    # Obtener datos del formulario
    producto_id = request.form.get('producto_id', type=int)
    peso = request.form.get('peso', type=float) or 0
    lote = request.form.get('lote', '').strip()
    fecha_fabricacion = request.form.get('fecha_fabricacion', '')
    fecha_expiracion = request.form.get('fecha_expiracion', '')

    producto = db.session.get(Producto, producto_id) if producto_id else None

    # Validar según tipo de producto
    if not producto:
        flash('Producto no válido', 'error')
        return redirect(url_for('detalles_pedido', pedido_id=pedido.id))

    if producto.se_pesa:
        # Manufactura: peso, lote, fechas obligatorios
        if not all([peso, lote, fecha_fabricacion, fecha_expiracion]):
            flash('Peso, lote, fecha fabricación y fecha expiración son requeridos', 'error')
            return redirect(url_for('detalles_pedido', pedido_id=pedido.id))

    # Actualizar el detalle
    detalle.producto_id = producto_id
    detalle.lote = lote or None
    detalle.fecha_fabricacion = fecha_fabricacion or None
    detalle.fecha_expiracion = fecha_expiracion or None
    if producto.se_pesa:
        detalle.peso = peso
    else:
        cajas = int(peso)  # el form envía cajas en el campo peso
        detalle.cajas = cajas
        detalle.subtotal = (detalle.precio_unitario or 0) * cajas
        detalle.fecha_expiracion = fecha_expiracion if fecha_expiracion else None

        # Sincronizar la línea original (la que muestra la tarjeta y de la que
        # se calcula el total del pedido) con la nueva cantidad. El modal edita
        # la línea de preparación; sin esta sincronización la tarjeta seguiría
        # mostrando la cantidad vieja.
        if not detalle.es_linea_pedido:
            original = DetallePedido.query.filter_by(
                pedido_id=pedido.id,
                producto_id=producto_id,
                es_linea_pedido=True,
            ).first()
            if original is not None:
                original.cajas = cajas
                original.cajas_pedidas = cajas
                original.subtotal = (original.precio_unitario or 0) * cajas

    db.session.commit()
    flash('Detalle actualizado correctamente.', 'success')
    return redirect(url_for('detalles_pedido', pedido_id=pedido.id))


# ---------------------------------------------------------------------
# Generar etiquetas a partir de los DetallePedido de un pedido concreto
# -> Genera PDF 4" x 2", una etiqueta por página (para PDF Direct)
# ---------------------------------------------------------------------
@app.route('/generar_etiqueta_detalle/<int:pedido_id>', methods=['GET', 'POST'])
@login_required
def generar_etiqueta_detalle(pedido_id):
    """
    Genera un PDF con etiquetas 4x2 (una por página) para PDF Direct.
    Soporta GET y POST para mejor compatibilidad con iOS.
    """
    try:
        pedido = Pedido.query.get_or_404(pedido_id)

        # ── Verificación de autorización IDOR ─────────────────────
        if not _user_can_manage_pedido(pedido):
            flash('No tienes permisos para ver este pedido', 'error')
            return redirect(url_for('lista_pedidos'))

        # Obtener parámetros desde GET o POST
        fecha_ini = request.args.get('fecha_inicio') if request.method == 'GET' else request.form.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin') if request.method == 'GET' else request.form.get('fecha_fin')

        # Validar fechas
        if not fecha_ini or not fecha_fin:
            error_msg = "Debe indicar fecha de inicio y fin"
            if request.method == 'GET' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({"error": error_msg}), 400
            flash(error_msg, "danger")
            return redirect(url_for('detalles_pedido', pedido_id=pedido_id))

        try:
            datetime.strptime(fecha_ini, '%Y-%m-%d')
            datetime.strptime(fecha_fin, '%Y-%m-%d')
        except ValueError:
            error_msg = "Formato de fecha inválido. Use YYYY-MM-DD"
            if request.method == 'GET' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({"error": error_msg}), 400
            flash("Formato de fecha inválido", "danger")
            return redirect(url_for('detalles_pedido', pedido_id=pedido_id))

        items = _build_label_items_for_pedido(pedido, fecha_ini, fecha_fin)

        if not items:
            error_msg = "No hay detalles en ese rango de fechas"
            if request.method == 'GET' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({"error": error_msg}), 404
            flash(error_msg, "warning")
            return redirect(url_for('detalles_pedido', pedido_id=pedido_id))

        # Crear PDF
        output, c = create_single_label_pdf()
        logo_path = get_logo_path(basedir)
        cliente_nombre = pedido.cliente.nombre if getattr(pedido, "cliente", None) else ""

        for item in items:
            draw_order_label(
                c, logo_path,
                client=cliente_nombre,
                product=item['producto_nombre'],
                temperature=item['temperatura'],
                lot=item['lote'],
                mfg_date=item['fecha_fabricacion'],
                exp_date=item['fecha_expiracion'],
                weight=item['peso_label']
            )
            c.showPage()

        c.save()
        output.seek(0)

        # Nombre del archivo
        nombre_cliente = cliente_nombre.replace(" ", "_").replace("/", "-") or "cliente"
        filename = f"etiquetas_{nombre_cliente}_{pedido_id}.pdf"

        # Response con headers optimizados para iOS
        response = make_response(send_file(
            output,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename
        ))
        response.headers['Content-Disposition'] = f'{"inline" if _is_ios_request() else "attachment"}; filename="{filename}"'
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        response.headers['X-Content-Type-Options'] = 'nosniff'

        return response

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error generando etiquetas: {e}")

        if request.method == 'GET' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({"error": "Error interno del servidor"}), 500

        flash("Error generando etiquetas. Intente de nuevo.", "danger")
        return redirect(url_for('detalles_pedido', pedido_id=pedido_id))


# ---------------------------------------------------------------------
# Generar etiquetas A4 (2 por página) a partir de los DetallePedido
# ---------------------------------------------------------------------
@app.route('/generar_etiqueta_detalle_a4/<int:pedido_id>', methods=['GET', 'POST'])
@login_required
def generar_etiqueta_detalle_a4(pedido_id):
    """
    Genera un PDF con etiquetas en formato A4 (2 etiquetas por página).
    """
    try:
        pedido = Pedido.query.get_or_404(pedido_id)

        # ── Verificación de autorización IDOR ─────────────────────
        if not _user_can_manage_pedido(pedido):
            flash('No tienes permisos para ver este pedido', 'error')
            return redirect(url_for('lista_pedidos'))

        # Obtener parámetros desde GET o POST
        fecha_ini = request.args.get('fecha_inicio') if request.method == 'GET' else request.form.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin') if request.method == 'GET' else request.form.get('fecha_fin')

        # Validar fechas
        if not fecha_ini or not fecha_fin:
            error_msg = "Debe indicar fecha de inicio y fin"
            if request.method == 'GET' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({"error": error_msg}), 400
            flash(error_msg, "danger")
            return redirect(url_for('detalles_pedido', pedido_id=pedido_id))

        try:
            datetime.strptime(fecha_ini, '%Y-%m-%d')
            datetime.strptime(fecha_fin, '%Y-%m-%d')
        except ValueError:
            error_msg = "Formato de fecha inválido. Use YYYY-MM-DD"
            if request.method == 'GET' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({"error": error_msg}), 400
            flash("Formato de fecha inválido", "danger")
            return redirect(url_for('detalles_pedido', pedido_id=pedido_id))

        items = _build_label_items_for_pedido(pedido, fecha_ini, fecha_fin)

        if not items:
            error_msg = "No hay detalles en ese rango de fechas"
            if request.method == 'GET' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({"error": error_msg}), 404
            flash(error_msg, "warning")
            return redirect(url_for('detalles_pedido', pedido_id=pedido_id))

        # Crear PDF A4
        output, c, page_width, page_height = create_a4_page_pdf()
        x_offset, y_top, y_bottom = get_a4_label_positions(page_width, page_height)
        logo_path = get_logo_path(basedir)
        cliente_nombre = pedido.cliente.nombre if getattr(pedido, "cliente", None) else ""

        etiqueta_contador = 0

        for item in items:
            # Determinar posición (arriba o abajo)
            y_offset = y_top if etiqueta_contador % 2 == 0 else y_bottom

            draw_order_label_a4(
                c, logo_path, cliente_nombre, item['producto_nombre'],
                item['temperatura'], item['lote'], item['fecha_fabricacion'], item['fecha_expiracion'], item['peso_label'],
                x_offset, y_offset
            )

            etiqueta_contador += 1

            # Nueva página después de 2 etiquetas
            if etiqueta_contador % 2 == 0:
                c.showPage()

        # Cerrar última página si quedó incompleta
        if etiqueta_contador % 2 != 0:
            c.showPage()

        c.save()
        output.seek(0)

        # Nombre del archivo
        nombre_cliente = (pedido.cliente.nombre if getattr(pedido, "cliente", None) else "cliente").replace(" ", "_").replace("/", "-")
        filename = f"etiquetas_{nombre_cliente}_{pedido_id}.pdf"

        # Crear response
        response = make_response(send_file(
            output,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename
        ))

        response.headers['Content-Disposition'] = f'{"inline" if _is_ios_request() else "attachment"}; filename="{filename}"'
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        response.headers['X-Content-Type-Options'] = 'nosniff'

        return response

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error generando etiquetas A4: {e}")

        if request.method == 'GET' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({"error": "Error interno del servidor"}), 500

        flash("Error generando etiquetas A4. Intente de nuevo.", "danger")
        return redirect(url_for('detalles_pedido', pedido_id=pedido_id))


@app.route('/pedidos/<int:pedido_id>/preparar', methods=['GET', 'POST'])
@login_required
def preparar_pedido(pedido_id):
    """Deprecated: redirect to detalles_pedido (Story 3-0)."""
    return redirect(url_for('detalles_pedido', pedido_id=pedido_id), code=301)


@app.route('/pedidos/<int:pedido_id>/marcar_preparado', methods=['POST'])
@login_required
@requiere_permiso_recurso('pedidos', 'editar')
def marcar_preparado(pedido_id):
    """Valida trazabilidad y marca pedido como preparado (Story 3-0)."""
    pedido = Pedido.query.get_or_404(pedido_id)

    # ── Verificación de autorización IDOR ─────────────────────
    if not _user_can_manage_pedido(pedido):
        flash('No tienes permisos para modificar este pedido', 'error')
        return redirect(url_for('lista_pedidos'))

    # ── Inmutabilidad post-facturación ─────────────────────────
    if pedido.estado == 'facturado':
        flash('No se puede modificar un pedido ya facturado', 'error')
        return redirect(url_for('detalles_pedido', pedido_id=pedido.id))

    if pedido.estado == 'preparado':
        flash('El pedido ya está marcado como preparado', 'info')
        return redirect(url_for('detalles_pedido', pedido_id=pedido.id))

    errores = _validar_preparacion_pedido(pedido)

    if errores:
        flash('No se puede marcar como preparado. Datos incompletos:', 'error')
        for err in errores:
            flash(err, 'error')
        return redirect(url_for('detalles_pedido', pedido_id=pedido.id))

    pedido.estado = 'preparado'
    _log_pedido_evento(pedido, 'preparado', 'Pedido marcado como preparado')
    db.session.commit()
    flash('Pedido marcado como preparado', 'success')
    return redirect(url_for('detalles_pedido', pedido_id=pedido.id))


N8N_WEBHOOK_URL = os.environ.get("N8N_WEBHOOK_URL")  # ponlo en tu .env
try:
    N8N_WEBHOOK_TIMEOUT = int(os.environ.get("N8N_WEBHOOK_TIMEOUT", 30))
except (ValueError, TypeError):
    N8N_WEBHOOK_TIMEOUT = 30

N8N_INVOICE_FETCH_WEBHOOK_URL = os.environ.get('N8N_INVOICE_FETCH_WEBHOOK_URL', '').strip()
try:
    N8N_INVOICE_FETCH_TIMEOUT = int(os.environ.get('N8N_INVOICE_FETCH_TIMEOUT', 20))
except (ValueError, TypeError):
    N8N_INVOICE_FETCH_TIMEOUT = 20


def _obtener_factura_qbo(invoice_id):
    """Pide a n8n la factura vigente en QuickBooks.

    Se consulta en vivo en lugar de guardar un snapshot al facturar porque las
    facturas se corrigen a mano en QBO cuando la lista de precios de la app
    está desactualizada. Devuelve None ante cualquier fallo; quien llama decide
    qué mostrar.
    """
    if not N8N_INVOICE_FETCH_WEBHOOK_URL:
        app.logger.warning('N8N_INVOICE_FETCH_WEBHOOK_URL no configurada')
        return None
    try:
        resp = requests.post(
            N8N_INVOICE_FETCH_WEBHOOK_URL,
            json={'invoice_id': str(invoice_id)},
            timeout=N8N_INVOICE_FETCH_TIMEOUT,
            headers=_webhook_headers(),
        )
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        app.logger.error(f'No se pudo obtener la factura {invoice_id} de QBO: {e}')
        return None


# Una diferencia por debajo de medio centavo es redondeo, no una corrección de
# precio: sin este umbral la pantalla se llena de ruido.
UMBRAL_DIFERENCIA_PRECIO = 0.005


def _comparar_precios_factura(pedido, factura_json):
    """Compara los precios de la factura vigente en QBO contra los de la app.

    Devuelve (filas, avisos).

    Todo se compara sobre `precio_base`, que es tax-exclusive: el
    `precio_unitario` de la línea se resuelve con
    `obtener_precio_producto_cliente(..., 'base')` y el `UnitPrice` de QBO
    tampoco lleva OB (QuickBooks lo aplica encima), así que las tres cifras
    son comparables sin conversión.
    """
    from utils.factura_pdf import extraer_datos_factura

    lineas_factura = (extraer_datos_factura(factura_json) or {}).get('lineas') or []

    # Precio que la app mandó a facturar, por producto. Se muestra al lado para
    # que se vea la corrección; la decisión de escribir se toma contra el vigente.
    #
    # La selección de líneas espeja la de `pedido_a_json`, que es lo que se envía
    # a facturar: los productos pesables se facturan desde su línea ORIGINAL (vía
    # sus CajaPesada) y el resto desde las líneas de preparación. Mirar solo las
    # de preparación deja fuera media facturación.
    facturado = {}
    for detalle in _pedido_detalles_pesables(pedido):
        if detalle.cajas_pesadas_count:
            facturado.setdefault(detalle.producto_id, float(detalle.precio_unitario or 0))

    for d in pedido.detalles:
        if d.es_linea_pedido:
            continue
        facturado.setdefault(d.producto_id, float(d.precio_unitario or 0))

    filas = []
    avisos = []
    productos_en_factura = set()
    por_producto = {}

    for linea in lineas_factura:
        qbo_id = linea.get('item_qbo_id')
        producto = Producto.query.filter_by(qbo_id=qbo_id).first() if qbo_id else None
        precio_qbo = float(linea.get('rate') or 0)
        qty = float(linea.get('qty') or 0)

        if producto is None:
            # Nunca descartar en silencio: la línea existe en la factura y el
            # usuario tiene que poder verla aunque no se pueda actuar sobre ella.
            filas.append({
                'producto_id': None,
                'nombre': linea.get('producto') or '(sin nombre)',
                'qty': qty,
                'precio_qbo': precio_qbo,
                'precio_facturado': None,
                'precio_vigente': None,
                'estado': 'sin_producto',
                'motivo': 'No corresponde a ningún producto de la app.',
            })
            continue

        productos_en_factura.add(producto.id)

        # Un producto puede venir en varias líneas de la misma factura: el payload
        # emite una línea por caja pesada. Si todas traen el mismo precio se suman
        # las cantidades; si traen precios distintos no hay un "precio nuevo" único
        # que aplicar, y hay que decirlo en vez de elegir uno al azar.
        anterior = por_producto.get(producto.id)
        if anterior is not None:
            anterior['qty'] += qty
            if abs(anterior['precio_qbo'] - precio_qbo) >= UMBRAL_DIFERENCIA_PRECIO:
                anterior['estado'] = 'precio_ambiguo'
                anterior['motivo'] = (
                    'La factura trae este producto con más de un precio; '
                    'corregilo en QuickBooks para que quede uno solo.'
                )
            continue

        precio_vigente = obtener_precio_producto_cliente(pedido.cliente_id, producto.id, 'base')
        heredado = False
        if precio_vigente is None:
            precio_vigente = obtener_precio_default_producto(producto.id, 'base')
            heredado = precio_vigente is not None
        else:
            # ¿El precio vigente es propio del cliente o lo hereda de una lista?
            heredado = PrecioClienteProducto.query.filter_by(
                cliente_id=pedido.cliente_id, producto_id=producto.id, activo=True,
            ).first() is None

        motivo = None
        if precio_qbo <= 0:
            # Escribir un precio 0 en la lista es justo lo que la guarda de
            # facturación impide del otro lado.
            estado = 'precio_invalido'
            motivo = 'La factura trae precio 0; no se puede usar como precio de lista.'
        elif (precio_vigente is not None
                and abs(precio_qbo - float(precio_vigente)) < UMBRAL_DIFERENCIA_PRECIO):
            # La lista ya coincide con la factura. Pero si el pedido se facturó a
            # otro precio, el problema no era la lista: el pedido nació mal
            # cobrado. No hay nada que actualizar y hay que decirlo igual, porque
            # si no "no hay diferencias" tapa una factura que sí difiere.
            facturado_prod = facturado.get(producto.id)
            if (facturado_prod is not None
                    and abs(facturado_prod - precio_qbo) >= UMBRAL_DIFERENCIA_PRECIO):
                estado = 'pedido_desfasado'
                motivo = (
                    f'El pedido se facturó a {facturado_prod:.2f} pero el precio '
                    f'de este cliente es {float(precio_vigente):.2f}, que es el que '
                    f'tiene la factura. La lista está bien: el precio quedó mal al '
                    f'cargar el pedido.'
                )
            else:
                estado = 'igual'
        else:
            # Sin precio vigente (producto nuevo) la diferencia también es real:
            # es justo el caso que dejó una línea en 0 en su momento.
            estado = 'difiere'

        fila = {
            'producto_id': producto.id,
            'nombre': producto.nombre,
            'qty': qty,
            'precio_qbo': precio_qbo,
            'precio_facturado': facturado.get(producto.id),
            'precio_vigente': float(precio_vigente) if precio_vigente is not None else None,
            'heredado': heredado,
            'estado': estado,
            'motivo': motivo,
        }
        filas.append(fila)
        por_producto[producto.id] = fila

    productos_del_pedido = set(facturado.keys())

    if not productos_del_pedido:
        avisos.append(
            'No se pudo determinar qué precios mandó la app a facturar para este '
            'pedido, así que solo se compara contra el precio vigente del cliente.'
        )
    else:
        faltantes = productos_del_pedido - productos_en_factura
        if faltantes:
            nombres = _nombres_de_productos(faltantes)
            avisos.append(
                'La factura de QuickBooks no cubre todas las líneas del pedido. '
                f'Sin contraparte en la factura: {nombres}. Puede haber una segunda '
                'factura para este pedido.'
            )

        sobrantes = productos_en_factura - productos_del_pedido
        if sobrantes:
            nombres = _nombres_de_productos(sobrantes)
            avisos.append(
                'La factura trae líneas que no están en el pedido: '
                f'{nombres}.'
            )

    return filas, avisos


def _nombres_de_productos(producto_ids):
    productos = Producto.query.filter(Producto.id.in_(list(producto_ids))).all()
    return ', '.join(sorted(p.nombre for p in productos)) or '(desconocidos)'


N8N_DRIVE_WEBHOOK_URL = os.environ.get('N8N_DRIVE_WEBHOOK_URL', '').strip()
# El archivado corre en línea, antes de devolver el PDF, así que su timeout se
# suma al de la consulta a QBO (N8N_INVOICE_FETCH_TIMEOUT = 20s). El router de
# Heroku corta la conexión a los 30s (H12), así que el peor caso combinado
# (20 + 8 = 28s) tiene que quedar por debajo de ese límite: si Drive tarda, el
# usuario igual recibe la factura que ya está generada en memoria.
try:
    N8N_DRIVE_TIMEOUT = int(os.environ.get('N8N_DRIVE_TIMEOUT', 8))
except (ValueError, TypeError):
    N8N_DRIVE_TIMEOUT = 8


def _archivar_factura_drive(pdf_bytes, filename):
    """Sube el PDF a Google Drive vía n8n. Best-effort: nunca lanza.

    Si falla, el usuario igual recibe su PDF; solo queda sin archivar.
    """
    if not N8N_DRIVE_WEBHOOK_URL:
        app.logger.info('N8N_DRIVE_WEBHOOK_URL no configurada; se omite el archivado')
        return False
    try:
        resp = requests.post(
            N8N_DRIVE_WEBHOOK_URL,
            json={
                'filename': filename,
                'pdf_base64': base64.b64encode(pdf_bytes).decode('ascii'),
            },
            timeout=N8N_DRIVE_TIMEOUT,
            headers=_webhook_headers(),
        )
        resp.raise_for_status()
        return True
    except Exception as e:
        app.logger.warning(f'No se pudo archivar {filename} en Drive: {e}')
        return False


@app.route('/pedidos/<int:pedido_id>/facturar', methods=['POST'])
@login_required
@requiere_permiso_recurso('pedidos', 'editar')
def facturar_pedido(pedido_id):
    pedido = Pedido.query.get_or_404(pedido_id)

    # ── Verificación de autorización IDOR ─────────────────────
    if not _user_can_manage_pedido(pedido):
        flash('No tienes permisos para facturar este pedido', 'error')
        return redirect(url_for('lista_pedidos'))

    # Sólo facturar si aún no fue facturado. Ojo: N8N no devuelve invoice_id, así
    # que invoice_id_qbo está NULL incluso en pedidos que sí se facturaron — no se
    # puede usar como prueba de "ya facturado". Sin esta guarda el pedido 1264 se
    # envió dos veces el 2026-08-14. El reintento exige confirmación explícita.
    if pedido.estado == 'facturado':
        if pedido.invoice_id_qbo:
            flash('El pedido ya está facturado.', 'info')
            return redirect(url_for('lista_pedidos'))
        if not request.form.get('reintentar'):
            flash(
                'Este pedido ya fue enviado a facturar. Verificá en QuickBooks '
                'antes de reenviarlo: si la factura ya existe, reenviar crea una '
                'factura duplicada.',
                'warning',
            )
            return redirect(url_for('lista_pedidos'))

    traz_errores = _validar_preparacion_pedido(pedido)
    if traz_errores:
        flash('No se puede facturar. Datos de trazabilidad incompletos:', 'error')
        for err in traz_errores:
            flash(err, 'error')
        return redirect(url_for('lista_pedidos'))

    payload = pedido_a_json(pedido)

    # ── Guard: datos que QBO necesita (item y precio en cada línea) ──
    datos_errores = _validar_datos_facturacion(payload)
    if datos_errores:
        flash('No se puede facturar. Datos incompletos para QuickBooks:', 'error')
        for err in datos_errores:
            flash(err, 'error')
        return redirect(url_for('lista_pedidos'))

    # ── Guard: verificar que N8N está configurado ──
    if not N8N_WEBHOOK_URL:
        app.logger.error(f'N8N_WEBHOOK_URL no configurada. No se puede facturar pedido {pedido_id}.')
        flash('Error de configuración: N8N_WEBHOOK_URL no está definida. Contacte al administrador.', 'danger')
        return redirect(url_for('lista_pedidos'))

    # ── Llamada al webhook N8N ──
    try:
        resp = requests.post(N8N_WEBHOOK_URL, json=payload, timeout=N8N_WEBHOOK_TIMEOUT, headers=_webhook_headers())
        resp.raise_for_status()
    except requests.Timeout:
        app.logger.error(f'Timeout al enviar pedido {pedido_id} a n8n ({N8N_WEBHOOK_TIMEOUT}s)')
        flash(f'Timeout — N8N no respondió en {N8N_WEBHOOK_TIMEOUT}s. Reintentar o verificar conexión.', 'danger')
        return redirect(url_for('lista_pedidos'))
    except requests.ConnectionError as e:
        app.logger.error(f'Error de conexión con n8n para pedido {pedido_id}: {e}')
        flash('Error de conexión con N8N. Verificar que el servicio está activo.', 'danger')
        return redirect(url_for('lista_pedidos'))
    except requests.HTTPError as e:
        err_resp = e.response
        status_code = err_resp.status_code
        try:
            error_body = err_resp.json()
            error_msg = error_body.get('message', error_body.get('error', str(error_body)))
        except Exception:
            error_msg = err_resp.text[:200] if err_resp.text else 'Sin detalle'
        app.logger.error(f'HTTP {status_code} de n8n para pedido {pedido_id}: {error_msg}')
        if 400 <= status_code < 500:
            flash(f'Error en facturación (HTTP {status_code}): {error_msg}', 'danger')
        else:
            flash('Error temporal en QuickBooks. Reintentar en unos momentos.', 'danger')
        return redirect(url_for('lista_pedidos'))
    except Exception as e:
        app.logger.error(f'Error inesperado al enviar pedido {pedido_id} a n8n: {e}')
        flash('Error al enviar a n8n. Intente de nuevo.', 'danger')
        return redirect(url_for('lista_pedidos'))

    # ── Extraer invoice_id de la respuesta N8N ──
    invoice_id = None
    doc_number = None
    try:
        invoice_id, doc_number = _extraer_invoice_id(resp.json())
    except Exception:
        app.logger.warning(f'No se pudo parsear JSON de respuesta n8n para pedido {pedido_id}')

    # ── Transacción atómica: marcar como facturado ──
    pedido.estado = 'facturado'
    pedido.invoice_id_qbo = invoice_id
    pedido.doc_number_qbo = doc_number
    pedido.fecha_facturacion = datetime.now(timezone.utc)
    _log_pedido_evento(
        pedido,
        'facturado',
        f'Pedido facturado{(": " + invoice_id) if invoice_id else " (QBO no confirmó número de factura)"}',
        meta={
            'invoice_id_qbo': invoice_id,
            'doc_number': doc_number,
            'qbo_confirmado': bool(invoice_id),
        },
    )
    try:
        db.session.commit()
    except Exception as e:
        app.logger.critical(f'CRITICAL: Webhook exitoso pero commit falló para pedido {pedido_id}: {e}')
        db.session.rollback()
        flash('Error interno al guardar. El pedido fue enviado a QuickBooks pero no se marcó como facturado. Contacte soporte.', 'danger')
        return redirect(url_for('lista_pedidos'))

    if invoice_id:
        if doc_number:
            flash(f'Factura {doc_number} generada (QBO {invoice_id})', 'success')
        else:
            flash(f'Factura generada: {invoice_id}', 'success')
    else:
        # Un 2xx de N8N no prueba que QBO haya creado la factura. Decir "generada
        # correctamente" acá fue lo que ocultó el fallo del 2026-08-14.
        flash(
            'Enviado a QuickBooks, pero QBO no confirmó el número de factura. '
            'Verificá en QuickBooks que la factura exista antes de reenviar.',
            'warning',
        )
    return redirect(url_for('lista_pedidos'))


def _sanitizar_para_archivo(texto):
    """Deja `texto` apto para un nombre de archivo (Drive, iOS, Web Share).

    Sólo se conservan letras (incluidas las acentuadas y ñ/ü de nombres de
    clientes de Curazao), dígitos, espacios, guiones y guiones bajos; el
    resto se descarta. Las corridas de espacio quedan como un solo guión
    bajo, y no sobran guiones bajos al principio ni al final.
    """
    if not texto:
        return ''
    limpio = re.sub(r'[^\w\s-]', '', texto, flags=re.UNICODE)
    limpio = re.sub(r'\s+', '_', limpio.strip())
    return limpio.strip('_')


def _truncar_en_guion_bajo(texto, largo):
    """Recorta `texto` a `largo` caracteres, cortando en un guión bajo en
    vez de a mitad de palabra cuando hay uno disponible dentro del límite."""
    if len(texto) <= largo:
        return texto
    cortado = texto[:largo]
    if '_' in cortado:
        cortado = cortado.rsplit('_', 1)[0]
    return cortado.rstrip('_')


def _nombre_archivo_factura(numero, cliente):
    """Nombre del PDF de la factura, con el cliente incluido para que sea
    identificable a simple vista en Drive (antes era sólo `Factura_<numero>`).

    El nombre de cliente viene de la factura de QBO (lo que se le imprimió
    de verdad al cliente), no de `pedido.cliente.nombre`.
    """
    cliente_limpio = _truncar_en_guion_bajo(_sanitizar_para_archivo(cliente), 40)
    if not cliente_limpio:
        return f'Factura_{numero}.pdf'
    return f'Factura_{numero}_{cliente_limpio}.pdf'


@app.route('/pedidos/<int:pedido_id>/factura.pdf')
@login_required
@requiere_permiso_recurso('pedidos', 'leer')
def factura_pdf(pedido_id):
    """PDF de la factura, con los datos vigentes en QuickBooks.

    Se consulta QBO en cada llamada en vez de guardar un snapshot: las facturas
    se corrigen a mano cuando la lista de precios está desactualizada, y el PDF
    tiene que reflejar lo que realmente se le cobró al cliente.
    """
    pedido = Pedido.query.get_or_404(pedido_id)

    if not _user_can_manage_pedido(pedido):
        abort(403)

    if not pedido.invoice_id_qbo:
        abort(404, description='Este pedido no tiene factura en QuickBooks.')

    factura = _obtener_factura_qbo(pedido.invoice_id_qbo)
    if not factura:
        abort(502, description='No se pudo obtener la factura desde QuickBooks.')

    from utils.factura_pdf import render_factura_pdf, extraer_datos_factura, _pick_invoice

    # Verificar que lo que devolvió n8n es de verdad la factura pedida. Como
    # todos los campos tienen valor por defecto, un payload vacío o de otra
    # factura se renderizaría igual (membrete completo, sin líneas, BALANCE DUE
    # 0.00) y se archivaría en Drive como si fuera legítimo.
    inv = _pick_invoice(factura)
    if str(inv.get('Id')) != str(pedido.invoice_id_qbo):
        app.logger.error(
            'QBO devolvió la factura Id=%s para el pedido %s '
            '(se esperaba invoice_id_qbo=%s)',
            inv.get('Id'), pedido_id, pedido.invoice_id_qbo,
        )
        abort(502, description='QuickBooks devolvió una factura que no corresponde a este pedido.')

    try:
        pdf = render_factura_pdf(factura)
        datos_factura = extraer_datos_factura(factura)
        numero = datos_factura['numero'] or pedido.doc_number_qbo or pedido.id
    except Exception as e:
        app.logger.error(f'Error al renderizar la factura del pedido {pedido_id}: {e}')
        abort(500, description='No se pudo generar el PDF de la factura.')

    # El nombre de cliente sale de la factura de QBO (lo impreso de verdad),
    # no de pedido.cliente.nombre, que puede haber cambiado desde entonces.
    filename = _nombre_archivo_factura(numero, datos_factura['cliente'])

    _archivar_factura_drive(pdf, filename)

    resp = send_file(
        BytesIO(pdf),
        mimetype='application/pdf',
        as_attachment=False,
        download_name=filename,
    )
    # Documento financiero: que no quede en la caché del navegador después
    # de cerrar sesión.
    resp.headers['Cache-Control'] = 'no-store, private'
    return resp


def _factura_vigente_del_pedido(pedido):
    """Trae la factura viva de QBO y verifica que sea la de este pedido.

    Mismas guardas que `factura_pdf`: sin `invoice_id_qbo` no hay nada que
    traer, y un payload vacío o de otra factura se rechaza en vez de usarse.
    """
    from utils.factura_pdf import _pick_invoice

    if not pedido.invoice_id_qbo:
        abort(404, description='Este pedido no tiene factura en QuickBooks.')

    factura = _obtener_factura_qbo(pedido.invoice_id_qbo)
    if not factura:
        abort(502, description='No se pudo obtener la factura desde QuickBooks.')

    inv = _pick_invoice(factura)
    if str(inv.get('Id')) != str(pedido.invoice_id_qbo):
        app.logger.error(
            'QBO devolvió la factura Id=%s para el pedido %s (se esperaba %s)',
            inv.get('Id'), pedido.id, pedido.invoice_id_qbo,
        )
        abort(502, description='QuickBooks devolvió una factura que no corresponde a este pedido.')

    return factura


@app.route('/pedidos/<int:pedido_id>/precios-factura')
@login_required
@requiere_permiso_recurso('precios', 'leer')
def revisar_precios_factura(pedido_id):
    """Compara los precios de la factura corregida en QBO contra los de la app.

    Cuando la lista está desactualizada el precio se corrige a mano en QBO y esa
    corrección no vuelve nunca a la app, así que hay que volver a corregirla en
    la factura siguiente. Esta pantalla es el camino de vuelta.
    """
    pedido = Pedido.query.get_or_404(pedido_id)

    if not _user_can_manage_pedido(pedido):
        abort(403)

    factura = _factura_vigente_del_pedido(pedido)
    filas, avisos = _comparar_precios_factura(pedido, factura)

    return render_template(
        'revisar_precios_factura.html',
        pedido=pedido,
        filas=filas,
        avisos=avisos,
        diferencias=[f for f in filas if f['estado'] == 'difiere'],
        desfasados=[f for f in filas if f['estado'] == 'pedido_desfasado'],
        no_aplicables=[
            f for f in filas
            if f['estado'] in ('sin_producto', 'precio_invalido', 'precio_ambiguo')
        ],
    )


@app.route('/pedidos/<int:pedido_id>/precios-factura/aplicar', methods=['POST'])
@login_required
@requiere_permiso_recurso('precios', 'editar')
def aplicar_precios_factura(pedido_id):
    """Escribe los precios confirmados en el precio del cliente.

    Escribe SOLO en `PrecioClienteProducto`: ahí el margen es 1.0, así que
    `precio_base` es directamente lo que se factura. En las listas generales hay
    filas con margen 1.2, donde escribir el UnitPrice crudo inflaría la factura
    siguiente un 20%.
    """
    pedido = Pedido.query.get_or_404(pedido_id)

    if not _user_can_manage_pedido(pedido):
        abort(403)

    seleccionados = set()
    for raw in request.form.getlist('aplicar'):
        try:
            seleccionados.add(int(raw))
        except (TypeError, ValueError):
            continue

    if not seleccionados:
        flash('No seleccionaste ningún precio para actualizar.', 'info')
        return redirect(url_for('detalles_pedido', pedido_id=pedido.id))

    # Los precios se vuelven a leer de QBO en vez de confiar en el formulario:
    # del form solo se toma QUÉ productos actualizar, nunca CON QUÉ valor.
    factura = _factura_vigente_del_pedido(pedido)
    filas, _ = _comparar_precios_factura(pedido, factura)
    aplicables = {
        f['producto_id']: f for f in filas
        if f['estado'] == 'difiere' and f['producto_id'] is not None
    }

    actualizados = 0
    for producto_id in sorted(seleccionados):
        fila = aplicables.get(producto_id)
        if not fila:
            continue

        nuevo = round(float(fila['precio_qbo']), 2)
        # Sin filtrar por `activo`: la unique constraint es (cliente_id,
        # producto_id), así que una fila desactivada haría fallar el insert.
        registro = PrecioClienteProducto.query.filter_by(
            cliente_id=pedido.cliente_id,
            producto_id=producto_id,
        ).first()

        anterior = float(registro.precio_base) if registro else None

        if registro is None:
            registro = PrecioClienteProducto(
                cliente_id=pedido.cliente_id,
                producto_id=producto_id,
                precio_base=nuevo,
                margen_jomar=1.0,
                margen_retail=1.2,
                activo=True,
            )
            db.session.add(registro)
        else:
            registro.precio_base = nuevo
            registro.activo = True

        registro.calcular_precios()

        _log_pedido_evento(
            pedido,
            'precio_actualizado',
            f"{fila['nombre']}: {anterior if anterior is not None else 'sin precio'} → {nuevo}",
            meta={
                'producto_id': producto_id,
                'precio_anterior': anterior,
                'precio_nuevo': nuevo,
                'invoice_id_qbo': pedido.invoice_id_qbo,
                'origen': 'factura_qbo',
            },
        )
        actualizados += 1

    if actualizados:
        db.session.commit()
        flash(
            f'{actualizados} precio(s) actualizados para {pedido.cliente.nombre}. '
            'Los próximos pedidos de este cliente los toman por defecto.',
            'success',
        )
    else:
        db.session.rollback()
        flash('No había diferencias vigentes que aplicar.', 'info')

    return redirect(url_for('detalles_pedido', pedido_id=pedido.id))


############################################
# RUTAS PARA SISTEMA DE PRECIOS
############################################

@app.route('/precios')
@login_required
@requiere_permiso_recurso('precios', 'leer')
def mostrar_precios():
    """Página principal del sistema de precios"""
    listas_precio = ListaPrecio.query.filter_by(activa=True).all()
    return render_template('precios/index.html', listas_precio=listas_precio)

# ---- LISTAS DE PRECIOS ----

@app.route('/precios/listas')
@login_required
@requiere_permiso_recurso('precios', 'leer')
def listas_precios():
    """Mostrar todas las listas de precios"""
    listas = ListaPrecio.query.order_by(ListaPrecio.es_default.desc(), ListaPrecio.nombre).all()
    return render_template('precios/listas.html', listas=listas)

@app.route('/precios/listas/nueva', methods=['GET', 'POST'])
@login_required
@requiere_permiso_recurso('precios', 'crear')
def nueva_lista_precio():
    """Crear nueva lista de precios"""
    if request.method == 'POST':
        try:
            nombre = request.form['nombre']
            descripcion = request.form.get('descripcion', '')
            
            nueva_lista = ListaPrecio(
                nombre=nombre,
                descripcion=descripcion
            )
            
            db.session.add(nueva_lista)
            db.session.commit()
            
            flash('Lista de precios creada exitosamente', 'success')
            return redirect(url_for('listas_precios'))
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Error al crear la lista de precios: {e}')
            flash('Error al crear la lista de precios. Intente de nuevo.', 'error')
    
    return render_template('precios/lista_form.html')

@app.route('/precios/listas/<int:lista_id>/editar', methods=['GET', 'POST'])
@login_required
@requiere_permiso_recurso('precios', 'editar')
def editar_lista_precio(lista_id):
    """Editar lista de precios"""
    lista = ListaPrecio.query.get_or_404(lista_id)
    
    if request.method == 'POST':
        try:
            lista.nombre = request.form['nombre']
            lista.descripcion = request.form.get('descripcion', '')
            
            db.session.commit()
            flash('Lista de precios actualizada exitosamente', 'success')
            return redirect(url_for('listas_precios'))
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Error al actualizar la lista de precios: {e}')
            flash('Error al actualizar la lista de precios. Intente de nuevo.', 'error')
    
    return render_template('precios/lista_form.html', lista=lista)

@app.route('/precios/listas/<int:lista_id>/eliminar', methods=['POST'])
@login_required
@requiere_permiso_recurso('precios', 'eliminar')
def eliminar_lista_precio(lista_id):
    """Eliminar lista de precios"""
    lista = ListaPrecio.query.get_or_404(lista_id)
    
    if lista.es_default:
        return jsonify({'error': 'No se puede eliminar la lista de precios por defecto'}), 400
    
    try:
        db.session.delete(lista)
        db.session.commit()
        return jsonify({'message': 'Lista de precios eliminada exitosamente'}), 200
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al eliminar la lista: {e}')
        return jsonify({'error': 'Error al eliminar la lista'}), 500

# ---- PRECIOS POR PRODUCTO EN LISTAS ----

@app.route('/precios/listas/<int:lista_id>/productos')
@login_required
def precios_lista_productos(lista_id):
    """Gestionar precios de productos en una lista específica"""
    lista = ListaPrecio.query.get_or_404(lista_id)
    productos = Producto.query.all()
    
    # Obtener precios existentes
    precios_existentes = db.session.query(PrecioProducto, Producto).join(
        Producto, PrecioProducto.producto_id == Producto.id
    ).filter(PrecioProducto.lista_precio_id == lista_id).order_by(Producto.nombre).all()

    # Proveedores únicos para filtro
    proveedores = sorted(set(
        p.proveedor for p in Producto.query.filter(Producto.proveedor.isnot(None)).all()
    ))

    # Productos que aún no están en la lista (para agregar nuevos)
    ids_en_lista = {pp.producto_id for pp, _ in precios_existentes}
    productos_disponibles = [p for p in productos if p.id not in ids_en_lista]

    return render_template('precios/lista_productos.html',
                         lista=lista,
                         productos=productos,
                         productos_disponibles=productos_disponibles,
                         precios_existentes=precios_existentes,
                         proveedores=proveedores)

@app.route('/precios/listas/<int:lista_id>/productos', methods=['POST'])
@login_required
@requiere_permiso_recurso('precios', 'editar')
def crear_precio_producto(lista_id):
    """Crear o actualizar precio de un producto en una lista"""
    try:
        producto_id = request.form['producto_id']
        precio_base = float(request.form['precio_base'])
        margen_jomar = float(request.form.get('margen_jomar', 1.0))
        margen_retail = float(request.form.get('margen_retail', 1.2))
        
        # Verificar si ya existe
        precio_existente = PrecioProducto.query.filter_by(
            lista_precio_id=lista_id,
            producto_id=producto_id
        ).first()
        
        if precio_existente:
            # Actualizar
            precio_existente.precio_base = precio_base
            precio_existente.margen_jomar = margen_jomar
            precio_existente.margen_retail = margen_retail
            precio_existente.calcular_precios()
            precio_existente.fecha_actualizacion = datetime.utcnow()
        else:
            # Crear nuevo
            nuevo_precio = PrecioProducto(
                lista_precio_id=lista_id,
                producto_id=producto_id,
                precio_base=precio_base,
                margen_jomar=margen_jomar,
                margen_retail=margen_retail
            )
            nuevo_precio.calcular_precios()
            db.session.add(nuevo_precio)
        
        db.session.commit()
        return jsonify({'message': 'Precio actualizado exitosamente'}), 200
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al actualizar precio: {e}')
        return jsonify({'error': 'Error al actualizar precio'}), 500

@app.route('/precios/listas/<int:lista_id>/productos/masivo', methods=['POST'])
@login_required
@requiere_permiso_recurso('precios', 'editar')
def actualizar_precios_masivo(lista_id):
    """Actualizar múltiples precios de productos en una lista (batch upsert)"""
    try:
        data = request.get_json()
        if not data or 'precios' not in data:
            return jsonify({'error': 'Datos no válidos'}), 400
        if not isinstance(data['precios'], list):
            return jsonify({'error': 'El campo "precios" debe ser una lista'}), 400

        lista = ListaPrecio.query.get_or_404(lista_id)
        actualizados = 0
        errores = []

        for idx, item in enumerate(data['precios'], start=1):
            if not isinstance(item, dict):
                errores.append(f'Fila {idx}: Formato inválido (se esperaba objeto JSON)')
                continue

            try:
                producto_id = int(item['producto_id'])
                precio_base = float(item['precio_base'])
                margen_jomar = float(item.get('margen_jomar', 1.0))
                margen_retail = float(item.get('margen_retail', 1.2))
            except (KeyError, ValueError, TypeError) as e:
                errores.append(f'Fila {idx}: {str(e)}')
                continue

            if precio_base <= 0:
                errores.append(f'Producto {producto_id}: precio_base debe ser mayor que 0')
                continue
            if margen_jomar <= 0:
                errores.append(f'Producto {producto_id}: margen_jomar debe ser mayor que 0')
                continue
            if margen_retail <= 0:
                errores.append(f'Producto {producto_id}: margen_retail debe ser mayor que 0')
                continue

            producto = db.session.get(Producto, producto_id)
            if not producto:
                errores.append(f'Producto {producto_id}: no existe')
                continue

            try:
                precio_existente = PrecioProducto.query.filter_by(
                    lista_precio_id=lista_id,
                    producto_id=producto_id
                ).first()

                if precio_existente:
                    precio_existente.precio_base = precio_base
                    precio_existente.margen_jomar = margen_jomar
                    precio_existente.margen_retail = margen_retail
                    precio_existente.calcular_precios()
                    precio_existente.fecha_actualizacion = datetime.utcnow()
                else:
                    nuevo_precio = PrecioProducto(
                        lista_precio_id=lista_id,
                        producto_id=producto_id,
                        precio_base=precio_base,
                        margen_jomar=margen_jomar,
                        margen_retail=margen_retail
                    )
                    nuevo_precio.calcular_precios()
                    db.session.add(nuevo_precio)

                actualizados += 1
            except Exception as e:
                errores.append(f'Producto {item.get("producto_id", "?")}: {str(e)}')

        if actualizados == 0 and errores:
            return jsonify({
                'error': 'No se actualizó ningún precio. Revise los datos enviados.',
                'errores': errores
            }), 400

        db.session.commit()
        return jsonify({
            'message': f'{actualizados} precios actualizados',
            'actualizados': actualizados,
            'errores': errores
        }), 200

    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error en actualización masiva: {e}')
        return jsonify({'error': 'Error al actualizar precios'}), 500

@app.route('/precios/productos/<int:precio_id>/eliminar', methods=['DELETE'])
@login_required
@requiere_permiso_recurso('precios', 'eliminar')
def eliminar_precio_producto(precio_id):
    """Eliminar precio de producto"""
    try:
        precio = PrecioProducto.query.get_or_404(precio_id)
        db.session.delete(precio)
        db.session.commit()
        return jsonify({'message': 'Precio eliminado exitosamente'}), 200
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al eliminar precio: {e}')
        return jsonify({'error': 'Error al eliminar precio'}), 500

# ---- ASIGNACIÓN DE LISTAS A CLIENTES ----

@app.route('/precios/clientes')
@login_required
@requiere_permiso_recurso('precios', 'leer')
def precios_clientes():
    """Gestionar listas de precios por cliente"""
    clientes = Cliente.query.all()
    listas = ListaPrecio.query.filter_by(activa=True).all()
    
    # Obtener asignaciones existentes
    asignaciones = db.session.query(ClienteListaPrecio, Cliente, ListaPrecio).join(
        Cliente, ClienteListaPrecio.cliente_id == Cliente.id
    ).join(
        ListaPrecio, ClienteListaPrecio.lista_precio_id == ListaPrecio.id
    ).filter(ClienteListaPrecio.activa == True).all()
    
    return render_template('precios/clientes.html', 
                         clientes=clientes, 
                         listas=listas,
                         asignaciones=asignaciones)

@app.route('/precios/clientes/asignar', methods=['POST'])
@login_required
@requiere_permiso_recurso('precios', 'editar')
def asignar_lista_cliente():
    """Asignar lista de precios a cliente"""
    try:
        cliente_id = request.form['cliente_id']
        lista_precio_id = request.form['lista_precio_id']
        
        # Verificar si ya existe una asignación activa
        asignacion_existente = ClienteListaPrecio.query.filter_by(
            cliente_id=cliente_id,
            activa=True
        ).first()
        
        if asignacion_existente:
            # Desactivar la asignación anterior
            asignacion_existente.activa = False
        
        # Crear nueva asignación
        nueva_asignacion = ClienteListaPrecio(
            cliente_id=cliente_id,
            lista_precio_id=lista_precio_id
        )
        
        db.session.add(nueva_asignacion)
        db.session.commit()
        
        return jsonify({'message': 'Lista de precios asignada exitosamente'}), 200
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al asignar lista: {e}')
        return jsonify({'error': 'Error al asignar lista'}), 500

@app.route('/precios/clientes/<int:asignacion_id>/eliminar', methods=['DELETE'])
@login_required
@requiere_permiso_recurso('precios', 'eliminar')
def eliminar_asignacion_cliente(asignacion_id):
    """Eliminar asignación de lista de precios a cliente"""
    try:
        asignacion = ClienteListaPrecio.query.get_or_404(asignacion_id)
        db.session.delete(asignacion)
        db.session.commit()
        return jsonify({'message': 'Asignación eliminada exitosamente'}), 200
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al eliminar asignación: {e}')
        return jsonify({'error': 'Error al eliminar asignación'}), 500

# ---- PRECIOS ESPECÍFICOS CLIENTE-PRODUCTO ----

@app.route('/precios/cliente-producto')
@login_required
def precios_cliente_producto():
    """Gestionar precios específicos por cliente-producto"""
    clientes = Cliente.query.all()
    productos = Producto.query.all()
    
    # Obtener precios específicos existentes
    precios_especificos = db.session.query(PrecioClienteProducto, Cliente, Producto).join(
        Cliente, PrecioClienteProducto.cliente_id == Cliente.id
    ).join(
        Producto, PrecioClienteProducto.producto_id == Producto.id
    ).filter(PrecioClienteProducto.activo == True).all()
    
    return render_template('precios/cliente_producto.html',
                         clientes=clientes,
                         productos=productos,
                         precios_especificos=precios_especificos)

@app.route('/precios/cliente-producto', methods=['POST'])
@login_required
@requiere_permiso_recurso('precios', 'editar')
def crear_precio_cliente_producto():
    """Crear precio específico cliente-producto"""
    try:
        cliente_id = request.form['cliente_id']
        producto_id = request.form['producto_id']
        precio_base = float(request.form['precio_base'])
        margen_jomar = float(request.form.get('margen_jomar', 1.0))
        margen_retail = float(request.form.get('margen_retail', 1.2))
        
        # Verificar si ya existe
        precio_existente = PrecioClienteProducto.query.filter_by(
            cliente_id=cliente_id,
            producto_id=producto_id
        ).first()
        
        if precio_existente:
            # Actualizar
            precio_existente.precio_base = precio_base
            precio_existente.margen_jomar = margen_jomar
            precio_existente.margen_retail = margen_retail
            precio_existente.calcular_precios()
            precio_existente.fecha_actualizacion = datetime.utcnow()
        else:
            # Crear nuevo
            nuevo_precio = PrecioClienteProducto(
                cliente_id=cliente_id,
                producto_id=producto_id,
                precio_base=precio_base,
                margen_jomar=margen_jomar,
                margen_retail=margen_retail
            )
            nuevo_precio.calcular_precios()
            db.session.add(nuevo_precio)
        
        db.session.commit()
        return jsonify({'message': 'Precio específico actualizado exitosamente'}), 200
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al actualizar precio específico: {e}')
        return jsonify({'error': 'Error al actualizar precio'}), 500

@app.route('/precios/cliente-producto/<int:precio_id>/eliminar', methods=['DELETE'])
@login_required
@requiere_permiso_recurso('precios', 'eliminar')
def eliminar_precio_cliente_producto(precio_id):
    """Eliminar precio específico cliente-producto"""
    try:
        precio = PrecioClienteProducto.query.get_or_404(precio_id)
        db.session.delete(precio)
        db.session.commit()
        return jsonify({'message': 'Precio específico eliminado exitosamente'}), 200
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al eliminar precio específico: {e}')
        return jsonify({'error': 'Error al eliminar precio'}), 500

# ---- API PARA OBTENER PRECIOS ----

@app.route('/api/precios/cliente/<int:cliente_id>/producto/<int:producto_id>')
@login_required
def api_precio_cliente_producto(cliente_id, producto_id):
    """API para obtener el precio de un producto específico para un cliente"""
    if not _user_can_view_cliente(cliente_id):
        return jsonify({'error': 'No autorizado'}), 403
    tipo = request.args.get('tipo', 'jomar')
    if tipo not in ('base', 'jomar', 'retail'):
        return jsonify({'error': 'Tipo de precio no válido'}), 400

    precio = obtener_precio_producto_cliente(cliente_id, producto_id, tipo)
    producto = db.session.get(Producto, producto_id)

    if not producto:
        return jsonify({'error': 'Producto no encontrado'}), 404

    return jsonify({
        'producto_id': producto_id,
        'producto_nombre': producto.nombre,
        'tipo_precio': tipo,
        'precio': float(precio) if precio is not None else None
    })

@app.route('/api/precios/cliente/<int:cliente_id>/productos')
@login_required
def api_precios_cliente_productos(cliente_id):
    """Precio de CADA producto para un cliente, resuelto por la misma jerarquía
    que usa el servidor al guardar el pedido.

    Antes esta API armaba el resultado por su cuenta: precios específicos, más
    los productos de la lista asignada al cliente. Un producto que no estuviera
    en esa lista **no venía en la respuesta**, así que el formulario nunca le
    actualizaba el precio y quedaba mostrando el de la lista default — mientras
    el servidor sí caía a la lista default y guardaba otro número. Form y
    servidor calculaban por caminos distintos y podían no coincidir.

    Ahora se recorre el catálogo completo y se resuelve con
    `obtener_precio_producto_cliente`, la misma cadena de
    `_resolver_precio_unitario_pedido`: lo que se ve es lo que se va a cobrar.
    """
    if not _user_can_view_cliente(cliente_id):
        return jsonify({'error': 'No autorizado'}), 403

    resultado = []
    for producto in Producto.query.all():
        # El precio sale del mismo resolutor que usa el servidor al guardar.
        precio = obtener_precio_producto_cliente(cliente_id, producto.id, 'base')
        if precio is None:
            precio = obtener_precio_default_producto(producto.id, 'base')

        fila, origen = _fila_precio_vigente(cliente_id, producto.id)

        if precio is None or fila is None:
            # Producto sin precio en ningún lado: se devuelve igual, con precio
            # nulo, para que el formulario lo muestre como "sin precio" en vez
            # de quedarse con el número que ya tenía y hacerlo pasar por válido.
            resultado.append({
                'id': producto.id, 'nombre': producto.nombre,
                'precio': None, 'tipo_precio': 'sin_precio',
                'precio_base': None, 'margen_jomar': None, 'margen_retail': None,
            })
            continue

        resultado.append({
            'id': producto.id,
            'nombre': producto.nombre,
            'precio': float(precio),
            'tipo_precio': origen,
            'precio_base': float(fila.precio_base),
            'margen_jomar': fila.margen_jomar,
            'margen_retail': fila.margen_retail,
        })

    resultado.sort(key=lambda x: x['nombre'])
    return jsonify(resultado)

# TAMBIÉN agregar esta nueva función para debugging:

@app.route('/api/precios/cliente/<int:cliente_id>/debug')
@login_required
def debug_precios_cliente(cliente_id):
    """API para debug - mostrar información detallada de precios de un cliente"""
    if not _user_can_view_cliente(cliente_id):
        return jsonify({'error': 'No autorizado'}), 403

    # Información del cliente
    cliente = Cliente.query.get_or_404(cliente_id)
    
    # Lista asignada al cliente
    cliente_lista = ClienteListaPrecio.query.filter_by(
        cliente_id=cliente_id,
        activa=True
    ).first()
    
    # Precios específicos
    precios_especificos_count = PrecioClienteProducto.query.filter_by(
        cliente_id=cliente_id,
        activo=True
    ).count()
    
    debug_info = {
        'cliente': {
            'id': cliente.id,
            'nombre': cliente.nombre
        },
        'lista_asignada': None,
        'precios_especificos_count': precios_especificos_count
    }
    
    if cliente_lista:
        # Contar productos en la lista asignada
        productos_en_lista = PrecioProducto.query.filter_by(
            lista_precio_id=cliente_lista.lista_precio_id,
            activo=True
        ).count()
        
        debug_info['lista_asignada'] = {
            'id': cliente_lista.lista_precio_id,
            'nombre': cliente_lista.lista_precio.nombre,
            'productos_count': productos_en_lista,
            'es_default': cliente_lista.lista_precio.es_default
        }
    
    # Lista por defecto
    lista_default = ListaPrecio.query.filter_by(es_default=True, activa=True).first()
    if lista_default:
        productos_en_default = PrecioProducto.query.filter_by(
            lista_precio_id=lista_default.id,
            activo=True
        ).count()
        
        debug_info['lista_default'] = {
            'id': lista_default.id,
            'nombre': lista_default.nombre,
            'productos_count': productos_en_default
        }
    
    return jsonify(debug_info)

@app.route('/api/precios/lista/<int:lista_id>')
@login_required
@requiere_permiso_recurso('precios', 'leer')
def api_precios_lista(lista_id):
    """API para obtener todos los precios de una lista"""
    precios = db.session.query(PrecioProducto, Producto).join(
        Producto, PrecioProducto.producto_id == Producto.id
    ).filter(
        PrecioProducto.lista_precio_id == lista_id,
        PrecioProducto.activo == True
    ).all()
    
    resultado = []
    for precio, producto in precios:
        resultado.append({
            'producto_id': producto.id,
            'producto_nombre': producto.nombre,
            'precio_base': precio.precio_base,
            'precio_jomar': precio.precio_jomar,
            'precio_retail': precio.precio_retail,
            'margen_jomar': precio.margen_jomar,
            'margen_retail': precio.margen_retail
        })
    
    return jsonify(resultado)

@app.route('/api/precios/cliente/<int:cliente_id>')
@login_required
def api_precios_cliente(cliente_id):
    """API para obtener todos los precios disponibles para un cliente"""
    if not _user_can_view_cliente(cliente_id):
        return jsonify({'error': 'No autorizado'}), 403
    resultado = []
    
    # 1. Precios específicos cliente-producto
    precios_especificos = db.session.query(PrecioClienteProducto, Producto).join(
        Producto, PrecioClienteProducto.producto_id == Producto.id
    ).filter(
        PrecioClienteProducto.cliente_id == cliente_id,
        PrecioClienteProducto.activo == True
    ).all()
    
    productos_procesados = set()
    
    for precio_esp, producto in precios_especificos:
        resultado.append({
            'producto_id': producto.id,
            'producto_nombre': producto.nombre,
            'precio_base': precio_esp.precio_base,
            'precio_jomar': precio_esp.precio_jomar,
            'precio_retail': precio_esp.precio_retail,
            'tipo_precio': 'específico',
            'margen_jomar': precio_esp.margen_jomar,
            'margen_retail': precio_esp.margen_retail
        })
        productos_procesados.add(producto.id)
    
    # 2. Productos de la lista asignada al cliente
    cliente_lista = ClienteListaPrecio.query.filter_by(
        cliente_id=cliente_id,
        activa=True
    ).first()
    
    if cliente_lista:
        precios_lista = db.session.query(PrecioProducto, Producto).join(
            Producto, PrecioProducto.producto_id == Producto.id
        ).filter(
            PrecioProducto.lista_precio_id == cliente_lista.lista_precio_id,
            PrecioProducto.activo == True,
            ~Producto.id.in_(productos_procesados)
        ).all()
        
        for precio_lista, producto in precios_lista:
            resultado.append({
                'producto_id': producto.id,
                'producto_nombre': producto.nombre,
                'precio_base': precio_lista.precio_base,
                'precio_jomar': precio_lista.precio_jomar,
                'precio_retail': precio_lista.precio_retail,
                'tipo_precio': 'lista_asignada',
                'margen_jomar': precio_lista.margen_jomar,
                'margen_retail': precio_lista.margen_retail,
                'lista_nombre': cliente_lista.lista_precio.nombre
            })
    
    return jsonify(resultado)

############################################
# Rutas de Productos
############################################


@app.route('/productos', methods=['GET', 'POST'])
@login_required
def productos():
    if request.method == 'POST':
        if isinstance(current_user, Vendedor) and not current_user.tiene_permiso('productos', 'crear'):
            return jsonify({'error': 'No tienes permisos para crear productos'}), 403
        try:
            nombre      = request.form['nombre']
            descripcion = request.form.get('descripcion', '')
            temperatura = request.form.get('temperatura', '')
            qbo_id      = request.form.get('qbo_id', '').strip() or None
            tax_rate    = float(request.form.get('tax_rate', 0.0))

            proveedor   = request.form.get('proveedor', '').strip() or None

            nuevo = Producto(
                nombre=nombre,
                descripcion=descripcion,
                temperatura=temperatura,
                qbo_id=qbo_id,
                tax_rate=tax_rate,
                se_pesa='se_pesa' in request.form,
                proveedor=proveedor
            )
            db.session.add(nuevo)
            db.session.commit()

            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({
                    'message': 'Producto creado correctamente',
                    'producto': {
                        'id': nuevo.id,
                        'nombre': nuevo.nombre,
                        'descripcion': nuevo.descripcion,
                        'temperatura': nuevo.temperatura,
                        'qbo_id': nuevo.qbo_id,
                        'tax_rate': nuevo.tax_rate
                    }
                })
            else:
                flash('Producto creado exitosamente.', 'success')
                return redirect(url_for('productos'))

        except IntegrityError as e:
            db.session.rollback()
            app.logger.warning(f"Producto duplicado al crear: {e}")
            # qbo_id es la única restricción UNIQUE de la tabla producto.
            mensaje = 'Ya existe un producto con ese QBO ID.'
            if qbo_id:
                existente = Producto.query.filter_by(qbo_id=qbo_id).first()
                if existente:
                    mensaje = f'El QBO ID {qbo_id} ya está asignado al producto "{existente.nombre}".'
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({'error': mensaje}), 400
            flash(mensaje, 'danger')
            return redirect(url_for('productos'))

        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error al crear producto: {e}")
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({'error': 'Error al crear el producto'}), 400
            flash('Error al crear el producto', 'danger')
            return redirect(url_for('productos'))

    # GET → listamos todos los productos ordenados
    todos = Producto.query.order_by(Producto.id.asc()).all()
    return render_template('productos.html', productos=todos)






@app.route('/productos/<int:producto_id>/editar', methods=['GET', 'POST'])
@login_required
@requiere_permiso_recurso('productos', 'editar')
def editar_producto(producto_id):
    producto = Producto.query.get_or_404(producto_id)
    if request.method == 'POST':
        producto.nombre      = request.form['nombre']
        producto.descripcion = request.form.get('descripcion', '')
        producto.temperatura = request.form.get('temperatura', '')
        producto.qbo_id      = request.form.get('qbo_id', '').strip() or None
        producto.tax_rate    = float(request.form.get('tax_rate', 0.0))
        producto.se_pesa     = 'se_pesa' in request.form
        producto.proveedor   = request.form.get('proveedor', '').strip() or None
        db.session.commit()
        flash('Producto actualizado correctamente.', 'success')
        return redirect(url_for('productos'))

    return render_template('editar_producto.html', producto=producto)




@app.route('/api/productos', methods=['GET'])
@login_required
def obtener_productos_api():
    productos = Producto.query.order_by(Producto.id).all()
    productos_data = []
    for p in productos:
        precio = obtener_precio_default_producto(p.id, 'base')
        productos_data.append({
            "id"         : p.id,
            "nombre"     : p.nombre,
            "descripcion": p.descripcion,
            "temperatura": p.temperatura,
            "qbo_id"     : p.qbo_id,
            "tax_rate"   : p.tax_rate,  # Nueva línea
            "precio"     : float(precio or 0)
        })
    return jsonify(productos_data)


@app.route('/productos/<int:producto_id>/eliminar', methods=['POST'])
@login_required
@requiere_permiso_recurso('productos', 'eliminar')
def eliminar_producto(producto_id):
    # The decorator above already filters by role permission. Restrict
    # further to super_admin since deleting a producto can affect every
    # cliente's price list and historical pedidos.
    if isinstance(current_user, Vendedor) and current_user.rol.nombre != 'super_admin':
        return jsonify({'error': 'Solo un super administrador puede eliminar productos.'}), 403
    try:
        producto = Producto.query.get_or_404(producto_id)
        db.session.delete(producto)
        db.session.commit()
        return jsonify({'message': 'Producto eliminado correctamente.'}), 200
    except Exception as e:
        app.logger.error(f"Error al eliminar el producto: {e}")
        return jsonify({'error': 'Error al eliminar el producto.'}), 500



@app.before_request
def override_method():
    if request.method == 'POST' and '_method' in request.form:
        method = request.form['_method'].upper()
        if method in ['PUT', 'DELETE']:
            request.environ['REQUEST_METHOD'] = method


# Rutas para Recepciones
@app.route('/recepciones')
@login_required
def mostrar_recepciones():
    productos = Producto.query.all()
    recepciones = Recepcion.query.all()
    ultima_recepcion = Recepcion.query.order_by(Recepcion.id.desc()).first()
    return render_template('recepciones.html', productos=productos, recepciones=recepciones, ultima_recepcion=ultima_recepcion)

@app.route('/api/recepciones', methods=['GET'])
@login_required
def obtener_recepciones_api():
    recepciones = Recepcion.query.order_by(Recepcion.id.desc()).limit(20).all()
    recepciones_data = [{
        'id': r.id,
        'producto_id': r.producto_id,
        'producto': r.producto.nombre if r.producto else 'undefined',
        'peso': r.peso,
        'recibido_en': r.recibido_en.strftime('%Y-%m-%d') if r.recibido_en else 'No disponible',
        'proveedor': r.proveedor,
        'numero_factura': r.numero_factura
    } for r in recepciones]
    return jsonify(recepciones_data)

@app.route('/recepciones', methods=['POST'])
@login_required
@requiere_permiso_recurso('importaciones', 'crear')
def crear_recepcion():
    try:
        producto_id = request.form['producto_id']
        peso = request.form['peso']
        proveedor = request.form['proveedor']
        numero_factura = request.form['numero_factura']
        fecha_recepcion = request.form['fecha_recepcion']
        try:
            recibido_en = datetime.strptime(fecha_recepcion, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({"error": "Formato de fecha inválido. Debe ser YYYY-MM-DD"}), 400
        nueva_recepcion = Recepcion(
            producto_id=producto_id,
            peso=peso,
            proveedor=proveedor,
            numero_factura=numero_factura,
            recibido_en=recibido_en
        )
        db.session.add(nueva_recepcion)
        db.session.commit()
        recepcion_data = nueva_recepcion.to_dict()
        return jsonify({
            "message": "Recepción registrada exitosamente",
            "recepcion": recepcion_data
        }), 200
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al registrar la recepción: {e}')
        return jsonify({"error": "Error al registrar la recepción"}), 500

@app.route('/recepciones/<int:id>', methods=['DELETE'])
@login_required
@requiere_permiso_recurso('importaciones', 'eliminar')
def eliminar_recepcion(id):
    recepcion = Recepcion.query.get_or_404(id)
    try:
        db.session.delete(recepcion)
        db.session.commit()
        return jsonify({"message": "Recepción eliminada con éxito"}), 200
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al eliminar la recepción: {e}')
        return jsonify({"error": "Error al eliminar la recepción"}), 500

# Rutas para Facturación
@app.route('/facturacion', methods=['GET', 'POST'])
@login_required
def facturacion():
    if request.method == 'GET':
        productos = Producto.query.all()
        clientes = Cliente.query.all()
        today = datetime.utcnow().date()
        previous_data = {
            'producto_id': session.get('producto_id', ''),
            'cliente_id': session.get('cliente_id', ''),
            'lote': session.get('lote', ''),
            'fecha_fabricacion': session.get('fecha_fabricacion', '')
        }
        if previous_data['cliente_id']:
            facturaciones = Facturacion.query.filter_by(cliente_id=previous_data['cliente_id']) \
                .filter(Facturacion.fecha_registro >= today) \
                .order_by(Facturacion.fecha_registro.desc()).all()
        else:
            facturaciones = []
        return render_template('facturacion.html', 
                               facturaciones=facturaciones, 
                               productos=productos, 
                               clientes=clientes,
                               previous_data=previous_data)
    return registrar_facturacion()

@app.route('/ultimos_facturaciones', methods=['GET'])
@login_required
@requiere_permiso_recurso('facturacion', 'leer')
def ultimos_facturaciones():
    cliente_id = request.args.get('cliente_id', type=int)
    hoy = datetime.utcnow().date()
    if cliente_id:
        facturaciones = db.session.query(Facturacion.id, Facturacion.peso, Facturacion.lote, 
                                         Facturacion.fecha_fabricacion, Facturacion.fecha_expiracion, 
                                         Facturacion.fecha_registro, Producto.nombre.label('producto_nombre'),
                                         Cliente.nombre.label('cliente_nombre')).\
            join(Producto, Facturacion.producto_id == Producto.id).\
            join(Cliente, Facturacion.cliente_id == Cliente.id).\
            filter(Facturacion.cliente_id == cliente_id).\
            filter(Facturacion.fecha_registro >= hoy).\
            order_by(Facturacion.fecha_registro.desc()).all()
    else:
        facturaciones = db.session.query(Facturacion.id, Facturacion.peso, Facturacion.lote, 
                                         Facturacion.fecha_fabricacion, Facturacion.fecha_expiracion, 
                                         Facturacion.fecha_registro, Producto.nombre.label('producto_nombre'),
                                         Cliente.nombre.label('cliente_nombre')).\
            join(Producto, Facturacion.producto_id == Producto.id).\
            join(Cliente, Facturacion.cliente_id == Cliente.id).\
            order_by(Facturacion.fecha_registro.desc()).limit(20).all()
    return jsonify([{
        'id': facturacion.id,
        'producto': facturacion.producto_nombre,
        'cliente': facturacion.cliente_nombre,
        'peso': facturacion.peso,
        'lote': facturacion.lote,
        'fecha_fabricacion': facturacion.fecha_fabricacion,
        'fecha_expiracion': facturacion.fecha_expiracion,
        'fecha_registro': facturacion.fecha_registro.strftime('%Y-%m-%d %H:%M')
    } for facturacion in facturaciones])

@app.route('/facturacion/registrar', methods=['POST'])
@login_required
@requiere_permiso_recurso('facturacion', 'crear')
def registrar_facturacion():
    try:
        producto_id = request.form['producto_id']
        cliente_id = request.form['cliente_id']
        peso = request.form['peso']
        lote = request.form['lote']
        fecha_fabricacion = request.form['fecha_fabricacion']
        fecha_fabricacion_date = datetime.strptime(fecha_fabricacion, '%Y-%m-%d')
        fecha_expiracion = fecha_fabricacion_date + timedelta(days=365)
        session['producto_id'] = producto_id
        session['cliente_id'] = cliente_id
        session['lote'] = lote
        session['fecha_fabricacion'] = fecha_fabricacion
        nueva_facturacion = Facturacion(
            producto_id=producto_id,
            cliente_id=cliente_id,
            peso=peso,
            lote=lote,
            fecha_fabricacion=fecha_fabricacion,
            fecha_expiracion=fecha_expiracion.strftime('%Y-%m-%d'),
            fecha_registro=datetime.utcnow()
        )
        db.session.add(nueva_facturacion)
        db.session.commit()
        return jsonify({"message": "Peso registrado exitosamente"}), 200
    except Exception as e:
        return jsonify({'error': 'Error interno del servidor'}), 500
    
@app.route('/facturacion/eliminar/<int:id>', methods=['DELETE'])
@login_required
@requiere_permiso_recurso('facturacion', 'eliminar')
def eliminar_facturacion(id):
    try:
        facturacion = Facturacion.query.get_or_404(id)
        db.session.delete(facturacion)
        db.session.commit()
        return jsonify({"message": "Facturación eliminada exitosamente"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": "Error interno del servidor"}), 500

# Rutas para Importaciones, Reportes y Etiquetas
@app.route('/formulario_importacion')
@login_required
def formulario_importacion():
    productos = Producto.query.all()
    facturas = db.session.query(
        Importacion.numero_factura,
        func.count(Importacion.id).label('cantidad_productos'),
        func.sum(Importacion.cantidad_total).label('cantidad_total'),
        func.sum(Importacion.precio_fob_unidad * Importacion.cantidad_total).label('total_fob'),
        func.sum(Importacion.flete).label('total_flete'),
        func.sum(Importacion.arancel).label('total_arancel'),
        func.sum(Importacion.costo_aduana).label('total_costo_aduana'),
        func.sum(Importacion.precio_jomar * Importacion.cantidad_total).label('total_precio_jomar'),
        func.max(Importacion.fecha_importacion).label('fecha_importacion')
    ).group_by(Importacion.numero_factura).all()
    return render_template('formulario_importacion.html', productos=productos, facturas=facturas)

@app.route('/registrar_importacion', methods=['POST'])
@login_required
@requiere_permiso_recurso('importaciones', 'crear')
def registrar_importacion():
    try:
        numero_factura = request.form.get('numero_factura')
        proveedor = request.form.get('proveedor')
        moneda = request.form.get('moneda')
        tipo_cambio_ang = float(request.form.get('tipo_cambio_ang') or 0)
        flete_total = float(request.form.get('flete_total') or 0)
        gastos_agente_aduanal = float(request.form.get('gastos_agente_aduanal') or 0)
        arancel_porcentaje = float(request.form.get('arancel') or 0) / 100
        ob_porcentaje = float(request.form.get('ob') or 0) / 100
        flete_local_total = float(request.form.get('flete_local') or 0)
        fecha_importacion_str = request.form.get('fecha_importacion')
        if fecha_importacion_str:
            fecha_importacion = datetime.strptime(fecha_importacion_str, '%Y-%m-%d')
        else:
            fecha_importacion = datetime.utcnow()
        productos = []
        total_fob_general = 0
        total_cif_ang_general = 0
        num_productos = len([key for key in request.form.keys() if 'productos' in key and '[producto]' in key])
        for i in range(num_productos):
            qty_total = float(request.form.get(f'productos[{i}][qty_total]', 0))
            price_fob = float(request.form.get(f'productos[{i}][price_fob]', 0))
            total_fob = qty_total * price_fob
            total_fob_general += total_fob
        for i in range(num_productos):
            qty_total = float(request.form.get(f'productos[{i}][qty_total]', 0))
            price_fob = float(request.form.get(f'productos[{i}][price_fob]', 0))
            total_fob = qty_total * price_fob
            flete_proporcional = (total_fob / total_fob_general) * flete_total if total_fob_general > 0 else 0
            total_cif = total_fob + flete_proporcional
            cif_ang = total_cif * tipo_cambio_ang
            total_cif_ang_general += cif_ang
        for i in range(num_productos):
            producto_id = int(request.form.get(f'productos[{i}][producto]', 0))
            und_caja = int(request.form.get(f'productos[{i}][und_caja]', 1))
            qty_total = float(request.form.get(f'productos[{i}][qty_total]', 0))
            price_fob = float(request.form.get(f'productos[{i}][price_fob]', 0))
            total_fob = qty_total * price_fob
            flete_proporcional = (total_fob / total_fob_general) * flete_total if total_fob_general > 0 else 0
            total_cif = total_fob + flete_proporcional
            cif_ang = total_cif * tipo_cambio_ang
            arancel_ang = cif_ang * arancel_porcentaje
            ob_ang = cif_ang * ob_porcentaje
            ob_45 = ob_ang * 0.045
            flete_local_proporcional = (cif_ang / total_cif_ang_general) * flete_local_total if total_cif_ang_general > 0 else 0
            gastos_aduanal_proporcional = (cif_ang / total_cif_ang_general) * gastos_agente_aduanal if total_cif_ang_general > 0 else 0
            costo_total_almacen_producto = cif_ang + arancel_ang + ob_ang + gastos_aduanal_proporcional + flete_local_proporcional - ob_45
            total_unidades = qty_total * und_caja
            costo_por_unidad_ang = costo_total_almacen_producto / total_unidades if total_unidades > 0 else 0
            precio_jomar = costo_por_unidad_ang
            precio_retail = precio_jomar * 1.2
            nueva_importacion = Importacion(
                numero_factura=numero_factura,
                producto_id=producto_id,
                cantidad_cajas=und_caja,
                cantidad_total=qty_total,
                precio_fob_unidad=price_fob,
                flete=flete_proporcional,
                arancel=arancel_ang,
                costo_aduana=gastos_aduanal_proporcional,
                precio_jomar=precio_jomar,
                precio_retail=precio_retail,
                fecha_importacion=fecha_importacion,
                cif_ang=cif_ang,
                ob_ang=ob_ang,
                flete_local=flete_local_proporcional,
                costo_total_almacen=costo_total_almacen_producto
            )
            db.session.add(nueva_importacion)
            productos.append(nueva_importacion)
        if not productos:
            flash("No se han proporcionado productos para la importación", "danger")
            return redirect(url_for('formulario_importacion'))
        db.session.commit()
        app.logger.info(f"Importación registrada con {len(productos)} productos.")
        flash("Importación registrada exitosamente", "success")
        return redirect(url_for('formulario_importacion'))
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error al registrar la importación: {e}", exc_info=True)
        flash("Error al registrar la importación. Intente de nuevo.", "danger")
        return redirect(url_for('formulario_importacion'))

@app.route('/reporte_factura/<numero_factura>', methods=['GET'])
@login_required
@requiere_permiso_recurso('importaciones', 'leer')
def reporte_factura(numero_factura):
    importaciones = db.session.query(Importacion, Producto).join(Producto).filter(Importacion.numero_factura == numero_factura).all()
    if not importaciones:
        return "No se encontraron importaciones para el número de factura proporcionado.", 404
    fecha_importacion = importaciones[0][0].fecha_importacion.strftime('%d/%m/%Y')
    buffer = BytesIO()
    page_width, page_height = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=0.25 * inch,
        rightMargin=0.25 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch
    )
    elements = []
    styles = getSampleStyleSheet()
    styleTitle = styles['Title']
    styleTitle.alignment = TA_LEFT
    style_cell = ParagraphStyle(
        name='CellStyle',
        fontSize=7,
        alignment=TA_CENTER,
        leading=8
    )
    style_header = ParagraphStyle(
        name='HeaderStyle',
        fontSize=7,
        alignment=TA_CENTER,
        leading=8,
        textColor=colors.whitesmoke,
        backColor=colors.grey
    )
    logo_path = os.path.join(basedir, 'static', 'logo_etiquetas.png')
    if os.path.exists(logo_path):
        logo_width = 50
        logo_height = 50
        logo = Image(logo_path, width=logo_width, height=logo_height)
        titulo_text = f"Reporte de Importación - Factura {_pdf_xe(numero_factura)}"
        titulo = Paragraph(titulo_text, styleTitle)
        desired_indent = 1.0 * inch
        data_title = [['', logo, titulo]]
        table_title = Table(
            data_title,
            colWidths=[desired_indent, logo_width, None]
        )
        table_title_style = TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ])
        table_title.setStyle(table_title_style)
        table_title.hAlign = 'LEFT'
        elements.append(table_title)
    else:
        desired_indent = 1.0 * inch
        titulo = Paragraph(f"Reporte de Importación - Factura {_pdf_xe(numero_factura)}", styleTitle)
        data_title = [['', titulo]]
        table_title = Table(
            data_title,
            colWidths=[desired_indent, None]
        )
        table_title.hAlign = 'LEFT'
        elements.append(table_title)
    style_fecha = ParagraphStyle(
        name='FechaStyle',
        fontSize=9,
        alignment=TA_LEFT,
        leading=12,
        leftIndent=1.0 * inch
    )
    fecha_paragraph = Paragraph(f"Fecha de Importación: {fecha_importacion}", style_fecha)
    elements.append(fecha_paragraph)
    elements.append(Spacer(1, 12))
    data = [[
        Paragraph("Producto", style_header),
        Paragraph("Qty", style_header),
        Paragraph("P FOB", style_header),
        Paragraph("T FOB", style_header),
        Paragraph("Flete", style_header),
        Paragraph("T CIF", style_header),
        Paragraph("CIF ANG", style_header),
        Paragraph("Arancel", style_header),
        Paragraph("OB ANG", style_header),
        Paragraph("OB 4.5%", style_header),
        Paragraph("GAA", style_header),
        Paragraph("Costo Alm.", style_header),
        Paragraph("Costo/U ANG", style_header)
    ]]
    total_qty = 0
    total_fob = 0
    total_flete = 0
    total_cif = 0
    total_cif_ang = 0
    total_arancel = 0
    total_ob_ang = 0
    total_ob_45 = 0
    total_gastos_agente_aduanal = 0
    total_costo_total_almacen = 0
    for imp, prod in importaciones:
        total_qty += imp.cantidad_total
        total_fob_producto = imp.cantidad_total * imp.precio_fob_unidad
        total_fob += total_fob_producto
        total_flete += imp.flete
        total_cif_producto = total_fob_producto + imp.flete
        total_cif += total_cif_producto
        total_cif_ang += imp.cif_ang or 0
        total_arancel += imp.arancel or 0
        total_ob_ang += imp.ob_ang or 0
        ob_dev_45 = (imp.ob_ang or 0) * 0.5
        total_ob_45 += ob_dev_45
        total_gastos_agente_aduanal += imp.costo_aduana or 0
        costo_total_almacen_producto = (
            (imp.cif_ang or 0) +
            (imp.arancel or 0) +
            (imp.ob_ang or 0) +
            (imp.costo_aduana or 0) +
            (imp.flete_local or 0) -
            ob_dev_45
        )
        total_costo_total_almacen += costo_total_almacen_producto
        unidades_por_caja = imp.cantidad_cajas or 1
        cantidad_total_unidades = imp.cantidad_total * unidades_por_caja
        if cantidad_total_unidades > 0:
            costo_por_unidad_ang = costo_total_almacen_producto / cantidad_total_unidades
        else:
            costo_por_unidad_ang = 0
        precio_jomar = costo_por_unidad_ang
        precio_retail = precio_jomar * 1.2
        data.append([
            Paragraph(_pdf_xe(prod.nombre), style_cell),
            Paragraph("{:,.2f}".format(imp.cantidad_total), style_cell),
            Paragraph("{:,.2f}".format(imp.precio_fob_unidad), style_cell),
            Paragraph("{:,.2f}".format(total_fob_producto), style_cell),
            Paragraph("{:,.2f}".format(imp.flete), style_cell),
            Paragraph("{:,.2f}".format(total_cif_producto), style_cell),
            Paragraph("{:,.2f}".format(imp.cif_ang or 0), style_cell),
            Paragraph("{:,.2f}".format(imp.arancel or 0), style_cell),
            Paragraph("{:,.2f}".format(imp.ob_ang or 0), style_cell),
            Paragraph("{:,.2f}".format(ob_dev_45), style_cell),
            Paragraph("{:,.2f}".format(imp.costo_aduana or 0), style_cell),
            Paragraph("{:,.2f}".format(costo_total_almacen_producto), style_cell),
            Paragraph("{:,.2f}".format(costo_por_unidad_ang), style_cell)
        ])
    data.append([
        Paragraph("Totales", style_header),
        Paragraph("{:,.2f}".format(total_qty), style_header),
        Paragraph("", style_header),
        Paragraph("{:,.2f}".format(total_fob), style_header),
        Paragraph("{:,.2f}".format(total_flete), style_header),
        Paragraph("{:,.2f}".format(total_cif), style_header),
        Paragraph("{:,.2f}".format(total_cif_ang), style_header),
        Paragraph("{:,.2f}".format(total_arancel), style_header),
        Paragraph("{:,.2f}".format(total_ob_ang), style_header),
        Paragraph("{:,.2f}".format(total_ob_45), style_header),
        Paragraph("{:,.2f}".format(total_gastos_agente_aduanal), style_header),
        Paragraph("{:,.2f}".format(total_costo_total_almacen), style_header),
        Paragraph("", style_header)
    ])
    column_widths = [
        2.0 * inch,
        0.5 * inch,
        0.5 * inch,
        0.6 * inch,
        0.5 * inch,
        0.6 * inch,
        0.6 * inch,
        0.6 * inch,
        0.6 * inch,
        0.6 * inch,
        0.6 * inch,
        0.6 * inch,
        0.8 * inch
    ]
    total_column_width = sum(column_widths)
    usable_width = page_width - doc.leftMargin - doc.rightMargin
    if total_column_width > usable_width:
        scale_factor = usable_width / total_column_width
        column_widths = [width * scale_factor for width in column_widths]
    table = Table(data, colWidths=column_widths, repeatRows=1)
    table_style = TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('LEFTPADDING', (0, 0), (-1, -1), 1),
        ('RIGHTPADDING', (0, 0), (-1, -1), 1),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ])
    for row_num in range(1, len(data) - 1):
        if (row_num % 2) == 1:
            bg_color = colors.HexColor('#f2f2f2')
            table_style.add('BACKGROUND', (0, row_num), (-1, row_num), bg_color)
    table.setStyle(table_style)
    elements.append(table)
    elements.append(Spacer(1, 12))
    doc.build(elements)
    buffer.seek(0)
    nombre_archivo = f"reporte_factura_{numero_factura}.pdf"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype='application/pdf'
    )

@app.route('/clientes', methods=['GET'])
@login_required
def mostrar_clientes():
    # Verificar si es vendedor del nuevo sistema
    if not isinstance(current_user, Vendedor):
        # Usuario del sistema anterior - mostrar todos
        clientes = (Cliente
                    .query
                    .order_by(Cliente.id.asc())
                    .all())
    else:
        # NUEVO: Filtrar por vendedor según su rol
        if current_user.rol.nombre == 'super_admin':
            # Super admin ve todos los clientes
            clientes = (Cliente
                        .query
                        .order_by(Cliente.id.asc())
                        .all())
        else:
            # Vendedor regular: solo ve SUS clientes asignados
            clientes = current_user.obtener_clientes_visibles()
            # Ordenar por ID para mantener consistencia
            clientes = sorted(clientes, key=lambda c: c.id)

    return render_template('clientes.html', clientes=clientes)


# TAMBIÉN modifica estas rutas si existen:

@app.route('/clientes/nuevo', methods=['POST'])
@login_required
@requiere_permiso_recurso('clientes', 'crear')
def nuevo_cliente():
    # VERIFICAR PERMISOS: Solo super_admin puede crear clientes
    if isinstance(current_user, Vendedor) and current_user.rol.nombre != 'super_admin':
        return jsonify({"error": "No tienes permisos para crear clientes"}), 403
    
    try:
        nombre = request.form['nombre']
        qbo_id = request.form.get('qbo_id') or None
        
        # Evitar duplicados
        if qbo_id and Cliente.query.filter_by(qbo_id=qbo_id).first():
            return jsonify({"error": "Ya existe un cliente con ese QBO ID"}), 400
        
        moneda = request.form.get('moneda', 'XCG').upper()
        if moneda not in ('XCG', 'USD'):
            moneda = 'XCG'

        nuevo_cliente = Cliente(nombre=nombre, qbo_id=qbo_id, moneda=moneda)
        db.session.add(nuevo_cliente)
        db.session.commit()
        return jsonify({
            "message": "Cliente registrado exitosamente",
            "cliente": nuevo_cliente.to_dict()
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": "Error interno del servidor"}), 500

@app.route('/clientes/<int:cliente_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_cliente(cliente_id):
    # VERIFICAR PERMISOS
    if isinstance(current_user, Vendedor) and current_user.rol.nombre != 'super_admin':
        # Vendedor regular: solo puede editar SUS clientes
        if not current_user.puede_ver_cliente(cliente_id):
            flash('No tienes permisos para editar este cliente', 'error')
            return redirect(url_for('mostrar_clientes'))
    
    cliente = Cliente.query.get_or_404(cliente_id)

    if request.method == 'POST':
        try:
            cliente.nombre = request.form['nombre']
            qbo_id = request.form.get('qbo_id') or None

            # Verificar unicidad si cambió
            if qbo_id != cliente.qbo_id and qbo_id \
                    and Cliente.query.filter_by(qbo_id=qbo_id).first():
                flash('Ya existe otro cliente con ese QBO ID', 'danger')
                return redirect(url_for('editar_cliente', cliente_id=cliente.id))
            
            cliente.qbo_id = qbo_id
            moneda = request.form.get('moneda', 'XCG').upper()
            if moneda in ('XCG', 'USD'):
                cliente.moneda = moneda
            db.session.commit()
            flash('Cliente actualizado', 'success')
            return redirect(url_for('mostrar_clientes'))
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Error al editar cliente: {e}')
            flash('Error al editar cliente. Intente de nuevo.', 'danger')
            return redirect(url_for('mostrar_clientes'))

    return render_template('cliente_form.html', cliente=cliente)

@app.route('/clientes/<int:cliente_id>', methods=['DELETE'])
@login_required
@requiere_permiso_recurso('clientes', 'eliminar')
def eliminar_cliente(cliente_id):
    # VERIFICAR PERMISOS: Solo super_admin puede eliminar clientes
    if isinstance(current_user, Vendedor) and current_user.rol.nombre != 'super_admin':
        return jsonify({"error": "No tienes permisos para eliminar clientes"}), 403
    
    try:
        cliente = db.session.get(Cliente, cliente_id)
        if not cliente:
            return jsonify({"error": "Cliente no encontrado"}), 404
        db.session.delete(cliente)
        db.session.commit()
        return jsonify({"message": "Cliente eliminado exitosamente"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": "Error interno del servidor"}), 500

# Rutas de CRM eliminadas

@app.route('/generar_reporte', methods=['GET'])
@login_required
def generar_reporte():
    try:
        numero_factura = request.args.get('numero_factura')
        proveedor = request.args.get('proveedor')
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        query = Recepcion.query
        if numero_factura:
            query = query.filter_by(numero_factura=numero_factura)
        if proveedor:
            query = query.filter_by(proveedor=proveedor)
        if fecha_inicio:
            query = query.filter(Recepcion.recibido_en >= fecha_inicio)
        if fecha_fin:
            query = query.filter(Recepcion.recibido_en <= fecha_fin)
        recepciones = query.all()
        data = [{
            'Producto': r.producto.nombre if r.producto else 'undefined',
            'Peso': r.peso,
            'Proveedor': r.proveedor,
            'Número de Factura': r.numero_factura,
            'Fecha de Recepción': r.recibido_en.strftime('%Y-%m-%d')
        } for r in recepciones]
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'strings_to_formulas': False, 'strings_to_urls': False})
        worksheet = workbook.add_worksheet('Reporte')
        bold_blue_format = workbook.add_format({'bold': True, 'color': 'blue'})
        bold_red_format = workbook.add_format({'bold': True, 'color': 'red'})
        bold_black_format = workbook.add_format({'bold': True, 'color': 'black'})
        number_format = workbook.add_format({'num_format': '0.00'})
        date_format = workbook.add_format({'num_format': 'yyyy-mm-dd'})
        alignment = workbook.add_format({'align': 'left'})
        worksheet.write('A1', "Proveedor:", bold_blue_format)
        worksheet.write('B1', proveedor, bold_black_format)
        worksheet.write('A2', "Factura", bold_blue_format)
        worksheet.write('B2', numero_factura, bold_black_format)
        worksheet.write('A3', "Fecha", bold_blue_format)
        worksheet.write('B3', fecha_inicio, bold_black_format)
        worksheet.write('A5', "Resumen de Productos", bold_blue_format)
        resumen_df = pd.DataFrame(data).groupby('Producto')['Peso'].sum().reset_index()
        resumen_df.columns = ['Producto', 'Peso Total']
        worksheet.write('A7', "Producto", bold_blue_format)
        worksheet.write('B7', "Peso Total", bold_blue_format)
        total_peso = resumen_df['Peso Total'].sum()
        row = 8
        for idx, item in resumen_df.iterrows():
            worksheet.write(row, 0, item['Producto'], bold_black_format)
            worksheet.write(row, 1, item['Peso Total'], number_format)
            row += 1
        worksheet.write(row, 0, "Total", bold_blue_format)
        worksheet.write(row, 1, total_peso, bold_blue_format)
        row += 2
        worksheet.write(row, 0, "Detalles de Recepciones", bold_blue_format)
        row += 1
        worksheet.write(row, 0, "Producto", bold_blue_format)
        worksheet.write(row, 1, "Peso", bold_blue_format)
        worksheet.write(row, 2, "Fecha", bold_blue_format)
        worksheet.write(row, 3, "Proveedor", bold_blue_format)
        worksheet.write(row, 4, "Número de Factura", bold_blue_format)
        row += 1
        for r in data:
            worksheet.write(row, 0, r['Producto'], bold_black_format)
            worksheet.write(row, 1, r['Peso'], number_format)
            worksheet.write(row, 2, r['Fecha de Recepción'], date_format)
            worksheet.write(row, 3, r['Proveedor'], bold_black_format)
            worksheet.write(row, 4, r['Número de Factura'], bold_black_format)
            row += 1
        workbook.close()
        output.seek(0)
        nombre_archivo = f"reporte_recepciones_{numero_factura}.xlsx"
        return send_file(output, as_attachment=True, download_name=nombre_archivo, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        app.logger.error(f"Error al generar el reporte: {e}")
        return jsonify({'error': 'Error interno del servidor'}), 500

try:
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
except locale.Error:
    logging.warning("No se pudo configurar el locale 'en_US.UTF-8'. Se usará el formato de números por defecto.")

@app.route('/generar_reporte_pesos', methods=['GET'])
@login_required
def generar_reporte_pesos():
    import openpyxl
    from openpyxl.styles import Font, Alignment
    cliente_nombre = request.args.get('cliente')
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    fecha_inicio_date = datetime.strptime(fecha_inicio, '%Y-%m-%d')
    fecha_fin_date = datetime.strptime(fecha_fin, '%Y-%m-%d')
    facturaciones = Facturacion.query.join(Cliente).filter(
        Cliente.nombre == cliente_nombre,
        Facturacion.fecha_registro.between(fecha_inicio_date, fecha_fin_date)
    ).all()
    if not facturaciones:
        return jsonify({'error': 'No se encontraron pesos registrados para el cliente y rango de fechas especificados'}), 404
    wb = openpyxl.Workbook()
    ws = wb.active
    bold_font = Font(bold=True, color="0000FF")
    red_bold_font = Font(bold=True, color="FF0000")
    alignment = Alignment(horizontal="left")
    ws['A1'] = "Cliente:"
    ws['B1'] = _excel_safe(cliente_nombre)
    ws['A2'] = "Fecha de Registro:"
    ws['B2'] = f"{fecha_inicio} - {fecha_fin}"
    ws['A1'].font = bold_font
    ws['A2'].font = bold_font
    ws['A1'].alignment = alignment
    ws['A2'].alignment = alignment
    row = 4
    producto_grupo = {}
    for facturacion in facturaciones:
        producto = facturacion.producto.nombre
        if (producto not in producto_grupo):
            producto_grupo[producto] = []
        producto_grupo[producto].append(f"{facturacion.peso:.2f}")
    total_general = 0
    producto_index = 1
    for producto, pesos in producto_grupo.items():
        ws[f'A{row}'] = f"Producto {producto_index}:"
        ws[f'B{row}'] = _excel_safe(producto)
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'].font = bold_font
        row += 1
        ws[f'A{row}'] = "Pesos:"
        ws[f'A{row}'].font = bold_font
        col = 'B'
        total_producto = 0
        count = 0
        for peso in pesos:
            ws[f'{col}{row}'] = peso
            total_producto += float(peso)
            col = chr(ord(col) + 1)
            count += 1
            if count % 3 == 0:
                row += 1
                col = 'B'
        row += 1
        ws[f'A{row}'] = "Total:"
        ws[f'B{row}'] = f"{total_producto:.2f}"
        ws[f'A{row}'].font = red_bold_font
        ws[f'B{row}'].font = red_bold_font
        total_general += total_producto
        row += 2
        producto_index += 1
    ws[f'A{row}'] = "Total General:"
    ws[f'B{row}'] = f"{total_general:.2f}"
    ws[f'A{row}'].font = red_bold_font
    ws[f'B{row}'].font = red_bold_font
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    nombre_archivo_cliente = cliente_nombre.replace(" ", "_").replace("/", "_")
    nombre_archivo = f"reporte_pesos_{nombre_archivo_cliente}_{fecha_inicio}_a_{fecha_fin}.xlsx"
    return send_file(output, as_attachment=True, download_name=nombre_archivo, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/generar_etiqueta', methods=['POST'])
@login_required
def generar_etiqueta():
    try:
        cliente_nombre = request.form.get('cliente')
        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin = request.form.get('fecha_fin')
        if not cliente_nombre or not fecha_inicio or not fecha_fin:
            return jsonify({"error": "Todos los campos son obligatorios"}), 400
        fecha_inicio_date = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        fecha_fin_date = datetime.strptime(fecha_fin, '%Y-%m-%d')
        facturaciones = Facturacion.query.join(Cliente).filter(
            Cliente.nombre == cliente_nombre,
            Facturacion.fecha_registro.between(fecha_inicio_date, fecha_fin_date)
        ).all()
        if not facturaciones:
            return jsonify({"error": "No se encontraron facturaciones para los criterios seleccionados"}), 404
        output = BytesIO()
        page_width, page_height = A4
        etiqueta_ancho = 100.16 / 25.4 * inch
        etiqueta_alto = 50.8 / 25.4 * inch
        x_offset = (page_width - etiqueta_ancho) / 2
        y_offset_top = page_height - etiqueta_alto + 3
        y_offset_bottom = y_offset_top - etiqueta_alto - 3
        c = canvas.Canvas(output, pagesize=A4)
        etiquetas_por_pagina = 2
        etiqueta_contador = 0
        basedir = os.path.abspath(os.path.dirname(__file__))
        logo_path = os.path.join(basedir, 'static', 'logo_etiquetas.png')
        if not os.path.exists(logo_path):
            return jsonify({"error": f"El archivo {logo_path} no existe"}), 500
        for facturacion in facturaciones:
            producto = facturacion.producto
            cliente_nombre = facturacion.cliente.nombre if facturacion.cliente else "N/A"
            producto_nombre = producto.nombre if producto else "N/A"
            temperatura = producto.temperatura if producto and producto.temperatura else "N/A"
            if etiqueta_contador % etiquetas_por_pagina == 0:
                y_offset = y_offset_top
            else:
                y_offset = y_offset_bottom
            shift_left = 25
            logo_shift_up = 20
            logo_shift_right = 20
            c.drawImage(logo_path, x_offset + 10 - shift_left + logo_shift_right, y_offset + 30 + logo_shift_up, width=1.2 * inch, height=1.2 * inch)
            c.setFont("Helvetica-Bold", 10)
            label_x = x_offset + 2.8 * inch - shift_left
            value_x = label_x + 0.2 * inch
            c.drawRightString(label_x, y_offset + 1.7 * inch, "Client:")
            c.drawRightString(label_x, y_offset + 1.5 * inch, "Lot:")
            c.drawRightString(label_x, y_offset + 1.3 * inch, "Manufactured:")
            c.drawRightString(label_x, y_offset + 1.1 * inch, "Expiration:")
            c.drawRightString(label_x, y_offset + 0.9 * inch, "When Kept at:")
            c.drawString(value_x, y_offset + 1.7 * inch, cliente_nombre)
            c.drawString(value_x, y_offset + 1.5 * inch, facturacion.lote)
            c.drawString(value_x, y_offset + 1.3 * inch, facturacion.fecha_fabricacion)
            c.drawString(value_x, y_offset + 1.1 * inch, facturacion.fecha_expiracion)
            c.drawString(value_x, y_offset + 0.9 * inch, temperatura)
            c.setFont("Helvetica-Bold", 14)
            c.drawRightString(label_x, y_offset + 0.5 * inch, f"Net Weight:")
            c.drawString(value_x, y_offset + 0.5 * inch, f"{facturacion.peso:.2f}")
            c.setFont("Helvetica-Bold", 18)
            c.drawCentredString(x_offset + (etiqueta_ancho / 2), y_offset + 0.15 * inch, producto_nombre)
            etiqueta_contador += 1
            if etiqueta_contador % etiquetas_por_pagina == 0:
                c.showPage()
        if etiqueta_contador % etiquetas_por_pagina != 0:
            c.showPage()
        c.save()
        output.seek(0)
        cliente_filename = cliente_nombre.replace(" ", "_").replace("/", "-")
        return send_file(output, as_attachment=True, download_name=f"etiquetas_{cliente_filename}.pdf", mimetype="application/pdf")
    except Exception as e:
        return jsonify({"error": "Error interno del servidor"}), 500

def _build_datos_etiqueta_vencimiento(form):
    """Normaliza el form de etiquetas de vencimiento a (datos, cantidad)."""
    fecha_fabricacion_date = datetime.strptime(form['fecha_fabricacion'], '%Y-%m-%d')
    fecha_expiracion = fecha_fabricacion_date + timedelta(days=365)
    producto = db.session.get(Producto, form['producto_id'])
    datos = {
        "nombre_producto": producto.nombre,
        "lote": form['lote'],
        "fecha_fabricacion": fecha_fabricacion_date.strftime('%d/%m/%Y'),
        "fecha_expiracion": fecha_expiracion.strftime('%d/%m/%Y'),
        "temperatura": producto.temperatura,
    }
    return datos, int(form['cantidad_etiquetas'])


@app.route('/etiquetas_vencimiento', methods=['GET', 'POST'])
@login_required
def etiquetas_vencimiento():
    """Genera etiquetas de vencimiento en formato A4 (2 por página)."""
    if request.method == 'POST':
        datos, cantidad = _build_datos_etiqueta_vencimiento(request.form)
        return generar_pdf_etiquetas(datos, cantidad)

    productos = Producto.query.all()
    return render_template('form_generar_etiquetas.html', productos=productos)


@app.route('/etiquetas_vencimiento_4x2', methods=['POST'])
@login_required
def etiquetas_vencimiento_4x2():
    """Genera etiquetas de vencimiento en formato 4"x2" (térmica, una por página)."""
    datos, cantidad = _build_datos_etiqueta_vencimiento(request.form)
    return generar_pdf_etiquetas_4x2(datos, cantidad)


def generar_pdf_etiquetas_4x2(datos, cantidad):
    """Genera PDF térmico 4"x2" (una etiqueta por página) con datos de vencimiento."""
    output, c = create_single_label_pdf()
    logo_path = get_logo_path(basedir)
    for _ in range(cantidad):
        draw_expiration_label(c, logo_path, datos, 0, 0)
        c.showPage()
    c.save()
    output.seek(0)
    return send_file(
        output,
        as_attachment=not _is_ios_request(),
        download_name="etiquetas_vencimiento_4x2.pdf",
        mimetype='application/pdf',
    )


def generar_pdf_etiquetas(datos, cantidad):
    """
    Genera PDF con etiquetas 4"x2" (dos por página, una arriba y otra abajo).
    Formato para etiquetas de vencimiento (sin cliente ni peso).
    """
    output, c, page_width, page_height = create_letter_page_pdf()
    logo_path = get_logo_path(basedir)
    x_centrado = get_centered_x(page_width)

    etiqueta_num = 0
    while etiqueta_num < cantidad:
        # Primera etiqueta: arriba
        y_pos_1 = page_height - LABEL_HEIGHT
        draw_expiration_label(c, logo_path, datos, x_centrado, y_pos_1)
        etiqueta_num += 1

        # Segunda etiqueta: abajo
        if etiqueta_num < cantidad:
            y_pos_2 = page_height - 2 * LABEL_HEIGHT
            draw_expiration_label(c, logo_path, datos, x_centrado, y_pos_2)
            etiqueta_num += 1

        # Nueva página si hay más etiquetas
        if etiqueta_num < cantidad:
            c.showPage()

    c.save()
    output.seek(0)
    return send_file(output, as_attachment=not _is_ios_request(), download_name="etiquetas_vencimiento.pdf", mimetype='application/pdf')


try:
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
except locale.Error:
    logging.warning("No se pudo configurar el locale 'en_US.UTF-8'. Se usará el formato de números por defecto.")



def validar_estructura_csv(csv_input, campos_requeridos):
    """Valida que el CSV tenga los campos requeridos"""
    try:
        primera_fila = next(csv_input)
        campos_csv = set(primera_fila.keys())
        campos_faltantes = set(campos_requeridos) - campos_csv
        
        if campos_faltantes:
            raise ValueError(f"Campos faltantes en CSV: {', '.join(campos_faltantes)}")
        
        # Resetear el iterator (necesario para pandas o re-leer el archivo)
        return True
    except StopIteration:
        raise ValueError("El archivo CSV está vacío")

def limpiar_codigo(codigo):
    """Limpia y normaliza códigos de productos/clientes"""
    if not codigo:
        return None
    return str(codigo).strip().upper()

def validar_precio(precio_str):
    """Valida y convierte precio a float"""
    try:
        precio = float(precio_str)
        if precio < 0:
            raise ValueError("El precio no puede ser negativo")
        return precio
    except (ValueError, TypeError):
        raise ValueError(f"Precio inválido: {precio_str}")

def validar_margen(margen_str, default=1.0):
    """Valida y convierte margen a float"""
    if not margen_str or margen_str.strip() == '':
        return default
    try:
        margen = float(margen_str)
        if margen <= 0:
            raise ValueError("El margen debe ser mayor a 0")
        return margen
    except (ValueError, TypeError):
        raise ValueError(f"Margen inválido: {margen_str}")

# Función mejorada para procesar precios por lista
def procesar_precios_por_lista_mejorado(csv_input, lista_precio_id, resultados):
    """
    Versión mejorada del procesamiento de precios por lista con más validaciones
    """
    if not lista_precio_id:
        raise ValueError("Se requiere seleccionar una lista de precios")
    
    lista = db.session.get(ListaPrecio, lista_precio_id)
    if not lista:
        raise ValueError("Lista de precios no encontrada")
    
    # Validar estructura del CSV
    campos_requeridos = ['codigo_producto', 'precio_base']
    
    # Contar total de filas para progress tracking
    filas_procesadas = 0
    batch_size = 100  # Procesar en lotes para mejor rendimiento
    
    for fila_num, fila in enumerate(csv_input, start=2):
        try:
            # Limpiar y validar código de producto
            codigo_producto = limpiar_codigo(fila.get('codigo_producto'))
            if not codigo_producto:
                resultados['errores'] += 1
                resultados['detalles'].append(f'Fila {fila_num}: Código de producto es obligatorio')
                continue
            
            # Buscar producto
            producto = Producto.query.filter_by(codigo=codigo_producto).first()
            if not producto:
                resultados['errores'] += 1
                resultados['detalles'].append(f'Fila {fila_num}: Producto {codigo_producto} no encontrado')
                continue
            
            # Validar y convertir precio base
            try:
                precio_base = validar_precio(fila.get('precio_base', ''))
            except ValueError as e:
                resultados['errores'] += 1
                resultados['detalles'].append(f'Fila {fila_num}: {str(e)}')
                continue
            
            # Obtener y validar márgenes
            try:
                margen_jomar = validar_margen(fila.get('margen_jomar', ''), 1.0)
                margen_retail = validar_margen(fila.get('margen_retail', ''), 1.2)
            except ValueError as e:
                resultados['errores'] += 1
                resultados['detalles'].append(f'Fila {fila_num}: {str(e)}')
                continue
            
            # Verificar límites razonables para márgenes
            if margen_jomar > 10 or margen_retail > 10:
                resultados['warnings'].append(f'Fila {fila_num}: Márgenes muy altos para producto {codigo_producto}')
            
            # Buscar precio existente o crear nuevo
            precio_existente = PrecioProducto.query.filter_by(
                lista_precio_id=lista_precio_id,
                producto_id=producto.id
            ).first()
            
            if precio_existente:
                # Actualizar existente
                precio_existente.precio_base = precio_base
                precio_existente.margen_jomar = margen_jomar
                precio_existente.margen_retail = margen_retail
                precio_existente.calcular_precios()
                precio_existente.fecha_actualizacion = datetime.utcnow()
                accion = 'actualizado'
            else:
                # Crear nuevo
                nuevo_precio = PrecioProducto(
                    lista_precio_id=lista_precio_id,
                    producto_id=producto.id,
                    precio_base=precio_base,
                    margen_jomar=margen_jomar,
                    margen_retail=margen_retail
                )
                nuevo_precio.calcular_precios()
                db.session.add(nuevo_precio)
                accion = 'creado'
            
            resultados['procesados'] += 1
            resultados['detalles'].append(f'Fila {fila_num}: Precio para {codigo_producto} {accion} exitosamente')
            
            filas_procesadas += 1
            
            # Commit en lotes para mejor rendimiento
            if filas_procesadas % batch_size == 0:
                db.session.flush()
                
        except Exception as e:
            resultados['errores'] += 1
            resultados['detalles'].append(f'Fila {fila_num}: Error inesperado - {str(e)}')
    
    return resultados

# Función para generar reporte de carga
@app.route('/precios/generar-reporte-carga', methods=['POST'])
@login_required
def generar_reporte_carga():
    """Genera un reporte detallado de la carga de precios"""
    try:
        datos = request.get_json()
        tipo_reporte = datos.get('tipo', 'lista_precios')
        lista_id = datos.get('lista_id')
        
        if tipo_reporte == 'lista_precios' and lista_id:
            lista = db.session.get(ListaPrecio, lista_id)
            if not lista:
                return jsonify({'error': 'Lista no encontrada'}), 404
            
            # Obtener precios de la lista
            precios = db.session.query(PrecioProducto, Producto).join(
                Producto, PrecioProducto.producto_id == Producto.id
            ).filter(PrecioProducto.lista_precio_id == lista_id).all()
            
            output = io.StringIO()
            fieldnames = ['codigo_producto', 'nombre_producto', 'precio_base', 
                         'precio_jomar', 'precio_retail', 'margen_jomar', 'margen_retail', 
                         'fecha_actualizacion']
            
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            
            for precio, producto in precios:
                writer.writerow({
                    'codigo_producto': _excel_safe(producto.codigo),
                    'nombre_producto': _excel_safe(producto.nombre),
                    'precio_base': precio.precio_base,
                    'precio_jomar': precio.precio_jomar,
                    'precio_retail': precio.precio_retail,
                    'margen_jomar': precio.margen_jomar,
                    'margen_retail': precio.margen_retail,
                    'fecha_actualizacion': precio.fecha_actualizacion.strftime('%Y-%m-%d %H:%M:%S') if precio.fecha_actualizacion else ''
                })
            
            csv_data = output.getvalue()
            output.close()
            
            response = make_response(csv_data)
            response.headers['Content-Type'] = 'text/csv; charset=utf-8'
            response.headers['Content-Disposition'] = f'attachment; filename=reporte_precios_{lista.nombre.replace(" ", "_")}.csv'
            
            return response
        
        return jsonify({'error': 'Tipo de reporte no soportado'}), 400
        
    except Exception as e:
        app.logger.error(f'Error generando reporte: {e}')
        return jsonify({'error': 'Error generando reporte'}), 500

# Función para validar CSV antes de procesarlo
@app.route('/precios/validar-csv', methods=['POST'])
@login_required
@requiere_permiso_recurso('precios', 'editar')
def validar_csv_precios():
    """Valida un CSV antes de procesarlo completamente"""
    # Tipos MIME válidos para archivos CSV
    ALLOWED_MIME_TYPES = ['text/csv', 'text/plain', 'application/csv', 'application/vnd.ms-excel']
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB máximo

    try:
        if 'archivo_csv' not in request.files:
            return jsonify({'error': 'No se encontró archivo CSV'}), 400

        archivo = request.files['archivo_csv']
        if archivo.filename == '':
            return jsonify({'error': 'No se seleccionó archivo'}), 400

        # Validar nombre de archivo con secure_filename
        filename = secure_filename(archivo.filename)
        if not filename.lower().endswith('.csv'):
            return jsonify({'error': 'El archivo debe ser CSV'}), 400

        # Validar tipo MIME
        if archivo.content_type and archivo.content_type not in ALLOWED_MIME_TYPES:
            app.logger.warning(f'Intento de validar archivo con MIME type no permitido: {archivo.content_type}')
            return jsonify({'error': 'Tipo de archivo no permitido'}), 400

        tipo_carga = request.form.get('tipo_carga')

        # Leer contenido y validar tamaño
        content = archivo.stream.read()
        if len(content) > MAX_FILE_SIZE:
            return jsonify({'error': 'El archivo excede el tamaño máximo permitido (10MB)'}), 400

        # Validar que el contenido es texto UTF-8 válido
        try:
            decoded_content = content.decode("UTF8")
        except UnicodeDecodeError:
            return jsonify({'error': 'El archivo debe estar codificado en UTF-8'}), 400

        # Leer primeras 10 filas para validación
        stream = io.StringIO(decoded_content, newline=None)
        csv_input = csv.DictReader(stream)
        
        validacion = {
            'valido': True,
            'errores': [],
            'warnings': [],
            'preview': [],
            'total_filas': 0
        }
        
        # Definir campos requeridos según tipo
        campos_requeridos = {
            'lista_precios': ['codigo_producto', 'precio_base'],
            'asignacion_clientes': ['codigo_cliente', 'nombre_lista_precio'],
            'precios_especificos': ['codigo_cliente', 'codigo_producto', 'precio_base']
        }
        
        if tipo_carga not in campos_requeridos:
            validacion['valido'] = False
            validacion['errores'].append('Tipo de carga no válido')
            return jsonify(validacion), 400
        
        # Validar encabezados
        try:
            primera_fila = next(csv_input)
            campos_csv = set(primera_fila.keys())
            campos_faltantes = set(campos_requeridos[tipo_carga]) - campos_csv
            
            if campos_faltantes:
                validacion['valido'] = False
                validacion['errores'].append(f'Campos faltantes: {", ".join(campos_faltantes)}')
            
            # Agregar primera fila al preview
            validacion['preview'].append(primera_fila)
            validacion['total_filas'] = 1
            
        except StopIteration:
            validacion['valido'] = False
            validacion['errores'].append('El archivo CSV está vacío')
            return jsonify(validacion), 400
        
        # Validar siguientes filas (máximo 9 más para preview)
        for i, fila in enumerate(csv_input):
            if i >= 9:  # Solo revisar 10 filas total
                break
                
            validacion['preview'].append(fila)
            validacion['total_filas'] += 1
            
            # Validaciones básicas según tipo
            if tipo_carga == 'lista_precios':
                if not fila.get('codigo_producto', '').strip():
                    validacion['warnings'].append(f'Fila {i+2}: Código producto vacío')
                
                try:
                    precio = float(fila.get('precio_base', 0))
                    if precio <= 0:
                        validacion['warnings'].append(f'Fila {i+2}: Precio base debe ser mayor a 0')
                except ValueError:
                    validacion['warnings'].append(f'Fila {i+2}: Precio base no es un número válido')
        
        # Contar filas totales (reset stream)
        archivo.stream.seek(0)
        stream = io.StringIO(archivo.stream.read().decode("UTF8"), newline=None)
        csv_input = csv.DictReader(stream)
        validacion['total_filas'] = sum(1 for _ in csv_input)
        
        return jsonify(validacion), 200
        
    except Exception as e:
        app.logger.error(f'Error validando archivo: {e}')
        return jsonify({'error': 'Error validando archivo'}), 500

# Agregar logging para mejor debugging
@app.route('/precios/log-carga', methods=['POST'])
@login_required  
def log_carga_precios():
    """Registra actividades de carga masiva para auditoría"""
    try:
        datos = request.get_json() or {}

        # Sanea el texto controlado por el usuario antes de loguear (log injection).
        def _log_safe(v, n=100):
            return str(v).replace('\n', ' ').replace('\r', ' ')[:n]

        # Aquí podrías agregar a una tabla de auditoría
        logging.info(f"Carga masiva ejecutada por usuario {_log_safe(current_user.username, 60)}: "
                    f"Tipo: {_log_safe(datos.get('tipo'))}, "
                    f"Registros: {datos.get('procesados', 0)}, "
                    f"Errores: {datos.get('errores', 0)}")
        
        return jsonify({'status': 'logged'}), 200
        
    except Exception as e:
        return jsonify({'error': 'Error interno del servidor'}), 500

# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('carga_masiva.log'),
        logging.StreamHandler()
    ]
)

@app.route('/precios/carga-masiva')
@login_required
@requiere_permiso_recurso('precios', 'editar')
def carga_masiva_precios():
    """Interfaz para carga masiva de precios mediante CSV"""
    listas = ListaPrecio.query.filter_by(activa=True).all()
    clientes = Cliente.query.all()
    productos = Producto.query.all()
    
    return render_template('precios/carga_masiva.html',
                         listas=listas,
                         listas_precios=listas,
                         clientes=clientes,
                         productos=productos)

@app.route('/precios/procesar-csv', methods=['POST'])
@login_required
@requiere_permiso_recurso('precios', 'editar')
def procesar_csv_precios():
    """Procesar archivo CSV con precios"""
    # Tipos MIME válidos para archivos CSV
    ALLOWED_MIME_TYPES = ['text/csv', 'text/plain', 'application/csv', 'application/vnd.ms-excel']
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB máximo

    try:
        if 'archivo_csv' not in request.files:
            return jsonify({'error': 'No se encontró archivo CSV'}), 400

        archivo = request.files['archivo_csv']
        if archivo.filename == '':
            return jsonify({'error': 'No se seleccionó archivo'}), 400

        # Validar nombre de archivo con secure_filename
        filename = secure_filename(archivo.filename)
        if not filename.lower().endswith('.csv'):
            return jsonify({'error': 'El archivo debe ser CSV'}), 400

        # Validar tipo MIME
        if archivo.content_type and archivo.content_type not in ALLOWED_MIME_TYPES:
            app.logger.warning(f'Intento de subir archivo con MIME type no permitido: {archivo.content_type}')
            return jsonify({'error': 'Tipo de archivo no permitido'}), 400

        # Leer contenido y validar tamaño
        content = archivo.stream.read()
        if len(content) > MAX_FILE_SIZE:
            return jsonify({'error': 'El archivo excede el tamaño máximo permitido (10MB)'}), 400

        tipo_carga = request.form.get('tipo_carga')
        lista_precio_id = request.form.get('lista_precio_id')

        # Validar que el contenido es texto UTF-8 válido
        try:
            decoded_content = content.decode("UTF8")
        except UnicodeDecodeError:
            return jsonify({'error': 'El archivo debe estar codificado en UTF-8'}), 400

        # Leer archivo CSV
        stream = io.StringIO(decoded_content, newline=None)
        csv_input = csv.DictReader(stream)
        
        resultados = {
            'procesados': 0,
            'errores': 0,
            'warnings': [],
            'detalles': []
        }
        
        if tipo_carga == 'lista_precios':
            resultados = procesar_precios_por_lista(csv_input, lista_precio_id, resultados)
        elif tipo_carga == 'asignacion_clientes':
            resultados = procesar_asignacion_clientes(csv_input, resultados)
        elif tipo_carga == 'precios_especificos':
            resultados = procesar_precios_especificos(csv_input, resultados)
        else:
            return jsonify({'error': 'Tipo de carga no válido'}), 400
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'mensaje': f'Procesamiento completado. {resultados["procesados"]} registros procesados, {resultados["errores"]} errores.',
            'resultados': resultados
        }), 200
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error procesando archivo: {e}')
        return jsonify({'error': 'Error procesando archivo'}), 500

def procesar_precios_por_lista(csv_input, lista_precio_id, resultados):
    """Procesar CSV para actualizar precios en una lista específica"""
    if not lista_precio_id:
        raise ValueError("Se requiere seleccionar una lista de precios")
    
    lista = db.session.get(ListaPrecio, lista_precio_id)
    if not lista:
        raise ValueError("Lista de precios no encontrada")
    
    for fila_num, fila in enumerate(csv_input, start=2):
        try:
            # Validar campos requeridos
            codigo_producto = fila.get('codigo_producto', '').strip()
            precio_base = fila.get('precio_base', '').strip()
            
            if not codigo_producto or not precio_base:
                resultados['errores'] += 1
                resultados['detalles'].append(f'Fila {fila_num}: Código producto y precio base son obligatorios')
                continue
            
            # CORRECCIÓN: Buscar producto por ID (ya que los valores en CSV son IDs)
            try:
                producto_id = int(codigo_producto)
                producto = db.session.get(Producto, producto_id)
            except ValueError:
                # Si no es un número, intentar buscar por qbo_id o nombre
                producto = Producto.query.filter(
                    (Producto.qbo_id == codigo_producto) | 
                    (Producto.nombre.ilike(f'%{codigo_producto}%'))
                ).first()
            
            if not producto:
                resultados['errores'] += 1
                resultados['detalles'].append(f'Fila {fila_num}: Producto {codigo_producto} no encontrado')
                continue
            
            # Validar precio base
            try:
                precio_base = float(precio_base)
                if precio_base < 0:
                    raise ValueError("Precio no puede ser negativo")
            except ValueError:
                resultados['errores'] += 1
                resultados['detalles'].append(f'Fila {fila_num}: Precio base inválido')
                continue
            
            # Obtener márgenes (opcionales, usar defaults si no se especifican)
            margen_jomar = float(fila.get('margen_jomar', 1.0) or 1.0)
            margen_retail = float(fila.get('margen_retail', 1.2) or 1.2)
            
            # Buscar precio existente o crear nuevo
            precio_existente = PrecioProducto.query.filter_by(
                lista_precio_id=lista_precio_id,
                producto_id=producto.id
            ).first()
            
            if precio_existente:
                # Actualizar existente
                precio_existente.precio_base = precio_base
                precio_existente.margen_jomar = margen_jomar
                precio_existente.margen_retail = margen_retail
                precio_existente.calcular_precios()
                precio_existente.fecha_actualizacion = datetime.utcnow()
                accion = 'actualizado'
            else:
                # Crear nuevo
                nuevo_precio = PrecioProducto(
                    lista_precio_id=lista_precio_id,
                    producto_id=producto.id,
                    precio_base=precio_base,
                    margen_jomar=margen_jomar,
                    margen_retail=margen_retail
                )
                nuevo_precio.calcular_precios()
                db.session.add(nuevo_precio)
                accion = 'creado'
            
            resultados['procesados'] += 1
            resultados['detalles'].append(f'Fila {fila_num}: Precio para producto {producto.nombre} (ID: {producto.id}) {accion} exitosamente')
            
        except Exception as e:
            resultados['errores'] += 1
            resultados['detalles'].append(f'Fila {fila_num}: Error - {str(e)}')
    
    return resultados

def procesar_asignacion_clientes(csv_input, resultados):
    """Procesar CSV para asignar listas de precios a clientes"""
    for fila_num, fila in enumerate(csv_input, start=2):
        try:
            codigo_cliente = fila.get('codigo_cliente', '').strip()
            nombre_lista = fila.get('nombre_lista_precio', '').strip()
            
            if not codigo_cliente or not nombre_lista:
                resultados['errores'] += 1
                resultados['detalles'].append(f'Fila {fila_num}: Código cliente y nombre lista son obligatorios')
                continue
            
            # Buscar cliente
            cliente = Cliente.query.filter_by(codigo=codigo_cliente).first()
            if not cliente:
                resultados['errores'] += 1
                resultados['detalles'].append(f'Fila {fila_num}: Cliente {codigo_cliente} no encontrado')
                continue
            
            # Buscar lista de precios
            lista = ListaPrecio.query.filter_by(nombre=nombre_lista, activa=True).first()
            if not lista:
                resultados['errores'] += 1
                resultados['detalles'].append(f'Fila {fila_num}: Lista {nombre_lista} no encontrada')
                continue
            
            # Verificar si ya existe asignación activa
            asignacion_existente = ClienteListaPrecio.query.filter_by(
                cliente_id=cliente.id,
                activa=True
            ).first()
            
            if asignacion_existente:
                # Desactivar asignación anterior
                asignacion_existente.activa = False
                resultados['warnings'].append(f'Cliente {codigo_cliente}: Lista anterior desactivada')
            
            # Crear nueva asignación
            nueva_asignacion = ClienteListaPrecio(
                cliente_id=cliente.id,
                lista_precio_id=lista.id
            )
            db.session.add(nueva_asignacion)
            
            resultados['procesados'] += 1
            resultados['detalles'].append(f'Fila {fila_num}: Lista {nombre_lista} asignada a cliente {codigo_cliente}')
            
        except Exception as e:
            resultados['errores'] += 1
            resultados['detalles'].append(f'Fila {fila_num}: Error - {str(e)}')
    
    return resultados

def procesar_precios_especificos(csv_input, resultados):
    """Procesar CSV para precios específicos cliente-producto"""
    for fila_num, fila in enumerate(csv_input, start=2):
        try:
            codigo_cliente = fila.get('codigo_cliente', '').strip()
            codigo_producto = fila.get('codigo_producto', '').strip()
            precio_base = fila.get('precio_base', '').strip()
            
            if not codigo_cliente or not codigo_producto or not precio_base:
                resultados['errores'] += 1
                resultados['detalles'].append(f'Fila {fila_num}: Cliente, producto y precio base son obligatorios')
                continue
            
            # Buscar cliente por ID o nombre
            try:
                cliente_id = int(codigo_cliente)
                cliente = db.session.get(Cliente, cliente_id)
            except ValueError:
                cliente = Cliente.query.filter(Cliente.nombre.ilike(f'%{codigo_cliente}%')).first()
            
            if not cliente:
                resultados['errores'] += 1
                resultados['detalles'].append(f'Fila {fila_num}: Cliente {codigo_cliente} no encontrado')
                continue
            
            # CORRECCIÓN: Buscar producto por ID (igual que arriba)
            try:
                producto_id = int(codigo_producto)
                producto = db.session.get(Producto, producto_id)
            except ValueError:
                producto = Producto.query.filter(
                    (Producto.qbo_id == codigo_producto) | 
                    (Producto.nombre.ilike(f'%{codigo_producto}%'))
                ).first()
            
            if not producto:
                resultados['errores'] += 1
                resultados['detalles'].append(f'Fila {fila_num}: Producto {codigo_producto} no encontrado')
                continue
            
            # Validar precio
            try:
                precio_base = float(precio_base)
                if precio_base < 0:
                    raise ValueError("Precio no puede ser negativo")
            except ValueError:
                resultados['errores'] += 1
                resultados['detalles'].append(f'Fila {fila_num}: Precio base inválido')
                continue
            
            # Obtener márgenes
            margen_jomar = float(fila.get('margen_jomar', 1.0) or 1.0)
            margen_retail = float(fila.get('margen_retail', 1.2) or 1.2)
            
            # Buscar precio específico existente o crear nuevo
            precio_existente = PrecioClienteProducto.query.filter_by(
                cliente_id=cliente.id,
                producto_id=producto.id,
                activo=True
            ).first()
            
            if precio_existente:
                # Actualizar existente
                precio_existente.precio_base = precio_base
                precio_existente.margen_jomar = margen_jomar
                precio_existente.margen_retail = margen_retail
                precio_existente.calcular_precios()
                precio_existente.fecha_actualizacion = datetime.utcnow()
                accion = 'actualizado'
            else:
                # Crear nuevo
                nuevo_precio = PrecioClienteProducto(
                    cliente_id=cliente.id,
                    producto_id=producto.id,
                    precio_base=precio_base,
                    margen_jomar=margen_jomar,
                    margen_retail=margen_retail
                )
                nuevo_precio.calcular_precios()
                db.session.add(nuevo_precio)
                accion = 'creado'
            
            resultados['procesados'] += 1
            resultados['detalles'].append(f'Fila {fila_num}: Precio específico {cliente.nombre}-{producto.nombre} {accion}')
            
        except Exception as e:
            resultados['errores'] += 1
            resultados['detalles'].append(f'Fila {fila_num}: Error - {str(e)}')
    
    return resultados

@app.route('/precios/descargar-plantilla/<tipo>')
@login_required
def descargar_plantilla_csv(tipo):
    """Descargar plantillas CSV para diferentes tipos de carga"""
    output = io.StringIO()
    
    if tipo == 'lista_precios':
        fieldnames = ['codigo_producto', 'precio_base', 'margen_jomar', 'margen_retail']
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        # Obtener algunos productos reales para el ejemplo
        productos_ejemplo = Producto.query.limit(3).all()
        if productos_ejemplo:
            for producto in productos_ejemplo:
                writer.writerow({
                    'codigo_producto': str(producto.id),  # Usar ID real
                    'precio_base': '25.50',
                    'margen_jomar': '1.0',
                    'margen_retail': '1.2'
                })
        else:
            # Fallback si no hay productos
            writer.writerow({
                'codigo_producto': '1',
                'precio_base': '25.50',
                'margen_jomar': '1.0',
                'margen_retail': '1.2'
            })
        filename = 'plantilla_precios_lista.csv'
        
    elif tipo == 'asignacion_clientes':
        fieldnames = ['codigo_cliente', 'nombre_lista_precio']
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        # Obtener ejemplos reales
        cliente_ejemplo = Cliente.query.first()
        lista_ejemplo = ListaPrecio.query.first()
        
        writer.writerow({
            'codigo_cliente': str(cliente_ejemplo.id) if cliente_ejemplo else '1',
            'nombre_lista_precio': _excel_safe(lista_ejemplo.nombre) if lista_ejemplo else 'Lista Mayorista'
        })
        filename = 'plantilla_asignacion_clientes.csv'
        
    elif tipo == 'precios_especificos':
        fieldnames = ['codigo_cliente', 'codigo_producto', 'precio_base', 'margen_jomar', 'margen_retail']
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        cliente_ejemplo = Cliente.query.first()
        producto_ejemplo = Producto.query.first()
        
        writer.writerow({
            'codigo_cliente': str(cliente_ejemplo.id) if cliente_ejemplo else '1',
            'codigo_producto': str(producto_ejemplo.id) if producto_ejemplo else '1',
            'precio_base': '23.75',
            'margen_jomar': '1.0',
            'margen_retail': '1.15'
        })
        filename = 'plantilla_precios_especificos.csv'
    else:
        return jsonify({'error': 'Tipo de plantilla no válido'}), 400
    
    # Crear respuesta con archivo CSV
    csv_data = output.getvalue()
    output.close()
    
    response = make_response(csv_data)
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

@app.route('/dev/primitives')
@login_required
def dev_primitives():
    """Admin-only showcase for Phase 1 glass foundation primitives.

    Living reference used by Phases 2-5. Renders every primitive in
    both light and dark themes (theme comes from the user's system
    setting via prefers-color-scheme)."""
    if not isinstance(current_user, Vendedor) or current_user.rol.nombre != 'super_admin':
        abort(403)
    return render_template('dev_primitives.html')


def _get_registro_config():
    """Devuelve la fila única de RegistroConfig; la crea con valores por
    defecto si aún no existe."""
    cfg = RegistroConfig.query.first()
    if cfg is None:
        cfg = RegistroConfig()
        db.session.add(cfg)
        db.session.commit()
    return cfg


# ───────────────────────── HACCP: Cámaras ─────────────────────────
_TIPOS_CAMARA = ('refrigeracion', 'congelacion')


def _parse_rango_camara(form):
    """Devuelve (nombre, tipo, temp_min, temp_max, error)."""
    nombre = (form.get('nombre') or '').strip()
    tipo = (form.get('tipo') or '').strip()
    if not nombre:
        return None, None, None, None, 'El nombre es obligatorio.'
    if tipo not in _TIPOS_CAMARA:
        return None, None, None, None, 'Tipo de cámara no válido.'
    try:
        temp_min = Decimal(str(form.get('temp_min')).replace(',', '.'))
        temp_max = Decimal(str(form.get('temp_max')).replace(',', '.'))
    except (InvalidOperation, TypeError, ValueError):
        return None, None, None, None, 'Temperaturas mínima y máxima deben ser números.'
    if temp_min > temp_max:
        return None, None, None, None, 'La temperatura mínima no puede ser mayor que la máxima.'
    if tipo == 'congelacion' and temp_max >= 0:
        return None, None, None, None, ('En cámaras de congelación las temperaturas deben ser '
                                        'bajo cero (la máxima debe ser menor a 0°C).')
    return nombre, tipo, temp_min, temp_max, None


def _registrado_en_from_form(form, field='registrado_en'):
    """Convierte el valor de un <input type="datetime-local"> (en hora local de
    negocio) a un datetime naive en UTC, para guardarlo igual que `default=utcnow`.
    Si está vacío o es inválido, devuelve None (→ se usará la hora actual)."""
    raw = (form.get(field) or '').strip()
    if not raw:
        return None
    for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%dT%H:%M:%S'):
        try:
            local = datetime.strptime(raw, fmt)
            break
        except ValueError:
            local = None
    if local is None:
        return None
    # Interpretar como hora local de negocio y pasar a UTC naive.
    local = local.replace(tzinfo=DASHBOARD_TIMEZONE)
    return local.astimezone(timezone.utc).replace(tzinfo=None)


def _hora_valida(s):
    """Normaliza 'HH:MM' o devuelve None."""
    s = (s or '').strip()
    if not s:
        return None
    try:
        datetime.strptime(s, '%H:%M')
        return s
    except ValueError:
        return None


@app.route('/registros/temperaturas/camaras')
@login_required
@requiere_permiso_recurso('registros', 'editar')
def camaras_list():
    camaras = Camara.query.order_by(Camara.nombre).all()
    responsables = Vendedor.query.filter_by(activo=True).order_by(Vendedor.nombre_completo).all()
    return render_template('registros/camaras.html', camaras=camaras, responsables=responsables)


@app.route('/registros/temperaturas/camaras/nueva', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'editar')
def camara_nueva():
    nombre, tipo, temp_min, temp_max, error = _parse_rango_camara(request.form)
    if error:
        flash(error, 'danger')
        return redirect(url_for('camaras_list'))
    db.session.add(Camara(nombre=nombre, tipo=tipo, temp_min=temp_min,
                          temp_max=temp_max, activa=True,
                          responsable_id=request.form.get('responsable_id', type=int),
                          ronda_am=_hora_valida(request.form.get('ronda_am')),
                          ronda_pm=_hora_valida(request.form.get('ronda_pm'))))
    db.session.commit()
    _audit('config', 'Creó cámara', nombre)
    flash('Cámara creada.', 'success')
    return redirect(url_for('camaras_list'))


@app.route('/registros/temperaturas/camaras/<int:camara_id>/editar', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'editar')
def camara_editar(camara_id):
    camara = Camara.query.get_or_404(camara_id)
    nombre, tipo, temp_min, temp_max, error = _parse_rango_camara(request.form)
    if error:
        flash(error, 'danger')
        return redirect(url_for('camaras_list'))
    camara.nombre, camara.tipo, camara.temp_min, camara.temp_max = nombre, tipo, temp_min, temp_max
    camara.responsable_id = request.form.get('responsable_id', type=int)
    camara.ronda_am = _hora_valida(request.form.get('ronda_am'))
    camara.ronda_pm = _hora_valida(request.form.get('ronda_pm'))
    db.session.commit()
    _audit('config', 'Editó cámara', nombre)
    flash('Cámara actualizada.', 'success')
    return redirect(url_for('camaras_list'))


@app.route('/registros/temperaturas/camaras/<int:camara_id>/toggle', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'editar')
def camara_toggle(camara_id):
    camara = Camara.query.get_or_404(camara_id)
    camara.activa = not camara.activa
    db.session.commit()
    flash('Cámara actualizada.', 'success')
    return redirect(url_for('camaras_list'))


def _camaras_con_lectura_hoy():
    """Set de camara_id con al menos una lectura HOY (día local de negocio).
    registrado_en se guarda en UTC; calculamos el rango UTC que corresponde al
    día local actual para no errar el indicador en horas límite."""
    ahora_local = datetime.now(DASHBOARD_TIMEZONE)
    inicio_local = ahora_local.replace(hour=0, minute=0, second=0, microsecond=0)
    inicio_utc = inicio_local.astimezone(timezone.utc).replace(tzinfo=None)
    fin_utc = (inicio_local + timedelta(days=1)).astimezone(timezone.utc).replace(tzinfo=None)
    filas = db.session.query(LecturaTemperatura.camara_id).filter(
        LecturaTemperatura.registrado_en >= inicio_utc,
        LecturaTemperatura.registrado_en < fin_utc,
    ).distinct().all()
    return {row[0] for row in filas}


@app.route('/registros/temperaturas')
@login_required
@requiere_permiso_recurso('registros', 'crear')
def temperaturas_index():
    camaras = Camara.query.filter_by(activa=True).order_by(Camara.nombre).all()

    # registrado_en se guarda en UTC naive; lo pasamos a hora local de negocio
    # para calcular fecha, ronda (AM/PM) y hora mostrada. Sin esta conversión,
    # una lectura tomada de mañana (p.ej. 8:00 local = 12:00 UTC) caería en el
    # bucket PM y no aparecería en la pestaña AM por defecto.
    def _a_local(dt):
        return dt.replace(tzinfo=timezone.utc).astimezone(DASHBOARD_TIMEZONE)

    ahora_local_dt = datetime.now(DASHBOARD_TIMEZONE)
    hoy = ahora_local_dt.date()

    # Ventana UTC que corresponde al día local de hoy (igual que _camaras_con_lectura_hoy).
    inicio_local = ahora_local_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    inicio_utc = inicio_local.astimezone(timezone.utc).replace(tzinfo=None)
    fin_utc = (inicio_local + timedelta(days=1)).astimezone(timezone.utc).replace(tzinfo=None)

    # Lecturas de hoy agrupadas por cámara y ronda (AM < 12:00, PM >= 12:00, hora local).
    lecturas_hoy = (LecturaTemperatura.query
                    .options(joinedload(LecturaTemperatura.registrado_por_vendedor))
                    .filter(LecturaTemperatura.registrado_en >= inicio_utc,
                            LecturaTemperatura.registrado_en < fin_utc)
                    .order_by(LecturaTemperatura.registrado_en.asc())
                    .all())
    lecturas_info = {cam.id: {'am': None, 'pm': None} for cam in camaras}
    for lec in lecturas_hoy:
        if lec.camara_id not in lecturas_info:
            continue
        local = _a_local(lec.registrado_en)
        ronda = 'am' if local.hour < 12 else 'pm'
        lecturas_info[lec.camara_id][ronda] = {
            'valor': float(lec.temperatura),
            'hora': local.strftime('%H:%M'),
            'fuera': bool(lec.fuera_de_rango),
            'por': (lec.registrado_por_vendedor.nombre_completo
                    if lec.registrado_por_vendedor else None),
        }
    con_lectura_hoy = {cid for cid, r in lecturas_info.items() if r['am'] or r['pm']}

    # Cumplimiento de los últimos 7 días (% de lecturas dentro de rango, por día local).
    desde = hoy - timedelta(days=6)
    desde_utc = (inicio_local - timedelta(days=6)).astimezone(timezone.utc).replace(tzinfo=None)
    lecturas_semana = (LecturaTemperatura.query
                       .filter(LecturaTemperatura.registrado_en >= desde_utc)
                       .all())
    por_dia = {}
    for lec in lecturas_semana:
        agg = por_dia.setdefault(_a_local(lec.registrado_en).date(), [0, 0])
        agg[1] += 1
        if not lec.fuera_de_rango:
            agg[0] += 1
    etiquetas = ['L', 'M', 'M', 'J', 'V', 'S', 'D']
    cumplimiento = []
    for i in range(7):
        d = desde + timedelta(days=i)
        ok, tot = por_dia.get(d, [0, 0])
        cumplimiento.append({'label': etiquetas[d.weekday()],
                             'pct': round(ok / tot * 100) if tot else None,
                             'tot': tot})
    dias_con = [c['pct'] for c in cumplimiento if c['pct'] is not None]
    cumplimiento_prom = round(sum(dias_con) / len(dias_con)) if dias_con else None

    es_admin = (not isinstance(current_user, Vendedor)) or current_user.tiene_permiso('registros', 'editar')
    ahora_local = datetime.now(DASHBOARD_TIMEZONE).strftime('%Y-%m-%dT%H:%M')
    return render_template('registros/temperaturas.html',
                           camaras=camaras,
                           lecturas_info=lecturas_info,
                           con_lectura_hoy=con_lectura_hoy,
                           cumplimiento=cumplimiento,
                           cumplimiento_prom=cumplimiento_prom,
                           hoy=hoy, ahora_local=ahora_local,
                           es_admin=es_admin)


@app.route('/registros/temperaturas/registrar', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'crear')
def temperatura_registrar():
    camara = Camara.query.filter_by(id=request.form.get('camara_id', type=int), activa=True).first()
    if camara is None:
        flash('Cámara no válida.', 'danger')
        return redirect(url_for('temperaturas_index'))
    try:
        temperatura = Decimal(str(request.form.get('temperatura')).replace(',', '.'))
    except (InvalidOperation, TypeError, ValueError):
        flash('La temperatura debe ser un número.', 'danger')
        return redirect(url_for('temperaturas_index'))

    fuera = camara.fuera_de_rango(temperatura)
    causa = (request.form.get('accion_causa') or '').strip()
    tomada = (request.form.get('accion_tomada') or '').strip()
    responsable = (request.form.get('accion_responsable') or '').strip()
    disposicion = (request.form.get('accion_disposicion') or '').strip()

    if fuera and (not tomada or not disposicion):
        flash(f'La lectura {temperatura}°C está fuera del rango de {camara.nombre} '
              f'({camara.temp_min}°C a {camara.temp_max}°C). Indica al menos la acción tomada '
              f'y la disposición del producto.', 'danger')
        return redirect(url_for('temperaturas_index'))

    momento = _registrado_en_from_form(request.form)
    lectura = LecturaTemperatura(
        camara_id=camara.id,
        temperatura=temperatura,
        registrado_por=current_user.id if isinstance(current_user, Vendedor) else None,
        fuera_de_rango=fuera,
        accion_causa=(causa or None) if fuera else None,
        accion_tomada=(tomada or None) if fuera else None,
        accion_responsable=(responsable or None) if fuera else None,
        accion_disposicion=(disposicion or None) if fuera else None,
    )
    if momento:
        lectura.registrado_en = momento
    db.session.add(lectura)
    db.session.commit()
    if fuera:
        _haccp_alerta('temp', f'{camara.nombre} fuera de rango',
                      f'{temperatura}°C (rango {camara.temp_min}° a {camara.temp_max}°)', tomada)
    else:
        _audit('temp', 'Registró temperatura', f'{camara.nombre} · {temperatura}°C')
    flash('Lectura registrada.' + (' (Fuera de rango — registrada con acción correctiva.)' if fuera else ''),
          'success' if not fuera else 'warning')
    return redirect(url_for('temperaturas_index'))


def _revision_que_cubre(fi, ff):
    """RevisionRegistro más reciente que cubre el período [fi, ff] ('YYYY-MM-DD').
    Si no hay fechas, devuelve la más reciente. None si no existe ninguna."""
    def _d(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date() if s else None
        except ValueError:
            return None
    d_fi, d_ff = _d(fi), _d(ff)
    q = RevisionRegistro.query
    if d_fi and d_ff:
        q = q.filter(RevisionRegistro.periodo_desde.isnot(None),
                     RevisionRegistro.periodo_hasta.isnot(None),
                     RevisionRegistro.periodo_desde <= d_fi,
                     RevisionRegistro.periodo_hasta >= d_ff)
    return q.order_by(RevisionRegistro.revisado_en.desc()).first()


def _filtrar_lecturas(args):
    """Aplica filtros opcionales (fecha_inicio, fecha_fin, camara_id) y
    devuelve las lecturas ordenadas por fecha desc. Acepta request.args o request.form."""
    q = LecturaTemperatura.query.options(
        joinedload(LecturaTemperatura.camara),
        joinedload(LecturaTemperatura.registrado_por_vendedor),
    )
    fi = args.get('fecha_inicio')
    ff = args.get('fecha_fin')
    cam = args.get('camara_id', type=int)
    if fi:
        try:
            q = q.filter(func.date(LecturaTemperatura.registrado_en) >= datetime.strptime(fi, '%Y-%m-%d').date())
        except ValueError:
            pass
    if ff:
        try:
            q = q.filter(func.date(LecturaTemperatura.registrado_en) <= datetime.strptime(ff, '%Y-%m-%d').date())
        except ValueError:
            pass
    if cam:
        q = q.filter(LecturaTemperatura.camara_id == cam)
    return q.order_by(LecturaTemperatura.registrado_en.desc()).all()


@app.route('/registros/temperaturas/historial')
@login_required
@requiere_permiso_recurso('registros', 'leer')
def temperaturas_historial():
    lecturas = _filtrar_lecturas(request.args)
    camaras = Camara.query.order_by(Camara.nombre).all()
    puede_verificar = (not isinstance(current_user, Vendedor)) or current_user.tiene_permiso('registros', 'editar')
    revision = _revision_que_cubre(request.args.get('fecha_inicio'), request.args.get('fecha_fin'))
    hoy = date.today()

    # Gráfico de tendencia: cámara con más excursiones (o la filtrada).
    candidatas = {}
    for l in lecturas:
        candidatas.setdefault(l.camara_id, []).append(l)
    cam_id = request.args.get('camara_id', type=int)
    serie_cam = cam_id if (cam_id and cam_id in candidatas) else (
        max(candidatas, key=lambda k: sum(1 for x in candidatas[k] if x.fuera_de_rango)) if candidatas else None)
    tendencia = None
    if serie_cam:
        cam_obj = next((l.camara for l in lecturas if l.camara_id == serie_cam), None)
        pts = sorted(candidatas[serie_cam], key=lambda x: x.registrado_en)
        vals = [float(x.temperatura) for x in pts]
        cmin = float(cam_obj.temp_min) if cam_obj else min(vals)
        cmax = float(cam_obj.temp_max) if cam_obj else max(vals)
        lo = min(cmin, min(vals)) - 1
        hi = max(cmax, max(vals)) + 1
        if hi <= lo:
            hi = lo + 1
        W, H = 300.0, 110.0
        n = len(pts)
        def _x(i):
            return round((i / (n - 1) * W) if n > 1 else W / 2, 1)
        def _y(v):
            return round(H - (v - lo) / (hi - lo) * H, 1)
        tendencia = {
            'nombre': cam_obj.nombre if cam_obj else '',
            'min': cmin, 'max': cmax, 'excursiones': sum(1 for x in pts if x.fuera_de_rango),
            'W': W, 'H': H,
            'band_top': _y(cmax), 'band_bot': _y(cmin),
            'poly': ' '.join(f'{_x(i)},{_y(v)}' for i, v in enumerate(vals)),
            'dots': [{'x': _x(i), 'y': _y(v), 'f': bool(pts[i].fuera_de_rango)} for i, v in enumerate(vals)],
        }

    return render_template('registros/temperaturas_historial.html',
                           lecturas=lecturas, camaras=camaras, filtros=request.args,
                           puede_verificar=puede_verificar, revision=revision, tendencia=tendencia,
                           hoy_iso=hoy.isoformat(),
                           fecha_7d=(hoy - timedelta(days=6)).isoformat(),
                           fecha_30d=(hoy - timedelta(days=29)).isoformat())


@app.route('/registros/temperaturas/revisar', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'editar')
def temperatura_revisar():
    fi = request.form.get('fecha_inicio') or None
    ff = request.form.get('fecha_fin') or None
    def _d(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date() if s else None
        except ValueError:
            return None
    db.session.add(RevisionRegistro(
        revisado_por=current_user.id if isinstance(current_user, Vendedor) else None,
        periodo_desde=_d(fi), periodo_hasta=_d(ff),
    ))
    db.session.commit()
    flash('Período marcado como revisado.', 'success')
    return redirect(url_for('temperaturas_historial', fecha_inicio=fi or '', fecha_fin=ff or ''))


def _pdf_xe(s):
    """Escapa caracteres XML para insertar texto dinámico en un Paragraph."""
    return str(s if s is not None else '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def _registro_pdf_tabla(headers, aligns, widths, estado_col, filas):
    """Tabla flowable estandarizada para los PDF de registros HACCP.

    - headers: list[str] — encabezados de columna
    - aligns:  list['L'|'R'|'C'] — alineación por columna
    - widths:  list[float] — anchos en puntos
    - estado_col: índice de la columna de estado (Sí/NO · Conforme/No conforme)
    - filas: list[dict] con:
        'cols'    : list[str] — valor por columna (texto plano; el de estado_col
                    es el texto del estado)
        'desvio'  : bool — fuera de rango / no conforme → rojo + sub-fila de alerta
        'detalle' : str|None — markup opcional; si existe, se agrega una sub-fila
                    de ancho completo (acción correctiva / observación)

    Zebra, grilla mínima (solo líneas horizontales), estado en color y filas
    normales de altura pareja (el texto largo va a la sub-fila).
    """
    NAVY = colors.HexColor('#1f2937')
    ZEBRA = colors.HexColor('#f8fafc')
    DESVIO = colors.HexColor('#fee2e2')
    LINE = colors.HexColor('#e2e8f0')
    DARK = colors.HexColor('#0f172a')
    amap = {'L': TA_LEFT, 'R': TA_RIGHT, 'C': TA_CENTER}

    base = {a: ParagraphStyle('c_' + a, fontSize=8, leading=11, alignment=amap[a],
                              textColor=DARK) for a in amap}
    head = {a: ParagraphStyle('h_' + a, fontSize=8, leading=10, fontName='Helvetica-Bold',
                              textColor=colors.white, alignment=amap[a]) for a in amap}
    ok_ps = ParagraphStyle('estado_ok', fontSize=8, leading=11, alignment=TA_CENTER,
                           fontName='Helvetica-Bold', textColor=colors.HexColor('#15803d'))
    no_ps = ParagraphStyle('estado_no', fontSize=8, leading=11, alignment=TA_CENTER,
                           fontName='Helvetica-Bold', textColor=colors.HexColor('#b91c1c'))
    sub_ps = ParagraphStyle('sub_detalle', fontSize=7.5, leading=10, alignment=TA_LEFT,
                            textColor=colors.HexColor('#475569'))

    ncols = len(headers)
    data = [[Paragraph(_pdf_xe(h), head[aligns[i]]) for i, h in enumerate(headers)]]
    cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), NAVY),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 7),
        ('RIGHTPADDING', (0, 0), (-1, -1), 7),
        ('LINEBELOW', (0, 0), (-1, 0), 1, NAVY),
    ]
    r = 1
    for idx, fila in enumerate(filas):
        row = []
        for i, val in enumerate(fila['cols']):
            if i == estado_col:
                row.append(Paragraph(_pdf_xe(val), no_ps if fila['desvio'] else ok_ps))
            else:
                row.append(Paragraph(_pdf_xe(val), base[aligns[i]]))
        data.append(row)
        bg = DESVIO if fila['desvio'] else (ZEBRA if idx % 2 else colors.white)
        cmds += [
            ('BACKGROUND', (0, r), (-1, r), bg),
            ('VALIGN', (0, r), (-1, r), 'MIDDLE'),
            ('TOPPADDING', (0, r), (-1, r), 5),
            ('BOTTOMPADDING', (0, r), (-1, r), 5),
        ]
        r += 1
        detalle = fila.get('detalle')
        if detalle:
            data.append([Paragraph(detalle, sub_ps)] + [''] * (ncols - 1))
            cmds += [
                ('SPAN', (0, r), (-1, r)),
                ('BACKGROUND', (0, r), (-1, r), DESVIO if fila['desvio'] else ZEBRA),
                ('LEFTPADDING', (0, r), (0, r), 16),
                ('TOPPADDING', (0, r), (-1, r), 1),
                ('BOTTOMPADDING', (0, r), (-1, r), 6),
            ]
            r += 1
        cmds.append(('LINEBELOW', (0, r - 1), (-1, r - 1), 0.5, LINE))

    tabla = Table(data, repeatRows=1, colWidths=widths)
    tabla.setStyle(TableStyle(cmds))
    return tabla


def _build_temperaturas_pdf(lecturas, fecha_inicio, fecha_fin, config, revision):
    """Construye el PDF tabular audit-ready del registro de temperaturas."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=0.4 * inch, rightMargin=0.4 * inch,
                            topMargin=0.5 * inch, bottomMargin=0.7 * inch)
    empresa_style = ParagraphStyle(name='reg_empresa', fontSize=10, leading=13,
                                   fontName='Helvetica-Bold', alignment=TA_LEFT,
                                   textColor=colors.HexColor('#1877ff'))
    titulo_style = ParagraphStyle(name='reg_titulo', fontSize=15, leading=18,
                                  fontName='Helvetica-Bold', alignment=TA_LEFT)
    sub_style = ParagraphStyle(name='reg_sub', fontSize=9, leading=12,
                               alignment=TA_LEFT, textColor=colors.HexColor('#475569'))

    if fecha_inicio or fecha_fin:
        periodo = f'Período: {fecha_inicio or "inicio"} a {fecha_fin or "hoy"}'
    else:
        periodo = 'Período: todas las fechas'
    generado = datetime.now(DASHBOARD_TIMEZONE).strftime('%Y-%m-%d %H:%M')

    encabezado = [
        Paragraph('Jomar Foods B.V.', empresa_style),
        Paragraph('Registro de temperaturas de cámaras', titulo_style),
        Paragraph(f'Documento: {config.codigo_documento} &middot; Versión: {config.version}', sub_style),
        Paragraph(f'{periodo} &middot; Generado: {generado}', sub_style),
    ]
    logo_path = os.path.join(basedir, 'static', 'logo_etiquetas.png')
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=52, height=52)
        head_tbl = Table([[logo, encabezado]], colWidths=[64, None])
        head_tbl.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        head_tbl.hAlign = 'LEFT'
        elements = [head_tbl, Spacer(1, 14)]
    else:
        elements = encabezado + [Spacer(1, 14)]

    def _detalle(l):
        partes = []
        if l.accion_causa: partes.append(f'Causa: {_pdf_xe(l.accion_causa)}')
        if l.accion_tomada: partes.append(f'Acción: {_pdf_xe(l.accion_tomada)}')
        if l.accion_responsable: partes.append(f'Resp.: {_pdf_xe(l.accion_responsable)}')
        if l.accion_disposicion: partes.append(f'Disposición: {_pdf_xe(l.accion_disposicion)}')
        cuerpo = ' · '.join(partes) or (_pdf_xe(l.accion_correctiva) if l.accion_correctiva else '')
        return f'<b>Acción correctiva</b> — {cuerpo}' if cuerpo else None

    headers = ['Fecha/Hora', 'Cámara', 'Tipo', 'Límite crítico (°C)', 'Lectura (°C)',
               'En rango', 'Responsable']
    aligns = ['L', 'L', 'L', 'R', 'R', 'C', 'L']
    widths = [88, 118, 92, 96, 78, 66, 242]
    filas = []
    for l in lecturas:
        filas.append({
            'cols': [
                _fmt_local(l.registrado_en),
                l.camara.nombre if l.camara else '—',
                'Refrigeración' if (l.camara and l.camara.tipo == 'refrigeracion') else 'Congelación',
                f'{l.camara.temp_min} a {l.camara.temp_max}' if l.camara else '—',
                str(l.temperatura),
                'NO' if l.fuera_de_rango else 'Sí',
                l.registrado_por_vendedor.nombre_completo if l.registrado_por_vendedor else '—',
            ],
            'desvio': bool(l.fuera_de_rango),
            'detalle': _detalle(l) if l.fuera_de_rango else None,
        })
    elements.append(_registro_pdf_tabla(headers, aligns, widths, 5, filas))

    elements.append(Spacer(1, 18))
    if revision:
        nombre = revision.revisado_por_vendedor.nombre_completo if revision.revisado_por_vendedor else '—'
        rev_txt = f'<b>Verificación:</b> Revisado por {nombre} el {_fmt_local(revision.revisado_en)}'
    else:
        rev_txt = '<b>Verificación:</b> Revisado por: ______________________      Fecha: __________'
    elements.append(Paragraph(rev_txt, sub_style))

    cal = config.termometro_calibrado_en.strftime('%Y-%m-%d') if config.termometro_calibrado_en else 'N/D'
    footer_left = (f'Frecuencia: {config.frecuencia_texto or "N/D"}   |   '
                   f'Instrumento: {config.termometro or "N/D"} (cal.: {cal})')
    footer_doc = f'{config.codigo_documento} v{config.version}'
    page_w = landscape(A4)[0]

    class _NumberedCanvas(canvas.Canvas):
        def __init__(self, *a, **k):
            canvas.Canvas.__init__(self, *a, **k)
            self._saved_states = []
        def showPage(self):
            self._saved_states.append(dict(self.__dict__))
            self._startPage()
        def save(self):
            total = len(self._saved_states)
            for st in self._saved_states:
                self.__dict__.update(st)
                self.setFont('Helvetica', 7)
                self.setFillColor(colors.HexColor('#475569'))
                self.drawString(0.4 * inch, 0.35 * inch, footer_left)
                self.drawRightString(page_w - 0.4 * inch, 0.35 * inch,
                                     f'{footer_doc}  ·  Página {self._pageNumber} de {total}')
                canvas.Canvas.showPage(self)
            canvas.Canvas.save(self)

    doc.build(elements, canvasmaker=_NumberedCanvas)
    buffer.seek(0)
    return buffer


def _build_xlsx(headers, rows, sheet_name, title):
    """Construye un .xlsx en memoria con encabezado en negrita y filas de datos.
    `rows` es una lista de listas de celdas (texto/numero)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    head_fill = PatternFill('solid', fgColor='1F2937')
    head_font = Font(bold=True, color='FFFFFF')
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(2, c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal='left')
    for row in rows:
        ws.append([_excel_safe(v) for v in row])
    for i, h in enumerate(headers, start=1):
        width = max(len(str(h)), *(len(str(r[i - 1])) for r in rows)) if rows else len(str(h))
        ws.column_dimensions[ws.cell(2, i).column_letter].width = min(max(width + 2, 10), 48)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route('/registros/temperaturas/export', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'leer')
def temperaturas_export():
    lecturas = _filtrar_lecturas(request.form)
    fi = request.form.get('fecha_inicio') or ''
    ff = request.form.get('fecha_fin') or ''
    if (request.form.get('formato') or 'pdf').lower() == 'excel':
        headers = ['Fecha', 'Hora', 'Cámara', 'Lectura (°C)', 'Estado', 'Registró',
                   'Causa', 'Acción tomada', 'Responsable acción', 'Disposición']
        rows = [[
            _fmt_local(l.registrado_en, '%Y-%m-%d'), _fmt_local(l.registrado_en, '%H:%M'),
            l.camara.nombre if l.camara else '', float(l.temperatura),
            'Fuera de rango' if l.fuera_de_rango else 'En rango',
            l.registrado_por_vendedor.nombre_completo if l.registrado_por_vendedor else '',
            l.accion_causa or '', l.accion_tomada or '', l.accion_responsable or '', l.accion_disposicion or '',
        ] for l in lecturas]
        buffer = _build_xlsx(headers, rows, 'Temperaturas', 'Historial de temperaturas')
        filename = f"registro_temperaturas_{fi or 'inicio'}_{ff or 'fin'}.xlsx"
        return make_response(send_file(buffer, as_attachment=True, download_name=filename,
                                       mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
    config = _get_registro_config()
    revision = _revision_que_cubre(fi, ff)
    buffer = _build_temperaturas_pdf(lecturas, fi, ff, config, revision)
    filename = f"registro_temperaturas_{fi or 'inicio'}_{ff or 'fin'}.pdf"
    response = make_response(send_file(buffer, mimetype='application/pdf',
                                       as_attachment=not _is_ios_request(),
                                       download_name=filename))
    response.headers['Content-Type'] = 'application/pdf'
    return response


@app.route('/registros/temperaturas/config', methods=['GET', 'POST'])
@login_required
@requiere_permiso_recurso('registros', 'editar')
def registro_config():
    cfg = _get_registro_config()
    if request.method == 'POST':
        cfg.codigo_documento = (request.form.get('codigo_documento') or '').strip() or 'FR-HACCP-TEMP-01'
        cfg.version = (request.form.get('version') or '').strip() or '1'
        cfg.frecuencia_texto = (request.form.get('frecuencia_texto') or '').strip() or '2 veces al día'
        cfg.termometro = (request.form.get('termometro') or '').strip() or None
        cal = (request.form.get('termometro_calibrado_en') or '').strip()
        try:
            cfg.termometro_calibrado_en = datetime.strptime(cal, '%Y-%m-%d').date() if cal else None
        except ValueError:
            cfg.termometro_calibrado_en = None
        cfg.actualizado_en = datetime.utcnow()
        db.session.commit()
        flash('Configuración guardada.', 'success')
        return redirect(url_for('registro_config'))
    return render_template('registros/config.html', cfg=cfg)


# ───────────────────────── HACCP: Limpieza ─────────────────────────
# Reservado para la validación de AreaLimpieza.tipo en _parse_area_limpieza.
_TIPOS_AREA_LIMPIEZA = ('equipo', 'espacio')


def _get_limpieza_config():
    """Devuelve la fila única de LimpiezaConfig; la crea con valores por defecto."""
    cfg = LimpiezaConfig.query.first()
    if cfg is None:
        cfg = LimpiezaConfig()
        db.session.add(cfg)
        db.session.commit()
    return cfg


def _parse_producto_limpieza(form):
    """Devuelve (nombre, dilucion, procedimiento, notas, error)."""
    nombre = (form.get('nombre') or '').strip()
    dilucion = (form.get('dilucion') or '').strip()
    if not nombre:
        return None, None, None, None, 'El nombre del producto es obligatorio.'
    if not dilucion:
        return None, None, None, None, 'La dilución es obligatoria.'
    procedimiento = (form.get('procedimiento') or '').strip() or None
    notas = (form.get('notas_seguridad') or '').strip() or None
    return nombre, dilucion, procedimiento, notas, None


@app.route('/registros/limpieza/productos')
@login_required
@requiere_permiso_recurso('registros', 'leer')
def productos_limpieza_index():
    productos = ProductoLimpieza.query.order_by(ProductoLimpieza.nombre).all()
    es_admin = (not isinstance(current_user, Vendedor)) or current_user.tiene_permiso('registros', 'editar')
    return render_template('registros/productos_limpieza.html', productos=productos, es_admin=es_admin)


@app.route('/registros/limpieza/productos/nuevo', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'editar')
def producto_limpieza_nuevo():
    nombre, dilucion, procedimiento, notas, error = _parse_producto_limpieza(request.form)
    if error:
        flash(error, 'danger')
        return redirect(url_for('productos_limpieza_index'))
    try:
        db.session.add(ProductoLimpieza(nombre=nombre, dilucion=dilucion,
                                        procedimiento=procedimiento, notas_seguridad=notas, activo=True))
        db.session.commit()
        flash('Producto creado.', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al crear producto de limpieza: {e}')
        flash('No se pudo crear el producto. Revisá los datos e intentá de nuevo.', 'danger')
    return redirect(url_for('productos_limpieza_index'))


@app.route('/registros/limpieza/productos/<int:producto_id>/editar', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'editar')
def producto_limpieza_editar(producto_id):
    producto = ProductoLimpieza.query.get_or_404(producto_id)
    nombre, dilucion, procedimiento, notas, error = _parse_producto_limpieza(request.form)
    if error:
        flash(error, 'danger')
        return redirect(url_for('productos_limpieza_index'))
    producto.nombre, producto.dilucion = nombre, dilucion
    producto.procedimiento, producto.notas_seguridad = procedimiento, notas
    try:
        db.session.commit()
        flash('Producto actualizado.', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error al editar producto de limpieza {producto_id}: {e}')
        flash('No se pudo actualizar el producto. Revisá los datos e intentá de nuevo.', 'danger')
    return redirect(url_for('productos_limpieza_index'))


@app.route('/registros/limpieza/productos/<int:producto_id>/toggle', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'editar')
def producto_limpieza_toggle(producto_id):
    producto = ProductoLimpieza.query.get_or_404(producto_id)
    producto.activo = not producto.activo
    db.session.commit()
    flash('Producto actualizado.', 'success')
    return redirect(url_for('productos_limpieza_index'))


def _parse_area_limpieza(form):
    """Devuelve (nombre, tipo, producto_id, sanitizante_id, metodo, frecuencia, error)."""
    nombre = (form.get('nombre') or '').strip()
    tipo = (form.get('tipo') or '').strip()
    if not nombre:
        return None, None, None, None, None, None, 'El nombre es obligatorio.'
    if tipo not in _TIPOS_AREA_LIMPIEZA:
        return None, None, None, None, None, None, 'Tipo de área no válido.'
    producto_id = form.get('producto_id', type=int) or None
    if producto_id and db.session.get(ProductoLimpieza, producto_id) is None:
        return None, None, None, None, None, None, 'El producto de limpieza seleccionado no existe.'
    sanitizante_id = form.get('sanitizante_id', type=int) or None
    if sanitizante_id and db.session.get(ProductoLimpieza, sanitizante_id) is None:
        return None, None, None, None, None, None, 'El sanitizante seleccionado no existe.'
    metodo = (form.get('metodo') or '').strip() or None
    frecuencia = (form.get('frecuencia_texto') or '').strip() or None
    return nombre, tipo, producto_id, sanitizante_id, metodo, frecuencia, None


@app.route('/registros/limpieza/areas')
@login_required
@requiere_permiso_recurso('registros', 'editar')
def areas_limpieza_list():
    areas = (AreaLimpieza.query
             .options(joinedload(AreaLimpieza.producto), joinedload(AreaLimpieza.sanitizante))
             .order_by(AreaLimpieza.nombre).all())
    productos = ProductoLimpieza.query.filter_by(activo=True).order_by(ProductoLimpieza.nombre).all()
    responsables = Vendedor.query.filter_by(activo=True).order_by(Vendedor.nombre_completo).all()
    puede_eliminar = (not isinstance(current_user, Vendedor)) or current_user.tiene_permiso('registros', 'eliminar')
    return render_template('registros/areas_limpieza.html', areas=areas, productos=productos, responsables=responsables, puede_eliminar=puede_eliminar)


@app.route('/registros/limpieza/areas/nueva', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'editar')
def area_limpieza_nueva():
    nombre, tipo, producto_id, sanitizante_id, metodo, frecuencia, error = _parse_area_limpieza(request.form)
    if error:
        flash(error, 'danger')
        return redirect(url_for('areas_limpieza_list'))
    db.session.add(AreaLimpieza(nombre=nombre, tipo=tipo, producto_id=producto_id,
                                sanitizante_id=sanitizante_id,
                                metodo=metodo, frecuencia_texto=frecuencia, activa=True,
                                responsable_id=request.form.get('responsable_id', type=int)))
    db.session.commit()
    _audit('config', 'Creó tarea de limpieza', nombre)
    flash('Área creada.', 'success')
    return redirect(url_for('areas_limpieza_list'))


@app.route('/registros/limpieza/areas/<int:area_id>/editar', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'editar')
def area_limpieza_editar(area_id):
    area = AreaLimpieza.query.get_or_404(area_id)
    nombre, tipo, producto_id, sanitizante_id, metodo, frecuencia, error = _parse_area_limpieza(request.form)
    if error:
        flash(error, 'danger')
        return redirect(url_for('areas_limpieza_list'))
    area.nombre, area.tipo, area.producto_id = nombre, tipo, producto_id
    area.sanitizante_id = sanitizante_id
    area.metodo, area.frecuencia_texto = metodo, frecuencia
    area.responsable_id = request.form.get('responsable_id', type=int)
    db.session.commit()
    _audit('config', 'Editó tarea de limpieza', nombre)
    flash('Área actualizada.', 'success')
    return redirect(url_for('areas_limpieza_list'))


@app.route('/registros/limpieza/areas/<int:area_id>/toggle', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'editar')
def area_limpieza_toggle(area_id):
    area = AreaLimpieza.query.get_or_404(area_id)
    area.activa = not area.activa
    db.session.commit()
    flash('Área actualizada.', 'success')
    return redirect(url_for('areas_limpieza_list'))


@app.route('/registros/limpieza/areas/<int:area_id>/eliminar', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'eliminar')
def area_limpieza_eliminar(area_id):
    area = AreaLimpieza.query.get_or_404(area_id)
    n = RegistroLimpieza.query.filter_by(area_id=area_id).count()
    if n > 0:
        flash(f'La tarea «{area.nombre}» tiene {n} registro(s) en el historial; no se puede '
              f'borrar. Desactívala en su lugar.', 'danger')
        return redirect(url_for('areas_limpieza_list'))
    nombre = area.nombre
    db.session.delete(area)
    db.session.commit()
    _audit('config', 'Eliminó tarea de limpieza', nombre)
    flash('Tarea eliminada.', 'success')
    return redirect(url_for('areas_limpieza_list'))


def _areas_con_registro_hoy():
    """Set de area_id con al menos un registro de limpieza HOY (día local de negocio)."""
    ahora_local = datetime.now(DASHBOARD_TIMEZONE)
    inicio_local = ahora_local.replace(hour=0, minute=0, second=0, microsecond=0)
    inicio_utc = inicio_local.astimezone(timezone.utc).replace(tzinfo=None)
    fin_utc = (inicio_local + timedelta(days=1)).astimezone(timezone.utc).replace(tzinfo=None)
    filas = db.session.query(RegistroLimpieza.area_id).filter(
        RegistroLimpieza.registrado_en >= inicio_utc,
        RegistroLimpieza.registrado_en < fin_utc,
    ).distinct().all()
    return {row[0] for row in filas}


@app.route('/registros/limpieza')
@login_required
@requiere_permiso_recurso('registros', 'crear')
def limpieza_index():
    areas = (AreaLimpieza.query.options(joinedload(AreaLimpieza.producto), joinedload(AreaLimpieza.sanitizante))
             .filter_by(activa=True).order_by(AreaLimpieza.nombre).all())
    hoy = date.today()

    # Registros de hoy por área (el más reciente del día gana).
    regs_hoy = (RegistroLimpieza.query
                .options(joinedload(RegistroLimpieza.registrado_por_vendedor))
                .filter(func.date(RegistroLimpieza.registrado_en) == hoy)
                .order_by(RegistroLimpieza.registrado_en.asc())
                .all())
    registros_info = {a.id: None for a in areas}
    for r in regs_hoy:
        if r.area_id in registros_info:
            registros_info[r.area_id] = {
                'conforme': bool(r.conforme),
                'hora': r.registrado_en.strftime('%H:%M'),
                'por': (r.registrado_por_vendedor.nombre_completo
                        if r.registrado_por_vendedor else None),
            }
    con_registro_hoy = {aid for aid, v in registros_info.items() if v}

    # Cumplimiento de los últimos 7 días (% de registros conformes).
    desde = hoy - timedelta(days=6)
    regs_semana = (RegistroLimpieza.query
                   .filter(func.date(RegistroLimpieza.registrado_en) >= desde)
                   .all())
    por_dia = {}
    for r in regs_semana:
        agg = por_dia.setdefault(r.registrado_en.date(), [0, 0])
        agg[1] += 1
        if r.conforme:
            agg[0] += 1
    etiquetas = ['L', 'M', 'M', 'J', 'V', 'S', 'D']
    cumplimiento = []
    for i in range(7):
        d = desde + timedelta(days=i)
        ok, tot = por_dia.get(d, [0, 0])
        cumplimiento.append({'label': etiquetas[d.weekday()],
                             'pct': round(ok / tot * 100) if tot else None,
                             'tot': tot})
    dias_con = [c['pct'] for c in cumplimiento if c['pct'] is not None]
    cumplimiento_prom = round(sum(dias_con) / len(dias_con)) if dias_con else None

    es_admin = (not isinstance(current_user, Vendedor)) or current_user.tiene_permiso('registros', 'editar')
    ahora_local = datetime.now(DASHBOARD_TIMEZONE).strftime('%Y-%m-%dT%H:%M')
    vendedores = (Vendedor.query.filter_by(activo=True)
                  .order_by(Vendedor.nombre_completo).all())
    operador_id = current_user.id if isinstance(current_user, Vendedor) else None
    return render_template('registros/limpieza.html', areas=areas,
                           registros_info=registros_info,
                           con_registro_hoy=con_registro_hoy,
                           cumplimiento=cumplimiento,
                           cumplimiento_prom=cumplimiento_prom,
                           hoy=hoy, ahora_local=ahora_local, es_admin=es_admin,
                           vendedores=vendedores, operador_id=operador_id)


@app.route('/registros/limpieza/registrar', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'crear')
def limpieza_registrar():
    area = AreaLimpieza.query.filter_by(id=request.form.get('area_id', type=int), activa=True).first()
    if area is None:
        flash('Área no válida.', 'danger')
        return redirect(url_for('limpieza_index'))
    conforme = (request.form.get('conforme') or 'si') != 'no'
    observacion = (request.form.get('observacion') or '').strip() or None
    causa = (request.form.get('accion_causa') or '').strip()
    tomada = (request.form.get('accion_tomada') or '').strip()
    responsable = (request.form.get('accion_responsable') or '').strip()
    disposicion = (request.form.get('accion_disposicion') or '').strip()

    if not conforme and (not tomada or not disposicion):
        flash(f'El registro de {area.nombre} es No conforme. Indica al menos la acción tomada '
              f'y la disposición.', 'danger')
        return redirect(url_for('limpieza_index'))

    # Concentración (ppm): obligatoria solo si el área tiene sanitizante.
    ppm_raw = (request.form.get('concentracion_ppm') or '').strip()
    ppm = None
    if ppm_raw:
        try:
            ppm = int(ppm_raw)
        except ValueError:
            flash('La concentración (ppm) debe ser un número entero.', 'danger')
            return redirect(url_for('limpieza_index'))
    if ppm is not None and ppm <= 0:
        flash('La concentración (ppm) debe ser un valor positivo.', 'danger')
        return redirect(url_for('limpieza_index'))
    requiere_ppm = area.sanitizante_id is not None
    if requiere_ppm and ppm is None:
        flash(f'Indica la concentración (ppm) de {area.sanitizante.nombre} para {area.nombre}.', 'danger')
        return redirect(url_for('limpieza_index'))
    if requiere_ppm and conforme and (ppm < 150 or ppm > 400):
        flash(f'ppm fuera de rango (150–400) en {area.nombre}: corrige y vuelve a medir, '
              f'o marca No conforme.', 'danger')
        return redirect(url_for('limpieza_index'))

    # Verificación independiente: persona distinta del operador.
    operador_id = current_user.id if isinstance(current_user, Vendedor) else None
    verificado_por_id = request.form.get('verificado_por', type=int)
    verificador = (Vendedor.query.filter_by(id=verificado_por_id, activo=True).first()
                   if verificado_por_id else None)
    if verificador is None:
        flash('Selecciona quién verificó la limpieza (persona distinta del operador).', 'danger')
        return redirect(url_for('limpieza_index'))
    if operador_id is not None and verificador.id == operador_id:
        flash('La verificación debe hacerla una persona distinta del operador.', 'danger')
        return redirect(url_for('limpieza_index'))

    # Método de verificación (opcional): visual | atp | hisopado.
    metodo = (request.form.get('metodo_verificacion') or '').strip().lower()
    if metodo not in ('visual', 'atp', 'hisopado'):
        metodo = None

    firma = (request.form.get('firma_png') or '').strip() or None
    if firma and not _firma_png_valida(firma):
        firma = None  # descartar payloads inválidos, no-PNG o excesivos

    momento = _registrado_en_from_form(request.form)
    registro = RegistroLimpieza(
        area_id=area.id,
        registrado_por=current_user.id if isinstance(current_user, Vendedor) else None,
        conforme=conforme,
        observacion=observacion,
        firma_png=firma,
        accion_causa=(causa or None) if not conforme else None,
        accion_tomada=(tomada or None) if not conforme else None,
        accion_responsable=(responsable or None) if not conforme else None,
        accion_disposicion=(disposicion or None) if not conforme else None,
        concentracion_ppm=ppm,
        verificado_por=verificador.id,
        metodo_verificacion=metodo,
    )
    if momento:
        registro.registrado_en = momento
    db.session.add(registro)
    db.session.commit()
    if not conforme:
        detalle_ppm = f' · ppm={ppm}' if ppm is not None else ''
        _haccp_alerta('clean', f'{area.nombre}: limpieza no conforme',
                      (observacion or 'Registro marcado no conforme') + detalle_ppm, tomada)
    else:
        _audit('clean', 'Firmó tarea de limpieza', area.nombre)
    flash('Limpieza registrada.' + (' (No conforme — registrada con acción correctiva.)' if not conforme else ''),
          'success' if conforme else 'warning')
    return redirect(url_for('limpieza_index'))


@app.route('/registros/limpieza/registro/<int:registro_id>/eliminar', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'eliminar')
def limpieza_registro_eliminar(registro_id):
    registro = RegistroLimpieza.query.get_or_404(registro_id)
    nombre = registro.area.nombre if registro.area else '—'
    cuando = _fmt_local(registro.registrado_en)
    db.session.delete(registro)
    db.session.commit()
    _audit('clean', 'Eliminó registro de limpieza', f'{nombre} · {cuando}')
    flash('Registro eliminado.', 'success')
    destino = request.referrer
    if not destino or urlparse(destino).netloc != urlparse(request.host_url).netloc:
        destino = url_for('limpieza_historial')
    return redirect(destino)


def _revision_limpieza_que_cubre(fi, ff):
    """RevisionLimpieza más reciente que cubre el período [fi, ff] ('YYYY-MM-DD')."""
    def _d(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date() if s else None
        except ValueError:
            return None
    d_fi, d_ff = _d(fi), _d(ff)
    q = RevisionLimpieza.query
    if d_fi and d_ff:
        q = q.filter(RevisionLimpieza.periodo_desde.isnot(None),
                     RevisionLimpieza.periodo_hasta.isnot(None),
                     RevisionLimpieza.periodo_desde <= d_fi,
                     RevisionLimpieza.periodo_hasta >= d_ff)
    return q.order_by(RevisionLimpieza.revisado_en.desc()).first()


def _filtrar_registros_limpieza(args):
    """Aplica filtros opcionales (fecha_inicio, fecha_fin, area_id) y devuelve
    los registros ordenados por fecha desc. Acepta request.args o request.form."""
    q = RegistroLimpieza.query.options(
        joinedload(RegistroLimpieza.area).joinedload(AreaLimpieza.producto),
        joinedload(RegistroLimpieza.area).joinedload(AreaLimpieza.sanitizante),
        joinedload(RegistroLimpieza.registrado_por_vendedor),
    )
    fi = args.get('fecha_inicio')
    ff = args.get('fecha_fin')
    area = args.get('area_id', type=int)
    if fi:
        try:
            q = q.filter(func.date(RegistroLimpieza.registrado_en) >= datetime.strptime(fi, '%Y-%m-%d').date())
        except ValueError:
            pass
    if ff:
        try:
            q = q.filter(func.date(RegistroLimpieza.registrado_en) <= datetime.strptime(ff, '%Y-%m-%d').date())
        except ValueError:
            pass
    if area:
        q = q.filter(RegistroLimpieza.area_id == area)
    return q.order_by(RegistroLimpieza.registrado_en.desc()).all()


@app.route('/registros/limpieza/historial')
@login_required
@requiere_permiso_recurso('registros', 'leer')
def limpieza_historial():
    registros = _filtrar_registros_limpieza(request.args)
    areas = AreaLimpieza.query.order_by(AreaLimpieza.nombre).all()
    puede_verificar = (not isinstance(current_user, Vendedor)) or current_user.tiene_permiso('registros', 'editar')
    puede_eliminar = (not isinstance(current_user, Vendedor)) or current_user.tiene_permiso('registros', 'eliminar')
    revision = _revision_limpieza_que_cubre(request.args.get('fecha_inicio'), request.args.get('fecha_fin'))
    hoy = date.today()
    return render_template('registros/limpieza_historial.html',
                           registros=registros, areas=areas, filtros=request.args,
                           puede_verificar=puede_verificar, puede_eliminar=puede_eliminar, revision=revision,
                           hoy_iso=hoy.isoformat(),
                           fecha_7d=(hoy - timedelta(days=6)).isoformat(),
                           fecha_30d=(hoy - timedelta(days=29)).isoformat())


@app.route('/registros/mi-turno')
@login_required
@requiere_permiso_recurso('registros', 'crear')
def mi_turno():
    """Vista operativa: todo lo que falta registrar HOY (temperaturas + limpieza),
    en una sola lista priorizada. Pensada para el operario en planta."""
    con_lectura = _camaras_con_lectura_hoy()
    con_registro = _areas_con_registro_hoy()
    camaras = Camara.query.filter_by(activa=True).order_by(Camara.nombre).all()
    areas = (AreaLimpieza.query.options(joinedload(AreaLimpieza.producto), joinedload(AreaLimpieza.sanitizante))
             .filter_by(activa=True).order_by(AreaLimpieza.nombre).all())

    pend_temp = [c for c in camaras if c.id not in con_lectura]
    pend_limp = [a for a in areas if a.id not in con_registro]
    total_pend = len(pend_temp) + len(pend_limp)
    total = len(camaras) + len(areas)
    hechas = total - total_pend
    return render_template('registros/mi_turno.html',
                           pend_temp=pend_temp, pend_limp=pend_limp,
                           hechas=hechas, total=total, hoy=date.today())


@app.route('/registros/haccp')
@login_required
@requiere_permiso_recurso('registros', 'leer')
def registros_haccp():
    """Vista de cumplimiento HACCP (solo lectura) — temperaturas + limpieza."""
    hoy = date.today()
    desde = hoy - timedelta(days=6)

    lecturas = (LecturaTemperatura.query
                .options(joinedload(LecturaTemperatura.camara))
                .filter(func.date(LecturaTemperatura.registrado_en) >= desde).all())
    temp_total = len(lecturas)
    temp_ok = sum(1 for l in lecturas if not l.fuera_de_rango)
    temp_pct = round(temp_ok / temp_total * 100) if temp_total else 100

    registros = (RegistroLimpieza.query
                 .options(joinedload(RegistroLimpieza.area))
                 .filter(func.date(RegistroLimpieza.registrado_en) >= desde).all())
    limp_total = len(registros)
    limp_ok = sum(1 for r in registros if r.conforme)
    limp_pct = round(limp_ok / limp_total * 100) if limp_total else 100

    glob = round((temp_pct + limp_pct) / 2)

    etiquetas = ['L', 'M', 'M', 'J', 'V', 'S', 'D']
    semana = []
    for i in range(7):
        d = desde + timedelta(days=i)
        lt = [l for l in lecturas if l.registrado_en.date() == d]
        rt = [r for r in registros if r.registrado_en.date() == d]
        semana.append({
            'label': etiquetas[d.weekday()],
            'temp': round(sum(1 for l in lt if not l.fuera_de_rango) / len(lt) * 100) if lt else None,
            'limp': round(sum(1 for r in rt if r.conforme) / len(rt) * 100) if rt else None,
        })

    incidencias = []
    for l in sorted((x for x in lecturas if x.fuera_de_rango), key=lambda x: x.registrado_en, reverse=True)[:6]:
        incidencias.append({'tipo': 'temp', 'titulo': l.camara.nombre if l.camara else 'Cámara',
                            'detalle': f'{l.temperatura}°C · fuera de rango',
                            'accion': l.accion_tomada or l.accion_disposicion,
                            'cuando': l.registrado_en.strftime('%d/%m · %H:%M')})
    for r in sorted((x for x in registros if not x.conforme), key=lambda x: x.registrado_en, reverse=True)[:6]:
        incidencias.append({'tipo': 'limp', 'titulo': r.area.nombre if r.area else 'Área',
                            'detalle': 'Registro no conforme',
                            'accion': r.accion_tomada or r.accion_disposicion,
                            'cuando': r.registrado_en.strftime('%d/%m · %H:%M')})
    incidencias = incidencias[:8]

    return render_template('registros/haccp.html',
                           temp_pct=temp_pct, limp_pct=limp_pct, glob=glob,
                           temp_total=temp_total, limp_total=limp_total,
                           semana=semana, incidencias=incidencias, hoy=hoy)


@app.route('/registros/limpieza/revisar', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'editar')
def limpieza_revisar():
    fi = request.form.get('fecha_inicio') or None
    ff = request.form.get('fecha_fin') or None
    def _d(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date() if s else None
        except ValueError:
            return None
    db.session.add(RevisionLimpieza(
        revisado_por=current_user.id if isinstance(current_user, Vendedor) else None,
        periodo_desde=_d(fi), periodo_hasta=_d(ff),
    ))
    db.session.commit()
    flash('Período marcado como revisado.', 'success')
    return redirect(url_for('limpieza_historial', fecha_inicio=fi or '', fecha_fin=ff or ''))


def _producto_proceso(area):
    """Texto del proceso de limpieza: 'Pooff → Sani-T-10' (paso 1 → paso 2)."""
    if not area:
        return '—'
    pasos = []
    if area.producto:
        pasos.append(area.producto.nombre)
    if area.sanitizante:
        pasos.append(area.sanitizante.nombre)
    return ' → '.join(pasos) if pasos else '—'


def _build_limpieza_pdf(registros, fecha_inicio, fecha_fin, config, revision):
    """Construye el PDF tabular audit-ready del registro de limpieza."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=0.4 * inch, rightMargin=0.4 * inch,
                            topMargin=0.5 * inch, bottomMargin=0.7 * inch)
    empresa_style = ParagraphStyle(name='reg_empresa', fontSize=10, leading=13,
                                   fontName='Helvetica-Bold', alignment=TA_LEFT,
                                   textColor=colors.HexColor('#1877ff'))
    titulo_style = ParagraphStyle(name='reg_titulo', fontSize=15, leading=18,
                                  fontName='Helvetica-Bold', alignment=TA_LEFT)
    sub_style = ParagraphStyle(name='reg_sub', fontSize=9, leading=12,
                               alignment=TA_LEFT, textColor=colors.HexColor('#475569'))

    if fecha_inicio or fecha_fin:
        periodo = f'Período: {fecha_inicio or "inicio"} a {fecha_fin or "hoy"}'
    else:
        periodo = 'Período: todas las fechas'
    generado = datetime.now(DASHBOARD_TIMEZONE).strftime('%Y-%m-%d %H:%M')

    encabezado = [
        Paragraph('Jomar Foods B.V.', empresa_style),
        Paragraph('Registro de limpieza y desinfección', titulo_style),
        Paragraph(f'Documento: {config.codigo_documento} &middot; Versión: {config.version}', sub_style),
        Paragraph(f'{periodo} &middot; Generado: {generado}', sub_style),
    ]
    logo_path = os.path.join(basedir, 'static', 'logo_etiquetas.png')
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=52, height=52)
        head_tbl = Table([[logo, encabezado]], colWidths=[64, None])
        head_tbl.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        head_tbl.hAlign = 'LEFT'
        elements = [head_tbl, Spacer(1, 14)]
    else:
        elements = encabezado + [Spacer(1, 14)]

    def _detalle(r):
        partes = []
        if r.accion_causa: partes.append(f'Causa: {_pdf_xe(r.accion_causa)}')
        if r.accion_tomada: partes.append(f'Acción: {_pdf_xe(r.accion_tomada)}')
        if r.accion_responsable: partes.append(f'Resp.: {_pdf_xe(r.accion_responsable)}')
        if r.accion_disposicion: partes.append(f'Disposición: {_pdf_xe(r.accion_disposicion)}')
        accion = ' · '.join(partes)
        bits = []
        if r.observacion: bits.append(f'<b>Obs.:</b> {_pdf_xe(r.observacion)}')
        if accion: bits.append(f'<b>Acción correctiva</b> — {accion}')
        return '   ·   '.join(bits) if bits else None

    headers = ['Fecha/Hora', 'Área', 'Tipo', 'Producto', 'ppm', 'Resultado', 'Responsable', 'Verificó', 'Método']
    aligns = ['L', 'L', 'L', 'L', 'C', 'C', 'L', 'L', 'L']
    widths = [78, 110, 48, 120, 36, 72, 102, 102, 62]
    filas = []
    for r in registros:
        filas.append({
            'cols': [
                _fmt_local(r.registrado_en),
                r.area.nombre if r.area else '—',
                'Equipo' if (r.area and r.area.tipo == 'equipo') else 'Espacio',
                _producto_proceso(r.area),
                str(r.concentracion_ppm) if r.concentracion_ppm is not None else '—',
                'No conforme' if not r.conforme else 'Conforme',
                r.registrado_por_vendedor.nombre_completo if r.registrado_por_vendedor else '—',
                r.verificado_por_vendedor.nombre_completo if r.verificado_por_vendedor else '—',
                r.metodo_verificacion.capitalize() if r.metodo_verificacion else '—',
            ],
            'desvio': not r.conforme,
            'detalle': _detalle(r),
        })
    elements.append(_registro_pdf_tabla(headers, aligns, widths, 5, filas))

    elements.append(Spacer(1, 18))
    if revision:
        nombre = revision.revisado_por_vendedor.nombre_completo if revision.revisado_por_vendedor else '—'
        rev_txt = f'<b>Verificación:</b> Revisado por {nombre} el {_fmt_local(revision.revisado_en)}'
    else:
        rev_txt = '<b>Verificación:</b> Revisado por: ______________________      Fecha: __________'
    elements.append(Paragraph(rev_txt, sub_style))

    footer_left = (f'Frecuencia: {config.frecuencia_texto or "N/D"}   |   '
                   f'Responsable de verificación: {config.responsable_verificacion or "N/D"}')
    footer_doc = f'{config.codigo_documento} v{config.version}'
    page_w = landscape(A4)[0]

    class _NumberedCanvas(canvas.Canvas):
        def __init__(self, *a, **k):
            canvas.Canvas.__init__(self, *a, **k)
            self._saved_states = []
        def showPage(self):
            self._saved_states.append(dict(self.__dict__))
            self._startPage()
        def save(self):
            total = len(self._saved_states)
            for st in self._saved_states:
                self.__dict__.update(st)
                self.setFont('Helvetica', 7)
                self.setFillColor(colors.HexColor('#475569'))
                self.drawString(0.4 * inch, 0.35 * inch, footer_left)
                self.drawRightString(page_w - 0.4 * inch, 0.35 * inch,
                                     f'{footer_doc}  ·  Página {self._pageNumber} de {total}')
                canvas.Canvas.showPage(self)
            canvas.Canvas.save(self)

    doc.build(elements, canvasmaker=_NumberedCanvas)
    buffer.seek(0)
    return buffer


@app.route('/registros/limpieza/export', methods=['POST'])
@login_required
@requiere_permiso_recurso('registros', 'leer')
def limpieza_export():
    registros = _filtrar_registros_limpieza(request.form)
    fi = request.form.get('fecha_inicio') or ''
    ff = request.form.get('fecha_fin') or ''
    if (request.form.get('formato') or 'pdf').lower() == 'excel':
        headers = ['Fecha', 'Hora', 'Área / tarea', 'Proceso (limpieza → sanitización)', 'ppm', 'Resultado',
                   'Registró', 'Verificó', 'Método verif.', 'Observación',
                   'Causa', 'Acción tomada', 'Responsable acción', 'Disposición']
        rows = [[
            _fmt_local(r.registrado_en, '%Y-%m-%d'), _fmt_local(r.registrado_en, '%H:%M'),
            r.area.nombre if r.area else '', _producto_proceso(r.area),
            r.concentracion_ppm if r.concentracion_ppm is not None else '',
            'Conforme' if r.conforme else 'No conforme',
            r.registrado_por_vendedor.nombre_completo if r.registrado_por_vendedor else '',
            r.verificado_por_vendedor.nombre_completo if r.verificado_por_vendedor else '',
            (r.metodo_verificacion or '').capitalize(),
            r.observacion or '', r.accion_causa or '', r.accion_tomada or '',
            r.accion_responsable or '', r.accion_disposicion or '',
        ] for r in registros]
        buffer = _build_xlsx(headers, rows, 'Limpieza', 'Historial de limpieza')
        filename = f"registro_limpieza_{fi or 'inicio'}_{ff or 'fin'}.xlsx"
        return make_response(send_file(buffer, as_attachment=True, download_name=filename,
                                       mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'))
    config = _get_limpieza_config()
    revision = _revision_limpieza_que_cubre(fi, ff)
    buffer = _build_limpieza_pdf(registros, fi, ff, config, revision)
    filename = f"registro_limpieza_{fi or 'inicio'}_{ff or 'fin'}.pdf"
    response = make_response(send_file(buffer, mimetype='application/pdf',
                                       as_attachment=not _is_ios_request(),
                                       download_name=filename))
    response.headers['Content-Type'] = 'application/pdf'
    return response


@app.route('/registros/limpieza/config', methods=['GET', 'POST'])
@login_required
@requiere_permiso_recurso('registros', 'editar')
def limpieza_config():
    cfg = _get_limpieza_config()
    if request.method == 'POST':
        cfg.codigo_documento = (request.form.get('codigo_documento') or '').strip() or 'FR-HACCP-LIMP-01'
        cfg.version = (request.form.get('version') or '').strip() or '1'
        cfg.frecuencia_texto = (request.form.get('frecuencia_texto') or '').strip() or 'Según programa de limpieza'
        cfg.responsable_verificacion = (request.form.get('responsable_verificacion') or '').strip() or None
        cfg.actualizado_en = datetime.utcnow()
        db.session.commit()
        flash('Configuración guardada.', 'success')
        return redirect(url_for('limpieza_config'))
    return render_template('registros/limpieza_config.html', cfg=cfg)


# Asegura columnas/tablas nuevas de HACCP en cada arranque (idempotente).
# Se ejecuta aquí, después de definir todos los modelos.
with app.app_context():
    _ensure_haccp_columns()
    _seed_catalogo_limpieza()


@app.route('/registros')
@login_required
@requiere_permiso_recurso('registros', 'leer')
def registros_index():
    """Hub HACCP: tarjetas Temperaturas / Limpieza / HACCP con estado de hoy."""
    hoy = date.today()

    camaras = Camara.query.filter_by(activa=True).all()
    cam_total = len(camaras)
    cam_con = len(_camaras_con_lectura_hoy())
    lec_hoy = (LecturaTemperatura.query
               .filter(func.date(LecturaTemperatura.registrado_en) == hoy).all())
    temp_alertas = sum(1 for l in lec_hoy if l.fuera_de_rango)

    areas = AreaLimpieza.query.filter_by(activa=True).all()
    area_total = len(areas)
    area_con = len(_areas_con_registro_hoy())
    reg_hoy = (RegistroLimpieza.query
               .filter(func.date(RegistroLimpieza.registrado_en) == hoy).all())
    limp_noconf = sum(1 for r in reg_hoy if not r.conforme)

    # Cumplimiento global 7 días (reusa la lógica de la vista HACCP).
    desde = hoy - timedelta(days=6)
    lec_semana = (LecturaTemperatura.query
                  .filter(func.date(LecturaTemperatura.registrado_en) >= desde).all())
    reg_semana = (RegistroLimpieza.query
                  .filter(func.date(RegistroLimpieza.registrado_en) >= desde).all())
    tp = round(sum(1 for l in lec_semana if not l.fuera_de_rango) / len(lec_semana) * 100) if lec_semana else 100
    lp = round(sum(1 for r in reg_semana if r.conforme) / len(reg_semana) * 100) if reg_semana else 100
    haccp_global = round((tp + lp) / 2)

    es_admin = (not isinstance(current_user, Vendedor)) or current_user.tiene_permiso('registros', 'editar')
    return render_template('registros/index.html',
                           cam_total=cam_total, cam_con=cam_con, temp_alertas=temp_alertas,
                           area_total=area_total, area_con=area_con, limp_noconf=limp_noconf,
                           haccp_global=haccp_global, es_admin=es_admin, hoy=hoy)


if __name__ == '__main__':
    # Configuración para desarrollo local
    if os.environ.get('FLASK_ENV') == 'development':
        ip_servidor = obtener_ip_servidor()
        app.logger.info(f"La aplicación está disponible en la IP: {ip_servidor}:{5002}")
        app.run(debug=True, host='0.0.0.0', port=5002)
    else:
        # Configuración para producción (Heroku)
        port = int(os.environ.get('PORT', 5000))
        app.run(host='0.0.0.0', port=port, debug=False)
